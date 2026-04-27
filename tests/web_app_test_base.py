from io import BytesIO
import unittest
from unittest import mock
from urllib.parse import quote
from zipfile import ZipFile

import inspect
import httpx
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import app.main as app_main


if "app" not in inspect.signature(httpx.Client.__init__).parameters:
    _original_httpx_client_init = httpx.Client.__init__

    def _compat_httpx_client_init(self, *args, app=None, **kwargs):
        return _original_httpx_client_init(self, *args, **kwargs)

    httpx.Client.__init__ = _compat_httpx_client_init


class WebAppTestCaseBase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(app_main.app)

    def visible_sheetnames(self, workbook):
        return [
            worksheet.title
            for worksheet in workbook.worksheets
            if worksheet.sheet_state == "visible"
        ]

    def find_row_by_value(self, worksheet, value, column=1):
        for row_index in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=row_index, column=column).value == value:
                return row_index
        raise AssertionError(f"value not found in worksheet column {column}: {value}")

    def summary_section_pairs(self, worksheet, stop_label="主要KPI"):
        pairs = {}
        for row_index in range(1, worksheet.max_row + 1):
            label = worksheet.cell(row=row_index, column=1).value
            if label == stop_label:
                break
            if label in (None, "", "項目"):
                continue
            value = worksheet.cell(row=row_index, column=2).value
            if value in (None, ""):
                continue
            pairs[label] = value
        return pairs

    def analyze_uploaded_csv(self, csv_text, analysis_keys=None, extra_data=None):
        response = self.client.post(
            "/api/analyze",
            data={
                "analysis_keys": analysis_keys or ["frequency", "pattern"],
                **(extra_data or {}),
            },
            files={"csv_file": ("custom_log.csv", BytesIO(csv_text.encode("utf-8")), "text/csv")},
        )
        self.assertEqual(200, response.status_code)
        return response.json()["run_id"]

    def build_duckdb_validation_csv(self):
        return "\n".join(
            [
                "case_id,activity,start_time",
                "C001,Start,2024-01-01 09:00:00",
                "C001,Review,2024-01-01 09:02:00",
                "C001,Rework,2024-01-01 09:17:00",
                "C001,Done,2024-01-01 09:20:00",
                "C002,Start,2024-01-01 10:00:00",
                "C002,Review,2024-01-01 10:03:00",
                "C002,Rework,2024-01-01 10:20:00",
                "C002,Review,2024-01-01 10:25:00",
                "C002,Done,2024-01-01 10:39:00",
            ]
        )

    def build_duckdb_validation_csv_with_variant(self):
        return "\n".join(
            [
                "case_id,activity,start_time,variant",
                "C001,Start,2024-01-01 09:00:00,V1",
                "C001,Review,2024-01-01 09:02:00,V1",
                "C001,Done,2024-01-01 09:09:00,V1",
                "C002,Start,2024-01-01 10:00:00,V2",
                "C002,Approve,2024-01-01 10:04:00,V2",
                "C002,Done,2024-01-01 10:12:00,V2",
            ]
        )

    def build_variant_collision_csv(self):
        return "\n".join(
            [
                "case_id,activity,start_time,variant",
                "C001,Start,2024-01-01 09:00:00,V1",
                "C001,Review,2024-01-01 09:05:00,V1",
                "C001,Done,2024-01-01 09:10:00,V1",
                "C002,Start,2024-01-02 09:00:00,V1",
                "C002,Review,2024-01-02 09:05:00,V1",
                "C002,Reminder,2024-01-02 09:08:00,V1",
                "C002,Done,2024-01-02 09:12:00,V1",
                "C003,Start,2024-01-03 09:00:00,V2",
                "C003,Review,2024-01-03 09:04:00,V2",
                "C003,Done,2024-01-03 09:09:00,V2",
            ]
        )
