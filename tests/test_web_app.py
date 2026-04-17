from io import BytesIO
import unittest
from unittest import mock
from urllib.parse import quote
from zipfile import ZipFile

import inspect
import httpx
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import web_app


if "app" not in inspect.signature(httpx.Client.__init__).parameters:
    _original_httpx_client_init = httpx.Client.__init__

    def _compat_httpx_client_init(self, *args, app=None, **kwargs):
        return _original_httpx_client_init(self, *args, **kwargs)

    httpx.Client.__init__ = _compat_httpx_client_init


class WebAppTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.client = TestClient(web_app.app)

    def visible_sheetnames(self, workbook):
        return [
            worksheet.title
            for worksheet in workbook.worksheets
            if worksheet.sheet_state == "visible"
        ]

    def find_row_by_value(self, worksheet, value, column=1):
        for row_index in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=row_index, column=column).value == value:
                return row_index
        raise AssertionError(f"value not found in worksheet column {column}: {value}")

    def summary_section_pairs(self, worksheet, stop_label="主要KPI"):
        pairs = {}
        for row_index in range(1, worksheet.max_row + 1):
            label = worksheet.cell(row=row_index, column=1).value
            if label == stop_label:
                break
            if label in (None, "", "項目"):
                continue
            value = worksheet.cell(row=row_index, column=2).value
            if value in (None, ""):
                continue
            pairs[label] = value
        return pairs

    def analyze_uploaded_csv(self, csv_text, analysis_keys=None, extra_data=None):
        response = self.client.post(
            "/api/analyze",
            data={
                "analysis_keys": analysis_keys or ["frequency", "pattern"],
                **(extra_data or {}),
            },
            files={"csv_file": ("custom_log.csv", BytesIO(csv_text.encode("utf-8")), "text/csv")},
        )
        self.assertEqual(200, response.status_code)
        return response.json()["run_id"]

    def build_duckdb_validation_csv(self):
        return "\n".join(
            [
                "case_id,activity,start_time",
                "C001,Start,2024-01-01 09:00:00",
                "C001,Review,2024-01-01 09:02:00",
                "C001,Rework,2024-01-01 09:17:00",
                "C001,Done,2024-01-01 09:20:00",
                "C002,Start,2024-01-01 10:00:00",
                "C002,Review,2024-01-01 10:03:00",
                "C002,Rework,2024-01-01 10:20:00",
                "C002,Review,2024-01-01 10:25:00",
                "C002,Done,2024-01-01 10:39:00",
            ]
        )

    def build_duckdb_validation_csv_with_variant(self):
        return "\n".join(
            [
                "case_id,activity,start_time,variant",
                "C001,Start,2024-01-01 09:00:00,V1",
                "C001,Review,2024-01-01 09:02:00,V1",
                "C001,Done,2024-01-01 09:09:00,V1",
                "C002,Start,2024-01-01 10:00:00,V2",
                "C002,Approve,2024-01-01 10:04:00,V2",
                "C002,Done,2024-01-01 10:12:00,V2",
            ]
        )

    def build_variant_collision_csv(self):
        return "\n".join(
            [
                "case_id,activity,start_time,variant",
                "C001,Start,2024-01-01 09:00:00,V1",
                "C001,Review,2024-01-01 09:05:00,V1",
                "C001,Done,2024-01-01 09:10:00,V1",
                "C002,Start,2024-01-02 09:00:00,V1",
                "C002,Review,2024-01-02 09:05:00,V1",
                "C002,Reminder,2024-01-02 09:08:00,V1",
                "C002,Done,2024-01-02 09:12:00,V1",
                "C003,Start,2024-01-03 09:00:00,V2",
                "C003,Review,2024-01-03 09:04:00,V2",
                "C003,Done,2024-01-03 09:09:00,V2",
            ]
        )

    def test_detail_script_is_served(self):
        response = self.client.get("/static/detail.js")

        self.assertEqual(200, response.status_code)
        self.assertIn("renderDetailPage", response.text)

    def test_index_page_hides_top_excel_export_ui(self):
        response = self.client.get("/")

        self.assertEqual(200, response.status_code)
        self.assertIn('<select name="case_id_column"', response.text)
        self.assertIn('<select name="activity_column"', response.text)
        self.assertIn('<select name="timestamp_column"', response.text)
        self.assertNotIn('id="export-excel-toggle"', response.text)
        self.assertNotIn("Excel ZIP", response.text)
        self.assertNotIn('id="select-output-directory-button"', response.text)
        self.assertIn("/static/style.css?v=", response.text)
        self.assertIn("/static/app.js?v=", response.text)

    def test_analysis_detail_page_uses_versioned_static_assets(self):
        response = self.client.get("/analysis/pattern")

        self.assertEqual(200, response.status_code)
        self.assertIn("/static/style.css?v=", response.text)
        self.assertIn("/static/detail.js?v=", response.text)

    def test_analysis_detail_api_returns_full_rows(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency", "transition", "pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        detail_response = self.client.get(f"/api/runs/{run_id}/analyses/frequency")
        self.assertEqual(200, detail_response.status_code)

        payload = detail_response.json()
        frequency_analysis = payload["analyses"]["frequency"]
        dashboard = payload["dashboard"]
        impact = payload["impact"]
        insights = payload["insights"]

        self.assertGreater(frequency_analysis["row_count"], 0)
        self.assertEqual(frequency_analysis["row_count"], len(frequency_analysis["rows"]))
        self.assertTrue(dashboard["has_data"])
        self.assertGreater(dashboard["total_cases"], 0)
        self.assertGreater(dashboard["total_records"], 0)
        self.assertGreater(dashboard["activity_type_count"], 0)
        self.assertIn("top10_variant_coverage_pct", dashboard)
        self.assertIn("top_bottleneck_transition_label", dashboard)
        self.assertTrue(impact["has_data"])
        self.assertTrue(impact["rows"])
        self.assertIn("impact_score", impact["rows"][0])
        self.assertEqual("rule_based", insights["mode"])
        self.assertTrue(insights["has_data"])
        self.assertGreaterEqual(len(insights["items"]), 3)
        self.assertTrue(any(item["id"] == "top_activity" for item in insights["items"]))

        pattern_detail_response = self.client.get(f"/api/runs/{run_id}/analyses/pattern")
        self.assertEqual(200, pattern_detail_response.status_code)
        pattern_payload = pattern_detail_response.json()
        pattern_insights = pattern_payload["insights"]
        self.assertTrue(any(item["id"] == "top_pattern" for item in pattern_insights["items"]))

    def test_analysis_detail_api_supports_row_limit(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency", "transition", "pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        detail_response = self.client.get(f"/api/runs/{run_id}/analyses/pattern?row_limit=2")
        self.assertEqual(200, detail_response.status_code)

        payload = detail_response.json()
        pattern_analysis = payload["analyses"]["pattern"]

        self.assertEqual(2, pattern_analysis["returned_row_count"])
        self.assertEqual(0, pattern_analysis["row_offset"])
        self.assertGreaterEqual(pattern_analysis["row_count"], pattern_analysis["returned_row_count"])
        self.assertEqual(2, len(pattern_analysis["rows"]))

    def test_analysis_detail_api_supports_row_offset_pagination_metadata(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        detail_response = self.client.get(
            f"/api/runs/{run_id}/analyses/pattern?row_limit=1&row_offset=1"
        )
        self.assertEqual(200, detail_response.status_code)

        payload = detail_response.json()
        pattern_analysis = payload["analyses"]["pattern"]

        self.assertEqual(1, pattern_analysis["returned_row_count"])
        self.assertEqual(1, pattern_analysis["row_offset"])
        self.assertEqual(2, pattern_analysis["page_start_row_number"])
        self.assertEqual(2, pattern_analysis["page_end_row_number"])
        self.assertTrue(pattern_analysis["has_previous_page"])
        self.assertTrue(pattern_analysis["has_next_page"])
        self.assertEqual(0, pattern_analysis["previous_row_offset"])
        self.assertEqual(2, pattern_analysis["next_row_offset"])

    def test_analysis_detail_api_supports_deferred_supplement_payload(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        detail_response = self.client.get(
            f"/api/runs/{run_id}/analyses/frequency"
            "?include_dashboard=false"
            "&include_impact=false"
            "&include_root_cause=false"
            "&include_insights=false"
        )
        self.assertEqual(200, detail_response.status_code)

        payload = detail_response.json()
        self.assertIsNone(payload["dashboard"])
        self.assertIsNone(payload["impact"])
        self.assertIsNone(payload["root_cause"])
        self.assertIsNone(payload["insights"])
        self.assertCountEqual(
            ["dashboard", "impact", "root_cause", "insights"],
            payload["deferred_sections"],
        )

    def test_analysis_detail_api_deferred_payload_skips_unfiltered_dataframe_copy(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        with mock.patch("web_app.filter_prepared_df", side_effect=AssertionError("filter_prepared_df should not run")):
            detail_response = self.client.get(
                f"/api/runs/{run_id}/analyses/frequency"
                "?include_dashboard=false"
                "&include_impact=false"
                "&include_root_cause=false"
                "&include_insights=false"
            )

        self.assertEqual(200, detail_response.status_code)

    def test_excel_archive_api_returns_zip_binary(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={
                "analysis_keys": ["frequency", "transition", "pattern"],
                "export_excel": "on",
            },
        )
        self.assertEqual(200, analyze_response.status_code)

        payload = analyze_response.json()
        self.assertIsNone(payload["analyses"]["frequency"]["excel_file"])

        run_id = payload["run_id"]
        excel_response = self.client.get(f"/api/runs/{run_id}/excel-archive")

        self.assertEqual(200, excel_response.status_code)
        self.assertEqual(
            "application/zip",
            excel_response.headers["content-type"],
        )
        self.assertIn("attachment;", excel_response.headers["content-disposition"])
        self.assertIn("filename*=UTF-8''", excel_response.headers["content-disposition"])
        self.assertIn("%E5%85%A8%E5%88%86%E6%9E%90%E3%83%AC%E3%83%9D%E3%83%BC%E3%83%88.zip", excel_response.headers["content-disposition"])
        self.assertTrue(excel_response.content.startswith(b"PK"))

        with ZipFile(BytesIO(excel_response.content)) as archive_file:
            archive_names = set(archive_file.namelist())
            self.assertEqual(4, len(archive_names))
            self.assertTrue(all(name.endswith(".xlsx") for name in archive_names))
            self.assertIn("ログ診断.xlsx", archive_names)
            self.assertIn("頻度分析.xlsx", archive_names)
            self.assertIn("前後処理分析.xlsx", archive_names)
            self.assertIn("処理順パターン分析.xlsx", archive_names)

    def test_excel_archive_api_includes_error_log_when_some_analyses_are_missing(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["frequency"],
        )

        excel_response = self.client.get(f"/api/runs/{run_id}/excel-archive")

        self.assertEqual(200, excel_response.status_code)
        with ZipFile(BytesIO(excel_response.content)) as archive_file:
            archive_names = set(archive_file.namelist())
            self.assertIn("ログ診断.xlsx", archive_names)
            self.assertIn("頻度分析.xlsx", archive_names)
            self.assertIn("エラーログ.txt", archive_names)
            error_text = archive_file.read("エラーログ.txt").decode("utf-8")
            self.assertIn("前後処理分析", error_text)
            self.assertIn("処理順パターン分析", error_text)


    def test_report_excel_export_api_returns_workbook(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-03 09:00:00,Sales,API,A",
                    "C002,Approve,2024-01-04 09:00:00,Sales,API,A",
                    "C003,Submit,2024-01-01 10:00:00,HR,Mail,B",
                    "C003,Reject,2024-01-02 10:00:00,HR,Mail,B",
                ]
            ),
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        with mock.patch(
            "web_app.build_excel_ai_summary",
            return_value={
                "title": "分析コメント",
                "mode": "ollama",
                "provider": "Ollama (qwen2.5:7b)",
                "generated_at": "2026-04-02T00:00:00+00:00",
                "period": "2024-01-01 09:00 〜 2024-01-04 09:00",
                "text": "【全体傾向】\nテスト用の分析コメントです。",
                "highlights": [
                    "Submit が中心のルートです。",
                    "Approve の前後を重点確認してください。",
                ],
                "recommended_actions": [
                    "Submit の前後処理を確認し、遷移パターンの改善ポイントを特定してください。",
                    "Approve の処理時間をケース別にドリルダウンで確認してください。",
                ],
                "note": "ローカルLLMで生成した解説を掲載しています。",
            },
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel"
                "?analysis_key=pattern"
                "&pattern_display_limit=10"
                "&filter_value_1=Sales"
                "&selected_transition_key=Submit__TO__Approve"
                "&case_id=C001"
            )

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn(
            quote("custom_log_処理順パターン分析レポート.xlsx"),
            response.headers["content-disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            ["サマリー", "分析コメント", "結論サマリー", "サマリーダッシュボード", "パターンサマリー", "処理順パターン分析", "パターン01詳細", "ドリルダウン", "ケース追跡"],
            self.visible_sheetnames(workbook),
        )
        summary_sheet = workbook["\u30b5\u30de\u30ea\u30fc"]
        self.assertEqual("\u30b5\u30de\u30ea\u30fc", summary_sheet["A1"].value)
        summary_pairs = self.summary_section_pairs(summary_sheet)
        self.assertNotIn("実行ID", summary_pairs)
        self.assertEqual("pattern", summary_pairs["分析種別"])
        self.assertEqual("処理順パターン分析", summary_pairs["分析名"])
        self.assertIn("Sales", str(summary_pairs["適用条件"]))
        self.assertEqual("2024-01-01 09:00 〜 2024-01-04 09:00", summary_pairs["分析期間"])
        self.assertEqual("group_a、group_b、group_c", summary_pairs["グルーピング条件"])
        self.assertEqual("Submit → Approve", summary_pairs["選択中遷移"])
        self.assertEqual("C001", summary_pairs["選択中ケースID"])
        applied_filters_row = self.find_row_by_value(summary_sheet, "適用条件")
        self.assertEqual("", summary_sheet.cell(row=applied_filters_row + 1, column=1).value or "")
        filter_note = str(summary_sheet.cell(row=applied_filters_row + 1, column=2).value)
        self.assertIn("※ 適用条件の種類:", filter_note)
        self.assertIn("期間フィルター: 開始日 / 終了日", filter_note)
        self.assertIn("グループ/カテゴリーフィルター①②③", filter_note)
        self.assertIn("アクティビティフィルター: 特定アクティビティを含む/除外", filter_note)
        self.assertEqual("分析期間", summary_sheet.cell(row=applied_filters_row + 2, column=1).value)
        grouping_row = self.find_row_by_value(summary_sheet, "グルーピング条件")
        self.assertEqual("", summary_sheet.cell(row=grouping_row + 1, column=1).value or "")
        grouping_note = str(summary_sheet.cell(row=grouping_row + 1, column=2).value)
        self.assertIn("カラムを選択し値を未選択にすると", grouping_note)
        self.assertIn("グルーピング軸（比較用）", grouping_note)
        summary_values = [summary_sheet.cell(row=row_index, column=1).value for row_index in range(1, summary_sheet.max_row + 1)]
        self.assertIn("主要KPI", summary_values)
        self.assertIn("分析ハイライト", summary_values)
        self.assertIn("グループ別比較", summary_values)

        group_table_row = self.find_row_by_value(summary_sheet, "グループ別比較")
        group_headers = [
            summary_sheet.cell(row=group_table_row + 2, column=column_index).value
            for column_index in range(1, 11)
        ]
        self.assertEqual(
            [
                "グルーピング軸",
                "値",
                "ケース数",
                "ケース比率(%)",
                "イベント数",
                "イベント比率(%)",
                "平均処理時間(分)",
                "中央値処理時間(分)",
                "最大処理時間(分)",
                "合計処理時間(分)",
            ],
            group_headers,
        )
        self.assertEqual("全体", summary_sheet.cell(row=group_table_row + 3, column=2).value)

        ai_sheet = workbook["分析コメント"]
        self.assertEqual("分析コメント", ai_sheet["A1"].value)
        self.assertEqual("対象分析", ai_sheet["A4"].value)
        self.assertEqual("\u51e6\u7406\u9806\u30d1\u30bf\u30fc\u30f3\u5206\u6790", ai_sheet["B4"].value)
        analysis_premise_row = self.find_row_by_value(ai_sheet, "分析前提")
        explanation_row = self.find_row_by_value(ai_sheet, "解説本文")
        recommended_actions_row = self.find_row_by_value(ai_sheet, "推奨アクション")
        terminology_row = self.find_row_by_value(ai_sheet, "用語説明")
        self.assertLess(analysis_premise_row, explanation_row)
        self.assertLess(explanation_row, recommended_actions_row)
        self.assertLess(recommended_actions_row, terminology_row)
        self.assertIn("処理時間は、同一ケース内で", str(ai_sheet.cell(row=analysis_premise_row + 1, column=1).value))
        self.assertIn("テスト用の分析コメント", str(ai_sheet.cell(row=explanation_row + 2, column=1).value))
        self.assertTrue(str(ai_sheet.cell(row=recommended_actions_row, column=1).fill.fgColor.rgb).endswith("E8EDF2"))
        self.assertIn(
            "Submit",
            str(ai_sheet.cell(row=recommended_actions_row + 1, column=2).value or ""),
        )
        self.assertEqual("用語", ai_sheet.cell(row=terminology_row + 1, column=1).value)
        self.assertEqual("説明", ai_sheet.cell(row=terminology_row + 1, column=2).value)
        self.assertEqual("ケース", ai_sheet.cell(row=terminology_row + 2, column=1).value)
        self.assertTrue(str(ai_sheet.cell(row=analysis_premise_row, column=1).fill.fgColor.rgb).endswith("E8EDF2"))
        self.assertTrue(str(ai_sheet.cell(row=terminology_row, column=1).fill.fgColor.rgb).endswith("F0F0F0"))
        ai_sheet_values = [
            ai_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, ai_sheet.max_row + 1)
        ]
        self.assertNotIn("補足・免責事項", ai_sheet_values)
        self.assertTrue(
            str(ai_sheet.cell(row=terminology_row + 2, column=1).fill.fgColor.rgb).endswith("F7F7F7")
        )
        self.assertTrue(
            str(ai_sheet.cell(row=terminology_row + 2, column=1).border.left.color.rgb).endswith("D0D0D0")
        )
        self.assertNotIn("\u983b\u5ea6\u5206\u6790", workbook.sheetnames)
        self.assertNotIn("\u30dc\u30c8\u30eb\u30cd\u30c3\u30af\u5206\u6790", workbook.sheetnames)
        self.assertNotIn("\u6539\u5584\u30a4\u30f3\u30d1\u30af\u30c8\u5206\u6790", workbook.sheetnames)

        conclusion_sheet = workbook["\u7d50\u8ad6\u30b5\u30de\u30ea\u30fc"]
        self.assertEqual("結論サマリー", conclusion_sheet["A1"].value)
        self.assertEqual("全体要約", conclusion_sheet["A4"].value)
        conclusion_values = [
            conclusion_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, conclusion_sheet.max_row + 1)
        ]
        self.assertIn("問題点3つ", conclusion_values)
        self.assertGreaterEqual(len(conclusion_sheet._charts), 2)

        dashboard_sheet = workbook["\u30b5\u30de\u30ea\u30fc\u30c0\u30c3\u30b7\u30e5\u30dc\u30fc\u30c9"]
        self.assertEqual("サマリーダッシュボード", dashboard_sheet["A1"].value)
        dashboard_values = [
            dashboard_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, dashboard_sheet.max_row + 1)
        ]
        self.assertIn("改善優先TOP3", dashboard_values)
        self.assertIn("問題点", dashboard_values)
        self.assertEqual(0, len(dashboard_sheet._charts))
        with ZipFile(BytesIO(response.content)) as archive:
            chart_xml_documents = {
                name: archive.read(name).decode("utf-8")
                for name in archive.namelist()
                if name.startswith("xl/charts/chart")
            }
        self.assertGreaterEqual(len(chart_xml_documents), 2)
        self.assertIn("<cat><strRef>", chart_xml_documents["xl/charts/chart1.xml"])
        self.assertIn("<cat><strRef>", chart_xml_documents["xl/charts/chart2.xml"])
        self.assertNotIn("<cat><numRef>", chart_xml_documents["xl/charts/chart1.xml"])
        self.assertNotIn("<cat><numRef>", chart_xml_documents["xl/charts/chart2.xml"])
        self.assertIn("showPercent val=\"1\"", chart_xml_documents["xl/charts/chart1.xml"])
        self.assertIn("平均処理時間の比較（長い順）", chart_xml_documents["xl/charts/chart2.xml"])
        self.assertIn("showVal val=\"1\"", chart_xml_documents["xl/charts/chart2.xml"])
        self.assertIn("<orientation val=\"maxMin\"/>", chart_xml_documents["xl/charts/chart2.xml"])
        self.assertIn("<crosses val=\"autoZero\"/>", chart_xml_documents["xl/charts/chart2.xml"])
        conclusion_chart_anchors = [
            (chart.anchor._from.col, chart.anchor._from.row)
            for chart in conclusion_sheet._charts
        ]
        self.assertEqual(2, len(conclusion_chart_anchors))
        self.assertEqual(conclusion_chart_anchors[0][1], conclusion_chart_anchors[1][1])
        self.assertGreater(conclusion_chart_anchors[1][0], conclusion_chart_anchors[0][0])

        pattern_summary_sheet = workbook["\u30d1\u30bf\u30fc\u30f3\u30b5\u30de\u30ea\u30fc"]
        self.assertEqual("\u30d1\u30bf\u30fc\u30f3\u30b5\u30de\u30ea\u30fc", pattern_summary_sheet["A1"].value)
        self.assertEqual("上位3パターン累積カバー率(%)", pattern_summary_sheet["A4"].value)
        pattern_summary_values = [
            pattern_summary_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, pattern_summary_sheet.max_row + 1)
        ]
        self.assertIn("上位10パターン", pattern_summary_values)
        self.assertIn("カバー率要約", pattern_summary_values)
        self.assertIn("最短処理パターン", pattern_summary_values)
        self.assertNotIn("上位3パターン", pattern_summary_values)
        self.assertNotIn("パターン比較（上位10件）", pattern_summary_values)
        self.assertIn("要確認パターン一覧", pattern_summary_values)
        self.assertIn("改善対象パターンTOP3", pattern_summary_values)

        pattern_sheet = workbook["\u51e6\u7406\u9806\u30d1\u30bf\u30fc\u30f3\u5206\u6790"]
        self.assertEqual("\u51e6\u7406\u9806\u30d1\u30bf\u30fc\u30f3\u5206\u6790", pattern_sheet["A1"].value)
        pattern_headers = [pattern_sheet.cell(row=3, column=column_index).value for column_index in range(1, pattern_sheet.max_column + 1)]
        self.assertIn("\u9806\u4f4d", pattern_headers)
        self.assertIn("パターン / バリアント", pattern_headers)
        self.assertIn("繰り返し", pattern_headers)
        self.assertIn("繰り返し回数", pattern_headers)
        self.assertIn("繰り返し率(%)", pattern_headers)
        self.assertIn("繰り返し率区分", pattern_headers)
        self.assertIn("確認区分", pattern_headers)
        self.assertIn("平均処理時間差分(分)", pattern_headers)
        self.assertIn("改善優先度スコア", pattern_headers)
        self.assertIn("全体影響度(%)", pattern_headers)
        self.assertIn("最短処理", pattern_headers)
        self.assertIn("標準偏差ケース処理時間(分)", pattern_headers)
        self.assertIn("75%点ケース処理時間(分)", pattern_headers)
        self.assertIn("90%点ケース処理時間(分)", pattern_headers)
        self.assertIn("95%点ケース処理時間(分)", pattern_headers)
        self.assertIn("簡易コメント", pattern_headers)
        self.assertIn("ステップ数", pattern_headers)
        self.assertIn("繰り返しアクティビティ", pattern_headers)
        self.assertIn("パターン", pattern_headers)
        comment_column_index = pattern_headers.index("簡易コメント") + 1
        comment_column_letter = get_column_letter(comment_column_index)
        self.assertFalse(pattern_sheet[f"{comment_column_letter}4"].alignment.wrap_text)
        self.assertGreaterEqual(pattern_sheet.column_dimensions[comment_column_letter].width, 48)

        pattern_detail_sheet = workbook["\u30d1\u30bf\u30fc\u30f301\u8a73\u7d30"]
        self.assertEqual("Pattern #1 詳細", pattern_detail_sheet["A1"].value)
        self.assertEqual("項目", pattern_detail_sheet["A3"].value)
        pattern_detail_values = [cell for row in pattern_detail_sheet.iter_rows(values_only=True) for cell in row if cell]
        self.assertIn("ステップ別所要時間", pattern_detail_values)

        drilldown_sheet = workbook["\u30c9\u30ea\u30eb\u30c0\u30a6\u30f3"]
        self.assertEqual("\u9077\u79fb\u30c9\u30ea\u30eb\u30c0\u30a6\u30f3: Submit \u2192 Approve", drilldown_sheet["A1"].value)
        self.assertEqual("\u30b1\u30fc\u30b9ID", drilldown_sheet["A3"].value)
        self.assertEqual("C001", drilldown_sheet["A4"].value)

        case_trace_sheet = workbook["\u30b1\u30fc\u30b9\u8ffd\u8de1"]
        self.assertEqual("\u30b1\u30fc\u30b9\u6982\u8981", case_trace_sheet["A1"].value)
        self.assertEqual("\u901a\u904e\u30a4\u30d9\u30f3\u30c8", case_trace_sheet["A10"].value)

    def test_report_excel_group_comparison_in_ai_text(self):
        """???????????AI????????????????????????????????"""
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,department",
                    "C001,Submit,2024-01-01 09:00:00,Sales",
                    "C001,Approve,2024-01-01 10:00:00,Sales",
                    "C002,Submit,2024-01-02 09:00:00,Sales",
                    "C002,Approve,2024-01-02 11:00:00,Sales",
                    "C003,Submit,2024-01-03 09:00:00,HR",
                    "C003,Approve,2024-01-03 09:30:00,HR",
                    "C004,Submit,2024-01-04 09:00:00,HR",
                    "C004,Approve,2024-01-04 09:45:00,HR",
                ]
            ),
            extra_data={
                "filter_column_1": "department",
            },
        )

        response = self.client.get(
            f"/api/runs/{run_id}/report-excel"
            "?analysis_key=frequency"
        )

        self.assertEqual(200, response.status_code)

        workbook = load_workbook(BytesIO(response.content))
        ai_sheet = workbook["\u5206\u6790\u30b3\u30e1\u30f3\u30c8"]

        all_values = []
        for row_index in range(1, ai_sheet.max_row + 1):
            cell_value = ai_sheet.cell(row=row_index, column=1).value
            if cell_value:
                all_values.append(str(cell_value))

        full_text = "\n".join(all_values)

        self.assertIn("\u30b0\u30eb\u30fc\u30d7\u9593\u6bd4\u8f03", full_text)
        self.assertTrue(
            "Sales" in full_text or "HR" in full_text,
            f"Group names not found in AI text: {full_text[:500]}",
        )

    def test_report_excel_export_api_falls_back_when_ollama_is_unavailable(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-02 09:00:00",
                    "C002,Submit,2024-01-03 09:00:00",
                    "C002,Approve,2024-01-04 09:00:00",
                ]
            )
        )

        with mock.patch(
            "web_app.request_ollama_insights_text",
            side_effect=web_app.httpx.ConnectError("offline"),
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=frequency"
            )

        self.assertEqual(200, response.status_code)
        self.assertIn(
            quote("custom_log_頻度分析レポート.xlsx"),
            response.headers["content-disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook["サマリー"]
        summary_pairs = self.summary_section_pairs(summary_sheet)
        summary_values = [
            summary_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, summary_sheet.max_row + 1)
        ]
        self.assertNotIn("実行ID", summary_pairs)
        self.assertNotIn("選択中アクティビティ", summary_pairs)
        self.assertNotIn("選択中遷移", summary_pairs)
        self.assertNotIn("選択中ケースID", summary_pairs)
        self.assertEqual("なし", summary_pairs["グルーピング条件"])
        applied_filters_row = self.find_row_by_value(summary_sheet, "適用条件")
        self.assertEqual("", summary_sheet.cell(row=applied_filters_row + 1, column=1).value or "")
        filter_note = str(summary_sheet.cell(row=applied_filters_row + 1, column=2).value)
        self.assertIn("※ 適用条件の種類:", filter_note)
        self.assertIn("期間フィルター: 開始日 / 終了日", filter_note)
        self.assertIn("グループ/カテゴリーフィルター①②③", filter_note)
        self.assertIn("アクティビティフィルター: 特定アクティビティを含む/除外", filter_note)
        grouping_row = self.find_row_by_value(summary_sheet, "グルーピング条件")
        self.assertEqual("", summary_sheet.cell(row=grouping_row + 1, column=1).value or "")
        grouping_note = str(summary_sheet.cell(row=grouping_row + 1, column=2).value)
        self.assertIn("カラムを選択し値を未選択にすると", grouping_note)
        self.assertIn("グルーピング軸（比較用）", grouping_note)
        self.assertEqual("分析ハイライト", summary_values[summary_values.index("分析ハイライト")])
        self.assertIn("最大ケース処理時間", summary_values)
        self.assertIn("バリアント総数", summary_values)
        self.assertIn("ユニークアクティビティ数", summary_values)
        self.assertIn("平均ケースあたりイベント数", summary_values)
        ai_sheet = workbook["分析コメント"]
        self.assertEqual("分析期間", ai_sheet["A5"].value)
        explanation_row = self.find_row_by_value(ai_sheet, "解説本文")
        self.assertTrue(str(ai_sheet.cell(row=explanation_row + 1, column=1).value).strip())
        ai_sheet_values = [
            ai_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, ai_sheet.max_row + 1)
        ]
        self.assertNotIn("補足・免責事項", ai_sheet_values)
        frequency_sheet = workbook["頻度分析"]
        frequency_headers = [
            frequency_sheet.cell(row=3, column=column_index).value
            for column_index in range(1, frequency_sheet.max_column + 1)
        ]
        for header in ("標準偏差(分)", "75%点(分)", "90%点(分)", "95%点(分)"):
            self.assertIn(header, frequency_headers)

    def test_analyze_api_applies_start_end_activity_filters_to_preview_and_detail(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-01 09:05:00",
                    "C001,Done,2024-01-01 09:10:00",
                    "C002,Intake,2024-01-01 10:00:00",
                    "C002,Approve,2024-01-01 10:05:00",
                    "C002,Done,2024-01-01 10:10:00",
                    "C003,Submit,2024-01-01 11:00:00",
                    "C003,Check,2024-01-01 11:05:00",
                    "C003,Reject,2024-01-01 11:10:00",
                ]
            ),
            analysis_keys=["frequency"],
            extra_data={
                "start_activity_values": "Submit",
                "end_activity_values": "Done",
            },
        )

        detail_response = self.client.get(f"/api/runs/{run_id}/analyses/frequency")
        self.assertEqual(200, detail_response.status_code)

        payload = detail_response.json()
        self.assertEqual(1, payload["case_count"])
        self.assertEqual(3, payload["event_count"])
        self.assertEqual("Submit", payload["applied_filters"]["start_activity_values"])
        self.assertEqual("Done", payload["applied_filters"]["end_activity_values"])

    def test_filter_options_api_returns_all_activity_names_for_start_end_filters(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-01 09:05:00",
                    "C001,Done,2024-01-01 09:10:00",
                    "C002,Intake,2024-01-01 10:00:00",
                    "C002,Done,2024-01-01 10:10:00",
                ]
            ),
            analysis_keys=["frequency"],
        )

        response = self.client.get(f"/api/runs/{run_id}/filter-options")
        self.assertEqual(200, response.status_code)

        payload = response.json()
        self.assertEqual(
            ["Approve", "Done", "Intake", "Submit"],
            payload["options"]["all_activity_names"],
        )

    def test_frequency_ai_fallback_uses_actual_activity_name_and_event_count(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-01 10:00:00",
                    "C002,Submit,2024-01-02 09:00:00",
                    "C002,Approve,2024-01-02 10:00:00",
                    "C003,Submit,2024-01-03 09:00:00",
                    "C003,Reject,2024-01-03 10:00:00",
                ]
            ),
            analysis_keys=["frequency"],
        )

        with mock.patch(
            "web_app.request_ollama_insights_text",
            side_effect=web_app.httpx.ConnectError("offline"),
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=frequency"
            )

        self.assertEqual(200, response.status_code)

        workbook = load_workbook(BytesIO(response.content))
        ai_sheet = workbook["分析コメント"]
        explanation_row = self.find_row_by_value(ai_sheet, "解説本文")
        explanation_text = "\n".join(
            str(ai_sheet.cell(row=row_index, column=1).value or "")
            for row_index in range(explanation_row + 1, ai_sheet.max_row + 1)
        )

        self.assertIn("Submit", explanation_text)
        self.assertIn("3 件", explanation_text)
        self.assertNotIn("「不明」", explanation_text)
        self.assertNotIn("0 件", explanation_text)

    def test_report_excel_export_api_returns_transition_workbook(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-02 09:00:00",
                    "C002,Submit,2024-01-03 09:00:00",
                    "C002,Reject,2024-01-04 09:00:00",
                ]
            ),
            analysis_keys=["transition"],
        )

        with mock.patch(
            "web_app.build_excel_ai_summary",
            return_value={
                "title": "分析コメント",
                "mode": "ollama",
                "provider": "Ollama (qwen2.5:7b)",
                "generated_at": "2026-04-02T00:00:00+00:00",
                "period": "2024-01-01 09:00 〜 2024-01-04 09:00",
                "text": "前後処理向けの分析コメントです。",
                "highlights": ["遷移の詰まりを確認してください。"],
                "note": "ローカルLLMで生成した解説を掲載しています。",
            },
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel"
                "?analysis_key=transition"
                "&selected_transition_key=Submit__TO__Approve"
            )

        self.assertEqual(200, response.status_code)
        self.assertIn(
            quote("custom_log_前後処理分析レポート.xlsx"),
            response.headers["content-disposition"],
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            ["サマリー", "分析コメント", "前後処理分析", "ボトルネック分析", "改善インパクト分析", "ドリルダウン"],
            self.visible_sheetnames(workbook),
        )
        summary_sheet = workbook["サマリー"]
        summary_pairs = self.summary_section_pairs(summary_sheet)
        self.assertNotIn("実行ID", summary_pairs)
        self.assertEqual("Submit → Approve", summary_pairs["選択中遷移"])
        self.assertEqual("未選択", summary_pairs["選択中アクティビティ"])
        self.assertEqual("未選択", summary_pairs["選択中ケースID"])
        self.assertNotIn("\u983b\u5ea6\u5206\u6790", workbook.sheetnames)
        self.assertNotIn("\u51e6\u7406\u9806\u30d1\u30bf\u30fc\u30f3\u5206\u6790", workbook.sheetnames)
        self.assertNotIn("\u30d0\u30ea\u30a2\u30f3\u30c8\u5206\u6790", workbook.sheetnames)
        transition_sheet = workbook["\u524d\u5f8c\u51e6\u7406\u5206\u6790"]
        self.assertEqual("\u524d\u5f8c\u51e6\u7406\u5206\u6790", transition_sheet["A1"].value)
        self.assertEqual("\u9806\u4f4d", transition_sheet["A3"].value)
        transition_headers = [
            transition_sheet.cell(row=3, column=column_index).value
            for column_index in range(1, transition_sheet.max_column + 1)
        ]
        for header in ("標準偏差(分)", "75%点(分)", "90%点(分)", "95%点(分)"):
            self.assertIn(header, transition_headers)

        bottleneck_sheet = workbook["\u30dc\u30c8\u30eb\u30cd\u30c3\u30af\u5206\u6790"]
        self.assertEqual("アクティビティボトルネック", bottleneck_sheet["A1"].value)
        self.assertEqual("平均所要時間", bottleneck_sheet["E3"].value)
        transition_title_row = next(
            row_index
            for row_index in range(1, bottleneck_sheet.max_row + 1)
            if bottleneck_sheet.cell(row=row_index, column=1).value == "遷移ボトルネック"
        )
        self.assertEqual("遷移ボトルネック", bottleneck_sheet.cell(row=transition_title_row, column=1).value)
        self.assertEqual("平均所要時間", bottleneck_sheet.cell(row=transition_title_row + 2, column=5).value)

        activity_data_rows = [
            row
            for row in bottleneck_sheet.iter_rows(min_row=4, max_row=transition_title_row - 1, values_only=True)
            if row[0] not in (None, "")
        ]
        self.assertTrue(activity_data_rows)
        self.assertTrue(all(row[0] not in (None, "") for row in activity_data_rows))
        self.assertTrue(all(row[1] not in (None, "") for row in activity_data_rows))
        self.assertTrue(all(row[4] not in (None, "") for row in activity_data_rows))
        self.assertTrue(all(row[5] not in (None, "") for row in activity_data_rows))
        self.assertTrue(all(row[6] not in (None, "") for row in activity_data_rows))

        transition_data_rows = [
            row
            for row in bottleneck_sheet.iter_rows(min_row=transition_title_row + 3, values_only=True)
            if row[0] not in (None, "")
        ]
        self.assertTrue(transition_data_rows)
        self.assertTrue(all(row[0] not in (None, "") for row in transition_data_rows))
        self.assertTrue(all(row[1] not in (None, "") for row in transition_data_rows))
        self.assertTrue(all(row[4] not in (None, "") for row in transition_data_rows))
        self.assertTrue(all(row[5] not in (None, "") for row in transition_data_rows))
        self.assertTrue(all(row[6] not in (None, "") for row in transition_data_rows))

    def test_query_group_summary_matches_build_group_summary_for_parquet(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web",
                    "C002,Submit,2024-01-03 09:00:00,Sales,API",
                    "C002,Approve,2024-01-04 09:00:00,Sales,API",
                    "C003,Submit,2024-01-05 09:00:00,HR,Mail",
                    "C003,Reject,2024-01-06 09:00:00,HR,Mail",
                ]
            ),
            analysis_keys=["frequency"],
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
            },
        )

        run_data = web_app.get_run_data(run_id)
        expected = run_data["result"]["group_summary"]
        actual = web_app.query_group_summary(
            run_data["prepared_parquet_path"],
            ["group_a", "group_b"],
        )
        self.assertEqual(expected, actual)

    def test_report_excel_export_api_groups_frequency_sheet_sections_in_memory_mode(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web",
                    "C002,Submit,2024-01-03 09:00:00,Sales,API",
                    "C002,Approve,2024-01-04 09:00:00,Sales,API",
                    "C003,Submit,2024-01-05 09:00:00,Sales,Web",
                    "C003,Approve,2024-01-06 09:00:00,Sales,Web",
                    "C004,Submit,2024-01-07 09:00:00,HR,Mail",
                    "C004,Reject,2024-01-08 09:00:00,HR,Mail",
                    "C005,Submit,2024-01-09 09:00:00,HR,Mail",
                    "C005,Approve,2024-01-10 09:00:00,HR,Mail",
                    "C006,Submit,2024-01-11 09:00:00,,Tokyo",
                    "C006,Approve,2024-01-12 09:00:00,,Tokyo",
                ]
            ),
            analysis_keys=["frequency"],
            extra_data={
                "filter_column_1": "group_a",
            },
        )

        with mock.patch(
            "web_app.build_excel_ai_summary",
            return_value={
                "title": "分析コメント",
                "mode": "ollama",
                "provider": "Ollama (qwen2.5:7b)",
                "generated_at": "2026-04-09T00:00:00+00:00",
                "period": "2024-01-01 09:00 〜 2024-01-12 09:00",
                "text": "既存集計からの要約です。",
                "highlights": ["頻度の高い活動を確認してください。"],
                "note": "ローカルLLMで生成した解説を掲載しています。",
            },
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=frequency"
            )

        self.assertEqual(200, response.status_code)
        workbook = load_workbook(BytesIO(response.content))
        frequency_sheet = workbook["頻度分析"]
        self.assertEqual("═══ 全体 ═══", frequency_sheet["A1"].value)
        self.assertEqual(True, frequency_sheet["A1"].font.bold)
        self.assertEqual(12, frequency_sheet["A1"].font.size)
        self.assertTrue(str(frequency_sheet["A1"].fill.fgColor.rgb).endswith("D9E1F2"))
        self.assertEqual("順位", frequency_sheet["A2"].value)
        group_header_rows = [
            row_index
            for row_index in range(1, frequency_sheet.max_row + 1)
            if str(frequency_sheet.cell(row=row_index, column=1).value or "").startswith("═══ グループ:")
        ]
        self.assertEqual(
            [
                "═══ グループ: Sales ═══",
                "═══ グループ: HR ═══",
                "═══ グループ: (未分類) ═══",
            ],
            [frequency_sheet.cell(row=row_index, column=1).value for row_index in group_header_rows],
        )
        first_group_row = group_header_rows[0]
        self.assertTrue(all((frequency_sheet.cell(row=first_group_row - offset, column=1).value in (None, "")) for offset in (1, 2, 3)))

    def test_report_excel_export_api_groups_frequency_sheet_sections_for_multiple_axes(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-03 09:00:00,HR,Mail,B",
                    "C002,Reject,2024-01-04 09:00:00,HR,Mail,B",
                ]
            ),
            analysis_keys=["frequency"],
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
            },
        )

        with mock.patch(
            "web_app.build_excel_ai_summary",
            return_value={
                "title": "分析コメント",
                "mode": "ollama",
                "provider": "Ollama (qwen2.5:7b)",
                "generated_at": "2026-04-09T00:00:00+00:00",
                "period": "2024-01-01 09:00 〜 2024-01-04 09:00",
                "text": "既存集計からの要約です。",
                "highlights": ["頻度の高い活動を確認してください。"],
                "note": "ローカルLLMで生成した解説を掲載しています。",
            },
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=frequency"
            )

        self.assertEqual(200, response.status_code)
        workbook = load_workbook(BytesIO(response.content))
        frequency_sheet = workbook["頻度分析"]
        group_headers = [
            frequency_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, frequency_sheet.max_row + 1)
            if str(frequency_sheet.cell(row=row_index, column=1).value or "").startswith("═══ グループ:")
        ]
        self.assertIn("═══ グループ: group_a=Sales, group_b=Web ═══", group_headers)
        self.assertIn("═══ グループ: group_a=HR, group_b=Mail ═══", group_headers)

    def test_report_excel_export_parquet_mode_includes_group_comparison_and_group_sections(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-03 09:00:00,HR,Mail,B",
                    "C002,Reject,2024-01-04 09:00:00,HR,Mail,B",
                ]
            ),
            analysis_keys=["frequency"],
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        run_data = web_app.get_run_data(run_id)
        self.assertIsNone(run_data["prepared_df"])

        with mock.patch(
            "web_app.build_excel_ai_summary",
            return_value={
                "title": "分析コメント",
                "mode": "ollama",
                "provider": "Ollama (qwen2.5:7b)",
                "generated_at": "2026-04-02T00:00:00+00:00",
                "period": "2024-01-01 09:00 〜 2024-01-04 09:00",
                "text": "既存集計からの要約です。",
                "highlights": ["頻度の高い活動を確認してください。"],
                "note": "ローカルLLMで生成した解説を掲載しています。",
            },
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=frequency"
            )

        self.assertEqual(200, response.status_code)
        workbook = load_workbook(BytesIO(response.content))
        summary_sheet = workbook["サマリー"]
        summary_pairs = self.summary_section_pairs(summary_sheet)
        summary_values = [
            summary_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, summary_sheet.max_row + 1)
        ]
        self.assertEqual("group_a、group_b、group_c", summary_pairs["グルーピング条件"])
        self.assertIn("分析ハイライト", summary_values)
        self.assertIn("グループ別比較", summary_values)
        group_table_row = self.find_row_by_value(summary_sheet, "グループ別比較")
        self.assertEqual("全体", summary_sheet.cell(row=group_table_row + 3, column=2).value)

        frequency_sheet = workbook["頻度分析"]
        self.assertEqual("═══ 全体 ═══", frequency_sheet["A1"].value)
        parquet_group_headers = [
            frequency_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, frequency_sheet.max_row + 1)
            if str(frequency_sheet.cell(row=row_index, column=1).value or "").startswith("═══ グループ:")
        ]
        self.assertEqual(
            [
                "═══ グループ: group_a=HR, group_b=Mail, group_c=B ═══",
                "═══ グループ: group_a=Sales, group_b=Web, group_c=A ═══",
            ],
            parquet_group_headers,
        )

    def test_ai_insights_api_restores_cached_output_across_page_reload(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-02 09:00:00",
                    "C002,Submit,2024-01-03 09:00:00",
                    "C002,Reject,2024-01-04 09:00:00",
                ]
            ),
            analysis_keys=["frequency", "pattern"],
        )

        empty_response = self.client.get(f"/api/runs/{run_id}/ai-insights/frequency")
        self.assertEqual(200, empty_response.status_code)
        self.assertFalse(empty_response.json()["generated"])

        with mock.patch(
            "web_app.request_ollama_insights_text",
            return_value="頻度分析向けの分析コメントです。",
        ):
            generate_response = self.client.post(f"/api/runs/{run_id}/ai-insights/frequency")

        self.assertEqual(200, generate_response.status_code)
        generate_payload = generate_response.json()
        self.assertTrue(generate_payload["generated"])
        self.assertFalse(generate_payload["cached"])
        self.assertEqual("頻度分析", generate_payload["analysis_name"])
        self.assertEqual("頻度分析向けの分析コメントです。", generate_payload["text"])

        restored_response = self.client.get(f"/api/runs/{run_id}/ai-insights/frequency")
        self.assertEqual(200, restored_response.status_code)
        restored_payload = restored_response.json()
        self.assertTrue(restored_payload["generated"])
        self.assertTrue(restored_payload["cached"])
        self.assertEqual("頻度分析向けの分析コメントです。", restored_payload["text"])

    def test_report_excel_export_uses_analysis_specific_ai_output(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,Submit,2024-01-01 09:00:00",
                    "C001,Approve,2024-01-02 09:00:00",
                    "C002,Submit,2024-01-03 09:00:00",
                    "C002,Reject,2024-01-04 09:00:00",
                ]
            ),
            analysis_keys=["frequency", "pattern"],
        )

        def _fake_ai(prompt):
            if "頻度分析" in prompt:
                return "frequency ai"
            if "処理順パターン分析" in prompt:
                return "pattern ai"
            return "generic ai"

        with mock.patch("web_app.request_ollama_insights_text", side_effect=_fake_ai):
            frequency_response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=frequency"
            )
            pattern_response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=pattern&pattern_display_limit=1"
            )

        self.assertEqual(200, frequency_response.status_code)
        self.assertEqual(200, pattern_response.status_code)

        frequency_workbook = load_workbook(BytesIO(frequency_response.content))
        pattern_workbook = load_workbook(BytesIO(pattern_response.content))
        self.assertEqual(
            ["サマリー", "分析コメント", "頻度分析", "ボトルネック分析", "改善インパクト分析"],
            self.visible_sheetnames(frequency_workbook),
        )
        self.assertEqual(["サマリー", "分析コメント", "結論サマリー", "サマリーダッシュボード", "パターンサマリー", "処理順パターン分析", "パターン01詳細"], self.visible_sheetnames(pattern_workbook))
        frequency_explanation_row = self.find_row_by_value(frequency_workbook["分析コメント"], "解説本文")
        pattern_explanation_row = self.find_row_by_value(pattern_workbook["分析コメント"], "解説本文")
        self.assertEqual("frequency ai", frequency_workbook["分析コメント"].cell(row=frequency_explanation_row + 1, column=1).value)
        self.assertEqual("pattern ai", pattern_workbook["分析コメント"].cell(row=pattern_explanation_row + 1, column=1).value)

    def test_pattern_report_excel_caps_detail_sheets_at_twenty(self):
        csv_lines = ["case_id,activity,start_time"]
        for case_index in range(1, 26):
            csv_lines.append(f"C{case_index:03d},Start_{case_index},2024-01-01 09:00:00")
            csv_lines.append(f"C{case_index:03d},End_{case_index},2024-01-01 10:00:00")

        run_id = self.analyze_uploaded_csv(
            "\n".join(csv_lines),
            analysis_keys=["pattern"],
        )

        with mock.patch(
            "web_app.request_ollama_insights_text",
            return_value="pattern ai",
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=pattern&pattern_display_limit=50"
            )

        self.assertEqual(200, response.status_code)
        workbook = load_workbook(BytesIO(response.content))
        detail_sheet_names = [
            sheet_name
            for sheet_name in workbook.sheetnames
            if sheet_name.startswith("\u30d1\u30bf\u30fc\u30f3") and sheet_name.endswith("詳細")
        ]
        self.assertEqual(20, len(detail_sheet_names))

    def test_pattern_detail_api_handles_variant_code_collisions(self):
        run_id = self.analyze_uploaded_csv(
            self.build_variant_collision_csv(),
            analysis_keys=["pattern"],
        )

        response = self.client.get(f"/api/runs/{run_id}/patterns/1")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertIn("Reminder", payload["pattern"])
        self.assertEqual(1, payload["case_count"])
        self.assertTrue(payload["step_metrics"])

    def test_pattern_report_excel_handles_variant_code_collisions(self):
        run_id = self.analyze_uploaded_csv(
            self.build_variant_collision_csv(),
            analysis_keys=["pattern"],
        )

        with mock.patch(
            "web_app.request_ollama_insights_text",
            return_value="pattern ai",
        ):
            response = self.client.get(
                f"/api/runs/{run_id}/report-excel?analysis_key=pattern&pattern_display_limit=10"
            )

        self.assertEqual(200, response.status_code)
        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("\u30d1\u30bf\u30fc\u30f301\u8a73\u7d30", workbook.sheetnames)
        self.assertIn("\u30d1\u30bf\u30fc\u30f302\u8a73\u7d30", workbook.sheetnames)

    def test_pattern_flow_api_accepts_exact_pattern_count(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency", "pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        flow_response = self.client.get(f"/api/runs/{run_id}/pattern-flow?pattern_count=1")

        self.assertEqual(200, flow_response.status_code)
        payload = flow_response.json()
        self.assertEqual(1, payload["pattern_window"]["requested_count"])
        self.assertEqual(1, payload["pattern_window"]["used_pattern_count"])
        self.assertGreaterEqual(len(payload["flow_data"]["nodes"]), 1)
        self.assertTrue(payload["flow_data"]["edges"])
        self.assertIn("avg_duration_text", payload["flow_data"]["edges"][0])
        self.assertGreater(payload["flow_data"]["edges"][0]["avg_duration_sec"], 0)

    def test_pattern_flow_api_reuses_cached_snapshot(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency", "pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        flow_path = (
            f"/api/runs/{run_id}/pattern-flow"
            "?pattern_percent=20&activity_percent=30&connection_percent=20"
        )

        with mock.patch(
            "web_app.create_pattern_flow_snapshot",
            wraps=web_app.create_pattern_flow_snapshot,
        ) as snapshot_mock:
            first_response = self.client.get(flow_path)
            second_response = self.client.get(flow_path)

        self.assertEqual(200, first_response.status_code)
        self.assertEqual(200, second_response.status_code)
        self.assertEqual(first_response.json(), second_response.json())
        self.assertEqual(1, snapshot_mock.call_count)

    def test_variant_list_api_returns_top_variants(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        variant_response = self.client.get(f"/api/runs/{run_id}/variants?limit=2")

        self.assertEqual(200, variant_response.status_code)
        payload = variant_response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertEqual(2, len(payload["variants"]))
        self.assertEqual(1, payload["variants"][0]["variant_id"])
        self.assertTrue(payload["variants"][0]["activities"])
        self.assertEqual(len(payload["variants"][0]["activities"]), payload["variants"][0]["activity_count"])
        self.assertEqual(0, payload["variants"][0]["pattern_index"])
        self.assertIn("repeat_flag", payload["variants"][0])
        self.assertGreater(payload["variants"][0]["avg_case_duration_sec"], 0)
        self.assertEqual(2, payload["coverage"]["displayed_variant_count"])
        self.assertGreater(payload["coverage"]["covered_case_count"], 0)
        self.assertLessEqual(payload["coverage"]["covered_case_count"], payload["filtered_case_count"])
        self.assertGreater(payload["coverage"]["ratio"], 0)
        self.assertLessEqual(payload["coverage"]["ratio"], 1)

    def test_variant_list_api_limit_zero_returns_all_variants(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        variant_response = self.client.get(f"/api/runs/{run_id}/variants?limit=0")

        self.assertEqual(200, variant_response.status_code)
        payload = variant_response.json()
        self.assertGreaterEqual(len(payload["variants"]), 3)
        self.assertEqual(len(payload["variants"]), payload["coverage"]["displayed_variant_count"])

    def test_variant_list_api_reuses_existing_pattern_analysis_when_unfiltered(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,受付,2024-01-01 09:00:00",
                    "C001,確認,2024-01-01 10:00:00",
                    "C001,完了,2024-01-01 11:00:00",
                    "C002,受付,2024-01-02 09:00:00",
                    "C002,確認,2024-01-02 10:00:00",
                    "C002,完了,2024-01-02 11:00:00",
                ]
            ),
            analysis_keys=["pattern"],
        )

        with mock.patch("web_app.create_variant_summary", side_effect=AssertionError("create_variant_summary should not run")):
            variant_response = self.client.get(f"/api/runs/{run_id}/variants?limit=0")

        self.assertEqual(200, variant_response.status_code)
        payload = variant_response.json()
        self.assertEqual(1, len(payload["variants"]))
        self.assertEqual(["受付", "確認", "完了"], payload["variants"][0]["activities"])

    def test_pattern_flow_api_supports_variant_filter(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["frequency", "pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        flow_response = self.client.get(f"/api/runs/{run_id}/pattern-flow?variant_id=1")

        self.assertEqual(200, flow_response.status_code)
        payload = flow_response.json()
        self.assertEqual(1, payload["selected_variant"]["variant_id"])
        self.assertEqual(0, payload["selected_variant"]["pattern_index"])
        self.assertEqual(1, payload["pattern_window"]["used_pattern_count"])
        self.assertTrue(payload["flow_data"]["nodes"])
        self.assertTrue(all("avg_duration_text" in edge for edge in payload["flow_data"]["edges"]))

    def test_pattern_flow_api_uses_lightweight_mode_for_large_datasets(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time",
                    "C001,受付,2024-01-01 09:00:00",
                    "C001,確認,2024-01-01 10:00:00",
                    "C001,完了,2024-01-01 11:00:00",
                    "C002,受付,2024-01-02 09:00:00",
                    "C002,確認,2024-01-02 10:00:00",
                    "C002,完了,2024-01-02 11:00:00",
                ]
            ),
            analysis_keys=["frequency", "pattern"],
        )

        with mock.patch.object(web_app, "LARGE_DATASET_FLOW_FAST_PATH_THRESHOLD", 1):
            with mock.patch("web_app.create_pattern_flow_snapshot", wraps=web_app.create_pattern_flow_snapshot) as wrapped_snapshot:
                flow_response = self.client.get(f"/api/runs/{run_id}/pattern-flow")

        self.assertEqual(200, flow_response.status_code)
        payload = flow_response.json()
        self.assertTrue(payload["is_large_dataset_optimized"])
        self.assertIsNone(wrapped_snapshot.call_args.kwargs["prepared_df"])

    def test_bottleneck_list_api_returns_ranked_activity_and_transition_rows(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        response = self.client.get(f"/api/runs/{run_id}/bottlenecks?limit=2")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertEqual(2, payload["limit"])
        self.assertEqual(2, len(payload["activity_bottlenecks"]))
        self.assertEqual(2, len(payload["transition_bottlenecks"]))
        self.assertEqual("\u78ba\u8a8d", payload["activity_bottlenecks"][0]["activity"])
        self.assertEqual(577.5, payload["activity_bottlenecks"][0]["avg_duration_sec"])
        self.assertEqual(0.16, payload["activity_bottlenecks"][0]["avg_duration_hours"])
        self.assertEqual("heat-5", payload["activity_heatmap"]["\u78ba\u8a8d"]["heat_class"])
        self.assertEqual("\u78ba\u8a8d", payload["transition_bottlenecks"][0]["from_activity"])
        self.assertEqual("\u5dee\u623b\u3057", payload["transition_bottlenecks"][0]["to_activity"])
        self.assertEqual("\u78ba\u8a8d__TO__\u5dee\u623b\u3057", payload["transition_bottlenecks"][0]["transition_key"])
        self.assertEqual(0.2, payload["transition_bottlenecks"][0]["avg_duration_hours"])
        self.assertEqual("heat-5", payload["transition_heatmap"]["\u78ba\u8a8d__TO__\u5dee\u623b\u3057"]["heat_class"])

    def test_transition_case_drilldown_api_returns_slowest_cases(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        response = self.client.get(
            f"/api/runs/{run_id}/transition-cases"
            "?from_activity=\u78ba\u8a8d&to_activity=\u5dee\u623b\u3057&limit=5"
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertEqual("\u78ba\u8a8d", payload["from_activity"])
        self.assertEqual("\u5dee\u623b\u3057", payload["to_activity"])
        self.assertEqual("\u78ba\u8a8d__TO__\u5dee\u623b\u3057", payload["transition_key"])
        self.assertEqual("\u78ba\u8a8d \u2192 \u5dee\u623b\u3057", payload["transition_label"])
        self.assertEqual(2, payload["returned_case_count"])
        self.assertEqual(2, len(payload["cases"]))
        self.assertEqual("C002", payload["cases"][0]["case_id"])
        self.assertEqual(1020.0, payload["cases"][0]["duration_sec"])
        self.assertEqual("17m 0s", payload["cases"][0]["duration_text"])

    def test_activity_case_drilldown_api_returns_slowest_cases(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        response = self.client.get(
            f"/api/runs/{run_id}/activity-cases"
            "?activity=\u78ba\u8a8d&limit=5"
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertEqual("\u78ba\u8a8d", payload["activity"])
        self.assertEqual(5, payload["returned_case_count"])
        self.assertEqual(5, len(payload["cases"]))
        self.assertEqual("C002", payload["cases"][0]["case_id"])
        self.assertEqual(1020.0, payload["cases"][0]["duration_sec"])
        self.assertEqual("\u5dee\u623b\u3057", payload["cases"][0]["next_activity"])

    def test_case_trace_api_returns_case_timeline(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        response = self.client.get(f"/api/runs/{run_id}/cases/C001")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertTrue(payload["found"])
        self.assertEqual("C001", payload["case_id"])
        self.assertEqual(4, payload["summary"]["event_count"])
        self.assertEqual(900.0, payload["summary"]["total_duration_sec"])
        self.assertEqual("15m 0s", payload["summary"]["total_duration_text"])
        self.assertEqual(4, len(payload["events"]))
        self.assertEqual(1, payload["events"][0]["sequence_no"])
        self.assertEqual(120.0, payload["events"][0]["wait_to_next_sec"])
        self.assertIsNone(payload["events"][-1]["wait_to_next_sec"])

    def test_case_trace_api_returns_not_found_payload(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        response = self.client.get(f"/api/runs/{run_id}/cases/C999")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertFalse(payload["found"])
        self.assertEqual("C999", payload["case_id"])
        self.assertIsNone(payload["summary"])
        self.assertEqual([], payload["events"])

    def test_analyze_api_persists_prepared_parquet_file(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["frequency"],
        )

        run_data = web_app.get_run_data(run_id)
        self.assertTrue(web_app.has_parquet_backing(run_data))
        self.assertTrue(web_app.Path(run_data["prepared_parquet_path"]).exists())
        self.assertTrue(
            str(run_data["prepared_parquet_path"]).endswith(f"{run_id}\\prepared.parquet")
        )

    def test_analyze_api_persists_log_diagnostic_inputs_for_excel_archive(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["frequency"],
        )

        run_data = web_app.get_run_data(run_id)
        self.assertTrue(web_app.Path(run_data["raw_csv_parquet_path"]).exists())
        self.assertTrue(
            str(run_data["raw_csv_parquet_path"]).endswith(f"{run_id}\\raw_upload.parquet")
        )
        self.assertIn("log_diagnostic_profile_payload", run_data)
        self.assertEqual(
            run_data["source_file_name"],
            run_data["log_diagnostic_profile_payload"]["source_file_name"],
        )
        self.assertIsNotNone(run_data["log_diagnostic_profile_payload"]["diagnostics"])

    def test_bottleneck_list_api_supports_uploaded_csv_with_duckdb_backing(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["pattern"],
        )

        response = self.client.get(f"/api/runs/{run_id}/bottlenecks?limit=2")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(2, len(payload["activity_bottlenecks"]))
        self.assertEqual("Review", payload["activity_bottlenecks"][0]["activity"])
        self.assertEqual(920.0, payload["activity_bottlenecks"][0]["avg_duration_sec"])
        self.assertEqual("Review", payload["transition_bottlenecks"][0]["from_activity"])
        self.assertEqual("Rework", payload["transition_bottlenecks"][0]["to_activity"])
        self.assertEqual(
            "Review__TO__Rework",
            payload["transition_bottlenecks"][0]["transition_key"],
        )

    def test_transition_and_activity_drilldown_support_uploaded_csv_with_duckdb_backing(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["pattern"],
        )

        transition_response = self.client.get(
            f"/api/runs/{run_id}/transition-cases?from_activity=Review&to_activity=Rework&limit=5"
        )
        self.assertEqual(200, transition_response.status_code)
        transition_payload = transition_response.json()
        self.assertEqual(2, transition_payload["returned_case_count"])
        self.assertEqual("C002", transition_payload["cases"][0]["case_id"])
        self.assertEqual(1020.0, transition_payload["cases"][0]["duration_sec"])

        activity_response = self.client.get(
            f"/api/runs/{run_id}/activity-cases?activity=Review&limit=5"
        )
        self.assertEqual(200, activity_response.status_code)
        activity_payload = activity_response.json()
        self.assertEqual(3, activity_payload["returned_case_count"])
        self.assertEqual("C002", activity_payload["cases"][0]["case_id"])
        self.assertEqual("Rework", activity_payload["cases"][0]["next_activity"])

    def test_case_trace_api_supports_uploaded_csv_with_duckdb_backing(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["pattern"],
        )

        response = self.client.get(f"/api/runs/{run_id}/cases/C001")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertTrue(payload["found"])
        self.assertEqual("C001", payload["case_id"])
        self.assertEqual(4, payload["summary"]["event_count"])
        self.assertEqual(1200.0, payload["summary"]["total_duration_sec"])
        self.assertEqual("20m 0s", payload["summary"]["total_duration_text"])
        self.assertEqual(4, len(payload["events"]))
        self.assertEqual("Review", payload["events"][1]["activity"])

    def test_pattern_flow_api_supports_uploaded_csv_with_duckdb_backing(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["frequency", "pattern"],
        )

        response = self.client.get(
            f"/api/runs/{run_id}/pattern-flow?pattern_count=2&activity_percent=100&connection_percent=100"
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(2, payload["pattern_window"]["requested_count"])
        self.assertEqual(2, payload["pattern_window"]["used_pattern_count"])
        self.assertTrue(payload["flow_data"]["nodes"])
        self.assertTrue(payload["flow_data"]["edges"])
        self.assertTrue(
            any(
                edge["source"] == "Start" and edge["target"] == "Review"
                for edge in payload["flow_data"]["edges"]
            )
        )
        self.assertTrue(all("avg_duration_text" in edge for edge in payload["flow_data"]["edges"]))

    def test_variant_pattern_flow_api_supports_uploaded_csv_with_duckdb_backing(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["frequency", "pattern"],
        )

        response = self.client.get(
            f"/api/runs/{run_id}/pattern-flow?variant_id=1&activity_percent=100&connection_percent=100"
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(1, payload["selected_variant"]["variant_id"])
        self.assertEqual(1, payload["pattern_window"]["used_pattern_count"])
        self.assertTrue(payload["flow_data"]["nodes"])
        self.assertTrue(payload["flow_data"]["edges"])
        self.assertTrue(all("avg_duration_text" in edge for edge in payload["flow_data"]["edges"]))

    def test_large_parquet_backed_run_can_release_prepared_df_and_still_render_pattern_flow(self):
        with mock.patch.object(web_app, "LARGE_DATASET_FLOW_FAST_PATH_THRESHOLD", 999999):
            run_id = self.analyze_uploaded_csv(
                self.build_duckdb_validation_csv(),
                analysis_keys=["frequency", "pattern"],
            )
            run_data = web_app.get_run_data(run_id)
            self.assertIsNone(run_data["prepared_df"])

            filter_options_response = self.client.get(f"/api/runs/{run_id}/filter-options")
            self.assertEqual(200, filter_options_response.status_code)
            self.assertEqual(run_data["filter_options"], filter_options_response.json()["options"])

            flow_response = self.client.get(
                f"/api/runs/{run_id}/pattern-flow?pattern_count=2&activity_percent=100&connection_percent=100"
            )

        self.assertEqual(200, flow_response.status_code)
        payload = flow_response.json()
        self.assertFalse(payload.get("is_large_dataset_optimized", False))
        self.assertTrue(payload["flow_data"]["edges"])
        self.assertTrue(all("avg_duration_text" in edge for edge in payload["flow_data"]["edges"]))

    def test_filter_options_api_returns_available_values(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-01 10:00:00,HR,Mail,B",
                    "C002,Reject,2024-01-03 10:00:00,HR,Mail,B",
                    "C003,Submit,2024-01-04 08:00:00,Sales,API,A",
                    "C003,Approve,2024-01-05 08:00:00,Sales,API,A",
                ]
            ),
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        response = self.client.get(f"/api/runs/{run_id}/filter-options")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("group_a", payload["column_settings"]["filters"][0]["column_name"])
        self.assertEqual(["HR", "Sales"], payload["options"]["filters"][0]["options"])
        self.assertEqual(["API", "Mail", "Web"], payload["options"]["filters"][1]["options"])
        self.assertEqual(["A", "B"], payload["options"]["filters"][2]["options"])

    def test_analysis_detail_api_supports_filters(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-01 10:00:00,HR,Mail,B",
                    "C002,Reject,2024-01-03 10:00:00,HR,Mail,B",
                    "C003,Submit,2024-01-04 08:00:00,Sales,API,A",
                    "C003,Approve,2024-01-05 08:00:00,Sales,API,A",
                ]
            ),
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        response = self.client.get(
            f"/api/runs/{run_id}/analyses/frequency?filter_value_1=Sales&date_from=2024-01-02"
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(2, payload["case_count"])
        self.assertEqual(3, payload["event_count"])
        self.assertEqual(2, payload["dashboard"]["total_cases"])
        self.assertEqual(3, payload["dashboard"]["total_records"])
        self.assertTrue(payload["impact"]["rows"])
        self.assertEqual(3, payload["root_cause"]["configured_group_count"])
        self.assertTrue(payload["root_cause"]["groups"][0]["rows"])
        self.assertEqual("Sales", payload["applied_filters"]["filter_value_1"])
        self.assertEqual("2024-01-02", payload["applied_filters"]["date_from"])

    def test_analysis_detail_api_supports_activity_filters(self):
        # Activity filter is case-level: keeps all events of cases that contain (or don't contain)
        # the target activity. C004 is a Sales case with no Submit → survives exclude-Submit filter.
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-03 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-02 10:00:00,HR,Mail,B",
                    "C002,Reject,2024-01-04 10:00:00,HR,Mail,B",
                    "C003,Submit,2024-01-04 08:00:00,Sales,API,A",
                    "C003,Approve,2024-01-05 08:00:00,Sales,API,A",
                    "C004,Approve,2024-01-06 08:00:00,Sales,API,A",
                ]
            ),
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        response = self.client.get(
            f"/api/runs/{run_id}/analyses/frequency?filter_value_1=Sales&activity_mode=exclude&activity_values=Submit"
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        # Sales cases that never pass through Submit: only C004 (1 case, 1 event)
        self.assertEqual(1, payload["case_count"])
        self.assertEqual(1, payload["event_count"])
        self.assertEqual(1, payload["dashboard"]["total_cases"])
        self.assertEqual(1, payload["dashboard"]["total_records"])
        self.assertEqual("exclude", payload["applied_filters"]["activity_mode"])
        self.assertEqual("Submit", payload["applied_filters"]["activity_values"])

    def test_bottleneck_and_variant_api_support_filters(self):
        run_id = self.analyze_uploaded_csv(
            "\n".join(
                [
                    "case_id,activity,start_time,group_a,group_b,group_c",
                    "C001,Submit,2024-01-01 09:00:00,Sales,Web,A",
                    "C001,Approve,2024-01-02 09:00:00,Sales,Web,A",
                    "C002,Submit,2024-01-01 10:00:00,HR,Mail,B",
                    "C002,Reject,2024-01-03 10:00:00,HR,Mail,B",
                    "C003,Submit,2024-01-04 08:00:00,Sales,API,A",
                    "C003,Approve,2024-01-05 08:00:00,Sales,API,A",
                ]
            ),
            extra_data={
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        variant_response = self.client.get(f"/api/runs/{run_id}/variants?filter_value_2=Web")
        bottleneck_response = self.client.get(f"/api/runs/{run_id}/bottlenecks?filter_value_3=A")

        self.assertEqual(200, variant_response.status_code)
        self.assertEqual(200, bottleneck_response.status_code)

        variant_payload = variant_response.json()
        bottleneck_payload = bottleneck_response.json()

        self.assertEqual(1, variant_payload["filtered_case_count"])
        self.assertEqual(2, variant_payload["filtered_event_count"])
        self.assertEqual("Web", variant_payload["applied_filters"]["filter_value_2"])
        self.assertEqual(2, bottleneck_payload["filtered_case_count"])
        self.assertEqual(4, bottleneck_payload["filtered_event_count"])
        self.assertEqual("A", bottleneck_payload["applied_filters"]["filter_value_3"])

    def test_pattern_detail_api_uses_pattern_index_not_display_text(self):
        analyze_response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": ["pattern"]},
        )
        self.assertEqual(200, analyze_response.status_code)

        run_id = analyze_response.json()["run_id"]
        run_data = web_app.get_run_data(run_id)
        pattern_rows = run_data["result"]["analyses"]["pattern"]["rows"]

        if pattern_rows:
            first_row = pattern_rows[0]
            for key in list(first_row.keys()):
                if "pattern" in str(key).lower() or "\u30d1\u30bf\u30fc\u30f3" in str(key) or "\u51e6\u7406\u9806" in str(key):
                    first_row[key] = ""

        response = self.client.get(f"/api/runs/{run_id}/patterns/0")

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(run_id, payload["run_id"])
        self.assertEqual(0, payload["pattern_index"])
        self.assertTrue(payload["pattern"])
        self.assertIn("repeat_flag", payload)
        self.assertIn("repeat_count", payload)
        self.assertIn("repeat_rate_pct", payload)
        self.assertIn("repeat_rate_band", payload)
        self.assertIn("review_flag", payload)
        self.assertIn("avg_case_duration_diff_min", payload)
        self.assertIn("improvement_priority_score", payload)
        self.assertIn("overall_impact_pct", payload)
        self.assertIn("fastest_pattern_flag", payload)
        self.assertIn("simple_comment", payload)
        self.assertEqual(payload["summary_row"].get("繰り返し", ""), payload["repeat_flag"])
        self.assertEqual(payload["summary_row"].get("繰り返し回数", 0), payload["repeat_count"])
        self.assertEqual(payload["summary_row"].get("繰り返し率(%)", 0), payload["repeat_rate_pct"])
        self.assertEqual(payload["summary_row"].get("繰り返し率区分", ""), payload["repeat_rate_band"])
        self.assertEqual(payload["summary_row"].get("確認区分", ""), payload["review_flag"])
        self.assertEqual(payload["summary_row"].get("平均処理時間差分(分)", 0), payload["avg_case_duration_diff_min"])
        self.assertEqual(payload["summary_row"].get("改善優先度スコア", 0), payload["improvement_priority_score"])
        self.assertEqual(payload["summary_row"].get("全体影響度(%)", 0), payload["overall_impact_pct"])
        self.assertEqual(payload["summary_row"].get("最短処理", ""), payload["fastest_pattern_flag"])
        self.assertEqual(payload["summary_row"].get("簡易コメント", ""), payload["simple_comment"])

    def test_pattern_detail_api_supports_parquet_backed_uploaded_csv_without_variant_column(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv(),
            analysis_keys=["pattern"],
        )

        run_data = web_app.get_run_data(run_id)
        self.assertIsNone(run_data["prepared_df"])

        detail_response = self.client.get(f"/api/runs/{run_id}/patterns/0")
        self.assertEqual(200, detail_response.status_code)
        detail_payload = detail_response.json()
        self.assertTrue(detail_payload["pattern"])
        self.assertTrue(detail_payload["step_metrics"])

        first_transition = detail_payload["step_metrics"][0]
        transition_response = self.client.get(
            f"/api/runs/{run_id}/transition-cases"
            f"?from_activity={quote(first_transition['activity'])}"
            f"&to_activity={quote(first_transition['next_activity'])}"
            "&pattern_index=0&limit=5"
        )
        self.assertEqual(200, transition_response.status_code)
        self.assertGreaterEqual(transition_response.json()["returned_case_count"], 1)

    def test_pattern_detail_api_supports_parquet_backed_uploaded_csv_with_variant_column(self):
        run_id = self.analyze_uploaded_csv(
            self.build_duckdb_validation_csv_with_variant(),
            analysis_keys=["pattern"],
        )

        run_data = web_app.get_run_data(run_id)
        self.assertIsNone(run_data["prepared_df"])

        detail_response = self.client.get(f"/api/runs/{run_id}/patterns/0")
        self.assertEqual(200, detail_response.status_code)
        detail_payload = detail_response.json()
        self.assertTrue(detail_payload["pattern"])
        self.assertTrue(detail_payload["step_metrics"])

        first_transition = detail_payload["step_metrics"][0]
        transition_response = self.client.get(
            f"/api/runs/{run_id}/transition-cases"
            f"?from_activity={quote(first_transition['activity'])}"
            f"&to_activity={quote(first_transition['next_activity'])}"
            "&pattern_index=0&limit=5"
        )
        self.assertEqual(200, transition_response.status_code)
        self.assertGreaterEqual(transition_response.json()["returned_case_count"], 1)

    def test_csv_headers_api_returns_sample_headers_without_upload(self):
        response = self.client.post("/api/csv-headers", data={})

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("sample_event_log.csv", payload["source_file_name"])
        self.assertEqual(["case_id", "activity", "start_time", "end_time"], payload["headers"])
        self.assertEqual("case_id", payload["default_selection"]["case_id_column"])
        self.assertEqual("case_id", payload["column_settings"]["case_id_column"])
        self.assertEqual(3, len(payload["column_settings"]["filters"]))
        self.assertEqual(3, len(payload["filter_options"]["filters"]))
        self.assertIsNone(payload["diagnostics"])

    def test_csv_headers_api_returns_uploaded_headers(self):
        csv_bytes = "request_id,step_name,event_at\nA001,Start,2024-01-01 09:00:00\n".encode("utf-8")

        response = self.client.post(
            "/api/csv-headers",
            files={"csv_file": ("custom_log.csv", BytesIO(csv_bytes), "text/csv")},
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("custom_log.csv", payload["source_file_name"])
        self.assertEqual(["request_id", "step_name", "event_at"], payload["headers"])
        self.assertEqual("", payload["default_selection"]["case_id_column"])

    def test_csv_headers_api_returns_filter_values_without_diagnostics(self):
        csv_bytes = "\n".join(
            [
                "case_no,step_name,event_at,division",
                "A001,Submit,2024-01-01 09:00:00,Sales",
                "A002,Approve,2024-01-02 10:00:00,HR",
                "A003,Review,2024-01-03 11:00:00,Sales",
            ]
        ).encode("utf-8")

        response = self.client.post(
            "/api/csv-headers",
            files={"csv_file": ("custom_log.csv", BytesIO(csv_bytes), "text/csv")},
            data={"filter_column_1": "division"},
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertIsNone(payload["diagnostics"])
        self.assertEqual("division", payload["filter_options"]["filters"][0]["column_name"])
        self.assertEqual(["HR", "Sales"], payload["filter_options"]["filters"][0]["options"])

    def test_analyze_api_rejects_duplicate_selected_columns(self):
        response = self.client.post(
            "/api/analyze",
            data={
                "case_id_column": "case_id",
                "activity_column": "case_id",
                "timestamp_column": "start_time",
                "analysis_keys": ["frequency"],
            },
        )

        self.assertEqual(400, response.status_code)
        payload = response.json()
        self.assertIn("ケースID列 / アクティビティ列 / タイムスタンプ列", payload["error"])


    def test_log_diagnostics_api_returns_uploaded_summary(self):
        csv_bytes = "\n".join(
            [
                "case_no,step_name,event_at,division",
                "A001,Submit,2024-01-01 09:00:00,Sales",
                "A001,Submit,2024-01-01 09:00:00,Sales",
                "A002,Approve,2024-01-02 10:00:00,HR",
            ]
        ).encode("utf-8")

        response = self.client.post(
            "/api/log-diagnostics",
            files={"csv_file": ("custom_log.csv", BytesIO(csv_bytes), "text/csv")},
            data={
                "case_id_column": "case_no",
                "activity_column": "step_name",
                "timestamp_column": "event_at",
            },
        )

        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("custom_log.csv", payload["source_file_name"])
        self.assertEqual(["case_no", "step_name", "event_at", "division"], payload["headers"])
        self.assertEqual(3, payload["diagnostics"]["record_count"])
        self.assertEqual(2, payload["diagnostics"]["activity_type_count"])
        self.assertEqual(3, payload["diagnostics"]["event_count"])
        self.assertEqual(1, payload["diagnostics"]["duplicate_row_count"])
        self.assertEqual("\u3042\u308a", payload["diagnostics"]["duplicate_status"])
        self.assertEqual(2, payload["diagnostics"]["deduplicated_record_count"])
        self.assertEqual(0.3333, payload["diagnostics"]["duplicate_rate"])
        self.assertEqual("case_no", payload["default_selection"]["case_id_column"])

    def test_log_diagnostics_excel_api_returns_summary_and_sample_sheets(self):
        csv_bytes = "\n".join(
            [
                "case_no,step_name,event_at,division",
                "A001,Submit,2024-01-01 09:00:00,Sales",
                "A001,Submit,2024-01-01 09:00:00,Sales",
                "A002,Approve,2024-01-02 10:00:00,HR",
                "A003,Review,2024-01-03 11:00:00,Sales",
            ]
        ).encode("utf-8")

        response = self.client.post(
            "/api/log-diagnostics-excel",
            files={"csv_file": ("custom_log.csv", BytesIO(csv_bytes), "text/csv")},
            data={
                "case_id_column": "case_no",
                "activity_column": "step_name",
                "timestamp_column": "event_at",
                "filter_column_1": "division",
                "sample_row_limit": "2",
            },
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )
        self.assertIn(
            quote("custom_log_ログ診断.xlsx"),
            response.headers["content-disposition"],
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(["ログ診断", "ログサンプル"], workbook.sheetnames)

        summary_sheet = workbook["ログ診断"]
        self.assertEqual("ログ診断サマリー", summary_sheet["A1"].value)
        self.assertEqual("元ファイル名", summary_sheet["A4"].value)
        self.assertEqual("custom_log.csv", summary_sheet["B4"].value)
        summary_values = [
            summary_sheet.cell(row=row_index, column=1).value
            for row_index in range(1, summary_sheet.max_row + 1)
        ]
        self.assertNotIn("フィルター候補", summary_values)
        self.assertNotIn("グループ/カテゴリー フィルター①", summary_values)

        sample_sheet = workbook["ログサンプル"]
        self.assertEqual("ログサンプル", sample_sheet["A1"].value)
        self.assertEqual("レコード順", sample_sheet["A3"].value)
        self.assertEqual(1, sample_sheet["A4"].value)
        self.assertEqual("A001", sample_sheet["B4"].value)
        self.assertEqual(2, sample_sheet["A5"].value)
        self.assertEqual("A001", sample_sheet["B5"].value)
        self.assertIsNone(sample_sheet["A6"].value)


    # -------------------------------------------------------------------------
    # 統計指標（標準偏差・パーセンタイル）のテスト
    # -------------------------------------------------------------------------

    def _run_sample_analysis(self, analysis_keys):
        """サンプルCSV（sample_event_log.csv）でデフォルト設定のまま分析を実行して run_id を返す。"""
        response = self.client.post(
            "/api/analyze",
            data={"analysis_keys": analysis_keys},
        )
        self.assertEqual(200, response.status_code)
        return response.json()["run_id"]

    def test_frequency_analysis_has_std_and_percentile_columns(self):
        """頻度分析の結果に標準偏差・P75/P90/P95 列が含まれること。"""
        run_id = self._run_sample_analysis(["frequency"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/frequency")
        self.assertEqual(200, response.status_code)

        rows = response.json()["analyses"]["frequency"]["rows"]
        self.assertTrue(rows, "頻度分析の結果行が空です。")

        first_row = rows[0]
        for col in ("標準偏差(分)", "75%点(分)", "90%点(分)", "95%点(分)"):
            self.assertIn(col, first_row, f"列 '{col}' が頻度分析の結果に存在しません。")

    def test_frequency_analysis_kakuin_statistics_match_expected(self):
        """確認アクティビティ（8件）の統計値がサンプルデータの手計算値と一致すること。

        サンプルデータの確認 duration_min（次イベント開始 - 自イベント開始）:
          C001:8, C002-1回目:17, C002-2回目:14, C003:7, C004:12, C005-1回目:7, C005-2回目:8, C006:4
          sorted = [4, 7, 7, 8, 8, 12, 14, 17]
          std  ≈ 4.31  (ddof=1)
          P75  = 12.5
          P90  = 14.9
          P95  = 15.95
        """
        run_id = self._run_sample_analysis(["frequency"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/frequency")
        rows = response.json()["analyses"]["frequency"]["rows"]

        kakuin_row = next((r for r in rows if r.get("アクティビティ") == "確認"), None)
        self.assertIsNotNone(kakuin_row, "確認アクティビティの行が見つかりません。")

        self.assertEqual(8, kakuin_row["イベント件数"])
        self.assertAlmostEqual(4.31, float(kakuin_row["標準偏差(分)"]), delta=0.15)
        self.assertAlmostEqual(12.5, float(kakuin_row["75%点(分)"]), delta=0.05)
        self.assertAlmostEqual(14.9, float(kakuin_row["90%点(分)"]), delta=0.05)
        self.assertAlmostEqual(15.95, float(kakuin_row["95%点(分)"]), delta=0.05)

    def test_frequency_analysis_std_is_dash_for_single_event_activity(self):
        """イベントが1件しかないアクティビティの標準偏差が '-' になること。"""
        csv_text = "\n".join([
            "case_id,activity,start_time",
            "C001,受付,2026-01-01 09:00:00",
            "C001,確認,2026-01-01 09:05:00",
            "C001,完了,2026-01-01 09:10:00",
        ])
        run_id = self.analyze_uploaded_csv(csv_text, analysis_keys=["frequency"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/frequency")
        self.assertEqual(200, response.status_code)

        rows = response.json()["analyses"]["frequency"]["rows"]
        # 全アクティビティがそれぞれ1件 → 標準偏差は "-" になる
        for row in rows:
            self.assertEqual("-", row["標準偏差(分)"],
                             f"アクティビティ '{row.get('アクティビティ')}' の標準偏差が '-' ではありません。")

    def test_transition_analysis_has_statistics_columns(self):
        """前後処理分析の結果に標準偏差・P75/P90/P95 等の統計列が含まれること。"""
        run_id = self._run_sample_analysis(["transition"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/transition")
        self.assertEqual(200, response.status_code)

        rows = response.json()["analyses"]["transition"]["rows"]
        self.assertTrue(rows, "前後処理分析の結果行が空です。")

        first_row = rows[0]
        self.assertEqual(17, len(first_row))
        for col in (
            "平均所要時間(分)",
            "中央値所要時間(分)",
            "標準偏差(分)",
            "最小所要時間(分)",
            "最大所要時間(分)",
            "75%点(分)",
            "90%点(分)",
            "95%点(分)",
            "ケース比率(%)",
            "前処理平均時間(分)",
            "後処理平均時間(分)",
        ):
            self.assertIn(col, first_row, f"列 '{col}' が前後処理分析の結果に存在しません。")

    def test_get_terminology_rows(self):
        from excel.common import get_terminology_rows

        freq_rows = get_terminology_rows("frequency")
        self.assertTrue(any(row["用語"] == "イベント比率(%)" for row in freq_rows))
        self.assertTrue(any(row["用語"] == "ケース比率(%)" for row in freq_rows))

        trans_rows = get_terminology_rows("transition")
        self.assertTrue(any(row["用語"] == "遷移" for row in trans_rows))
        self.assertTrue(any(row["用語"] == "ケース比率(%)" for row in trans_rows))
        self.assertTrue(any(row["用語"] == "前処理平均時間(分)" for row in trans_rows))
        self.assertTrue(any(row["用語"] == "後処理平均時間(分)" for row in trans_rows))
        self.assertFalse(any(row["用語"] == "イベント比率(%)" for row in trans_rows))

        pattern_rows = get_terminology_rows("pattern")
        self.assertTrue(any(row["用語"] == "処理順パターン" for row in pattern_rows))
        self.assertTrue(any(row["用語"] == "累積カバー率(%)" for row in pattern_rows))
        self.assertTrue(any(row["用語"] == "繰り返し率(%)" for row in pattern_rows))
        self.assertTrue(any(row["用語"] == "改善優先度スコア" for row in pattern_rows))
        self.assertTrue(any(row["用語"] == "全体影響度(%)" for row in pattern_rows))

        unknown_rows = get_terminology_rows("unknown")
        self.assertEqual(4, len(unknown_rows))

    def test_pattern_analysis_has_std_and_percentile_columns(self):
        """処理順パターン分析の結果に標準偏差・P75/P90/P95 列が含まれること。"""
        run_id = self._run_sample_analysis(["pattern"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/pattern")
        self.assertEqual(200, response.status_code)

        rows = response.json()["analyses"]["pattern"]["rows"]
        self.assertTrue(rows, "処理順パターン分析の結果行が空です。")

        first_row = rows[0]
        for col in (
            "繰り返し",
            "繰り返し回数",
            "繰り返し率(%)",
            "繰り返し率区分",
            "確認区分",
            "平均処理時間差分(分)",
            "改善優先度スコア",
            "全体影響度(%)",
            "最短処理",
            "簡易コメント",
            "ステップ数",
            "繰り返しアクティビティ",
            "累積カバー率(%)",
            "標準偏差ケース処理時間(分)",
            "75%点ケース処理時間(分)",
            "90%点ケース処理時間(分)",
            "95%点ケース処理時間(分)",
        ):
            self.assertIn(col, first_row, f"列 '{col}' が処理順パターン分析の結果に存在しません。")
        self.assertGreater(first_row["ステップ数"], 0)
        self.assertIsInstance(first_row["繰り返しアクティビティ"], str)

    def test_pattern_analysis_assigns_repeat_flags_and_evaluations(self):
        csv_text = "\n".join([
            "case_id,activity,start_time",
            "C001,受付,2026-01-01 09:00:00",
            "C001,確認,2026-01-01 09:05:00",
            "C001,完了,2026-01-01 09:10:00",
            "C002,受付,2026-01-01 10:00:00",
            "C002,確認,2026-01-01 10:06:00",
            "C002,完了,2026-01-01 10:12:00",
            "C003,受付,2026-01-01 11:00:00",
            "C003,確認,2026-01-01 11:05:00",
            "C003,差戻し,2026-01-01 11:10:00",
            "C003,確認,2026-01-01 11:18:00",
            "C003,完了,2026-01-01 11:28:00",
            "C004,受付,2026-01-01 12:00:00",
            "C004,確認,2026-01-01 12:05:00",
            "C004,保留,2026-01-01 12:10:00",
            "C004,再確認,2026-01-01 12:20:00",
            "C004,承認,2026-01-01 12:30:00",
            "C004,完了,2026-01-01 12:40:00",
        ])
        run_id = self.analyze_uploaded_csv(csv_text, analysis_keys=["pattern"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/pattern")
        self.assertEqual(200, response.status_code)

        rows = response.json()["analyses"]["pattern"]["rows"]
        pattern_by_route = {
            row["処理順パターン"]: row
            for row in rows
        }

        self.assertEqual("", pattern_by_route["受付→確認→完了"]["繰り返し"])
        self.assertEqual(0, pattern_by_route["受付→確認→完了"]["繰り返し回数"])
        self.assertEqual(0.0, pattern_by_route["受付→確認→完了"]["繰り返し率(%)"])
        self.assertEqual("0〜10%", pattern_by_route["受付→確認→完了"]["繰り返し率区分"])
        self.assertEqual("", pattern_by_route["受付→確認→完了"]["確認区分"])
        self.assertEqual("○", pattern_by_route["受付→確認→完了"]["最短処理"])
        self.assertEqual(3, pattern_by_route["受付→確認→完了"]["ステップ数"])
        self.assertEqual("", pattern_by_route["受付→確認→完了"]["繰り返しアクティビティ"])
        self.assertEqual(50.0, pattern_by_route["受付→確認→完了"]["累積カバー率(%)"])
        self.assertLess(float(pattern_by_route["受付→確認→完了"]["平均処理時間差分(分)"]), 0)
        self.assertEqual(0.0, float(pattern_by_route["受付→確認→完了"]["改善優先度スコア"]))
        self.assertIn("安定", pattern_by_route["受付→確認→完了"]["簡易コメント"])
        self.assertEqual("○", pattern_by_route["受付→確認→差戻し→確認→完了"]["繰り返し"])
        self.assertEqual(1, pattern_by_route["受付→確認→差戻し→確認→完了"]["繰り返し回数"])
        self.assertEqual(20.0, pattern_by_route["受付→確認→差戻し→確認→完了"]["繰り返し率(%)"])
        self.assertEqual("10〜30%", pattern_by_route["受付→確認→差戻し→確認→完了"]["繰り返し率区分"])
        self.assertEqual("", pattern_by_route["受付→確認→差戻し→確認→完了"]["確認区分"])
        self.assertEqual(5, pattern_by_route["受付→確認→差戻し→確認→完了"]["ステップ数"])
        self.assertEqual("確認", pattern_by_route["受付→確認→差戻し→確認→完了"]["繰り返しアクティビティ"])
        self.assertGreater(float(pattern_by_route["受付→確認→差戻し→確認→完了"]["平均処理時間差分(分)"]), 0)
        self.assertGreater(float(pattern_by_route["受付→確認→差戻し→確認→完了"]["改善優先度スコア"]), 0)
        self.assertGreaterEqual(float(pattern_by_route["受付→確認→差戻し→確認→完了"]["全体影響度(%)"]), 0)
        self.assertIn("改善候補", pattern_by_route["受付→確認→差戻し→確認→完了"]["簡易コメント"])
        self.assertEqual("", pattern_by_route["受付→確認→保留→再確認→承認→完了"]["繰り返し"])
        self.assertEqual(0, pattern_by_route["受付→確認→保留→再確認→承認→完了"]["繰り返し回数"])
        self.assertEqual(0.0, pattern_by_route["受付→確認→保留→再確認→承認→完了"]["繰り返し率(%)"])
        self.assertEqual("0〜10%", pattern_by_route["受付→確認→保留→再確認→承認→完了"]["繰り返し率区分"])
        self.assertEqual("", pattern_by_route["受付→確認→保留→再確認→承認→完了"]["確認区分"])
        self.assertEqual(6, pattern_by_route["受付→確認→保留→再確認→承認→完了"]["ステップ数"])
        self.assertEqual("", pattern_by_route["受付→確認→保留→再確認→承認→完了"]["繰り返しアクティビティ"])
        self.assertGreater(float(pattern_by_route["受付→確認→保留→再確認→承認→完了"]["平均処理時間差分(分)"]), 0)
        self.assertEqual(0.0, float(pattern_by_route["受付→確認→保留→再確認→承認→完了"]["改善優先度スコア"]))
        self.assertIn("平均超過", pattern_by_route["受付→確認→保留→再確認→承認→完了"]["簡易コメント"])

    def test_pattern_analysis_std_is_dash_for_single_case_pattern(self):
        """ケース数が1件のパターンの標準偏差が '-' になること。"""
        csv_text = "\n".join([
            "case_id,activity,start_time",
            "C001,受付,2026-01-01 09:00:00",
            "C001,確認,2026-01-01 09:05:00",
            "C001,完了,2026-01-01 09:10:00",
            "C002,受付,2026-01-01 10:00:00",
            "C002,確認,2026-01-01 10:05:00",
            "C002,差戻し,2026-01-01 10:08:00",
            "C002,確認,2026-01-01 10:12:00",
            "C002,完了,2026-01-01 10:18:00",
        ])
        run_id = self.analyze_uploaded_csv(csv_text, analysis_keys=["pattern"])
        response = self.client.get(f"/api/runs/{run_id}/analyses/pattern")
        self.assertEqual(200, response.status_code)

        rows = response.json()["analyses"]["pattern"]["rows"]
        # C001 と C002 で異なるパターン → それぞれ1件 → 標準偏差は "-"
        for row in rows:
            self.assertEqual("-", row["標準偏差ケース処理時間(分)"],
                             f"パターン '{row.get('処理順パターン')}' の標準偏差が '-' ではありません。")


if __name__ == "__main__":
    unittest.main()




