from io import BytesIO
from pathlib import Path
import unittest

import pandas as pd
from openpyxl import load_workbook

from core.analysis_service import (
    build_group_summary,
    create_activity_case_drilldown,
    create_case_trace_details,
    create_log_diagnostics,
    create_rule_based_insights,
    create_transition_case_drilldown,
    create_variant_flow_snapshot,
    create_variant_summary,
    detect_group_columns,
    filter_by_start_end_activity,
    filter_prepared_df,
    get_filter_options,
    normalize_filter_column_settings,
    normalize_filter_params,
)
from core.data_loader import load_and_prepare_data, prepare_event_log
from excel.basic_exporter import _build_summary_sheet_df, build_excel_bytes
from core.分析.前後処理分析.transition_analysis import create_transition_analysis
from core.分析.処理順パターン分析.pattern_analysis import create_pattern_analysis
from core.分析.頻度分析.frequency_analysis import create_frequency_analysis


ROOT_DIR = Path(__file__).resolve().parents[1]
SAMPLE_FILE = ROOT_DIR / "sample_event_log.csv"



class GroupingTestCase(unittest.TestCase):
    """グルーピングモード機能のテスト"""

    @classmethod
    def setUpClass(cls):
        # グルーピングテスト用の合成データを直接作成する
        cls.prepared_df = pd.DataFrame({
            "case_id": ["C1", "C1", "C1", "C2", "C2", "C3", "C3", "C4", "C4"],
            "activity": ["受付", "確認", "完了", "受付", "確認", "受付", "完了", "受付", "確認"],
            "sequence_no": [1, 2, 3, 1, 2, 1, 2, 1, 2],
            "start_time": pd.to_datetime([
                "2024-01-01 09:00", "2024-01-01 09:10", "2024-01-01 09:30",
                "2024-01-02 09:00", "2024-01-02 09:15",
                "2024-01-03 09:00", "2024-01-03 09:20",
                "2024-01-04 09:00", "2024-01-04 09:05",
            ]),
            "next_time": pd.to_datetime([
                "2024-01-01 09:10", "2024-01-01 09:30", "2024-01-01 09:30",
                "2024-01-02 09:15", "2024-01-02 09:15",
                "2024-01-03 09:20", "2024-01-03 09:20",
                "2024-01-04 09:05", "2024-01-04 09:05",
            ]),
            "timestamp": pd.to_datetime([
                "2024-01-01 09:00", "2024-01-01 09:10", "2024-01-01 09:30",
                "2024-01-02 09:00", "2024-01-02 09:15",
                "2024-01-03 09:00", "2024-01-03 09:20",
                "2024-01-04 09:00", "2024-01-04 09:05",
            ]),
            "duration_min": [10.0, 20.0, 0.0, 15.0, 0.0, 20.0, 0.0, 5.0, 0.0],
            "duration_sec": [600, 1200, 0, 900, 0, 1200, 0, 300, 0],
            "group_a": ["Sales", "Sales", "Sales", "Sales", "Sales", "Mfg", "Mfg", "Mfg", "Mfg"],
            "group_b": ["Tokyo", "Tokyo", "Tokyo", "Osaka", "Osaka", "Tokyo", "Tokyo", "Osaka", "Osaka"],
        })

    # ── detect_group_columns ────────────────────────────────────────────────

    def test_detect_group_columns_returns_col_without_value(self):
        result = detect_group_columns(
            {"filter_value_1": None, "filter_value_2": None},
            {"filter_column_1": "group_a", "filter_column_2": "group_b"},
        )
        self.assertEqual(["group_a", "group_b"], result)

    def test_detect_group_columns_excludes_col_with_value(self):
        result = detect_group_columns(
            {"filter_value_1": "Sales", "filter_value_2": None},
            {"filter_column_1": "group_a", "filter_column_2": "group_b"},
        )
        self.assertEqual(["group_b"], result)

    def test_detect_group_columns_returns_empty_when_no_col_set(self):
        result = detect_group_columns({}, {})
        self.assertEqual([], result)

    # ── create_frequency_analysis (group_columns) ──────────────────────────

    def test_frequency_analysis_group_columns_adds_group_col(self):
        result = create_frequency_analysis(self.prepared_df, group_columns=["group_a"])
        self.assertIn("group_a", result.columns)
        self.assertIn("activity", result.columns)
        # グループ列の値が2種類あること
        self.assertEqual({"Sales", "Mfg"}, set(result["group_a"].tolist()))

    def test_frequency_analysis_no_group_unchanged(self):
        result = create_frequency_analysis(self.prepared_df)
        self.assertNotIn("group_a", result.columns)

    # ── create_transition_analysis (group_columns) ─────────────────────────

    def test_transition_analysis_group_columns_adds_group_col(self):
        result = create_transition_analysis(self.prepared_df, group_columns=["group_a"])
        self.assertIn("group_a", result.columns)
        self.assertIn("from_activity", result.columns)
        self.assertIn("to_activity", result.columns)

    def test_transition_analysis_no_group_unchanged(self):
        result = create_transition_analysis(self.prepared_df)
        self.assertNotIn("group_a", result.columns)

    # ── create_pattern_analysis (group_columns) ────────────────────────────

    def test_pattern_analysis_group_columns_adds_group_col(self):
        result = create_pattern_analysis(self.prepared_df, group_columns=["group_a"])
        self.assertIn("group_a", result.columns)
        self.assertIn("pattern", result.columns)

    def test_pattern_analysis_no_group_unchanged(self):
        result = create_pattern_analysis(self.prepared_df)
        self.assertNotIn("group_a", result.columns)

    # ── build_group_summary ────────────────────────────────────────────────

    def test_build_group_summary_single_column(self):
        summary = build_group_summary(self.prepared_df, ["group_a"])
        self.assertIn("__meta__", summary)
        self.assertIn("group_a", summary)
        col_data = summary["group_a"]
        self.assertIn("Sales", col_data)
        self.assertIn("Mfg", col_data)
        self.assertEqual(4, summary["__meta__"]["total_case_count"])
        self.assertEqual(9, summary["__meta__"]["total_event_count"])
        self.assertEqual(17.5, summary["__meta__"]["avg_duration_min"])
        self.assertEqual(17.5, summary["__meta__"]["median_duration_min"])
        self.assertEqual(30.0, summary["__meta__"]["max_duration_min"])
        self.assertEqual(70.0, summary["__meta__"]["total_duration_min"])
        self.assertEqual(2, col_data["Sales"]["case_count"])
        self.assertEqual(5, col_data["Sales"]["event_count"])
        self.assertEqual(50.0, col_data["Sales"]["case_ratio_pct"])
        self.assertEqual(55.56, col_data["Sales"]["event_ratio_pct"])
        self.assertEqual(22.5, col_data["Sales"]["avg_duration_min"])
        self.assertEqual(22.5, col_data["Sales"]["median_duration_min"])
        self.assertEqual(30.0, col_data["Sales"]["max_duration_min"])
        self.assertEqual(45.0, col_data["Sales"]["total_duration_min"])
        for entry in col_data.values():
            self.assertIn("case_count", entry)
            self.assertIn("case_ratio_pct", entry)
            self.assertIn("event_count", entry)
            self.assertIn("event_ratio_pct", entry)
            self.assertIn("median_duration_min", entry)
            self.assertIn("max_duration_min", entry)
            self.assertIn("total_duration_min", entry)
            self.assertGreater(entry["case_count"], 0)
            self.assertGreater(entry["event_count"], 0)

    def test_build_group_summary_multiple_columns(self):
        summary = build_group_summary(self.prepared_df, ["group_a", "group_b"])
        self.assertIn("__meta__", summary)
        self.assertIn("group_a", summary)
        self.assertIn("group_b", summary)
        self.assertIn("Tokyo", summary["group_b"])
        for col in ("group_a", "group_b"):
            self.assertAlmostEqual(100.0, sum(entry["case_ratio_pct"] for entry in summary[col].values()), places=2)
            self.assertAlmostEqual(100.0, sum(entry["event_ratio_pct"] for entry in summary[col].values()), places=2)

    def test_build_group_summary_empty_group_columns(self):
        summary = build_group_summary(self.prepared_df, [])
        self.assertEqual({}, summary)

    def test_build_group_summary_missing_column_skipped(self):
        summary = build_group_summary(self.prepared_df, ["nonexistent_col"])
        self.assertEqual({}, summary)

    def test_build_summary_sheet_df_adds_total_row_and_expands_columns(self):
        summary = build_group_summary(self.prepared_df, ["group_a"])

        summary_df = _build_summary_sheet_df(summary, ["group_a"])

        self.assertEqual(
            [
                "グルーピング軸",
                "値",
                "ケース数",
                "ケース比率(%)",
                "イベント数",
                "イベント比率(%)",
                "平均所要時間(分)",
                "中央値所要時間(分)",
                "最大所要時間(分)",
                "合計所要時間(分)",
            ],
            summary_df.columns.tolist(),
        )
        self.assertEqual(3, len(summary_df))
        self.assertEqual("全体", summary_df.iloc[0]["値"])
        self.assertEqual("", summary_df.iloc[0]["グルーピング軸"])
        self.assertEqual(100.0, summary_df.iloc[0]["ケース比率(%)"])
        self.assertEqual(100.0, summary_df.iloc[0]["イベント比率(%)"])
        self.assertEqual(17.5, summary_df.iloc[0]["平均所要時間(分)"])
        self.assertEqual(["group_a", "group_a"], summary_df["グルーピング軸"].iloc[1:].tolist())
        self.assertTrue(summary_df["ケース数"].iloc[1:].is_monotonic_decreasing)

    def test_build_summary_sheet_df_keeps_group_suffix_for_multiple_axes(self):
        summary = build_group_summary(self.prepared_df, ["group_a", "group_b"])

        summary_df = _build_summary_sheet_df(summary, ["group_a", "group_b"])

        axis_labels = set(summary_df["グルーピング軸"].iloc[1:].tolist())
        self.assertIn("group_a（グループ①）", axis_labels)
        self.assertIn("group_b（グループ②）", axis_labels)

    def test_build_excel_bytes_writes_summary_sheet_for_grouping_mode(self):
        summary = build_group_summary(self.prepared_df, ["group_a"])
        detail_df = create_frequency_analysis(self.prepared_df, group_columns=["group_a"])

        excel_bytes = build_excel_bytes(
            detail_df,
            "頻度分析",
            group_columns=["group_a"],
            group_summary=summary,
        )

        workbook = load_workbook(BytesIO(excel_bytes))
        self.assertIn("サマリー", workbook.sheetnames)
        summary_sheet = workbook["サマリー"]
        headers = [summary_sheet.cell(row=1, column=column_index).value for column_index in range(1, 11)]
        self.assertEqual(
            [
                "グルーピング軸",
                "値",
                "ケース数",
                "ケース比率(%)",
                "イベント数",
                "イベント比率(%)",
                "平均所要時間(分)",
                "中央値所要時間(分)",
                "最大所要時間(分)",
                "合計所要時間(分)",
            ],
            headers,
        )
        self.assertEqual("全体", summary_sheet["B2"].value)
        self.assertEqual(100.0, summary_sheet["D2"].value)
        self.assertEqual("group_a", summary_sheet["A3"].value)

    def test_build_excel_bytes_omits_summary_sheet_without_grouping(self):
        detail_df = create_frequency_analysis(self.prepared_df)

        excel_bytes = build_excel_bytes(detail_df, "頻度分析")

        workbook = load_workbook(BytesIO(excel_bytes))
        self.assertEqual(["頻度分析"], workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
