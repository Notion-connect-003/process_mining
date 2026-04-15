from io import BytesIO
from pathlib import Path
import unittest

import pandas as pd
from openpyxl import load_workbook

from 共通スクリプト.analysis_service import (
    build_group_summary,
    create_activity_case_drilldown,
    create_case_trace_details,
    create_log_diagnostics,
    create_rule_based_insights,
    create_transition_case_drilldown,
    create_variant_flow_snapshot,
    create_variant_summary,
    detect_group_columns,
    filter_by_start_end_activity,
    filter_prepared_df,
    get_filter_options,
    normalize_filter_column_settings,
    normalize_filter_params,
)
from 共通スクリプト.data_loader import load_and_prepare_data, prepare_event_log
from 共通スクリプト.Excel出力.excel_exporter import _build_summary_sheet_df, build_excel_bytes
from 共通スクリプト.分析.前後処理分析.transition_analysis import create_transition_analysis
from 共通スクリプト.分析.処理順パターン分析.pattern_analysis import create_pattern_analysis
from 共通スクリプト.分析.頻度分析.frequency_analysis import create_frequency_analysis


ROOT_DIR = Path(__file__).resolve().parents[1]
SAMPLE_FILE = ROOT_DIR / "sample_event_log.csv"


class ProcessMiningTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.prepared_df = load_and_prepare_data(
            file_path=SAMPLE_FILE,
            case_id_column="case_id",
            activity_column="activity",
            timestamp_column="start_time",
        )

    def test_prepare_event_log_adds_analysis_columns(self):
        expected_columns = {
            "case_id",
            "activity",
            "timestamp",
            "start_time",
            "next_time",
            "duration_sec",
            "duration_min",
            "sequence_no",
            "event_count_in_case",
        }
        self.assertTrue(expected_columns.issubset(set(self.prepared_df.columns.tolist())))
        self.assertIn("end_time", self.prepared_df.columns.tolist())

        first_row = self.prepared_df.iloc[0]
        self.assertEqual("C001", first_row["case_id"])
        self.assertEqual("受付", first_row["activity"])
        self.assertEqual(120.0, first_row["duration_sec"])
        self.assertEqual(2.0, first_row["duration_min"])
        self.assertEqual(1, first_row["sequence_no"])
        self.assertEqual(4, first_row["event_count_in_case"])

        last_case_row = self.prepared_df[self.prepared_df["case_id"] == "C001"].iloc[-1]
        self.assertEqual(0.0, last_case_row["duration_sec"])
        self.assertEqual(last_case_row["start_time"], last_case_row["next_time"])

    def test_prepare_event_log_raises_when_required_column_is_missing(self):
        invalid_df = pd.DataFrame(
            [
                {"case_id": "C001", "activity": "受付"},
            ]
        )

        with self.assertRaisesRegex(ValueError, "入力CSVに必要な列がありません"):
            prepare_event_log(
                df=invalid_df,
                case_id_column="case_id",
                activity_column="activity",
                timestamp_column="start_time",
            )

    def test_frequency_analysis_returns_expected_summary(self):
        result = create_frequency_analysis(self.prepared_df)

        self.assertEqual(["確認", "受付", "完了", "承認", "差戻し"], result["activity"].tolist())

        top_row = result.iloc[0]
        self.assertEqual("確認", top_row["activity"])
        self.assertEqual(8, top_row["event_count"])
        self.assertEqual(6, top_row["case_count"])
        self.assertEqual(77.0, top_row["total_duration_min"])
        self.assertEqual(30.77, top_row["event_ratio_pct"])

    def test_transition_analysis_returns_expected_transitions(self):
        result = create_transition_analysis(self.prepared_df)

        first_row = result.iloc[0]
        self.assertEqual("受付", first_row["from_activity"])
        self.assertEqual("確認", first_row["to_activity"])
        self.assertEqual(6, first_row["transition_count"])
        self.assertEqual(6, first_row["case_count"])
        self.assertEqual(30.0, first_row["transition_ratio_pct"])

        confirm_to_completion = result[
            (result["from_activity"] == "確認") & (result["to_activity"] == "完了")
        ].iloc[0]
        self.assertEqual(2, confirm_to_completion["transition_count"])
        self.assertEqual(0.0, confirm_to_completion["to_total_duration_min"])

    def test_pattern_analysis_returns_expected_patterns(self):
        result = create_pattern_analysis(self.prepared_df)

        self.assertEqual(3, len(result))

        first_row = result.iloc[0]
        self.assertEqual("受付→確認→完了", first_row["pattern"])
        self.assertEqual(2, first_row["case_count"])
        self.assertEqual(6.5, first_row["avg_case_duration_min"])
        self.assertEqual(33.33, first_row["case_ratio_pct"])

    def test_variant_summary_returns_ranked_variants(self):
        variants = create_variant_summary(self.prepared_df, limit=10)

        self.assertEqual(3, len(variants))
        self.assertEqual(1, variants[0]["variant_id"])
        self.assertEqual(["受付", "確認", "完了"], variants[0]["activities"])
        self.assertEqual(3, variants[0]["activity_count"])
        self.assertEqual(2, variants[0]["count"])
        self.assertEqual(0.3333, variants[0]["ratio"])
        self.assertGreater(variants[0]["avg_case_duration_sec"], 0)

    def test_rule_based_insights_returns_key_points(self):
        dashboard = {
            "has_data": True,
            "total_cases": 6,
            "total_records": len(self.prepared_df),
            "top10_variant_coverage_pct": 100.0,
        }
        bottleneck_summary = {
            "activity_bottlenecks": [{"activity": "確認", "avg_duration_sec": 577.5}],
            "transition_bottlenecks": [{"from_activity": "確認", "to_activity": "差戻し", "avg_duration_sec": 720.0}],
            "activity_heatmap": {},
            "transition_heatmap": {},
        }
        impact_summary = {
            "has_data": True,
            "rows": [{
                "rank": 1,
                "transition_key": "確認__TO__承認",
                "transition_label": "確認 → 承認",
                "avg_duration_text": "10m 30s",
                "case_count": 4,
            }],
        }
        insights = create_rule_based_insights(
            self.prepared_df,
            dashboard_summary=dashboard,
            bottleneck_summary=bottleneck_summary,
            impact_summary=impact_summary,
            max_items=5,
        )

        self.assertEqual("rule_based", insights["mode"])
        self.assertTrue(insights["has_data"])
        self.assertGreaterEqual(len(insights["items"]), 3)
        self.assertLessEqual(len(insights["items"]), 5)
        self.assertEqual("scope", insights["items"][0]["id"])
        self.assertIn("対象は", insights["items"][0]["text"])
        self.assertTrue(any(item["id"] == "top_impact_transition" for item in insights["items"]))

    def test_rule_based_insights_support_analysis_specific_content(self):
        frequency_rows = create_frequency_analysis(self.prepared_df).to_dict(orient="records")
        pattern_rows = create_pattern_analysis(self.prepared_df).to_dict(orient="records")
        minimal_dashboard = {
            "has_data": True,
            "total_cases": self.prepared_df["case_id"].nunique(),
            "total_records": len(self.prepared_df),
            "top10_variant_coverage_pct": 100.0,
        }

        frequency_insights = create_rule_based_insights(
            self.prepared_df,
            analysis_key="frequency",
            analysis_rows=frequency_rows,
            dashboard_summary=minimal_dashboard,
            max_items=5,
        )
        pattern_insights = create_rule_based_insights(
            self.prepared_df,
            analysis_key="pattern",
            analysis_rows=pattern_rows,
            dashboard_summary=minimal_dashboard,
            max_items=5,
        )

        self.assertTrue(any(item["id"] == "top_activity" for item in frequency_insights["items"]))
        self.assertTrue(any(item["id"] == "event_distribution" for item in frequency_insights["items"]))
        self.assertTrue(any(item["id"] == "top_pattern" for item in pattern_insights["items"]))
        self.assertTrue(any(item["id"] == "pattern_variability" for item in pattern_insights["items"]))

    def test_transition_case_drilldown_returns_slowest_cases(self):
        rows = create_transition_case_drilldown(
            self.prepared_df,
            from_activity="確認",
            to_activity="差戻し",
            limit=5,
        )

        self.assertEqual(2, len(rows))
        self.assertEqual("C002", rows[0]["case_id"])
        self.assertEqual(1020.0, rows[0]["duration_sec"])
        self.assertEqual("17m 0s", rows[0]["duration_text"])
        self.assertEqual("C005", rows[1]["case_id"])
        self.assertEqual(420.0, rows[1]["duration_sec"])

    def test_activity_case_drilldown_returns_slowest_cases(self):
        rows = create_activity_case_drilldown(
            self.prepared_df,
            activity="確認",
            limit=5,
        )

        self.assertEqual(5, len(rows))
        self.assertEqual("C002", rows[0]["case_id"])
        self.assertEqual(1020.0, rows[0]["duration_sec"])
        self.assertEqual("17m 0s", rows[0]["duration_text"])
        self.assertEqual("差戻し", rows[0]["next_activity"])

    def test_variant_flow_snapshot_returns_single_variant_graph(self):
        snapshot = create_variant_flow_snapshot(
            self.prepared_df,
            "受付→確認→承認→完了",
            activity_percent=100,
            connection_percent=100,
        )

        self.assertEqual(1, snapshot["pattern_window"]["used_pattern_count"])
        self.assertTrue(snapshot["flow_data"]["nodes"])
        self.assertTrue(snapshot["flow_data"]["edges"])


    def test_pattern_flow_snapshot_filters_top_patterns_nodes_and_edges(self):
        import importlib.util

        module_path = ROOT_DIR / "共通スクリプト" / "analysis_service.py"
        spec = importlib.util.spec_from_file_location("analysis_service_local", module_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)

        pattern_rows = [
            {"ケース数": 90, "処理順パターン": "申請受付→内容確認→処理完了"},
            {"ケース数": 70, "処理順パターン": "申請受付→一次承認→処理完了"},
            {"ケース数": 50, "処理順パターン": "申請受付→差戻し→再提出→処理完了"},
            {"ケース数": 30, "処理順パターン": "申請受付→自動処理→処理完了"},
        ]

        snapshot = module.create_pattern_flow_snapshot(
            pattern_rows=pattern_rows,
            pattern_percent=50,
            activity_percent=60,
            connection_percent=50,
            pattern_cap=4,
        )

        self.assertEqual(2, snapshot["pattern_window"]["used_pattern_count"])
        self.assertEqual(4, snapshot["activity_window"]["available_activity_count"])
        self.assertEqual(4, snapshot["connection_window"]["available_connection_count"])
        self.assertTrue(snapshot["flow_data"]["nodes"])
        self.assertTrue(snapshot["flow_data"]["edges"])
        self.assertTrue(
            all("layer" in node and "orderScore" in node for node in snapshot["flow_data"]["nodes"])
        )

    def test_case_trace_details_returns_case_timeline(self):
        trace = create_case_trace_details(self.prepared_df, "C001")

        self.assertTrue(trace["found"])
        self.assertEqual("C001", trace["case_id"])
        self.assertEqual(4, trace["summary"]["event_count"])
        self.assertEqual(900.0, trace["summary"]["total_duration_sec"])
        self.assertEqual("15m 0s", trace["summary"]["total_duration_text"])
        self.assertEqual(4, len(trace["events"]))
        self.assertEqual(1, trace["events"][0]["sequence_no"])
        self.assertEqual(trace["events"][1]["activity"], trace["events"][0]["next_activity"])
        self.assertEqual(120.0, trace["events"][0]["wait_to_next_sec"])
        self.assertEqual("", trace["events"][-1]["wait_to_next_text"])
        self.assertIsNone(trace["events"][-1]["next_activity"])

    def test_case_trace_details_returns_not_found_payload(self):
        trace = create_case_trace_details(self.prepared_df, "C999")

        self.assertFalse(trace["found"])
        self.assertEqual("C999", trace["case_id"])
        self.assertIsNone(trace["summary"])
        self.assertEqual([], trace["events"])

    def test_create_log_diagnostics_reports_duplicate_counts(self):
        raw_df = pd.DataFrame(
            [
                {"case_id": "C001", "activity": "申請", "start_time": "2024-01-01 09:00:00", "group_a": "Sales"},
                {"case_id": "C001", "activity": "申請", "start_time": "2024-01-01 09:00:00", "group_a": "Sales"},
                {"case_id": "C002", "activity": "承認", "start_time": "2024-01-02 10:00:00", "group_a": "HR"},
            ]
        )

        diagnostics = create_log_diagnostics(
            raw_df,
            case_id_column="case_id",
            activity_column="activity",
            timestamp_column="start_time",
        )

        self.assertEqual(3, diagnostics["record_count"])
        self.assertEqual(2, diagnostics["activity_type_count"])
        self.assertEqual(1, diagnostics["duplicate_row_count"])
        self.assertEqual("あり", diagnostics["duplicate_status"])
        self.assertEqual(2, diagnostics["deduplicated_record_count"])
        self.assertEqual(0.3333, diagnostics["duplicate_rate"])

    def test_normalize_filter_params_trims_blank_values(self):
        normalized = normalize_filter_params(
            date_from=" 2024-01-01 ",
            date_to="",
            filter_value_1=" Sales ",
            filter_value_2="   ",
            filter_value_3=None,
            activity_mode=" Exclude ",
            activity_values=[" Submit ", "", "Approve", "Submit"],
        )

        self.assertEqual("2024-01-01", normalized["date_from"])
        self.assertIsNone(normalized["date_to"])
        self.assertEqual("Sales", normalized["filter_value_1"])
        self.assertIsNone(normalized["filter_value_2"])
        self.assertIsNone(normalized["filter_value_3"])
        self.assertEqual("exclude", normalized["activity_mode"])
        self.assertEqual("Submit,Approve", normalized["activity_values"])

    def test_normalize_filter_params_trims_start_end_activity_values(self):
        normalized = normalize_filter_params(
            start_activity_values=[" Submit ", "", "Review", "Submit"],
            end_activity_values=" Done , Reject , Done ",
        )

        self.assertEqual("Submit,Review", normalized["start_activity_values"])
        self.assertEqual("Done,Reject", normalized["end_activity_values"])

    def test_filter_by_start_end_activity_filters_cases_by_boundary_events(self):
        df = pd.DataFrame(
            [
                {"case_id": "C001", "activity": "Submit", "timestamp": pd.Timestamp("2024-01-01 09:00:00")},
                {"case_id": "C001", "activity": "Approve", "timestamp": pd.Timestamp("2024-01-01 09:05:00")},
                {"case_id": "C001", "activity": "Done", "timestamp": pd.Timestamp("2024-01-01 09:10:00")},
                {"case_id": "C002", "activity": "Intake", "timestamp": pd.Timestamp("2024-01-01 10:00:00")},
                {"case_id": "C002", "activity": "Approve", "timestamp": pd.Timestamp("2024-01-01 10:05:00")},
                {"case_id": "C002", "activity": "Done", "timestamp": pd.Timestamp("2024-01-01 10:10:00")},
                {"case_id": "C003", "activity": "Submit", "timestamp": pd.Timestamp("2024-01-01 11:00:00")},
                {"case_id": "C003", "activity": "Check", "timestamp": pd.Timestamp("2024-01-01 11:05:00")},
                {"case_id": "C003", "activity": "Reject", "timestamp": pd.Timestamp("2024-01-01 11:10:00")},
            ]
        )

        start_filtered = filter_by_start_end_activity(
            df,
            "case_id",
            "activity",
            "timestamp",
            start_activities=["Submit"],
        )
        self.assertEqual({"C001", "C003"}, set(start_filtered["case_id"].unique()))

        end_filtered = filter_by_start_end_activity(
            df,
            "case_id",
            "activity",
            "timestamp",
            end_activities=["Done"],
        )
        self.assertEqual({"C001", "C002"}, set(end_filtered["case_id"].unique()))

        boundary_filtered = filter_by_start_end_activity(
            df,
            "case_id",
            "activity",
            "timestamp",
            start_activities=["Submit"],
            end_activities=["Done"],
        )
        self.assertEqual({"C001"}, set(boundary_filtered["case_id"].unique()))

    def test_filter_prepared_df_filters_by_date_and_attributes(self):
        raw_df = pd.DataFrame(
            [
                {"case": "C001", "step": "Submit", "ts": "2024-01-01 09:00:00", "group_a": "Sales", "group_b": "Web", "group_c": "A"},
                {"case": "C001", "step": "Approve", "ts": "2024-01-03 09:00:00", "group_a": "Sales", "group_b": "Web", "group_c": "A"},
                {"case": "C002", "step": "Submit", "ts": "2024-01-02 10:00:00", "group_a": "HR", "group_b": "Mail", "group_c": "B"},
                {"case": "C002", "step": "Reject", "ts": "2024-01-04 10:00:00", "group_a": "HR", "group_b": "Mail", "group_c": "B"},
            ]
        )
        prepared_df = prepare_event_log(
            df=raw_df,
            case_id_column="case",
            activity_column="step",
            timestamp_column="ts",
        )

        filtered_df = filter_prepared_df(
            prepared_df,
            {
                "date_from": "2024-01-02",
                "date_to": "2024-01-03",
                "filter_value_1": "Sales",
                "filter_value_2": "Web",
                "filter_value_3": "A",
            },
            {
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        self.assertEqual(1, len(filtered_df))
        self.assertEqual(["C001"], filtered_df["case_id"].tolist())
        self.assertEqual(["Approve"], filtered_df["activity"].tolist())

    def test_filter_prepared_df_filters_by_activity_mode(self):
        raw_df = pd.DataFrame(
            [
                {"case": "C001", "step": "Submit", "ts": "2024-01-01 09:00:00", "group_a": "Sales"},
                {"case": "C001", "step": "Approve", "ts": "2024-01-02 09:00:00", "group_a": "Sales"},
                {"case": "C002", "step": "Submit", "ts": "2024-01-01 10:00:00", "group_a": "HR"},
                {"case": "C002", "step": "Reject", "ts": "2024-01-02 10:00:00", "group_a": "HR"},
            ]
        )
        prepared_df = prepare_event_log(
            df=raw_df,
            case_id_column="case",
            activity_column="step",
            timestamp_column="ts",
        )

        included_df = filter_prepared_df(
            prepared_df,
            {
                "activity_mode": "include",
                "activity_values": "Submit",
            },
            {},
        )
        excluded_df = filter_prepared_df(
            prepared_df,
            {
                "activity_mode": "exclude",
                "activity_values": "Submit",
            },
            {},
        )

        self.assertEqual(["Submit", "Submit"], included_df["activity"].tolist())
        self.assertEqual(["Approve", "Reject"], excluded_df["activity"].tolist())

    def test_filter_prepared_df_filters_by_start_and_end_activity(self):
        raw_df = pd.DataFrame(
            [
                {"case": "C001", "step": "Submit", "ts": "2024-01-01 09:00:00"},
                {"case": "C001", "step": "Approve", "ts": "2024-01-01 09:05:00"},
                {"case": "C001", "step": "Done", "ts": "2024-01-01 09:10:00"},
                {"case": "C002", "step": "Intake", "ts": "2024-01-01 10:00:00"},
                {"case": "C002", "step": "Approve", "ts": "2024-01-01 10:05:00"},
                {"case": "C002", "step": "Done", "ts": "2024-01-01 10:10:00"},
                {"case": "C003", "step": "Submit", "ts": "2024-01-01 11:00:00"},
                {"case": "C003", "step": "Check", "ts": "2024-01-01 11:05:00"},
                {"case": "C003", "step": "Reject", "ts": "2024-01-01 11:10:00"},
            ]
        )
        prepared_df = prepare_event_log(
            df=raw_df,
            case_id_column="case",
            activity_column="step",
            timestamp_column="ts",
        )

        filtered_df = filter_prepared_df(
            prepared_df,
            {
                "start_activity_values": "Submit",
                "end_activity_values": "Done",
            },
            {},
        )

        self.assertEqual({"C001"}, set(filtered_df["case_id"].unique()))
        self.assertEqual(["Submit", "Approve", "Done"], filtered_df["activity"].tolist())

    def test_get_filter_options_returns_sorted_unique_values(self):
        raw_df = pd.DataFrame(
            [
                {"case": "C001", "step": "Submit", "ts": "2024-01-01 09:00:00", "group_a": "Sales", "group_b": "Web", "group_c": "A"},
                {"case": "C002", "step": "Review", "ts": "2024-01-01 10:00:00", "group_a": "HR", "group_b": "Mail", "group_c": "B"},
                {"case": "C003", "step": "Approve", "ts": "2024-01-01 11:00:00", "group_a": "Sales", "group_b": "API", "group_c": "A"},
            ]
        )
        prepared_df = prepare_event_log(
            df=raw_df,
            case_id_column="case",
            activity_column="step",
            timestamp_column="ts",
        )

        options = get_filter_options(
            prepared_df,
            {
                "filter_column_1": "group_a",
                "filter_column_2": "group_b",
                "filter_column_3": "group_c",
            },
        )

        self.assertEqual(["HR", "Sales"], options["filters"][0]["options"])
        self.assertEqual(["API", "Mail", "Web"], options["filters"][1]["options"])
        self.assertEqual(["A", "B"], options["filters"][2]["options"])
        self.assertEqual(["Approve", "Review", "Submit"], options["all_activity_names"])

    def test_normalize_filter_column_settings_accepts_stored_slot_shape(self):
        normalized = normalize_filter_column_settings(
            filter_value_1={"column_name": "group_a", "label": "分類1"},
            filter_value_2={"column_name": "group_b", "label": "分類2"},
        )

        self.assertEqual("group_a", normalized["filter_value_1"]["column_name"])
        self.assertEqual("分類1", normalized["filter_value_1"]["label"])
        self.assertEqual("group_b", normalized["filter_value_2"]["column_name"])
        self.assertEqual("分類2", normalized["filter_value_2"]["label"])
        self.assertIsNone(normalized["filter_value_3"]["column_name"])


class GroupingTestCase(unittest.TestCase):
    """グルーピングモード機能のテスト"""

    @classmethod
    def setUpClass(cls):
        # グルーピングテスト用の合成データを直接作成する
        cls.prepared_df = pd.DataFrame({
            "case_id": ["C1", "C1", "C1", "C2", "C2", "C3", "C3", "C4", "C4"],
            "activity": ["受付", "確認", "完了", "受付", "確認", "受付", "完了", "受付", "確認"],
            "sequence_no": [1, 2, 3, 1, 2, 1, 2, 1, 2],
            "start_time": pd.to_datetime([
                "2024-01-01 09:00", "2024-01-01 09:10", "2024-01-01 09:30",
                "2024-01-02 09:00", "2024-01-02 09:15",
                "2024-01-03 09:00", "2024-01-03 09:20",
                "2024-01-04 09:00", "2024-01-04 09:05",
            ]),
            "next_time": pd.to_datetime([
                "2024-01-01 09:10", "2024-01-01 09:30", "2024-01-01 09:30",
                "2024-01-02 09:15", "2024-01-02 09:15",
                "2024-01-03 09:20", "2024-01-03 09:20",
                "2024-01-04 09:05", "2024-01-04 09:05",
            ]),
            "timestamp": pd.to_datetime([
                "2024-01-01 09:00", "2024-01-01 09:10", "2024-01-01 09:30",
                "2024-01-02 09:00", "2024-01-02 09:15",
                "2024-01-03 09:00", "2024-01-03 09:20",
                "2024-01-04 09:00", "2024-01-04 09:05",
            ]),
            "duration_min": [10.0, 20.0, 0.0, 15.0, 0.0, 20.0, 0.0, 5.0, 0.0],
            "duration_sec": [600, 1200, 0, 900, 0, 1200, 0, 300, 0],
            "group_a": ["Sales", "Sales", "Sales", "Sales", "Sales", "Mfg", "Mfg", "Mfg", "Mfg"],
            "group_b": ["Tokyo", "Tokyo", "Tokyo", "Osaka", "Osaka", "Tokyo", "Tokyo", "Osaka", "Osaka"],
        })

    # ── detect_group_columns ────────────────────────────────────────────────

    def test_detect_group_columns_returns_col_without_value(self):
        result = detect_group_columns(
            {"filter_value_1": None, "filter_value_2": None},
            {"filter_column_1": "group_a", "filter_column_2": "group_b"},
        )
        self.assertEqual(["group_a", "group_b"], result)

    def test_detect_group_columns_excludes_col_with_value(self):
        result = detect_group_columns(
            {"filter_value_1": "Sales", "filter_value_2": None},
            {"filter_column_1": "group_a", "filter_column_2": "group_b"},
        )
        self.assertEqual(["group_b"], result)

    def test_detect_group_columns_returns_empty_when_no_col_set(self):
        result = detect_group_columns({}, {})
        self.assertEqual([], result)

    # ── create_frequency_analysis (group_columns) ──────────────────────────

    def test_frequency_analysis_group_columns_adds_group_col(self):
        result = create_frequency_analysis(self.prepared_df, group_columns=["group_a"])
        self.assertIn("group_a", result.columns)
        self.assertIn("activity", result.columns)
        # グループ列の値が2種類あること
        self.assertEqual({"Sales", "Mfg"}, set(result["group_a"].tolist()))

    def test_frequency_analysis_no_group_unchanged(self):
        result = create_frequency_analysis(self.prepared_df)
        self.assertNotIn("group_a", result.columns)

    # ── create_transition_analysis (group_columns) ─────────────────────────

    def test_transition_analysis_group_columns_adds_group_col(self):
        result = create_transition_analysis(self.prepared_df, group_columns=["group_a"])
        self.assertIn("group_a", result.columns)
        self.assertIn("from_activity", result.columns)
        self.assertIn("to_activity", result.columns)

    def test_transition_analysis_no_group_unchanged(self):
        result = create_transition_analysis(self.prepared_df)
        self.assertNotIn("group_a", result.columns)

    # ── create_pattern_analysis (group_columns) ────────────────────────────

    def test_pattern_analysis_group_columns_adds_group_col(self):
        result = create_pattern_analysis(self.prepared_df, group_columns=["group_a"])
        self.assertIn("group_a", result.columns)
        self.assertIn("pattern", result.columns)

    def test_pattern_analysis_no_group_unchanged(self):
        result = create_pattern_analysis(self.prepared_df)
        self.assertNotIn("group_a", result.columns)

    # ── build_group_summary ────────────────────────────────────────────────

    def test_build_group_summary_single_column(self):
        summary = build_group_summary(self.prepared_df, ["group_a"])
        self.assertIn("__meta__", summary)
        self.assertIn("group_a", summary)
        col_data = summary["group_a"]
        self.assertIn("Sales", col_data)
        self.assertIn("Mfg", col_data)
        self.assertEqual(4, summary["__meta__"]["total_case_count"])
        self.assertEqual(9, summary["__meta__"]["total_event_count"])
        self.assertEqual(17.5, summary["__meta__"]["avg_duration_min"])
        self.assertEqual(17.5, summary["__meta__"]["median_duration_min"])
        self.assertEqual(30.0, summary["__meta__"]["max_duration_min"])
        self.assertEqual(70.0, summary["__meta__"]["total_duration_min"])
        self.assertEqual(2, col_data["Sales"]["case_count"])
        self.assertEqual(5, col_data["Sales"]["event_count"])
        self.assertEqual(50.0, col_data["Sales"]["case_ratio_pct"])
        self.assertEqual(55.56, col_data["Sales"]["event_ratio_pct"])
        self.assertEqual(22.5, col_data["Sales"]["avg_duration_min"])
        self.assertEqual(22.5, col_data["Sales"]["median_duration_min"])
        self.assertEqual(30.0, col_data["Sales"]["max_duration_min"])
        self.assertEqual(45.0, col_data["Sales"]["total_duration_min"])
        for entry in col_data.values():
            self.assertIn("case_count", entry)
            self.assertIn("case_ratio_pct", entry)
            self.assertIn("event_count", entry)
            self.assertIn("event_ratio_pct", entry)
            self.assertIn("median_duration_min", entry)
            self.assertIn("max_duration_min", entry)
            self.assertIn("total_duration_min", entry)
            self.assertGreater(entry["case_count"], 0)
            self.assertGreater(entry["event_count"], 0)

    def test_build_group_summary_multiple_columns(self):
        summary = build_group_summary(self.prepared_df, ["group_a", "group_b"])
        self.assertIn("__meta__", summary)
        self.assertIn("group_a", summary)
        self.assertIn("group_b", summary)
        self.assertIn("Tokyo", summary["group_b"])
        for col in ("group_a", "group_b"):
            self.assertAlmostEqual(100.0, sum(entry["case_ratio_pct"] for entry in summary[col].values()), places=2)
            self.assertAlmostEqual(100.0, sum(entry["event_ratio_pct"] for entry in summary[col].values()), places=2)

    def test_build_group_summary_empty_group_columns(self):
        summary = build_group_summary(self.prepared_df, [])
        self.assertEqual({}, summary)

    def test_build_group_summary_missing_column_skipped(self):
        summary = build_group_summary(self.prepared_df, ["nonexistent_col"])
        self.assertEqual({}, summary)

    def test_build_summary_sheet_df_adds_total_row_and_expands_columns(self):
        summary = build_group_summary(self.prepared_df, ["group_a"])

        summary_df = _build_summary_sheet_df(summary, ["group_a"])

        self.assertEqual(
            [
                "グルーピング軸",
                "値",
                "ケース数",
                "ケース比率(%)",
                "イベント数",
                "イベント比率(%)",
                "平均処理時間(分)",
                "中央値処理時間(分)",
                "最大処理時間(分)",
                "合計処理時間(分)",
            ],
            summary_df.columns.tolist(),
        )
        self.assertEqual(3, len(summary_df))
        self.assertEqual("全体", summary_df.iloc[0]["値"])
        self.assertEqual("", summary_df.iloc[0]["グルーピング軸"])
        self.assertEqual(100.0, summary_df.iloc[0]["ケース比率(%)"])
        self.assertEqual(100.0, summary_df.iloc[0]["イベント比率(%)"])
        self.assertEqual(17.5, summary_df.iloc[0]["平均処理時間(分)"])
        self.assertEqual(["group_a", "group_a"], summary_df["グルーピング軸"].iloc[1:].tolist())
        self.assertTrue(summary_df["ケース数"].iloc[1:].is_monotonic_decreasing)

    def test_build_summary_sheet_df_keeps_group_suffix_for_multiple_axes(self):
        summary = build_group_summary(self.prepared_df, ["group_a", "group_b"])

        summary_df = _build_summary_sheet_df(summary, ["group_a", "group_b"])

        axis_labels = set(summary_df["グルーピング軸"].iloc[1:].tolist())
        self.assertIn("group_a（グループ①）", axis_labels)
        self.assertIn("group_b（グループ②）", axis_labels)

    def test_build_excel_bytes_writes_summary_sheet_for_grouping_mode(self):
        summary = build_group_summary(self.prepared_df, ["group_a"])
        detail_df = create_frequency_analysis(self.prepared_df, group_columns=["group_a"])

        excel_bytes = build_excel_bytes(
            detail_df,
            "頻度分析",
            group_columns=["group_a"],
            group_summary=summary,
        )

        workbook = load_workbook(BytesIO(excel_bytes))
        self.assertIn("サマリー", workbook.sheetnames)
        summary_sheet = workbook["サマリー"]
        headers = [summary_sheet.cell(row=1, column=column_index).value for column_index in range(1, 11)]
        self.assertEqual(
            [
                "グルーピング軸",
                "値",
                "ケース数",
                "ケース比率(%)",
                "イベント数",
                "イベント比率(%)",
                "平均処理時間(分)",
                "中央値処理時間(分)",
                "最大処理時間(分)",
                "合計処理時間(分)",
            ],
            headers,
        )
        self.assertEqual("全体", summary_sheet["B2"].value)
        self.assertEqual(100.0, summary_sheet["D2"].value)
        self.assertEqual("group_a", summary_sheet["A3"].value)

    def test_build_excel_bytes_omits_summary_sheet_without_grouping(self):
        detail_df = create_frequency_analysis(self.prepared_df)

        excel_bytes = build_excel_bytes(detail_df, "頻度分析")

        workbook = load_workbook(BytesIO(excel_bytes))
        self.assertEqual(["頻度分析"], workbook.sheetnames)


if __name__ == "__main__":
    unittest.main()
