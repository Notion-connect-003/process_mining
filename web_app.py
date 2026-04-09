from collections import OrderedDict
from io import BytesIO
from datetime import datetime, timezone
import inspect
import json
from pathlib import Path
import shutil
import unicodedata
from urllib.parse import quote
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import duckdb
import httpx
import pandas as pd
import uvicorn
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from 共通スクリプト.Excel出力.excel_exporter import (
    build_excel_bytes,
    build_summary_sheet_df,
    convert_analysis_result_to_records,
)
from 共通スクリプト.analysis_service import (
    DEFAULT_ANALYSIS_KEYS,
    FLOW_PATTERN_CASE_COUNT_COLUMN,
    FLOW_PATTERN_COLUMN,
    analyze_prepared_event_log,
    build_group_summary,
    detect_group_columns,
    create_analysis_records,
    create_bottleneck_summary,
    create_dashboard_summary,
    create_impact_summary,
    create_log_diagnostics,
    create_rule_based_insights,
    create_root_cause_summary,
    filter_prepared_df,
    filter_prepared_df_by_pattern,
    get_filter_options,
    create_variant_summary,
    create_pattern_flow_snapshot,
    create_pattern_bottleneck_details,
    get_available_analysis_definitions,
    load_prepared_event_log,
    merge_filter_params,
    normalize_filter_params,
    normalize_filter_column_settings,
    select_pattern_rows_for_flow,
)
from 共通スクリプト.duckdb_service import (
    _build_scoped_relation_cte,
    _format_stddev_column,
    _get_parquet_column_names,
    _quote_identifier,
    persist_prepared_parquet,
    query_activity_case_drilldown,
    query_analysis_records,
    query_bottleneck_summary,
    query_case_trace_details,
    query_dashboard_summary,
    query_filter_options,
    query_filtered_meta,
    query_group_summary,
    query_impact_summary,
    query_pattern_bottleneck_details,
    query_period_text,
    query_root_cause_summary,
    query_transition_records_for_patterns,
    query_transition_case_drilldown,
    query_variant_summary,
)


BASE_DIR = Path(__file__).resolve().parent
SAMPLE_FILE = BASE_DIR / "process_mining_sample_10000行.csv"
RUN_STORAGE_DIR = BASE_DIR / "storage" / "runs"
MAX_STORED_RUNS = 5
PREVIEW_ROW_COUNT = 10
DEFAULT_LOG_DIAGNOSTIC_SAMPLE_ROW_LIMIT = 3000
MAX_LOG_DIAGNOSTIC_SAMPLE_ROW_LIMIT = 50000
PROCESS_FLOW_PATTERN_CAP = 300
MAX_PATTERN_FLOW_CACHE = 24
LARGE_DATASET_FLOW_FAST_PATH_THRESHOLD = 1_000_000
FILTER_PARAM_NAMES = (
    "date_from",
    "date_to",
    "filter_value_1",
    "filter_value_2",
    "filter_value_3",
    "activity_mode",
    "activity_values",
)
FILTER_COLUMN_NAMES = ("filter_column_1", "filter_column_2", "filter_column_3")
FILTER_LABEL_NAMES = ("filter_label_1", "filter_label_2", "filter_label_3")
REPORT_SHEET_NAMES = {
    "summary": "サマリー",
    "ai_insights": "分析コメント",
    "pattern_conclusion": "結論サマリー",
    "pattern_dashboard": "サマリーダッシュボード",
    "frequency": "頻度分析",
    "transition": "前後処理分析",
    "pattern": "処理順パターン分析",
    "pattern_summary": "パターンサマリー",
    "variant": "バリアント分析",
    "bottleneck": "ボトルネック分析",
    "impact": "改善インパクト分析",
    "drilldown": "ドリルダウン",
    "case_trace": "ケース追跡",
}
LOG_DIAGNOSTIC_SHEET_NAMES = {
    "summary": "ログ診断",
    "sample": "ログサンプル",
}
REPORT_HEADER_LABELS = {
    "run_id": "実行ID",
    "analysis_key": "分析種別",
    "analysis_name": "分析名",
    "source_file_name": "元ファイル名",
    "analysis_executed_at": "分析実行日時",
    "exported_at": "出力日時",
    "case_count": "対象ケース数",
    "event_count": "対象イベント数",
    "applied_filters": "適用条件",
    "selected_variant": "選択中バリアント",
    "selected_activity": "選択中アクティビティ",
    "selected_transition": "選択中遷移",
    "selected_case_id": "選択中ケースID",
    "rank": "順位",
    "pattern_variant": "パターン / バリアント",
    "repeat_flag": "繰り返し",
    "repeat_count": "繰り返し回数",
    "repeat_rate_pct": "繰り返し率(%)",
    "repeat_rate_band": "繰り返し率区分",
    "review_flag": "確認区分",
    "avg_case_duration_diff_min": "平均処理時間差分(分)",
    "improvement_priority_score": "改善優先度スコア",
    "overall_impact_pct": "全体影響度(%)",
    "fastest_pattern_flag": "最短処理",
    "simple_comment": "簡易コメント",
    "variant_id": "バリアントID",
    "count": "件数",
    "case_count": "対象ケース数",
    "ratio": "比率",
    "cumulative_case_ratio_pct": "累積カバー率(%)",
    "pattern": "パターン",
    "activity_count": "アクティビティ数",
    "avg_case_duration": "平均ケース処理時間",
    "avg_case_duration_min": "平均ケース処理時間(分)",
    "std_case_duration_min": "標準偏差ケース処理時間(分)",
    "min_case_duration_min": "最短処理時間(分)",
    "max_case_duration_min": "最長処理時間(分)",
    "p75_case_duration_min": "75%点ケース処理時間(分)",
    "p90_case_duration_min": "90%点ケース処理時間(分)",
    "p95_case_duration_min": "95%点ケース処理時間(分)",
    "avg_duration": "平均所要時間",
    "avg_duration_text": "平均所要時間",
    "avg_duration_min": "平均所要時間(分)",
    "median_duration_min": "中央値所要時間(分)",
    "std_duration_min": "標準偏差(分)",
    "min_duration_min": "最小所要時間(分)",
    "median_duration_text": "中央値所要時間",
    "max_duration": "最大所要時間",
    "max_duration_text": "最大所要時間",
    "max_duration_min": "最大所要時間(分)",
    "total_duration_min": "合計所要時間(分)",
    "p75_duration_min": "75%点(分)",
    "p90_duration_min": "90%点(分)",
    "p95_duration_min": "95%点(分)",
    "impact_score": "改善インパクト",
    "impact_share_pct": "改善インパクト比率(%)",
    "wait_share_pct": "構成比(%)",
    "case_id": "ケースID",
    "from_time": "開始時刻",
    "to_time": "終了時刻",
    "activity": "アクティビティ",
    "next_activity": "次アクティビティ",
    "sequence_no": "ステップ順",
    "transition": "遷移",
    "transition_label": "遷移",
    "duration_text": "所要時間",
    "total_duration": "総処理時間",
    "start_time": "開始時刻",
    "end_time": "終了時刻",
}
APPLIED_FILTERS_NOTE_TEXT = "\n".join(
    [
        "※ 適用条件の種類:",
        "  • 期間フィルター: 開始日 / 終了日",
        "  • グループ/カテゴリーフィルター①②③: CSVの任意カラムで絞り込み（例: 部署=営業部）",
        "  • アクティビティフィルター: 特定アクティビティを含む/除外",
    ]
)
GROUPING_CONDITION_NOTE_TEXT = (
    "※ カラムを選択し値を未選択にすると、そのカラムがグルーピング軸（比較用）になります"
)
ANALYSIS_PRECONDITIONS_TEXT = "\n".join(
    [
        "• 処理時間は、同一ケース内で当該アクティビティの開始時刻から次のアクティビティの開始時刻までの差分として算出しています。",
        "• ケース内の最終アクティビティは、次のイベントが存在しないため処理時間が0分となります。",
        "• 処理時間が0分のイベントも集計対象に含まれています。統計値（平均・中央値等）に影響する点にご留意ください。",
        "• 本分析はフィルター適用後のデータに基づいています。適用条件の詳細はサマリーシートをご参照ください。",
    ]
)
TERMINOLOGY_ROWS = [
    {
        "用語": "ケース",
        "説明": "業務プロセスの1つの実行単位（例: 1件の注文、1件の申請）",
    },
    {
        "用語": "アクティビティ",
        "説明": "ケース内で実行される個々の作業ステップ（例: 申請、承認、支払）",
    },
    {
        "用語": "イベント",
        "説明": "特定のケースで特定のアクティビティが実行された1回の記録",
    },
    {
        "用語": "処理時間",
        "説明": "あるアクティビティの開始から次のアクティビティの開始までの所要時間",
    },
    {
        "用語": "イベント比率(%)",
        "説明": "全イベント数に対する当該アクティビティのイベント数の割合",
    },
]
EXCEL_TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
EXCEL_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
EXCEL_SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="EFF5FB")
EXCEL_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E7F6")
EXCEL_GROUP_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
EXCEL_ASSUMPTION_SECTION_FILL = PatternFill(fill_type="solid", fgColor="E8EDF2")
EXCEL_MUTED_SECTION_FILL = PatternFill(fill_type="solid", fgColor="F0F0F0")
EXCEL_HEADER_FILL = PatternFill(fill_type="solid", fgColor="EDF2F7")
EXCEL_LABEL_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
EXCEL_ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="FBFDFF")
EXCEL_TEXT_BLOCK_FILL = PatternFill(fill_type="solid", fgColor="F7FAFE")
EXCEL_TITLE_BORDER = Border(
    left=Side(style="thin", color="1F4E78"),
    right=Side(style="thin", color="1F4E78"),
    top=Side(style="thin", color="1F4E78"),
    bottom=Side(style="thin", color="1F4E78"),
)
EXCEL_THIN_BORDER = Border(
    left=Side(style="thin", color="D6DEE8"),
    right=Side(style="thin", color="D6DEE8"),
    top=Side(style="thin", color="D6DEE8"),
    bottom=Side(style="thin", color="D6DEE8"),
)
EXCEL_MUTED_FONT = Font(size=10, color="5B6B82")
EXCEL_NOTE_FONT = Font(size=9, color="5B6B82")
EXCEL_BODY_FONT = Font(size=10, color="1F2937")
EXCEL_BOLD_FONT = Font(bold=True, size=10, color="1F2937")
EXCEL_GROUP_SECTION_FONT = Font(bold=True, size=12, color="1F2937")

DEFAULT_HEADERS = {
    "case_id_column": "case_id",
    "activity_column": "activity",
    "timestamp_column": "timestamp",
}
CASE_ID_COLUMN_CANDIDATES = [
    "case_id",
    "Case ID",
    "CASE_ID",
    "caseId",
    "CaseID",
    "caseid",
    "ケースID",
    "案件ID",
    "case_no",
    "CaseNo",
    "case_number",
    "CaseNumber",
]
ACTIVITY_COLUMN_CANDIDATES = [
    "activity",
    "Activity",
    "ACTIVITY",
    "activity_name",
    "ActivityName",
    "task",
    "Task",
    "event",
    "Event",
    "action",
    "Action",
    "アクティビティ",
    "アクティビティ名",
    "処理",
    "工程",
    "ステップ",
]
TIMESTAMP_COLUMN_CANDIDATES = [
    "timestamp",
    "Timestamp",
    "TIMESTAMP",
    "start_time",
    "StartTime",
    "start_timestamp",
    "日時",
    "タイムスタンプ",
    "開始日時",
    "event_timestamp",
    "EventTimestamp",
    "time",
    "Time",
    "datetime",
    "DateTime",
    "date",
    "Date",
]
COLUMN_CANDIDATES = {
    "case_id_column": CASE_ID_COLUMN_CANDIDATES,
    "activity_column": ACTIVITY_COLUMN_CANDIDATES,
    "timestamp_column": TIMESTAMP_COLUMN_CANDIDATES,
}
COLUMN_DISPLAY_LABELS = {
    "case_id_column": "ケースID",
    "activity_column": "アクティビティ",
    "timestamp_column": "タイムスタンプ",
}

app = FastAPI(title="Process Mining Workbench")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Starlette <0.41: TemplateResponse(name, context)  context must have "request" key
# Starlette >=0.41: TemplateResponse(request, name, context)
_STARLETTE_OLD_TEMPLATE_API = (
    list(inspect.signature(templates.TemplateResponse).parameters.keys())[0] == "name"
)


def _template_response(request: Request, name: str, context: dict):
    ctx = {"request": request, **context}
    if _STARLETTE_OLD_TEMPLATE_API:
        return templates.TemplateResponse(name, ctx)
    return templates.TemplateResponse(request, name, ctx)


RUN_STORE = OrderedDict()


def _normalize_header_name(value):
    return str(value or "").strip()


def _build_header_lookup(headers):
    exact_lookup = {}
    casefold_lookup = {}

    for header in headers:
        normalized_header = _normalize_header_name(header)
        if not normalized_header:
            continue
        exact_lookup.setdefault(normalized_header, normalized_header)
        casefold_lookup.setdefault(normalized_header.casefold(), normalized_header)

    return exact_lookup, casefold_lookup


def suggest_column_name(headers, field_name, preferred_name=""):
    exact_lookup, casefold_lookup = _build_header_lookup(headers)
    requested_name = _normalize_header_name(preferred_name)

    if requested_name:
        exact_match = exact_lookup.get(requested_name)
        if exact_match:
            return exact_match

        casefold_match = casefold_lookup.get(requested_name.casefold())
        if casefold_match:
            return casefold_match

    for candidate_name in COLUMN_CANDIDATES.get(field_name, []):
        normalized_candidate = _normalize_header_name(candidate_name)
        if not normalized_candidate:
            continue

        exact_match = exact_lookup.get(normalized_candidate)
        if exact_match:
            return exact_match

        casefold_match = casefold_lookup.get(normalized_candidate.casefold())
        if casefold_match:
            return casefold_match

    return ""


def resolve_required_column_name(headers, field_name, preferred_name=""):
    resolved_name = suggest_column_name(headers, field_name, preferred_name)
    if resolved_name:
        return resolved_name

    requested_name = (
        _normalize_header_name(preferred_name)
        or DEFAULT_HEADERS.get(field_name)
        or (COLUMN_CANDIDATES.get(field_name) or [""])[0]
    )
    available_headers = ", ".join(
        normalized_header
        for normalized_header in (_normalize_header_name(header) for header in headers)
        if normalized_header
    ) or "(なし)"
    field_label = COLUMN_DISPLAY_LABELS.get(field_name, field_name)
    raise ValueError(
        f"{field_label}列 '{requested_name}' が見つかりません。"
        f"CSVのヘッダーを確認してください。利用可能な列: {available_headers}"
    )


def build_column_selection_payload(headers):
    return {
        "headers": headers,
        "default_selection": {
            field_name: suggest_column_name(
                headers,
                field_name,
                default_header,
            )
            for field_name, default_header in DEFAULT_HEADERS.items()
        },
    }


def validate_selected_columns(case_id_column, activity_column, timestamp_column):
    selected_columns = {
        "ケースID": case_id_column,
        "アクティビティ": activity_column,
        "タイムスタンプ": timestamp_column,
    }

    missing_fields = [
        field_label
        for field_label, column_name in selected_columns.items()
        if not str(column_name or "").strip()
    ]
    if missing_fields:
        raise ValueError(f"次の列を選択してください: {' / '.join(missing_fields)}")

    normalized_columns = [column_name.strip() for column_name in selected_columns.values()]
    if len(set(normalized_columns)) != len(normalized_columns):
        raise ValueError("ケースID列 / アクティビティ列 / タイムスタンプ列にはそれぞれ異なる列を選択してください。")


def validate_filter_column_settings(filter_column_settings):
    selected_filter_columns = [
        filter_config["column_name"]
        for filter_config in filter_column_settings.values()
        if filter_config["column_name"]
    ]

    if len(selected_filter_columns) != len(set(selected_filter_columns)):
        raise ValueError("グループ/カテゴリー フィルター①〜③ にはそれぞれ異なる列を選択してください。")


def read_raw_log_dataframe(file_source):
    if hasattr(file_source, "seek"):
        file_source.seek(0)

    try:
        raw_df = pd.read_csv(file_source, dtype=str, keep_default_na=False)
    finally:
        if hasattr(file_source, "seek"):
            file_source.seek(0)

    return raw_df


def resolve_profile_file_source(form):
    uploaded_file = form.get("csv_file")

    if uploaded_file and uploaded_file.filename:
        uploaded_file.file.seek(0)
        return uploaded_file.file, uploaded_file.filename

    return SAMPLE_FILE, SAMPLE_FILE.name


def get_static_version():
    static_dir = BASE_DIR / "static"
    return str(
        max(
            entry.stat().st_mtime_ns
            for entry in static_dir.iterdir()
            if entry.is_file()
        )
    )


def get_run_storage_dir(run_id):
    return RUN_STORAGE_DIR / str(run_id)


def get_run_prepared_parquet_path(run_id):
    return get_run_storage_dir(run_id) / "prepared.parquet"


def cleanup_run_storage(run_id):
    run_storage_dir = get_run_storage_dir(run_id)
    if run_storage_dir.exists():
        shutil.rmtree(run_storage_dir, ignore_errors=True)


def save_run_data(
    source_file_name,
    selected_analysis_keys,
    prepared_df,
    result,
    column_settings,
    base_filter_params,
):
    run_id = uuid4().hex
    prepared_row_count = int(len(prepared_df))
    pattern_rows = ((result or {}).get("analyses", {}).get("pattern") or {}).get("rows", [])
    filter_options = get_filter_options(
        prepared_df,
        filter_column_settings=column_settings,
    )
    prepared_parquet_path = get_run_prepared_parquet_path(run_id)
    persist_prepared_parquet(prepared_df, prepared_parquet_path)
    RUN_STORE[run_id] = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "source_file_name": source_file_name,
        "selected_analysis_keys": selected_analysis_keys,
        "prepared_df": None,
        "prepared_row_count": prepared_row_count,
        "prepared_parquet_path": str(prepared_parquet_path),
        "result": result,
        "column_settings": column_settings,
        "base_filter_params": base_filter_params,
        "pattern_index_entries": build_pattern_index_entries_from_rows(pattern_rows),
        "pattern_flow_cache": OrderedDict(),
        "variant_cache": {},
        "bottleneck_cache": {},
        "dashboard_cache": {},
        "root_cause_cache": {},
        "impact_cache": {},
        "insights_cache": {},
        "ai_insights_cache": {},
        "analysis_cache": {},
        "filter_options": filter_options,
    }
    RUN_STORE.move_to_end(run_id)

    while len(RUN_STORE) > MAX_STORED_RUNS:
        removed_run_id, _ = RUN_STORE.popitem(last=False)
        cleanup_run_storage(removed_run_id)

    return run_id


def get_run_data(run_id):
    run_data = RUN_STORE.get(run_id)

    if not run_data:
        raise HTTPException(status_code=404, detail="実行データが見つかりません。")

    RUN_STORE.move_to_end(run_id)
    return run_data


def get_run_group_columns(run_data):
    return list(
        (run_data.get("result") or {}).get("group_columns")
        or run_data.get("group_columns")
        or []
    )


def get_run_variant_pattern(run_data, variant_id=None, pattern_index=None, filter_params=None):
    if pattern_index is not None:
        _, _, _, pattern = get_pattern_summary_row(run_data, pattern_index)
        return pattern

    if variant_id is not None:
        variant_item = get_variant_item(run_data, variant_id, filter_params=filter_params)
        return variant_item["pattern"]

    return None


def get_request_filter_params(request: Request):
    return normalize_filter_params(
        **{
            filter_name: request.query_params.get(filter_name)
            for filter_name in FILTER_PARAM_NAMES
        }
    )


def get_form_filter_params(form):
    return normalize_filter_params(
        **{
            filter_name: form.get(filter_name)
            for filter_name in FILTER_PARAM_NAMES
        }
    )


def get_form_filter_column_settings(form):
    raw_settings = {
        setting_name: form.get(setting_name)
        for setting_name in (*FILTER_COLUMN_NAMES, *FILTER_LABEL_NAMES)
    }
    return normalize_filter_column_settings(**raw_settings)


def get_effective_filter_params(run_data, filter_params=None):
    return merge_filter_params(run_data.get("base_filter_params"), filter_params)


def build_filter_cache_key(filter_params):
    normalized_filters = normalize_filter_params(**(filter_params or {}))
    return tuple(
        normalized_filters.get(filter_name)
        for filter_name in FILTER_PARAM_NAMES
    )


def get_filtered_meta_for_run(run_data, filter_params=None, variant_pattern=None):
    return query_filtered_meta(
        run_data["prepared_parquet_path"],
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    )


def build_column_settings_payload(column_settings):
    raw_column_settings = column_settings or {}
    filter_slot_names = ("filter_value_1", "filter_value_2", "filter_value_3")

    if any(isinstance(raw_column_settings.get(filter_slot_name), dict) for filter_slot_name in filter_slot_names):
        default_filter_settings = normalize_filter_column_settings()
        normalized_filter_settings = {
            filter_slot_name: {
                "column_name": str((raw_column_settings.get(filter_slot_name) or {}).get("column_name") or "").strip() or None,
                "label": str((raw_column_settings.get(filter_slot_name) or {}).get("label") or "").strip()
                or default_filter_settings[filter_slot_name]["label"],
            }
            for filter_slot_name in filter_slot_names
        }
    else:
        normalized_filter_settings = normalize_filter_column_settings(**raw_column_settings)

    return {
        "case_id_column": str(raw_column_settings.get("case_id_column") or "").strip(),
        "activity_column": str(raw_column_settings.get("activity_column") or "").strip(),
        "timestamp_column": str(raw_column_settings.get("timestamp_column") or "").strip(),
        "filters": [
            {
                "slot": filter_key,
                **normalized_filter_settings[filter_key],
            }
            for filter_key in normalized_filter_settings
        ],
    }


def build_analysis_payload(analysis, row_limit=None, row_offset=0):
    total_row_count = len(analysis["rows"])
    safe_row_offset = max(0, int(row_offset or 0))

    if row_limit is None:
        safe_row_limit = total_row_count
        rows = analysis["rows"][safe_row_offset:]
    else:
        safe_row_limit = max(0, int(row_limit))
        rows = analysis["rows"][safe_row_offset : safe_row_offset + safe_row_limit]

    page_end_row_number = safe_row_offset + len(rows)
    has_previous_page = safe_row_offset > 0
    has_next_page = page_end_row_number < total_row_count
    previous_row_offset = max(0, safe_row_offset - safe_row_limit) if has_previous_page else None
    next_row_offset = page_end_row_number if has_next_page else None

    return {
        "analysis_name": analysis["analysis_name"],
        "sheet_name": analysis["sheet_name"],
        "output_file_name": analysis.get("output_file_name"),
        "row_count": total_row_count,
        "returned_row_count": len(rows),
        "row_offset": safe_row_offset,
        "page_size": safe_row_limit,
        "page_start_row_number": safe_row_offset + 1 if rows else 0,
        "page_end_row_number": page_end_row_number,
        "has_previous_page": has_previous_page,
        "has_next_page": has_next_page,
        "previous_row_offset": previous_row_offset,
        "next_row_offset": next_row_offset,
        "rows": rows,
        "excel_file": analysis["excel_file"],
    }


def extract_pattern_text_from_row(row):
    for column_name, value in (row or {}).items():
        normalized_column_name = str(column_name or "").lower()
        normalized_value = str(value or "").strip()
        if not normalized_value:
            continue
        if "pattern" in normalized_column_name or "パターン" in str(column_name or ""):
            return normalized_value
    return ""


def build_pattern_index_entries_from_rows(pattern_rows):
    return [
        {
            "index": int(index),
            "pattern": extract_pattern_text_from_row(row),
        }
        for index, row in enumerate(pattern_rows or [])
    ]


def get_pattern_index_for_pattern(run_data, pattern_text):
    normalized_pattern = str(pattern_text or "").strip()
    if not normalized_pattern:
        return None

    pattern_analysis = run_data["result"]["analyses"].get("pattern")
    if not pattern_analysis:
        return None

    pattern_index_entries = run_data.get("pattern_index_entries")
    if pattern_index_entries is None:
        pattern_index_entries = build_pattern_index_entries_from_rows(pattern_analysis.get("rows", []))
        run_data["pattern_index_entries"] = pattern_index_entries

    for pattern_index, entry in enumerate(pattern_index_entries):
        if str(entry.get("pattern") or "").strip() == normalized_pattern:
            return pattern_index

    for pattern_index, row in enumerate(pattern_analysis.get("rows", [])):
        for column_name, value in row.items():
            normalized_column_name = str(column_name or "").lower()
            normalized_value = str(value or "").strip()
            if not normalized_value:
                continue
            if "pattern" not in normalized_column_name and "パターン" not in str(column_name or ""):
                continue
            if normalized_value == normalized_pattern:
                return pattern_index

    return None


def build_variant_response_item(variant_item, run_data=None):
    activities = variant_item["activities"]
    repeated_activity_count = len(activities) - len({str(activity or "").strip() for activity in activities if str(activity or "").strip()})
    repeat_flag = variant_item.get("repeat_flag", "○" if repeated_activity_count > 0 else "")
    return {
        "variant_id": variant_item["variant_id"],
        "activities": activities,
        "activity_count": variant_item.get("activity_count", len(activities)),
        "pattern": variant_item.get("pattern", ""),
        "pattern_index": (
            get_pattern_index_for_pattern(run_data, variant_item.get("pattern"))
            if run_data is not None
            else None
        ),
        "count": variant_item["count"],
        "ratio": variant_item["ratio"],
        "avg_case_duration_sec": variant_item.get("avg_case_duration_sec", 0.0),
        "avg_case_duration_text": variant_item.get("avg_case_duration_text", "0s"),
        "repeat_flag": repeat_flag,
    }


def build_variant_coverage_payload(total_case_count, variant_items):
    covered_case_count = sum(int(variant_item["count"]) for variant_item in variant_items)
    return {
        "displayed_variant_count": len(variant_items),
        "covered_case_count": covered_case_count,
        "total_case_count": int(total_case_count),
        "ratio": round(covered_case_count / total_case_count, 4) if total_case_count else 0.0,
    }


def sanitize_workbook_sheet_name(sheet_name):
    invalid_characters = set('[]:*?/\\')
    normalized_name = "".join("_" if character in invalid_characters else character for character in str(sheet_name or "").strip())
    normalized_name = normalized_name or "Sheet"
    return normalized_name[:31]


def sanitize_file_name_component(value):
    invalid_characters = set('<>:"/\\|?*')
    normalized_value = "".join("_" if character in invalid_characters else character for character in str(value or "").strip())
    return normalized_value.strip(" .") or "analysis"


def resolve_analysis_display_name(analysis_key, analysis_name=""):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    if normalized_analysis_key in REPORT_SHEET_NAMES:
        return REPORT_SHEET_NAMES[normalized_analysis_key]
    return str(analysis_name or normalized_analysis_key or "分析").strip() or "分析"


def build_analysis_excel_file_name(source_file_name, analysis_key, analysis_name="", suffix=""):
    source_stem = sanitize_file_name_component(Path(str(source_file_name or "analysis")).stem)
    display_name = sanitize_file_name_component(resolve_analysis_display_name(analysis_key, analysis_name))
    normalized_suffix = sanitize_file_name_component(suffix) if str(suffix or "").strip() else ""
    return f"{source_stem}_{display_name}{normalized_suffix}.xlsx"


def normalize_excel_cell_value(value):
    if value is None:
        return ""

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, (list, tuple, set)):
        return " / ".join(str(item) for item in value)

    if isinstance(value, dict):
        return ", ".join(f"{key}={normalize_excel_cell_value(item_value)}" for key, item_value in value.items())

    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass

    return value


def style_excel_cell(cell, *, font=None, fill=None, alignment=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def merge_excel_row(worksheet, row_index, column_count):
    safe_column_count = max(1, int(column_count or 1))
    if safe_column_count > 1:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=safe_column_count,
        )


def estimate_wrapped_row_height(text, column_count=1, min_height=22, max_height=240):
    safe_text = str(text or "")
    if not safe_text:
        return min_height

    approx_chars_per_line = max(24, int(max(1, column_count)) * 18)
    logical_lines = 0
    for raw_line in safe_text.splitlines() or [""]:
        line_length = max(1, len(raw_line))
        logical_lines += max(1, (line_length + approx_chars_per_line - 1) // approx_chars_per_line)

    return max(min_height, min(max_height, 16 + logical_lines * 15))


def initialize_excel_worksheet(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90


def estimate_excel_text_width(value):
    normalized_value = str(normalize_excel_cell_value(value))
    if not normalized_value:
        return 0

    max_line_width = 0
    for line in normalized_value.splitlines() or [""]:
        line_width = 0
        for character in line:
            line_width += 2 if unicodedata.east_asian_width(character) in {"F", "W", "A"} else 1
        max_line_width = max(max_line_width, line_width)

    return max_line_width


def get_autosize_ignored_cells(worksheet):
    ignored_cells = set()
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_col == merged_range.max_col and merged_range.min_row == merged_range.max_row:
            continue

        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                ignored_cells.add((row_index, column_index))

    return ignored_cells


def autosize_worksheet_columns(worksheet, min_width=12, max_width=120):
    ignored_cells = get_autosize_ignored_cells(worksheet)
    min_width_overrides = getattr(worksheet, "_codex_min_column_widths", {})

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        measured_width = min_width

        for row_index in range(1, worksheet.max_row + 1):
            if (row_index, column_index) in ignored_cells:
                continue

            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value in (None, ""):
                continue
            measured_width = max(
                measured_width,
                estimate_excel_text_width(cell_value),
            )

        worksheet.column_dimensions[column_letter].width = max(
            min_width,
            min(max_width, max(measured_width + 3, float(min_width_overrides.get(column_letter, min_width)))),
        )


def append_table_to_worksheet(
    worksheet,
    title,
    rows,
    headers,
    start_row=1,
    description="",
    no_wrap_headers=None,
    min_column_widths=None,
):
    current_row = start_row
    column_count = max(1, len(headers))
    normalized_no_wrap_headers = {
        str(header)
        for header in (no_wrap_headers or [])
    }
    merge_excel_row(worksheet, current_row, column_count)
    title_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        title_cell,
        font=EXCEL_TITLE_FONT,
        fill=EXCEL_TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_TITLE_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    if description:
        merge_excel_row(worksheet, current_row, column_count)
        description_cell = worksheet.cell(row=current_row, column=1, value=description)
        style_excel_cell(
            description_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_SUBTITLE_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(description, column_count)
        current_row += 1

    header_row = current_row
    for column_index, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=header not in normalized_no_wrap_headers),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    if not rows:
        merge_excel_row(worksheet, current_row, column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value="表示できるデータがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    data_start_row = current_row
    for row_index, row in enumerate(rows, start=0):
        for column_index, header in enumerate(headers, start=1):
            body_cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
            fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(wrap_text=header not in normalized_no_wrap_headers, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

    if min_column_widths:
        width_overrides = dict(getattr(worksheet, "_codex_min_column_widths", {}))
        for column_index, header in enumerate(headers, start=1):
            if header in min_column_widths:
                width_overrides[get_column_letter(column_index)] = max(
                    float(width_overrides.get(get_column_letter(column_index), 0)),
                    float(min_column_widths[header]),
                )
        worksheet._codex_min_column_widths = width_overrides

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(column_count)}{current_row - 1}"

    return current_row + 1


def append_key_value_rows(worksheet, title, rows, start_row=1, description=""):
    current_row = start_row
    merge_excel_row(worksheet, current_row, 2)
    title_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        title_cell,
        font=EXCEL_TITLE_FONT,
        fill=EXCEL_TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_TITLE_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    if description:
        merge_excel_row(worksheet, current_row, 2)
        description_cell = worksheet.cell(row=current_row, column=1, value=description)
        style_excel_cell(
            description_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_SUBTITLE_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(description, 2)
        current_row += 1

    for column_index, header in enumerate(("項目", "値"), start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    data_start_row = current_row
    for row_index, row in enumerate(rows, start=0):
        if isinstance(row, dict):
            label = row.get("label", "")
            value = row.get("value", "")
            row_style = row.get("style", "default")
        else:
            label, value = row
            row_style = "default"
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        label_cell = worksheet.cell(row=current_row, column=1, value=label)
        value_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(value))
        if row_style == "note":
            note_fill = EXCEL_SUBTITLE_FILL if fill is None else fill
            style_excel_cell(
                label_cell,
                font=EXCEL_NOTE_FONT,
                fill=note_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
            style_excel_cell(
                value_cell,
                font=EXCEL_NOTE_FONT,
                fill=note_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        else:
            style_excel_cell(
                label_cell,
                font=EXCEL_BOLD_FONT,
                fill=EXCEL_LABEL_FILL if fill is None else fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
            style_excel_cell(
                value_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = max(
            20,
            estimate_wrapped_row_height(value, 1, min_height=20, max_height=120),
        )
        current_row += 1

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref and rows:
        worksheet.auto_filter.ref = f"A{data_start_row - 1}:B{current_row - 1}"

    return current_row + 1


def append_bullet_rows(worksheet, title, items, start_row=1, column_count=6, empty_text="表示できる要点がありません。"):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_BOLD_FONT,
        fill=EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    if not items:
        merge_excel_row(worksheet, current_row, safe_column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value=empty_text)
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    for index, item in enumerate(items, start=1):
        bullet_cell = worksheet.cell(row=current_row, column=1, value=f"{index}.")
        style_excel_cell(
            bullet_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_ALT_ROW_FILL if index % 2 == 0 else None,
            alignment=Alignment(horizontal="center", vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        if safe_column_count > 1:
            worksheet.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=safe_column_count,
            )
        text_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(item))
        style_excel_cell(
            text_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_ALT_ROW_FILL if index % 2 == 0 else None,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(item, safe_column_count - 1)
        current_row += 1

    return current_row + 1


def append_text_block_to_worksheet(worksheet, title, text, start_row=1, column_count=6):
    current_row = start_row
    safe_column_count = max(1, int(column_count or 1))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_BOLD_FONT,
        fill=EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    merge_excel_row(worksheet, current_row, safe_column_count)
    body_cell = worksheet.cell(row=current_row, column=1, value=normalize_excel_cell_value(text) or "解説を表示できませんでした。")
    style_excel_cell(
        body_cell,
        font=EXCEL_BODY_FONT,
        fill=EXCEL_TEXT_BLOCK_FILL,
        alignment=Alignment(wrap_text=True, vertical="top"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(body_cell.value, safe_column_count, min_height=88, max_height=360)

    return current_row + 2


def append_custom_text_section_to_worksheet(
    worksheet,
    title,
    text,
    start_row=1,
    column_count=6,
    header_fill=None,
    body_fill=None,
    empty_text="表示できる内容がありません。",
):
    current_row = start_row
    safe_column_count = max(1, int(column_count or 1))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_BOLD_FONT,
        fill=header_fill or EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    merge_excel_row(worksheet, current_row, safe_column_count)
    body_cell = worksheet.cell(
        row=current_row,
        column=1,
        value=normalize_excel_cell_value(text) or empty_text,
    )
    style_excel_cell(
        body_cell,
        font=EXCEL_BODY_FONT,
        fill=body_fill or EXCEL_TEXT_BLOCK_FILL,
        alignment=Alignment(wrap_text=True, vertical="top"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(
        body_cell.value,
        safe_column_count,
        min_height=54,
        max_height=240,
    )

    return current_row + 2


def append_definition_table_to_worksheet(worksheet, title, rows, start_row=1, column_count=6, header_fill=None):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_BOLD_FONT,
        fill=header_fill or EXCEL_MUTED_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    headers = ("用語", "説明")
    for column_index, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    for row_index, row in enumerate(rows or [], start=0):
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        term_cell = worksheet.cell(row=current_row, column=1, value=normalize_excel_cell_value(row.get("用語")))
        description_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(row.get("説明")))
        style_excel_cell(
            term_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_LABEL_FILL if fill is None else fill,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        style_excel_cell(
            description_cell,
            font=EXCEL_BODY_FONT,
            fill=fill,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = max(
            20,
            estimate_wrapped_row_height(row.get("説明"), 1, min_height=20, max_height=96),
        )
        current_row += 1

    return current_row + 1


def build_ranked_rows(rows, rank_key="rank"):
    ranked_rows = []
    for index, row in enumerate(rows, start=1):
        ranked_rows.append({
            rank_key: index,
            **row,
        })
    return ranked_rows


def _normalize_group_section_value(value):
    if pd.isna(value):
        return "(未分類)"
    normalized_value = str(value).strip()
    return normalized_value or "(未分類)"


def _write_section_header(worksheet, row_index, title, column_count=10):
    safe_column_count = max(10, int(column_count or 10))
    for column_index in range(1, safe_column_count + 1):
        header_cell = worksheet.cell(row=row_index, column=column_index)
        if column_index == 1:
            header_cell.value = f"═══ {title} ═══"
            font = EXCEL_GROUP_SECTION_FONT
            alignment = Alignment(horizontal="left", vertical="center")
        else:
            font = EXCEL_GROUP_SECTION_FONT
            alignment = Alignment(horizontal="left", vertical="center")
        style_excel_cell(
            header_cell,
            font=font,
            fill=EXCEL_GROUP_SECTION_FILL,
            alignment=alignment,
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[row_index].height = 22
    return row_index + 1


def _iter_groups(prepared_df, grouping_columns):
    valid_columns = [column_name for column_name in (grouping_columns or []) if column_name in prepared_df.columns]
    if not valid_columns or prepared_df.empty:
        return

    working_df = prepared_df.copy()
    for column_name in valid_columns:
        working_df[column_name] = working_df[column_name].apply(_normalize_group_section_value)

    grouped_entries = []
    for group_key, group_df in working_df.groupby(valid_columns, dropna=False, sort=False):
        normalized_key = group_key if isinstance(group_key, tuple) else (group_key,)
        if len(valid_columns) == 1:
            group_name = str(normalized_key[0])
        else:
            group_name = ", ".join(
                f"{column_name}={value}"
                for column_name, value in zip(valid_columns, normalized_key)
            )
        case_count = int(group_df["case_id"].nunique()) if "case_id" in group_df.columns else int(len(group_df))
        grouped_entries.append((case_count, group_name, group_df.copy()))

    for _, group_name, group_df in sorted(grouped_entries, key=lambda item: (-item[0], item[1])):
        yield group_name, group_df


def _iter_groups_from_parquet(
    parquet_path,
    grouping_columns,
    filter_params=None,
    filter_column_settings=None,
    variant_pattern=None,
):
    valid_columns = [
        str(column_name)
        for column_name in (grouping_columns or [])
        if str(column_name) in set(_get_parquet_column_names(parquet_path))
    ]
    if not valid_columns:
        return

    connection = duckdb.connect()
    try:
        cte_sql, params, relation_name = _build_scoped_relation_cte(
            parquet_path,
            filter_params=filter_params,
            filter_column_settings=filter_column_settings,
            variant_pattern=variant_pattern,
        )
        case_group_columns_sql = ",\n                ".join(
            [
                f"COALESCE(MAX(NULLIF(TRIM(CAST({_quote_identifier(column_name)} AS VARCHAR)), '')), '(未分類)') AS {_quote_identifier(column_name)}"
                for column_name in valid_columns
            ]
        )
        group_select_sql = ", ".join(_quote_identifier(column_name) for column_name in valid_columns)
        group_order_sql = ", ".join(_quote_identifier(column_name) for column_name in valid_columns)
        group_list_rows = connection.execute(
            f"""
            {cte_sql},
            case_groups AS (
                SELECT
                    case_id,
                    {case_group_columns_sql}
                FROM {relation_name}
                GROUP BY case_id
            )
            SELECT
                {group_select_sql},
                COUNT(*) AS case_count
            FROM case_groups
            GROUP BY {group_select_sql}
            ORDER BY case_count DESC, {group_order_sql}
            """,
            params,
        ).fetchall()
        if not group_list_rows:
            return

        frequency_display_columns = (
            get_available_analysis_definitions()
            .get("frequency", {})
            .get("config", {})
            .get("display_columns", {})
        )

        for row in group_list_rows:
            group_values = [str(value or "").strip() or "(未分類)" for value in row[: len(valid_columns)]]
            if len(valid_columns) == 1:
                group_name = group_values[0]
            else:
                group_name = ", ".join(
                    f"{column_name}={group_value}"
                    for column_name, group_value in zip(valid_columns, group_values)
                )

            group_conditions_sql = " AND ".join(
                f"{_quote_identifier(column_name)} = ?"
                for column_name in valid_columns
            )
            group_df = connection.execute(
                f"""
                {cte_sql},
                case_groups AS (
                    SELECT
                        case_id,
                        {case_group_columns_sql}
                    FROM {relation_name}
                    GROUP BY case_id
                ),
                selected_cases AS (
                    SELECT case_id
                    FROM case_groups
                    WHERE {group_conditions_sql}
                ),
                group_scoped AS (
                    SELECT scoped.*
                    FROM {relation_name} AS scoped
                    INNER JOIN selected_cases USING (case_id)
                )
                SELECT
                    activity,
                    COUNT(*) AS event_count,
                    COUNT(DISTINCT case_id) AS case_count,
                    ROUND(SUM(duration_min), 2) AS total_duration_min,
                    ROUND(AVG(duration_min), 2) AS avg_duration_min,
                    ROUND(MEDIAN(duration_min), 2) AS median_duration_min,
                    CASE WHEN COUNT(*) > 1 THEN ROUND(STDDEV_SAMP(duration_min), 2) ELSE NULL END AS std_duration_min,
                    ROUND(MIN(duration_min), 2) AS min_duration_min,
                    ROUND(MAX(duration_min), 2) AS max_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.75), 2) AS p75_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.90), 2) AS p90_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.95), 2) AS p95_duration_min,
                    ROUND(
                        COUNT(*) * 100.0 / NULLIF((SELECT COUNT(*) FROM group_scoped), 0),
                        2
                    ) AS event_ratio_pct
                FROM group_scoped
                GROUP BY activity
                ORDER BY event_count DESC, activity ASC
                """,
                params + group_values,
            ).df()
            if "std_duration_min" in group_df.columns:
                group_df["std_duration_min"] = _format_stddev_column(group_df["std_duration_min"])
            yield group_name, convert_analysis_result_to_records(
                group_df,
                frequency_display_columns,
            )
    finally:
        connection.close()


def _write_frequency_data(worksheet, rows, start_row):
    frequency_rows = build_ranked_rows(rows or [], rank_key=REPORT_HEADER_LABELS["rank"])
    frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
    header_row = start_row
    for column_index, header in enumerate(frequency_headers, start=1):
        header_cell = worksheet.cell(row=header_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[header_row].height = 22

    if not frequency_rows:
        merge_excel_row(worksheet, header_row + 1, len(frequency_headers))
        empty_cell = worksheet.cell(row=header_row + 1, column=1, value="表示できるデータがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[header_row + 1].height = 22
        return header_row + 3

    data_start_row = header_row + 1
    current_row = data_start_row
    for row_index, row in enumerate(frequency_rows, start=0):
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        for column_index, header in enumerate(frequency_headers, start=1):
            body_cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(frequency_headers))}{current_row - 1}"

    return current_row + 1


def localize_report_headers(headers):
    return [REPORT_HEADER_LABELS.get(header, header) for header in headers]


def localize_report_rows(rows, headers):
    localized_headers = localize_report_headers(headers)
    localized_rows = []

    for row in rows:
        localized_rows.append(
            {
                localized_header: row.get(header)
                for header, localized_header in zip(headers, localized_headers)
            }
        )

    return localized_rows, localized_headers


def build_filter_summary_text(filter_params, column_settings):
    normalized_filters = normalize_filter_params(**(filter_params or {}))
    column_payload = build_column_settings_payload(column_settings)
    filter_label_map = {
        filter_item["slot"]: filter_item["label"]
        for filter_item in column_payload.get("filters", [])
    }
    summary_items = []

    if normalized_filters.get("date_from"):
        summary_items.append(f"開始日: {normalized_filters['date_from']}")
    if normalized_filters.get("date_to"):
        summary_items.append(f"終了日: {normalized_filters['date_to']}")

    for filter_slot in ("filter_value_1", "filter_value_2", "filter_value_3"):
        if normalized_filters.get(filter_slot):
            summary_items.append(
                f"{filter_label_map.get(filter_slot, filter_slot)}: {normalized_filters[filter_slot]}"
            )

    activity_values = normalized_filters.get("activity_values")
    if activity_values:
        activity_label = "アクティビティ 含む" if normalized_filters.get("activity_mode") != "exclude" else "アクティビティ 除外"
        summary_items.append(f"{activity_label}: {activity_values}")

    return " / ".join(summary_items) if summary_items else "未適用"


def build_transition_display_label(row):
    if not row:
        return ""

    transition_label = str(row.get("transition_label") or row.get("transition") or "").strip()
    if transition_label:
        return transition_label

    from_activity = str(row.get("from_activity") or row.get("activity") or "").strip()
    to_activity = str(row.get("to_activity") or row.get("next_activity") or "").strip()
    if from_activity and to_activity:
        return f"{from_activity} → {to_activity}"
    return from_activity or to_activity


def format_duration_text_for_report(duration_sec):
    total_seconds = max(0, int(round(float(duration_sec or 0))))
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)
    parts = []

    if days:
        parts.append(f"{days}d")
    if hours or days:
        parts.append(f"{hours}h")
    if minutes or hours or days:
        parts.append(f"{minutes}m")
    parts.append(f"{seconds}s")
    return " ".join(parts)


def build_bottleneck_export_rows(rows, label_key):
    export_rows = []

    for index, row in enumerate(rows or [], start=1):
        export_rows.append(
            {
                "rank": index,
                label_key: (
                    build_transition_display_label(row)
                    if label_key == "transition_label"
                    else str(row.get(label_key) or "").strip()
                ),
                "count": int(row.get("count") or 0),
                "case_count": int(row.get("case_count") or 0),
                "avg_duration_text": format_duration_text_for_report(row.get("avg_duration_sec")),
                "median_duration_text": format_duration_text_for_report(row.get("median_duration_sec")),
                "max_duration_text": format_duration_text_for_report(row.get("max_duration_sec")),
            }
        )

    return export_rows


def resolve_pattern_detail_sheet_count(pattern_display_limit, available_count):
    normalized_limit = str(pattern_display_limit or "10").strip().lower()

    if normalized_limit == "all":
        requested_count = available_count
    else:
        try:
            requested_count = int(normalized_limit)
        except (TypeError, ValueError):
            requested_count = 10

    requested_count = max(0, requested_count)
    if requested_count > 20:
        requested_count = 20

    return min(available_count, requested_count)


def build_pattern_overview_rows(pattern_rows, variant_items, pattern_column_label, analysis_definitions):
    pattern_config = analysis_definitions.get("pattern", {}).get("config", {})
    pattern_display_columns = pattern_config.get("display_columns", {})
    repeat_flag_label = pattern_display_columns.get("repeat_flag", "繰り返し")
    repeat_count_label = pattern_display_columns.get("repeat_count", "繰り返し回数")
    repeat_rate_label = pattern_display_columns.get("repeat_rate_pct", "繰り返し率(%)")
    repeat_rate_band_label = pattern_display_columns.get("repeat_rate_band", "繰り返し率区分")
    review_flag_label = pattern_display_columns.get("review_flag", "確認区分")
    avg_case_duration_diff_label = pattern_display_columns.get("avg_case_duration_diff_min", "平均処理時間差分(分)")
    improvement_priority_score_label = pattern_display_columns.get("improvement_priority_score", "改善優先度スコア")
    overall_impact_pct_label = pattern_display_columns.get("overall_impact_pct", "全体影響度(%)")
    fastest_pattern_flag_label = pattern_display_columns.get("fastest_pattern_flag", "最短処理")
    simple_comment_label = pattern_display_columns.get("simple_comment", "簡易コメント")
    case_count_label = pattern_display_columns.get("case_count", "ケース数")
    case_ratio_label = pattern_display_columns.get("case_ratio_pct", "ケース比率(%)")
    cumulative_case_ratio_label = pattern_display_columns.get("cumulative_case_ratio_pct", "累積カバー率(%)")
    avg_case_duration_label = pattern_display_columns.get("avg_case_duration_min", "平均ケース処理時間(分)")
    std_case_duration_label = pattern_display_columns.get("std_case_duration_min", "標準偏差ケース処理時間(分)")
    min_case_duration_label = pattern_display_columns.get("min_case_duration_min", "最小ケース処理時間(分)")
    max_case_duration_label = pattern_display_columns.get("max_case_duration_min", "最大ケース処理時間(分)")
    p75_case_duration_label = pattern_display_columns.get("p75_case_duration_min", "75%点ケース処理時間(分)")
    p90_case_duration_label = pattern_display_columns.get("p90_case_duration_min", "90%点ケース処理時間(分)")
    p95_case_duration_label = pattern_display_columns.get("p95_case_duration_min", "95%点ケース処理時間(分)")
    variant_by_pattern = {
        str(variant_item.get("pattern") or "").strip(): variant_item
        for variant_item in (variant_items or [])
    }
    overview_rows = []

    for index, pattern_row in enumerate(pattern_rows or [], start=1):
        pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
        matched_variant = variant_by_pattern.get(pattern_text, {})
        variant_id = matched_variant.get("variant_id")
        overview_rows.append(
            {
                "rank": index,
                "pattern_variant": (
                    f"Pattern #{index} / Variant #{variant_id}"
                    if variant_id
                    else f"Pattern #{index}"
                ),
                "repeat_flag": pattern_row.get(repeat_flag_label, ""),
                "repeat_count": pattern_row.get(repeat_count_label, 0),
                "repeat_rate_pct": pattern_row.get(repeat_rate_label, 0),
                "repeat_rate_band": pattern_row.get(repeat_rate_band_label, ""),
                "review_flag": pattern_row.get(review_flag_label, ""),
                "avg_case_duration_diff_min": pattern_row.get(avg_case_duration_diff_label, 0),
                "improvement_priority_score": pattern_row.get(improvement_priority_score_label, 0),
                "overall_impact_pct": pattern_row.get(overall_impact_pct_label, 0),
                "fastest_pattern_flag": pattern_row.get(fastest_pattern_flag_label, ""),
                "simple_comment": pattern_row.get(simple_comment_label, ""),
                "count": pattern_row.get(case_count_label, 0),
                "ratio": pattern_row.get(case_ratio_label, 0),
                "cumulative_case_ratio_pct": pattern_row.get(cumulative_case_ratio_label, 0),
                "avg_case_duration_min": pattern_row.get(avg_case_duration_label, 0),
                "std_case_duration_min": pattern_row.get(std_case_duration_label, 0),
                "min_case_duration_min": pattern_row.get(min_case_duration_label, 0),
                "max_case_duration_min": pattern_row.get(max_case_duration_label, 0),
                "p75_case_duration_min": pattern_row.get(p75_case_duration_label, 0),
                "p90_case_duration_min": pattern_row.get(p90_case_duration_label, 0),
                "p95_case_duration_min": pattern_row.get(p95_case_duration_label, 0),
                "pattern": pattern_text,
            }
        )

    return overview_rows


def coerce_report_number(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def build_pattern_export_summary(pattern_rows, pattern_display_columns):
    repeat_flag_label = pattern_display_columns.get("repeat_flag", "繰り返し")
    repeat_count_label = pattern_display_columns.get("repeat_count", "繰り返し回数")
    repeat_rate_label = pattern_display_columns.get("repeat_rate_pct", "繰り返し率(%)")
    repeat_rate_band_label = pattern_display_columns.get("repeat_rate_band", "繰り返し率区分")
    review_flag_label = pattern_display_columns.get("review_flag", "確認区分")
    avg_case_duration_diff_label = pattern_display_columns.get("avg_case_duration_diff_min", "平均処理時間差分(分)")
    improvement_priority_score_label = pattern_display_columns.get("improvement_priority_score", "改善優先度スコア")
    overall_impact_pct_label = pattern_display_columns.get("overall_impact_pct", "全体影響度(%)")
    fastest_pattern_flag_label = pattern_display_columns.get("fastest_pattern_flag", "最短処理")
    simple_comment_label = pattern_display_columns.get("simple_comment", "簡易コメント")
    case_count_label = pattern_display_columns.get("case_count", "ケース数")
    case_ratio_label = pattern_display_columns.get("case_ratio_pct", "ケース比率(%)")
    cumulative_case_ratio_label = pattern_display_columns.get("cumulative_case_ratio_pct", "累積カバー率(%)")
    avg_case_duration_label = pattern_display_columns.get("avg_case_duration_min", "平均ケース処理時間(分)")
    min_case_duration_label = pattern_display_columns.get("min_case_duration_min", "最小ケース処理時間(分)")
    max_case_duration_label = pattern_display_columns.get("max_case_duration_min", "最大ケース処理時間(分)")
    pattern_label = pattern_display_columns.get("pattern", "処理順パターン")

    comparison_rows = []
    repeated_patterns = []
    improvement_targets = []
    for index, pattern_row in enumerate(pattern_rows or [], start=1):
        comparison_row = {
            "順位": index,
            "繰り返し": pattern_row.get(repeat_flag_label, ""),
            "繰り返し回数": pattern_row.get(repeat_count_label, 0),
            "繰り返し率(%)": pattern_row.get(repeat_rate_label, 0),
            "繰り返し率区分": pattern_row.get(repeat_rate_band_label, ""),
            "件数": pattern_row.get(case_count_label, 0),
            "全体比率(%)": pattern_row.get(case_ratio_label, 0),
            "平均処理時間(分)": pattern_row.get(avg_case_duration_label, 0),
            "平均処理時間差分(分)": pattern_row.get(avg_case_duration_diff_label, 0),
            "改善優先度スコア": pattern_row.get(improvement_priority_score_label, 0),
            "全体影響度(%)": pattern_row.get(overall_impact_pct_label, 0),
            "最短処理": pattern_row.get(fastest_pattern_flag_label, ""),
            "最短処理時間(分)": pattern_row.get(min_case_duration_label, 0),
            "最長処理時間(分)": pattern_row.get(max_case_duration_label, 0),
            "確認区分": pattern_row.get(review_flag_label, ""),
            "簡易コメント": pattern_row.get(simple_comment_label, ""),
            "パターン": pattern_row.get(pattern_label, ""),
        }
        comparison_rows.append(comparison_row)
        if str(pattern_row.get(review_flag_label, "")).strip() != "要確認":
            pass
        else:
            repeated_patterns.append(comparison_row)
        if (
            coerce_report_number(pattern_row.get(repeat_rate_label, 0)) >= 10.0
            and coerce_report_number(pattern_row.get(avg_case_duration_diff_label, 0)) > 0
        ):
            improvement_targets.append(comparison_row)

    top3_rows = comparison_rows[:3]
    top3_coverage_pct = (
        coerce_report_number((pattern_rows or [])[2].get(cumulative_case_ratio_label, 0))
        if len(pattern_rows or []) >= 3
        else (
            coerce_report_number((pattern_rows or [])[-1].get(cumulative_case_ratio_label, 0))
            if pattern_rows
            else 0.0
        )
    )
    top10_coverage_pct = (
        coerce_report_number((pattern_rows or [])[9].get(cumulative_case_ratio_label, 0))
        if len(pattern_rows or []) >= 10
        else (
            coerce_report_number((pattern_rows or [])[-1].get(cumulative_case_ratio_label, 0))
            if pattern_rows
            else 0.0
        )
    )
    repeated_case_ratio_pct = round(
        sum(coerce_report_number(problem_row["全体比率(%)"]) for problem_row in repeated_patterns),
        2,
    )
    improvement_targets = sorted(
        improvement_targets,
        key=lambda row: (
            -coerce_report_number(row.get("改善優先度スコア"), 0),
            -coerce_report_number(row.get("繰り返し率(%)"), 0),
            -coerce_report_number(row.get("平均処理時間差分(分)"), 0),
            -coerce_report_number(row.get("件数"), 0),
            row.get("パターン", ""),
        ),
    )[:3]
    fastest_pattern = min(
        comparison_rows,
        key=lambda row: (
            coerce_report_number(row.get("平均処理時間(分)"), float("inf")),
            row.get("順位", 0),
        ),
        default=None,
    )
    coverage_summary_text = (
        f"上位3パターンで {round(top3_coverage_pct, 2):.2f}%、"
        f"上位10パターンで {round(top10_coverage_pct, 2):.2f}% をカバーしています。"
    )

    return {
        "top_patterns": top3_rows,
        "comparison_rows": comparison_rows[:10],
        "repeated_patterns": repeated_patterns,
        "top3_coverage_pct": round(top3_coverage_pct, 2),
        "top10_coverage_pct": round(top10_coverage_pct, 2),
        "coverage_summary_text": coverage_summary_text,
        "repeated_case_ratio_pct": repeated_case_ratio_pct,
        "fastest_pattern": fastest_pattern or {},
        "improvement_targets": improvement_targets,
    }


def calculate_pattern_time_impact_minutes(pattern_row):
    return round(
        max(0.0, coerce_report_number(pattern_row.get("平均処理時間差分(分)"), 0.0))
        * max(0.0, coerce_report_number(pattern_row.get("件数"), 0.0)),
        2,
    )


def build_pattern_issue_row(pattern_row):
    repeat_rate_pct = coerce_report_number(pattern_row.get("繰り返し率(%)"), 0.0)
    duration_diff_min = coerce_report_number(pattern_row.get("平均処理時間差分(分)"), 0.0)
    pattern_text = str(pattern_row.get("パターン") or "").strip()

    if repeat_rate_pct >= 30:
        issue_text = f"繰り返し率が {repeat_rate_pct:.2f}% と高く、手戻りが多いパターンです。"
        cause_text = "差戻しや再確認が発生しやすく、同じ工程を複数回通過している可能性があります。"
        action_text = "差戻し発生条件を洗い出し、一次判定や入力チェックの前倒しを検討してください。"
    elif repeat_rate_pct >= 10:
        issue_text = f"繰り返し率が {repeat_rate_pct:.2f}% あり、再作業が混在しています。"
        cause_text = "一部ケースで確認や承認のやり直しが発生し、処理が伸びている可能性があります。"
        action_text = "繰り返しが起きる工程の条件分岐を整理し、再実行の発生源を減らしてください。"
    else:
        issue_text = f"繰り返しは少ないものの、平均処理時間が全体平均より {duration_diff_min:.2f} 分長いパターンです。"
        cause_text = "特定工程の待ちや滞留により、パターン全体の処理時間が長くなっている可能性があります。"
        action_text = "ボトルネック工程の担当・承認・待機条件を見直し、滞留時間の短縮を優先してください。"

    return {
        "問題点": issue_text,
        "原因": cause_text,
        "改善案": action_text,
        "期待効果（時間短縮・分）": calculate_pattern_time_impact_minutes(pattern_row),
        "対象パターン": pattern_text,
    }


def build_pattern_conclusion_summary(pattern_summary):
    comparison_rows = pattern_summary.get("comparison_rows", [])
    improvement_targets = pattern_summary.get("improvement_targets", [])
    repeated_patterns = pattern_summary.get("repeated_patterns", [])
    fastest_pattern = pattern_summary.get("fastest_pattern", {}) or {}
    top_issue_candidates = improvement_targets[:3] if improvement_targets else comparison_rows[:3]
    issue_rows = [build_pattern_issue_row(row) for row in top_issue_candidates]
    total_impact_minutes = round(
        sum(calculate_pattern_time_impact_minutes(row) for row in improvement_targets),
        2,
    )
    overall_summary = (
        f"上位10パターンで {coerce_report_number(pattern_summary.get('top10_coverage_pct', 0.0)):.2f}% をカバーし、"
        f"要確認パターンは {len(repeated_patterns)} 件、改善対象TOP3で約 "
        f"{round(sum(calculate_pattern_time_impact_minutes(row) for row in improvement_targets[:3]), 2):.2f} 分の短縮余地があります。"
    )

    return {
        "overall_summary": overall_summary,
        "issue_rows": issue_rows,
        "total_impact_minutes": total_impact_minutes,
        "total_impact_hours": round(total_impact_minutes / 60.0, 2),
        "fastest_pattern": fastest_pattern,
        "improvement_targets": improvement_targets[:3],
    }


def build_pattern_dashboard_summary(pattern_summary, pattern_conclusion):
    top3_rows = pattern_summary.get("improvement_targets") or pattern_summary.get("comparison_rows", [])[:3]
    problem_points = [
        issue_row.get("問題点")
        for issue_row in pattern_conclusion.get("issue_rows", [])
        if str(issue_row.get("問題点") or "").strip()
    ]
    if not problem_points and pattern_summary.get("comparison_rows"):
        problem_points = [
            str(row.get("簡易コメント") or "").strip()
            for row in pattern_summary.get("comparison_rows", [])[:3]
            if str(row.get("簡易コメント") or "").strip()
        ]

    return {
        "overall_summary": pattern_conclusion.get("overall_summary", ""),
        "top3_rows": top3_rows,
        "problem_points": problem_points[:3],
        "top10_coverage_pct": pattern_summary.get("top10_coverage_pct", 0),
        "total_impact_minutes": pattern_conclusion.get("total_impact_minutes", 0),
    }


def _set_chart_str_categories(chart, labels_ref):
    """Set chart categories as strRef so Excel treats text labels correctly."""
    label_formula = str(labels_ref)
    for series in chart.ser:
        series.cat = AxDataSource(strRef=StrRef(f=label_formula))


def _ensure_chart_data_sheet(workbook):
    sheet_name = "_chart_data"
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    data_sheet = workbook.create_sheet(title=sheet_name)
    data_sheet.sheet_state = "hidden"
    return data_sheet


def _write_chart_data_block(workbook, block_label, comparison_rows, columns):
    data_sheet = _ensure_chart_data_sheet(workbook)
    start_row = (data_sheet.max_row or 0) + 2

    data_sheet.cell(row=start_row, column=1, value=block_label)
    header_row = start_row + 1
    for col_offset, col_name in enumerate(columns):
        data_sheet.cell(row=header_row, column=1 + col_offset, value=col_name)

    for index, row in enumerate(comparison_rows, start=1):
        data_row = header_row + index
        data_sheet.cell(row=data_row, column=1, value=f"Pattern #{row.get('順位', index)}")
        for col_offset, col_name in enumerate(columns):
            if col_offset == 0:
                continue
            data_sheet.cell(
                row=data_row, column=1 + col_offset,
                value=coerce_report_number(row.get(col_name), 0.0),
            )

    return data_sheet, header_row, header_row + len(comparison_rows)


def build_excel_anchor(column_letter, row_number):
    return f"{column_letter}{max(1, int(row_number or 1))}"


def sort_pattern_rows_by_avg_duration_desc(rows):
    return sorted(
        rows,
        key=lambda row: (
            -coerce_report_number(row.get("平均処理時間(分)"), 0.0),
            row.get("順位", 0),
        ),
    )


def append_pattern_dashboard_pie_chart(workbook, worksheet, comparison_rows, anchor="A1"):
    if not comparison_rows:
        return

    data_sheet, header_row, max_row = _write_chart_data_block(
        workbook, "dashboard_pie", comparison_rows, ["パターン", "件数"],
    )

    pie_chart = PieChart()
    pie_chart.title = "上位10パターンの割合"
    pie_chart.style = 10
    pie_chart.height = 14.0
    pie_chart.width = 16.0
    pie_data = Reference(data_sheet, min_col=2, min_row=header_row, max_row=max_row)
    pie_labels = Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    _set_chart_str_categories(pie_chart, pie_labels)
    pie_chart.dLbls = DataLabelList(
        showCatName=False,
        showVal=False,
        showPercent=True,
        showSerName=False,
        showLegendKey=False,
        showLeaderLines=True,
        dLblPos="bestFit",
        separator="\n",
    )
    worksheet.add_chart(pie_chart, anchor)


def append_pattern_conclusion_charts(workbook, worksheet, comparison_rows, pie_anchor="A1", bar_anchor="A20"):
    if not comparison_rows:
        return

    data_sheet, header_row, max_row = _write_chart_data_block(
        workbook, "conclusion_charts", comparison_rows,
        ["パターン", "件数"],
    )
    bar_rows = sort_pattern_rows_by_avg_duration_desc(comparison_rows)
    bar_data_sheet, bar_header_row, bar_max_row = _write_chart_data_block(
        workbook,
        "conclusion_duration_desc",
        bar_rows,
        ["パターン", "平均処理時間(分)"],
    )

    pie_chart = PieChart()
    pie_chart.title = "上位10パターンの割合"
    pie_chart.style = 10
    pie_chart.height = 14.0
    pie_chart.width = 16.0
    pie_data = Reference(data_sheet, min_col=2, min_row=header_row, max_row=max_row)
    pie_labels = Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    _set_chart_str_categories(pie_chart, pie_labels)
    pie_chart.dLbls = DataLabelList(
        showCatName=False,
        showVal=False,
        showPercent=True,
        showSerName=False,
        showLegendKey=False,
        showLeaderLines=True,
        dLblPos="bestFit",
        separator="\n",
    )
    worksheet.add_chart(pie_chart, pie_anchor)

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 10
    bar_chart.title = "平均処理時間の比較（長い順）"
    bar_chart.height = 14.0
    bar_chart.width = 16.0
    bar_chart.legend = None
    bar_chart.varyColors = False
    bar_chart.gapWidth = 60
    bar_chart.x_axis.scaling.orientation = "maxMin"
    bar_chart.x_axis.tickLblPos = "low"
    bar_chart.y_axis.crosses = "autoZero"
    bar_chart.y_axis.delete = False
    bar_data = Reference(bar_data_sheet, min_col=2, min_row=bar_header_row, max_row=bar_max_row)
    bar_labels = Reference(bar_data_sheet, min_col=1, min_row=bar_header_row + 1, max_row=bar_max_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    _set_chart_str_categories(bar_chart, bar_labels)
    bar_chart.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        dLblPos="outEnd",
    )
    worksheet.add_chart(bar_chart, bar_anchor)


def append_pattern_detail_sheet(
    workbook,
    filtered_df,
    pattern_row,
    pattern_rank,
    pattern_column_label,
    analysis_definitions,
    variant_item=None,
    pattern_detail=None,
):
    pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
    if not pattern_text:
        return

    detail = pattern_detail or create_pattern_bottleneck_details(filtered_df, pattern_text)
    sheet_name = sanitize_workbook_sheet_name(f"パターン{pattern_rank:02d}詳細")
    detail_sheet = workbook.create_sheet(title=sheet_name)
    initialize_excel_worksheet(detail_sheet)

    variant_label = (
        f"Variant #{variant_item['variant_id']}"
        if variant_item and variant_item.get("variant_id")
        else "該当なし"
    )
    bottleneck_transition = detail.get("bottleneck_transition") or {}
    next_row = append_key_value_rows(
        detail_sheet,
        f"Pattern #{pattern_rank} 詳細",
        [
            ("パターン / バリアント", f"Pattern #{pattern_rank} / {variant_label}"),
            ("繰り返し", pattern_row.get("繰り返し", "")),
            ("繰り返し回数", pattern_row.get("繰り返し回数", 0)),
            ("繰り返し率(%)", pattern_row.get("繰り返し率(%)", 0)),
            ("繰り返し率区分", pattern_row.get("繰り返し率区分", "")),
            ("確認区分", pattern_row.get("確認区分", "")),
            ("平均処理時間差分(分)", pattern_row.get("平均処理時間差分(分)", 0)),
            ("改善優先度スコア", pattern_row.get("改善優先度スコア", 0)),
            ("全体影響度(%)", pattern_row.get("全体影響度(%)", 0)),
            ("最短処理", pattern_row.get("最短処理", "")),
            ("簡易コメント", pattern_row.get("簡易コメント", "")),
            ("ケース数", detail.get("case_count", 0)),
            ("ケース比率(%)", detail.get("case_ratio_pct", 0)),
            ("平均ケース処理時間(分)", detail.get("avg_case_duration_min", 0)),
            ("中央値ケース処理時間(分)", detail.get("median_case_duration_min", 0)),
            ("最小ケース処理時間(分)", detail.get("min_case_duration_min", 0)),
            ("最大ケース処理時間(分)", detail.get("max_case_duration_min", 0)),
            ("代表ルート", pattern_text),
            ("ボトルネック遷移", bottleneck_transition.get("transition_label", "該当なし")),
            ("ボトルネック平均所要時間(分)", bottleneck_transition.get("avg_duration_min", 0)),
        ],
        description="上位パターンのケース概要、代表ルート、ボトルネック遷移をまとめています。",
    )

    step_metric_rows, step_metric_headers = localize_report_rows(
        [
            {
                "rank": index,
                "sequence_no": step_metric["sequence_no"],
                "transition": step_metric["transition_label"],
                "case_count": step_metric["case_count"],
                "avg_duration_min": step_metric["avg_duration_min"],
                "median_duration_min": step_metric["median_duration_min"],
                "min_duration_min": step_metric["min_duration_min"],
                "max_duration_min": step_metric["max_duration_min"],
                "total_duration_min": step_metric["total_duration_min"],
                "wait_share_pct": step_metric["wait_share_pct"],
            }
            for index, step_metric in enumerate(detail.get("step_metrics", []), start=1)
        ],
        [
            "rank",
            "sequence_no",
            "transition",
            "case_count",
            "avg_duration_min",
            "median_duration_min",
            "min_duration_min",
            "max_duration_min",
            "total_duration_min",
            "wait_share_pct",
        ],
    )
    next_row = append_table_to_worksheet(
        detail_sheet,
        "ステップ別所要時間",
        step_metric_rows,
        step_metric_headers,
        start_row=next_row,
        description="各ステップの所要時間と全体に占める比率を比較できます。",
        no_wrap_headers=["遷移"],
        min_column_widths={"遷移": 32},
    )

    case_example_rows, case_example_headers = localize_report_rows(
        [
            {
                "rank": index,
                "case_id": case_example["case_id"],
                "start_time": case_example["start_time"],
                "end_time": case_example["end_time"],
                "total_duration_min": case_example["case_total_duration_min"],
            }
            for index, case_example in enumerate(detail.get("case_examples", []), start=1)
        ],
        ["rank", "case_id", "start_time", "end_time", "total_duration_min"],
    )
    append_table_to_worksheet(
        detail_sheet,
        "代表ケース",
        case_example_rows,
        case_example_headers,
        start_row=next_row,
        description="このパターンに属するケースのうち、総処理時間が長いケースを上位から掲載しています。",
    )


def serialize_ai_prompt_rows(rows, max_items=5):
    serialized_rows = []

    for row in list(rows or [])[: max(0, int(max_items or 0))]:
        if isinstance(row, dict):
            serialized_rows.append(
                {
                    str(key): normalize_excel_cell_value(value)
                    for key, value in row.items()
                }
            )
        else:
            serialized_rows.append(normalize_excel_cell_value(row))

    return json.dumps(serialized_rows, ensure_ascii=False, indent=2)


def build_analysis_ai_prompt(ai_context):
    analysis_key = str(ai_context["analysis_key"]).strip().lower()
    analysis_name = ai_context["analysis_name"]
    focus_map = {
        "frequency": {
            "focus": "件数が集中しているアクティビティと、平均処理時間が長いアクティビティを分けて解釈してください。",
            "priority": "負荷集中、入力不備、担当偏り、差戻し起点を優先して原因仮説を述べてください。",
            "actions": "入口制御、担当割付、事前チェック、差戻し削減に直結する改善アクションを提案してください。",
        },
        "transition": {
            "focus": "前後遷移の処理時間、引き継ぎの詰まり、ループや差戻しを中心に解釈してください。",
            "priority": "最も遅い遷移と改善インパクトが大きい遷移を分けて説明してください。",
            "actions": "承認待ち、差戻し、バッチ処理、手作業の受け渡しを減らす改善アクションを提案してください。",
        },
        "pattern": {
            "focus": "標準ルートと例外ルートの違い、分岐の多さ、例外パターンの発生理由を解釈してください。",
            "priority": "最頻出パターンと、時間が長い例外パターンを比較して説明してください。",
            "actions": "標準化、例外削減、分岐条件の見直しに直結する改善アクションを提案してください。",
        },
    }
    focus_config = focus_map.get(
        analysis_key,
        {
            "focus": "主要な傾向、滞留箇所、改善優先度を解釈してください。",
            "priority": "影響が大きい箇所を数値付きで説明してください。",
            "actions": "現場で着手しやすい改善アクションを提案してください。",
        },
    )

    return f"""あなたはプロセスマイニング結果を業務改善に落とし込むアナリストです。
以下は「{analysis_name}」の詳細画面に対応する分析データです。現場担当者が画面を切り替えても同じ解釈を再利用できるよう、論点がぶれない説明にしてください。

## この分析で重視する視点
- {focus_config['focus']}
- {focus_config['priority']}
- {focus_config['actions']}

## 基本情報
- 分析名: {analysis_name}
- 総ケース数: {int(ai_context['dashboard_summary'].get('total_cases', 0)):,}
- 総イベント数: {int(ai_context['dashboard_summary'].get('total_records', 0)):,}
- 分析期間: {ai_context['period_text']}
- 平均ケース処理時間: {ai_context['dashboard_summary'].get('avg_case_duration_text', '0s')}
- 上位10バリアントカバー率: {float(ai_context['dashboard_summary'].get('top10_variant_coverage_pct', 0.0)):.2f}%

## 現在の分析結果上位
{serialize_ai_prompt_rows(ai_context['analysis_rows'], max_items=7)}

## アクティビティボトルネック
{serialize_ai_prompt_rows(ai_context['bottleneck_summary'].get('activity_bottlenecks', []), max_items=5)}

## 遷移ボトルネック
{serialize_ai_prompt_rows(ai_context['bottleneck_summary'].get('transition_bottlenecks', []), max_items=5)}

## 改善インパクト上位
{serialize_ai_prompt_rows(ai_context['impact_summary'].get('rows', []), max_items=5)}

## Root Cause 候補
{serialize_ai_prompt_rows(ai_context['root_cause_summary'].get('rows', []), max_items=5)}

## 主要バリアント
{serialize_ai_prompt_rows(ai_context['variant_items'], max_items=5)}

## ルールベース要点
{serialize_ai_prompt_rows([item.get('text', '') for item in ai_context['insights_summary'].get('items', [])], max_items=5)}

## 回答形式
【1. 全体サマリー】
2〜3文で全体像を要約してください。

【2. この分析で読むべきポイント】
この分析ならではの見方で、重要点を2つ説明してください。

【3. 考えられる原因】
現場で起こりやすい原因を2〜3つ挙げてください。

【4. 改善アクション】
すぐできることを1つ、中期的な改善を1つ提案してください。

【5. 次に見るべきこと】
次に確認すべき切り口を1つ提案してください。

専門用語は使いすぎず、現場担当者が読みやすい自然な日本語で書いてください。
"""


def build_ai_fallback_text(ai_context):
    analysis_key = str(ai_context["analysis_key"]).strip().lower()
    analysis_name = ai_context["analysis_name"]
    dashboard_summary = ai_context["dashboard_summary"]
    period_text = ai_context["period_text"]
    insights_summary = ai_context["insights_summary"]
    impact_summary = ai_context["impact_summary"]
    bottleneck_summary = ai_context["bottleneck_summary"]
    analysis_rows = ai_context["analysis_rows"]

    top_impact_row = impact_summary["rows"][0] if impact_summary.get("rows") else None
    top_transition_bottleneck = (
        bottleneck_summary["transition_bottlenecks"][0]
        if bottleneck_summary.get("transition_bottlenecks")
        else None
    )
    top_activity_bottleneck = (
        bottleneck_summary["activity_bottlenecks"][0]
        if bottleneck_summary.get("activity_bottlenecks")
        else None
    )
    top_transition_bottleneck_label = build_transition_display_label(top_transition_bottleneck)
    top_row = analysis_rows[0] if analysis_rows else {}

    highlight_lines = [f"- {item['text']}" for item in insights_summary.get("items", [])]
    if not highlight_lines:
        highlight_lines.append("- 既存集計から明確なハイライトを抽出できませんでした。")

    if analysis_key == "frequency":
        priority_text = (
            f"件数の中心は「{top_row.get('アクティビティ', '不明')}」で、"
            f"{normalize_excel_cell_value(top_row.get('イベント件数', 0))} 件です。"
            if top_row
            else "件数集中の中心アクティビティは特定できませんでした。"
        )
        action_lines = [
            (
                f"件数が集中する「{top_row.get('アクティビティ', '対象アクティビティ')}」について、"
                "受付経路や担当者別件数を比較してください。"
            ),
            (
                f"平均所要時間が長い「{top_activity_bottleneck['activity']}」の前後で、"
                "入力不備や差戻しが発生していないか確認してください。"
                if top_activity_bottleneck
                else "上位アクティビティの担当別処理時間を比較してください。"
            ),
        ]
    elif analysis_key == "transition":
        priority_text = (
            f"最も優先度が高い遷移候補は「{top_transition_bottleneck_label}」です。"
            if top_transition_bottleneck_label
            else "優先度が高い遷移候補は特定できませんでした。"
        )
        action_lines = [
            (
                f"「{top_transition_bottleneck_label}」の前後で、承認待ちや引き継ぎ待ちの内訳を確認してください。"
                if top_transition_bottleneck_label
                else "遷移単位で担当待ちの内訳を確認してください。"
            ),
            (
                f"改善インパクトが高い「{top_impact_row['transition_label']}」から先に改善対象を絞ってください。"
                if top_impact_row
                else "差戻しや再提出を含む遷移を優先して確認してください。"
            ),
        ]
    elif analysis_key == "pattern":
        priority_text = (
            f"最頻出パターンは「{top_row.get('処理順パターン', top_row.get('パターン', '不明'))}」です。"
            if top_row
            else "標準ルートは特定できませんでした。"
        )
        action_lines = [
            "最頻出パターンと時間が長い例外パターンを比較し、分岐条件を整理してください。",
            "差戻しや再提出を含むパターンを優先して、標準ルートへ寄せられるか確認してください。",
        ]
    else:
        priority_text = (
            f"改善インパクト最大の遷移は「{top_impact_row['transition_label']}」で、平均所要時間は {top_impact_row['avg_duration_text']} です。"
            if top_impact_row
            else "改善インパクト上位の遷移は検出されませんでした。"
        )
        action_lines = [
            (
                f"「{top_transition_bottleneck_label}」の前後で、担当待ち・承認待ち・差戻し理由の内訳を確認してください。"
                if top_transition_bottleneck_label
                else "上位パターンと例外パターンを比較し、どこで処理が分岐しているかを確認してください。"
            ),
            (
                f"アクティビティ「{top_activity_bottleneck['activity']}」について、担当者別の件数と平均所要時間を比較してください。"
                if top_activity_bottleneck
                else "改善インパクトが高い遷移を優先して、滞留の主因を確認してください。"
            ),
        ]

    return "\n".join(
        [
            "【全体サマリー】",
            (
                f"{analysis_name} を中心に確認すると、対象は "
                f"{int(dashboard_summary.get('total_cases', 0)):,} ケース / "
                f"{int(dashboard_summary.get('total_records', 0)):,} イベントです。"
            ),
            f"分析期間は {period_text} です。",
            "",
            "【重要ポイント】",
            *highlight_lines,
            "",
            "【この分析で優先して見るべき点】",
            priority_text,
            "",
            "【次のアクション】",
            *[f"- {action_line}" for action_line in action_lines[:2]],
        ]
    )


def request_ollama_insights_text(prompt, model="qwen2.5:7b"):
    timeout = httpx.Timeout(connect=5.0, read=60.0, write=10.0, pool=5.0)
    with httpx.Client(timeout=timeout) as client:
        response = client.post(
            "http://localhost:11434/api/generate",
            json={"model": model, "prompt": prompt, "stream": False},
        )
        response.raise_for_status()
        payload = response.json()
        return str(payload.get("response") or "").strip()


def build_empty_ai_summary(analysis_key, analysis_name):
    return {
        "title": REPORT_SHEET_NAMES["ai_insights"],
        "analysis_key": analysis_key,
        "analysis_name": analysis_name,
        "generated": False,
        "cached": False,
        "mode": "idle",
        "provider": "",
        "generated_at": "",
        "period": "",
        "text": "",
        "highlights": [],
        "note": "まだ生成していません。現在の分析条件に対する分析コメントを生成すると、画面を切り替えても保持されます。",
    }


def get_cached_ai_summary(run_data, analysis_key, filter_params=None):
    cache_key = (
        str(analysis_key or "").strip().lower(),
        build_filter_cache_key(filter_params),
        None,
    )
    cached_payload = run_data.setdefault("ai_insights_cache", {}).get(cache_key)
    if cached_payload is None:
        return None
    return {
        **cached_payload,
        "generated": True,
        "cached": True,
    }


def build_ai_context_summary(
    run_data,
    analysis_key,
    filter_params=None,
    prepared_df=None,
    variant_pattern=None,
    analysis=None,
    dashboard_summary=None,
    impact_summary=None,
    bottleneck_summary=None,
    root_cause_summary=None,
    insights_summary=None,
    variant_items=None,
    analysis_name=None,
):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    analysis_definitions = get_available_analysis_definitions()
    resolved_analysis = analysis or get_analysis_data(
        run_data,
        normalized_analysis_key,
        filter_params=filter_params,
        variant_pattern=variant_pattern,
    )
    resolved_prepared_df = prepared_df
    if resolved_prepared_df is None:
        if normalized_analysis_key in {"frequency", "transition", "pattern"}:
            resolved_prepared_df = pd.DataFrame(
                columns=["case_id", "activity", "duration_sec", "start_time", "next_time", "timestamp"]
            )
        else:
            resolved_prepared_df = pd.DataFrame(
                columns=["case_id", "activity", "duration_sec", "start_time", "next_time", "timestamp"]
            )
    resolved_dashboard_summary = dashboard_summary or get_dashboard_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=resolved_prepared_df,
        variant_pattern=variant_pattern,
    )
    resolved_impact_summary = impact_summary or get_impact_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=resolved_prepared_df,
        variant_pattern=variant_pattern,
    )
    resolved_bottleneck_summary = bottleneck_summary
    if resolved_bottleneck_summary is None:
        if variant_pattern:
            resolved_bottleneck_summary = query_bottleneck_summary(
                run_data["prepared_parquet_path"],
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
                variant_pattern=variant_pattern,
                limit=None,
            )
        else:
            resolved_bottleneck_summary = get_bottleneck_summary(
                run_data,
                filter_params=filter_params,
            )
    resolved_root_cause_summary = root_cause_summary or get_root_cause_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=resolved_prepared_df,
        variant_pattern=variant_pattern,
    )
    resolved_analysis_name = (
        analysis_name
        or resolved_analysis.get("analysis_name")
        or analysis_definitions.get(normalized_analysis_key, {}).get("config", {}).get("analysis_name", analysis_key)
    )
    period_text = query_period_text(
        run_data["prepared_parquet_path"],
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    )
    resolved_insights_summary = insights_summary or get_rule_based_insights_summary(
        run_data,
        normalized_analysis_key,
        analysis_rows=resolved_analysis.get("rows"),
        filter_params=filter_params,
        prepared_df=resolved_prepared_df,
        variant_pattern=variant_pattern,
        dashboard_summary=resolved_dashboard_summary,
        impact_summary=resolved_impact_summary,
    )

    return {
        "analysis_key": normalized_analysis_key,
        "analysis_name": resolved_analysis_name,
        "analysis_rows": list(resolved_analysis.get("rows", []))[:10],
        "dashboard_summary": resolved_dashboard_summary,
        "impact_summary": resolved_impact_summary,
        "bottleneck_summary": resolved_bottleneck_summary,
        "root_cause_summary": resolved_root_cause_summary,
        "variant_items": (
            list(variant_items)[:5]
            if variant_items is not None
            else list(
                get_variant_items(
                    run_data,
                    filter_params=filter_params,
                    variant_pattern=variant_pattern,
                )
            )[:5]
        ),
        "period_text": period_text,
        "insights_summary": resolved_insights_summary,
    }


def build_ai_insights_summary(
    run_data,
    analysis_key,
    filter_params=None,
    prepared_df=None,
    variant_pattern=None,
    analysis=None,
    dashboard_summary=None,
    impact_summary=None,
    bottleneck_summary=None,
    root_cause_summary=None,
    insights_summary=None,
    variant_items=None,
    analysis_name=None,
    force_refresh=False,
    use_cache=True,
):
    cache_key = (
        str(analysis_key or "").strip().lower(),
        build_filter_cache_key(filter_params),
        str(variant_pattern or "").strip() or None,
    )
    cache = run_data.setdefault("ai_insights_cache", {})

    if use_cache and not force_refresh and cache_key in cache:
        return {
            **cache[cache_key],
            "generated": True,
            "cached": True,
        }

    ai_context = build_ai_context_summary(
        run_data=run_data,
        analysis_key=analysis_key,
        filter_params=filter_params,
        prepared_df=prepared_df,
        variant_pattern=variant_pattern,
        analysis=analysis,
        dashboard_summary=dashboard_summary,
        impact_summary=impact_summary,
        bottleneck_summary=bottleneck_summary,
        root_cause_summary=root_cause_summary,
        insights_summary=insights_summary,
        variant_items=variant_items,
        analysis_name=analysis_name,
    )
    generated_at = datetime.now(timezone.utc).isoformat()
    fallback_text = build_ai_fallback_text(ai_context)

    if not ai_context["dashboard_summary"].get("has_data"):
        payload = {
            "title": REPORT_SHEET_NAMES["ai_insights"],
            "analysis_key": ai_context["analysis_key"],
            "analysis_name": ai_context["analysis_name"],
            "mode": "rule_based",
            "provider": "",
            "generated_at": generated_at,
            "period": ai_context["period_text"],
            "text": fallback_text,
            "highlights": [item["text"] for item in ai_context["insights_summary"].get("items", [])],
            "note": "対象データがないため、既存集計からの要約のみを表示しています。",
        }
        if use_cache:
            cache[cache_key] = payload
        return {
            **payload,
            "generated": True,
            "cached": False,
        }

    try:
        ai_text = request_ollama_insights_text(build_analysis_ai_prompt(ai_context))
        if ai_text:
            payload = {
                "title": REPORT_SHEET_NAMES["ai_insights"],
                "analysis_key": ai_context["analysis_key"],
                "analysis_name": ai_context["analysis_name"],
                "mode": "ollama",
                "provider": "",
                "generated_at": generated_at,
                "period": ai_context["period_text"],
                "text": ai_text,
                "highlights": [item["text"] for item in ai_context["insights_summary"].get("items", [])],
                "note": "現在の分析条件に対応する分析コメントを保存しました。画面を切り替えても同じ条件なら再表示されます。",
            }
            if use_cache:
                cache[cache_key] = payload
            return {
                **payload,
                "generated": True,
                "cached": False,
            }
    except httpx.ConnectError:
            error_message = "分析コメントを生成できなかったため、既存集計からの要約を掲載しています。"
    except Exception as exc:
                error_message = f"分析コメントの生成に失敗したため、ルールベース要約を掲載しています。({exc})"
    else:
                error_message = "分析コメントを生成できなかったため、ルールベース要約を掲載しています。"

    payload = {
        "title": REPORT_SHEET_NAMES["ai_insights"],
        "analysis_key": ai_context["analysis_key"],
        "analysis_name": ai_context["analysis_name"],
        "mode": "rule_based",
        "provider": "",
        "generated_at": generated_at,
        "period": ai_context["period_text"],
        "text": fallback_text,
        "highlights": [item["text"] for item in ai_context["insights_summary"].get("items", [])],
        "note": error_message,
    }
    if use_cache:
        cache[cache_key] = payload
    return {
        **payload,
        "generated": True,
        "cached": False,
    }


def get_analysis_export_sheet_keys(analysis_key):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    sheet_keys = ["summary", "ai_insights"]

    if normalized_analysis_key == "frequency":
        sheet_keys.append("frequency")
    elif normalized_analysis_key == "transition":
        sheet_keys.extend(["transition", "bottleneck", "impact"])
    elif normalized_analysis_key == "pattern":
        sheet_keys.extend(["pattern_summary", "pattern"])
    else:
        sheet_keys.append(normalized_analysis_key)

    return sheet_keys


def build_detail_summary_kpi_rows(
    analysis_key,
    analysis_rows,
    dashboard_summary,
    impact_summary,
    bottleneck_summary,
    prepared_df=None,
    variant_items=None,
):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    top_row = analysis_rows[0] if analysis_rows else {}
    top_transition_bottleneck = (
        bottleneck_summary["transition_bottlenecks"][0]
        if bottleneck_summary.get("transition_bottlenecks")
        else None
    )
    top_transition_bottleneck_label = build_transition_display_label(top_transition_bottleneck) or "該当なし"
    top_impact_row = impact_summary["rows"][0] if impact_summary.get("rows") else None

    common_rows = [
        ("平均ケース処理時間", dashboard_summary.get("avg_case_duration_text", "0s")),
        ("中央値ケース処理時間", dashboard_summary.get("median_case_duration_text", "0s")),
    ]

    if normalized_analysis_key == "frequency":
        frequency_rows = common_rows + [
            ("最大ケース処理時間", dashboard_summary.get("max_case_duration_text", "0s")),
            ("最多アクティビティ", top_row.get("アクティビティ", "該当なし") if top_row else "該当なし"),
            ("最多アクティビティ件数", normalize_excel_cell_value(top_row.get("イベント件数", 0)) if top_row else 0),
            ("上位10バリアントカバー率", f"{float(dashboard_summary.get('top10_variant_coverage_pct', 0.0)):.2f}%"),
        ]
        if variant_items is not None:
            frequency_rows.append(("バリアント総数", len(variant_items)))

        unique_activity_count = dashboard_summary.get("unique_activity_count")
        if unique_activity_count in (None, ""):
            unique_activity_count = dashboard_summary.get("activity_type_count")
        if unique_activity_count in (None, "") and prepared_df is not None and "activity" in prepared_df.columns:
            unique_activity_count = int(prepared_df["activity"].nunique())
        if unique_activity_count not in (None, ""):
            frequency_rows.append(("ユニークアクティビティ数", int(unique_activity_count)))

        total_cases = int(dashboard_summary.get("total_cases", 0) or 0)
        total_records = int(dashboard_summary.get("total_records", 0) or 0)
        if total_cases > 0:
            frequency_rows.append(("平均ケースあたりイベント数", round(total_records / total_cases, 2)))

        return frequency_rows

    if normalized_analysis_key == "transition":
        top_transition_label = (
            f"{top_row.get('前処理アクティビティ名', '')} → {top_row.get('後処理アクティビティ名', '')}".strip(" →")
            if top_row
            else "該当なし"
        )
        return common_rows + [
            ("主要遷移", top_transition_label or "該当なし"),
            ("最大ボトルネック遷移", top_transition_bottleneck_label),
            ("最大改善インパクト遷移", top_impact_row.get("transition_label", "該当なし") if top_impact_row else "該当なし"),
        ]

    if normalized_analysis_key == "pattern":
        return common_rows + [
            ("最頻出パターン", top_row.get("処理順パターン", top_row.get("パターン", "該当なし")) if top_row else "該当なし"),
            ("最頻出パターン比率", f"{float(top_row.get('ケース比率(%)', 0.0)):.2f}%" if top_row else "0.00%"),
            ("上位10バリアントカバー率", f"{float(dashboard_summary.get('top10_variant_coverage_pct', 0.0)):.2f}%"),
        ]

    return common_rows + [
        ("最大ケース処理時間", dashboard_summary.get("max_case_duration_text", "0s")),
        ("最大ボトルネック遷移", top_transition_bottleneck_label),
    ]


def build_excel_ai_summary(
    run_data,
    analysis_key,
    analysis_name,
    filter_params,
    prepared_df,
    variant_pattern,
    dashboard_summary,
    impact_summary,
    bottleneck_summary,
    analysis=None,
    root_cause_summary=None,
    insights_summary=None,
    variant_items=None,
    use_cache=True,
):
    return build_ai_insights_summary(
        run_data=run_data,
        analysis_key=analysis_key,
        filter_params=filter_params,
        prepared_df=prepared_df,
        variant_pattern=variant_pattern,
        analysis=analysis,
        dashboard_summary=dashboard_summary,
        impact_summary=impact_summary,
        bottleneck_summary=bottleneck_summary,
        root_cause_summary=root_cause_summary,
        insights_summary=insights_summary,
        variant_items=variant_items,
        analysis_name=analysis_name,
        force_refresh=False,
        use_cache=use_cache,
    )


def parse_transition_selection(selected_transition_key):
    normalized_key = str(selected_transition_key or "").strip()
    if "__TO__" not in normalized_key:
        return "", ""
    from_activity, to_activity = normalized_key.split("__TO__", 1)
    return from_activity.strip(), to_activity.strip()


def build_detail_export_workbook_bytes(
    run_id,
    run_data,
    analysis_key,
    filter_params,
    pattern_display_limit="10",
    variant_id=None,
    selected_activity="",
    selected_transition_key="",
    case_id="",
    drilldown_limit=20,
):
    analysis_definitions = get_available_analysis_definitions()
    analysis_name = analysis_definitions.get(analysis_key, {}).get("config", {}).get("analysis_name", analysis_key)
    export_sheet_keys = set(get_analysis_export_sheet_keys(analysis_key))
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["summary"])
    initialize_excel_worksheet(summary_sheet)

    variant_pattern = get_run_variant_pattern(
        run_data,
        variant_id=variant_id,
        filter_params=filter_params,
    )
    filtered_meta = get_filtered_meta_for_run(
        run_data,
        filter_params=filter_params,
        variant_pattern=variant_pattern,
    )
    selected_analysis = get_analysis_data(
        run_data,
        analysis_key,
        filter_params=filter_params,
        variant_pattern=variant_pattern,
    )
    export_variant_items = get_variant_items(
        run_data,
        filter_params=filter_params,
        variant_pattern=variant_pattern,
    )
    if variant_pattern:
        bottleneck_summary = query_bottleneck_summary(
            run_data["prepared_parquet_path"],
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
            limit=None,
        )
    else:
        bottleneck_summary = get_bottleneck_summary(
            run_data,
            variant_id=variant_id,
            filter_params=filter_params,
        )
    dashboard_summary = get_dashboard_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=None,
        variant_pattern=variant_pattern,
    )
    impact_summary = get_impact_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=None,
        variant_pattern=variant_pattern,
    )
    root_cause_summary = get_root_cause_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=None,
        variant_pattern=variant_pattern,
    )
    insights_summary = get_rule_based_insights_summary(
        run_data,
        analysis_key=analysis_key,
        analysis_rows=selected_analysis.get("rows", []),
        filter_params=filter_params,
        prepared_df=None,
        variant_pattern=variant_pattern,
        dashboard_summary=dashboard_summary,
        bottleneck_summary=bottleneck_summary,
        impact_summary=impact_summary,
    )
    ai_summary = build_excel_ai_summary(
        run_data=run_data,
        analysis_key=analysis_key,
        analysis_name=analysis_name,
        filter_params=filter_params,
        prepared_df=None,
        variant_pattern=variant_pattern,
        dashboard_summary=dashboard_summary,
        impact_summary=impact_summary,
        bottleneck_summary=bottleneck_summary,
        analysis=selected_analysis,
        root_cause_summary=root_cause_summary,
        insights_summary=insights_summary,
        variant_items=export_variant_items[:5],
        use_cache=variant_id is None,
    )

    from_activity, to_activity = parse_transition_selection(selected_transition_key)
    selected_transition_label = f"{from_activity} → {to_activity}" if from_activity and to_activity else str(selected_transition_key or "").strip()
    period_text = query_period_text(
        run_data["prepared_parquet_path"],
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    )
    group_columns = get_run_group_columns(run_data)
    grouping_text = "、".join(group_columns) if group_columns else "なし"
    summary_rows = [
        (REPORT_HEADER_LABELS["analysis_key"], analysis_key),
        (REPORT_HEADER_LABELS["analysis_name"], analysis_name),
        (REPORT_HEADER_LABELS["source_file_name"], run_data["source_file_name"]),
        (REPORT_HEADER_LABELS["analysis_executed_at"], run_data.get("created_at", "")),
        (REPORT_HEADER_LABELS["exported_at"], datetime.now(timezone.utc).isoformat()),
        (REPORT_HEADER_LABELS["case_count"], filtered_meta["case_count"]),
        (REPORT_HEADER_LABELS["event_count"], filtered_meta["event_count"]),
        (REPORT_HEADER_LABELS["applied_filters"], build_filter_summary_text(filter_params, run_data.get("column_settings"))),
        {"label": "", "value": APPLIED_FILTERS_NOTE_TEXT, "style": "note"},
        ("分析期間", period_text),
        ("グルーピング条件", grouping_text),
        {"label": "", "value": GROUPING_CONDITION_NOTE_TEXT, "style": "note"},
    ]
    if variant_id is not None:
        summary_rows.append((REPORT_HEADER_LABELS["selected_variant"], f"#{variant_id}"))
    if analysis_key != "frequency":
        summary_rows.extend(
            [
                (REPORT_HEADER_LABELS["selected_activity"], selected_activity or "未選択"),
                (REPORT_HEADER_LABELS["selected_transition"], selected_transition_label or "未選択"),
                (REPORT_HEADER_LABELS["selected_case_id"], case_id or "未選択"),
            ]
        )
    next_row = append_key_value_rows(
        summary_sheet,
        REPORT_SHEET_NAMES["summary"],
        summary_rows,
        description="対象範囲、選択条件、出力時点の情報をまとめています。",
    )
    kpi_rows = build_detail_summary_kpi_rows(
        analysis_key,
        selected_analysis.get("rows", []),
        dashboard_summary,
        impact_summary,
        bottleneck_summary,
        prepared_df=None,
        variant_items=export_variant_items,
    )
    next_row = append_key_value_rows(
        summary_sheet,
        "主要KPI",
        kpi_rows,
        start_row=next_row,
        description=f"{analysis_name} で優先して見たい代表値をまとめています。",
    )
    next_row = append_bullet_rows(
        summary_sheet,
        "分析ハイライト",
        ai_summary.get("highlights", []),
        start_row=next_row,
        column_count=4,
    )
    if group_columns:
        group_summary = query_group_summary(
            run_data["prepared_parquet_path"],
            group_columns,
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
        )
        if group_summary:
            group_summary_df = build_summary_sheet_df(group_summary, group_columns)
            next_row = append_table_to_worksheet(
                summary_sheet,
                "グループ別比較",
                group_summary_df.to_dict(orient="records"),
                list(group_summary_df.columns),
                start_row=next_row + 2,
                description="グルーピング条件ごとのケース数・処理時間の比較です。",
            )

    ai_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["ai_insights"]))
    initialize_excel_worksheet(ai_sheet)
    ai_meta_rows = [
        ("対象分析", analysis_name),
        ("分析期間", ai_summary.get("period", "不明")),
        ("出力時刻", ai_summary.get("generated_at", "")),
    ]
    next_row = append_key_value_rows(
        ai_sheet,
        REPORT_SHEET_NAMES["ai_insights"],
        ai_meta_rows,
        description="現在の分析条件に対応する分析コメント、または既存集計からの要約を掲載します。",
    )
    next_row = append_custom_text_section_to_worksheet(
        ai_sheet,
        "分析前提",
        ANALYSIS_PRECONDITIONS_TEXT,
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_ASSUMPTION_SECTION_FILL,
    )
    next_row = append_text_block_to_worksheet(
        ai_sheet,
        "解説本文",
        ai_summary.get("text", ""),
        start_row=next_row,
        column_count=6,
    )
    next_row = append_definition_table_to_worksheet(
        ai_sheet,
        "用語説明",
        TERMINOLOGY_ROWS,
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_MUTED_SECTION_FILL,
    )
    append_custom_text_section_to_worksheet(
        ai_sheet,
        "補足・免責事項",
        ai_summary.get("note", ""),
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_MUTED_SECTION_FILL,
        body_fill=EXCEL_LABEL_FILL,
    )

    if "frequency" in export_sheet_keys:
        frequency_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["frequency"]))
        initialize_excel_worksheet(frequency_sheet)
        if not group_columns:
            frequency_rows = build_ranked_rows(selected_analysis["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
            frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
            append_table_to_worksheet(
                frequency_sheet,
                REPORT_SHEET_NAMES["frequency"],
                frequency_rows,
                frequency_headers,
                description="アクティビティごとの件数、ケース数、処理時間の代表値を確認できます。",
            )
        else:
            overall_frequency_rows = query_analysis_records(
                run_data["prepared_parquet_path"],
                "frequency",
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
                variant_pattern=variant_pattern,
            )["rows"]

            current_row = 1
            max_frequency_columns = max(10, len(overall_frequency_rows[0]) + 1 if overall_frequency_rows else 10)
            current_row = _write_section_header(
                frequency_sheet,
                current_row,
                "全体",
                column_count=max_frequency_columns,
            )
            current_row = _write_frequency_data(
                frequency_sheet,
                overall_frequency_rows,
                current_row,
            )

            for group_name, group_frequency_rows in _iter_groups_from_parquet(
                run_data["prepared_parquet_path"],
                group_columns,
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
                variant_pattern=variant_pattern,
            ) or []:
                current_row += 3
                group_column_count = max(10, len(group_frequency_rows[0]) + 1 if group_frequency_rows else max_frequency_columns)
                current_row = _write_section_header(
                    frequency_sheet,
                    current_row,
                    f"グループ: {group_name}",
                    column_count=group_column_count,
                )
                current_row = _write_frequency_data(
                    frequency_sheet,
                    group_frequency_rows,
                    current_row,
                )

    if "transition" in export_sheet_keys:
        transition_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["transition"]))
        initialize_excel_worksheet(transition_sheet)
        transition_rows = build_ranked_rows(selected_analysis["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
        transition_headers = list(transition_rows[0].keys()) if transition_rows else [REPORT_HEADER_LABELS["rank"]]
        append_table_to_worksheet(
            transition_sheet,
            REPORT_SHEET_NAMES["transition"],
            transition_rows,
            transition_headers,
            description="前後遷移ごとの件数、ケース数、平均所要時間を確認できます。",
        )

    if "pattern" in export_sheet_keys:
        pattern_config = analysis_definitions.get("pattern", {}).get("config", {})
        pattern_display_columns = pattern_config.get("display_columns", {})
        pattern_column_label = pattern_display_columns.get("pattern", "処理順パターン")
        pattern_summary = build_pattern_export_summary(
            selected_analysis["rows"],
            pattern_display_columns,
        )
        pattern_conclusion = build_pattern_conclusion_summary(pattern_summary)
        pattern_dashboard = build_pattern_dashboard_summary(pattern_summary, pattern_conclusion)
        conclusion_sheet = workbook.create_sheet(
            title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["pattern_conclusion"])
        )
        initialize_excel_worksheet(conclusion_sheet)
        conclusion_next_row = append_key_value_rows(
            conclusion_sheet,
            REPORT_SHEET_NAMES["pattern_conclusion"],
            [
                ("全体要約", pattern_conclusion["overall_summary"]),
                ("改善による時間インパクト(分)", pattern_conclusion["total_impact_minutes"]),
                ("改善による時間インパクト(時間)", pattern_conclusion["total_impact_hours"]),
                ("最短処理パターン", pattern_conclusion["fastest_pattern"].get("パターン", "該当なし")),
                ("最短処理パターン平均処理時間(分)", pattern_conclusion["fastest_pattern"].get("平均処理時間(分)", 0)),
            ],
            description="処理順パターン分析の結論、改善優先度、想定時間効果をまとめています。",
        )
        conclusion_next_row = append_table_to_worksheet(
            conclusion_sheet,
            "問題点3つ",
            pattern_conclusion["issue_rows"],
            ["問題点", "原因", "改善案", "期待効果（時間短縮・分）", "対象パターン"],
            start_row=conclusion_next_row,
            description="改善対象パターンTOP3を中心に、問題点・原因・改善案・期待効果を整理しています。",
            no_wrap_headers=["対象パターン"],
            min_column_widths={"問題点": 40, "原因": 38, "改善案": 42, "対象パターン": 72},
        )
        append_pattern_conclusion_charts(
            workbook,
            conclusion_sheet,
            pattern_summary["comparison_rows"],
            pie_anchor=build_excel_anchor("A", conclusion_next_row + 1),
            bar_anchor=build_excel_anchor("C", conclusion_next_row + 1),
        )
        dashboard_sheet = workbook.create_sheet(
            title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["pattern_dashboard"])
        )
        initialize_excel_worksheet(dashboard_sheet)
        dashboard_next_row = append_key_value_rows(
            dashboard_sheet,
            REPORT_SHEET_NAMES["pattern_dashboard"],
            [
                ("全体要約", pattern_dashboard["overall_summary"]),
                ("上位10パターン累積カバー率(%)", pattern_dashboard["top10_coverage_pct"]),
                ("改善による時間インパクト(分)", pattern_dashboard["total_impact_minutes"]),
            ],
            description="処理順パターン分析の主要サマリーをダッシュボード形式でまとめています。",
        )
        dashboard_next_row = append_table_to_worksheet(
            dashboard_sheet,
            "改善優先TOP3",
            pattern_dashboard["top3_rows"],
            ["順位", "パターン", "改善優先度スコア", "全体影響度(%)", "繰り返し率(%)", "平均処理時間差分(分)", "簡易コメント"],
            start_row=dashboard_next_row,
            description="改善優先度スコアが高い上位3パターンです。",
            no_wrap_headers=["パターン", "簡易コメント"],
            min_column_widths={"パターン": 72, "簡易コメント": 36},
        )
        dashboard_next_row = append_bullet_rows(
            dashboard_sheet,
            "問題点",
            pattern_dashboard["problem_points"],
            start_row=dashboard_next_row,
            column_count=6,
            empty_text="抽出できる問題点はありません。",
        )
        if "pattern_summary" in export_sheet_keys:
            pattern_summary_sheet = workbook.create_sheet(
                title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["pattern_summary"])
            )
            initialize_excel_worksheet(pattern_summary_sheet)
            next_row = append_key_value_rows(
                pattern_summary_sheet,
                REPORT_SHEET_NAMES["pattern_summary"],
                [
                    ("上位3パターン累積カバー率(%)", pattern_summary["top3_coverage_pct"]),
                    ("上位10パターン累積カバー率(%)", pattern_summary["top10_coverage_pct"]),
                    ("カバー率要約", pattern_summary["coverage_summary_text"]),
                    ("最短処理パターン", pattern_summary["fastest_pattern"].get("パターン", "該当なし")),
                    ("最短処理パターン平均処理時間(分)", pattern_summary["fastest_pattern"].get("平均処理時間(分)", 0)),
                    ("要確認パターン数", len(pattern_summary["repeated_patterns"])),
                    ("要確認パターン影響比率(%)", pattern_summary["repeated_case_ratio_pct"]),
                    ("要確認判定基準", "繰り返し率 30%以上"),
                    ("改善対象抽出基準", "繰り返し率 10%以上 かつ 平均処理時間差分がプラス"),
                ],
                description="処理順パターンの上位カバー率と、繰り返し率が高いパターンをまとめています。",
            )
            next_row = append_table_to_worksheet(
                pattern_summary_sheet,
                "上位10パターン",
                pattern_summary["comparison_rows"],
                ["順位", "繰り返し", "繰り返し回数", "繰り返し率(%)", "繰り返し率区分", "件数", "全体比率(%)", "平均処理時間(分)", "平均処理時間差分(分)", "改善優先度スコア", "全体影響度(%)", "最短処理", "最短処理時間(分)", "最長処理時間(分)", "確認区分", "簡易コメント", "パターン"],
                start_row=next_row,
                description="件数上位10パターンの比率・処理時間・繰り返し率を比較できます。",
                no_wrap_headers=["パターン", "簡易コメント"],
                min_column_widths={"パターン": 72, "簡易コメント": 48},
            )
            next_row = append_table_to_worksheet(
                pattern_summary_sheet,
                "要確認パターン一覧",
                pattern_summary["repeated_patterns"],
                ["順位", "繰り返し", "繰り返し回数", "繰り返し率(%)", "繰り返し率区分", "件数", "全体比率(%)", "平均処理時間(分)", "平均処理時間差分(分)", "改善優先度スコア", "全体影響度(%)", "最短処理", "最短処理時間(分)", "最長処理時間(分)", "確認区分", "簡易コメント", "パターン"],
                start_row=next_row,
                description="繰り返し率が高く、確認を優先したいパターンを一覧化しています。",
                no_wrap_headers=["パターン", "簡易コメント"],
                min_column_widths={"パターン": 72, "簡易コメント": 48},
            )
            append_table_to_worksheet(
                pattern_summary_sheet,
                "改善対象パターンTOP3",
                pattern_summary["improvement_targets"],
                ["順位", "繰り返し", "繰り返し回数", "繰り返し率(%)", "繰り返し率区分", "件数", "全体比率(%)", "平均処理時間(分)", "平均処理時間差分(分)", "改善優先度スコア", "全体影響度(%)", "最短処理", "確認区分", "簡易コメント", "パターン"],
                start_row=next_row,
                description="繰り返し率が一定以上で、平均処理時間も全体平均より長い改善候補パターンです。",
                no_wrap_headers=["パターン", "簡易コメント"],
                min_column_widths={"パターン": 72, "簡易コメント": 48},
            )

        pattern_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["pattern"]))
        initialize_excel_worksheet(pattern_sheet)
        pattern_rows, pattern_headers = localize_report_rows(
            build_pattern_overview_rows(
                selected_analysis["rows"],
                export_variant_items,
                pattern_column_label,
                analysis_definitions,
            ),
            [
                "rank",
                "pattern_variant",
                "repeat_flag",
                "repeat_count",
                "repeat_rate_pct",
                "repeat_rate_band",
                "review_flag",
                "count",
                "ratio",
                "cumulative_case_ratio_pct",
                "avg_case_duration_min",
                "avg_case_duration_diff_min",
                "improvement_priority_score",
                "overall_impact_pct",
                "fastest_pattern_flag",
                "std_case_duration_min",
                "min_case_duration_min",
                "max_case_duration_min",
                "p75_case_duration_min",
                "p90_case_duration_min",
                "p95_case_duration_min",
                "simple_comment",
                "pattern",
            ],
        )
        append_table_to_worksheet(
            pattern_sheet,
            REPORT_SHEET_NAMES["pattern"],
            pattern_rows,
            pattern_headers,
            description="パターン / バリアントを 1 つの一覧にまとめ、繰り返し回数・繰り返し率・確認区分・件数・比率・累積カバー率・処理時間の代表値と代表ルートを比較できます。",
            no_wrap_headers=["パターン", "簡易コメント"],
            min_column_widths={"パターン": 80, "簡易コメント": 48},
        )
        variant_by_pattern = {
            str(variant_item.get("pattern") or "").strip(): variant_item
            for variant_item in export_variant_items
        }
        pattern_detail_count = resolve_pattern_detail_sheet_count(
            pattern_display_limit,
            len(selected_analysis.get("rows", [])),
        )
        for pattern_rank, pattern_row in enumerate(selected_analysis.get("rows", [])[:pattern_detail_count], start=1):
            pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
            pattern_detail = None
            if pattern_text:
                pattern_detail = query_pattern_bottleneck_details(
                    run_data["prepared_parquet_path"],
                    pattern_text,
                    filter_params=filter_params,
                    filter_column_settings=run_data.get("column_settings"),
                    scope_variant_pattern=variant_pattern,
                )
            append_pattern_detail_sheet(
                workbook,
                None,
                pattern_row,
                pattern_rank,
                pattern_column_label,
                analysis_definitions,
                variant_item=variant_by_pattern.get(str(pattern_row.get(pattern_column_label) or "").strip()),
                pattern_detail=pattern_detail,
            )

    if "bottleneck" in export_sheet_keys:
        bottleneck_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["bottleneck"]))
        initialize_excel_worksheet(bottleneck_sheet)
        activity_bottleneck_rows, activity_bottleneck_headers = localize_report_rows(
            build_bottleneck_export_rows(
                bottleneck_summary["activity_bottlenecks"],
                "activity",
            ),
            ["rank", "activity", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
        )
        next_row = append_table_to_worksheet(
            bottleneck_sheet,
            "アクティビティボトルネック",
            activity_bottleneck_rows,
            activity_bottleneck_headers,
            description="アクティビティ単位で所要時間が大きい箇所を並べています。",
        )
        transition_bottleneck_rows, transition_bottleneck_headers = localize_report_rows(
            build_bottleneck_export_rows(
                bottleneck_summary["transition_bottlenecks"],
                "transition_label",
            ),
            ["rank", "transition_label", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
        )
        append_table_to_worksheet(
            bottleneck_sheet,
            "遷移ボトルネック",
            transition_bottleneck_rows,
            transition_bottleneck_headers,
            start_row=next_row,
            description="前後遷移ごとの平均所要時間・中央値・最大値を比較できます。",
        )

    if "impact" in export_sheet_keys:
        impact_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["impact"]))
        initialize_excel_worksheet(impact_sheet)
        impact_rows, impact_headers = localize_report_rows(
            [
                {
                    "rank": impact_row["rank"],
                    "transition": impact_row["transition_label"],
                    "case_count": impact_row["case_count"],
                    "avg_duration": impact_row["avg_duration_text"],
                    "max_duration": impact_row["max_duration_text"],
                    "impact_score": impact_row["impact_score"],
                    "impact_share_pct": impact_row["impact_share_pct"],
                }
                for impact_row in impact_summary["rows"]
            ],
            ["rank", "transition", "case_count", "avg_duration", "max_duration", "impact_score", "impact_share_pct"],
        )
        append_table_to_worksheet(
            impact_sheet,
            REPORT_SHEET_NAMES["impact"],
            impact_rows,
            impact_headers,
            description="改善インパクトが高い遷移を優先順位付きで確認できます。",
        )

    selected_activity_name = str(selected_activity or "").strip()
    drilldown_rows = []
    drilldown_title = REPORT_SHEET_NAMES["drilldown"]
    if from_activity and to_activity:
        drilldown_title = f"遷移ドリルダウン: {from_activity} → {to_activity}"
        drilldown_rows = query_transition_case_drilldown(
            run_data["prepared_parquet_path"],
            from_activity=from_activity,
            to_activity=to_activity,
            limit=max(0, int(drilldown_limit)),
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
        )
    elif selected_activity_name:
        drilldown_title = f"アクティビティドリルダウン: {selected_activity_name}"
        drilldown_rows = query_activity_case_drilldown(
            run_data["prepared_parquet_path"],
            activity=selected_activity_name,
            limit=max(0, int(drilldown_limit)),
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
        )
    if drilldown_rows:
        drilldown_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["drilldown"]))
        initialize_excel_worksheet(drilldown_sheet)
        drilldown_rows, drilldown_headers = localize_report_rows(
            drilldown_rows,
            ["case_id", "activity", "next_activity", "duration_text", "from_time", "to_time"],
        )
        append_table_to_worksheet(
            drilldown_sheet,
            drilldown_title,
            drilldown_rows,
            drilldown_headers,
            description="選択中アクティビティ / 遷移に該当するケースの明細です。",
        )

    normalized_case_id = str(case_id or "").strip()
    if normalized_case_id:
        case_trace = query_case_trace_details(
            run_data["prepared_parquet_path"],
            normalized_case_id,
        )
        if case_trace.get("found"):
            case_trace_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["case_trace"]))
            initialize_excel_worksheet(case_trace_sheet)
            next_row = append_key_value_rows(
                case_trace_sheet,
                "ケース概要",
                [
                    (REPORT_HEADER_LABELS["case_id"], case_trace["case_id"]),
                    (REPORT_HEADER_LABELS["event_count"], case_trace["summary"]["event_count"]),
                    (REPORT_HEADER_LABELS["total_duration"], case_trace["summary"]["total_duration_text"]),
                    (REPORT_HEADER_LABELS["start_time"], case_trace["summary"]["start_time"]),
                    (REPORT_HEADER_LABELS["end_time"], case_trace["summary"]["end_time"]),
                ],
                description="指定したケースの概要情報です。",
            )
            case_trace_event_rows, case_trace_event_headers = localize_report_rows(
                case_trace["events"],
                ["case_id", "activity", "next_activity", "start_time", "end_time", "duration_text"],
            )
            append_table_to_worksheet(
                case_trace_sheet,
                "通過イベント",
                case_trace_event_rows,
                case_trace_event_headers,
                start_row=next_row,
                description="ケース内で通過したイベントを時系列順に並べています。",
            )

    for worksheet in workbook.worksheets:
        autosize_worksheet_columns(worksheet)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()


def get_filter_options_payload(run_data):
    filter_options = run_data.get("filter_options")
    if filter_options is None:
        prepared_parquet_path = run_data.get("prepared_parquet_path")
        if not prepared_parquet_path:
            raise HTTPException(
                status_code=500,
                detail="フィルター候補を再構築できませんでした。",
            )
        filter_options = query_filter_options(
            prepared_parquet_path,
            filter_column_settings=run_data.get("column_settings"),
        )
        run_data["filter_options"] = filter_options

    return filter_options


def _to_int(value, default=0):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return int(default)


def _to_float(value, default=0.0):
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return float(default)


def build_variant_items_from_pattern_rows(pattern_rows):
    if not pattern_rows:
        return []

    total_cases = sum(
        _to_int(row.get("ケース数", row.get("case_count", 0)))
        for row in pattern_rows
    )
    items = []

    for index, row in enumerate(pattern_rows, start=1):
        pattern_text = str(
            row.get("処理順パターン")
            or row.get("pattern")
            or ""
        ).strip()
        activities = [step.strip() for step in pattern_text.split("→") if step.strip()]
        case_count = _to_int(row.get("ケース数", row.get("case_count", 0)))
        ratio_pct = row.get("ケース比率(%)", row.get("case_ratio_pct"))
        avg_case_duration_min = _to_float(
            row.get("平均ケース処理時間(分)", row.get("平均ケース時間(分)", row.get("avg_case_duration_min", 0.0)))
        )

        if ratio_pct in (None, ""):
            ratio = round(case_count / total_cases, 4) if total_cases else 0.0
        else:
            ratio = round(_to_float(ratio_pct) / 100, 4)

        avg_case_duration_sec = round(avg_case_duration_min * 60, 2)
        items.append(
            {
                "variant_id": index,
                "activities": activities,
                "activity_count": len(activities),
                "pattern": pattern_text,
                "count": case_count,
                "ratio": ratio,
                "avg_case_duration_sec": avg_case_duration_sec,
                "avg_case_duration_text": format_duration_text_for_report(avg_case_duration_sec),
                "repeat_flag": "○" if len(set(activities)) < len(activities) else "",
            }
        )

    return items


def is_unfiltered_request(filter_cache_key):
    return all(filter_value is None for filter_value in filter_cache_key)


def get_variant_items(run_data, filter_params=None, variant_pattern=None):
    cache_key = (build_filter_cache_key(filter_params), str(variant_pattern or "").strip() or None)
    variant_cache = run_data.setdefault("variant_cache", {})

    if cache_key not in variant_cache:
        filter_cache_key = cache_key[0]
        if variant_pattern is None and is_unfiltered_request(filter_cache_key):
            pattern_analysis = run_data["result"]["analyses"].get("pattern")
            pattern_rows = pattern_analysis.get("rows", []) if pattern_analysis else []
            if pattern_rows:
                variant_cache[cache_key] = build_variant_items_from_pattern_rows(pattern_rows)
                return variant_cache[cache_key]

        variant_cache[cache_key] = query_variant_summary(
            run_data["prepared_parquet_path"],
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
            limit=None,
        )

    return variant_cache[cache_key]


def get_variant_item(run_data, variant_id, filter_params=None):
    safe_variant_id = int(variant_id)

    for variant_item in get_variant_items(run_data, filter_params=filter_params):
        if variant_item["variant_id"] == safe_variant_id:
            return variant_item

    raise HTTPException(status_code=404, detail="バリアントが見つかりません。")


def get_pattern_summary_row(run_data, pattern_index):
    pattern_analysis = run_data["result"]["analyses"].get("pattern")

    if not pattern_analysis:
        raise HTTPException(status_code=400, detail="処理順パターン分析を利用できません。")

    pattern_rows = pattern_analysis["rows"]
    safe_pattern_index = int(pattern_index)
    if safe_pattern_index < 0 or safe_pattern_index >= len(pattern_rows):
        raise HTTPException(status_code=404, detail="パターン番号が見つかりません。")

    pattern_index_entries = run_data.get("pattern_index_entries")
    if pattern_index_entries is None:
        pattern_index_entries = build_pattern_index_entries_from_rows(pattern_rows)
        run_data["pattern_index_entries"] = pattern_index_entries

    if safe_pattern_index >= len(pattern_index_entries):
        raise HTTPException(status_code=404, detail="パターン番号が見つかりません。")

    summary_row = pattern_rows[safe_pattern_index]
    pattern_entry = pattern_index_entries[safe_pattern_index]
    pattern = str(pattern_entry.get("pattern") or "").strip() or extract_pattern_text_from_row(summary_row)

    if not pattern:
        raise HTTPException(status_code=500, detail="パターン文字列を解決できませんでした。")

    return pattern_analysis, summary_row, pattern_entry, pattern


def get_analysis_data(run_data, analysis_key, filter_params=None, variant_pattern=None):
    normalized_filter_key = build_filter_cache_key(filter_params)
    analysis_cache = run_data.setdefault("analysis_cache", {})

    if variant_pattern is None and all(filter_value is None for filter_value in normalized_filter_key):
        analysis = run_data["result"]["analyses"].get(analysis_key)
        if analysis:
            return analysis

        cache_key = ("analysis", analysis_key, normalized_filter_key, None)
        if cache_key not in analysis_cache:
            analysis_cache[cache_key] = query_analysis_records(
                run_data["prepared_parquet_path"],
                analysis_key=analysis_key,
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
                variant_pattern=None,
            )
        return analysis_cache[cache_key]

    cache_key = ("analysis", analysis_key, normalized_filter_key, str(variant_pattern or "").strip() or None)

    if cache_key not in analysis_cache:
        analysis_cache[cache_key] = query_analysis_records(
            run_data["prepared_parquet_path"],
            analysis_key=analysis_key,
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
        )

    return analysis_cache[cache_key]


def get_bottleneck_summary(run_data, variant_id=None, pattern_index=None, filter_params=None):
    cache_key = (
        "bottleneck",
        None if variant_id is None else int(variant_id),
        None if pattern_index is None else int(pattern_index),
        build_filter_cache_key(filter_params),
    )
    cache = run_data.setdefault("bottleneck_cache", {})

    if cache_key not in cache:
        cache[cache_key] = query_bottleneck_summary(
            run_data["prepared_parquet_path"],
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=get_run_variant_pattern(
                run_data,
                variant_id=variant_id,
                pattern_index=pattern_index,
                filter_params=filter_params,
            ),
            limit=None,
        )

    return cache[cache_key]


def get_dashboard_summary(run_data, filter_params=None, prepared_df=None, variant_pattern=None):
    cache_key = (build_filter_cache_key(filter_params), str(variant_pattern or "").strip() or None)
    cache = run_data.setdefault("dashboard_cache", {})

    if cache_key not in cache:
        if variant_pattern:
            resolved_bottleneck_summary = query_bottleneck_summary(
                run_data["prepared_parquet_path"],
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
                variant_pattern=variant_pattern,
                limit=None,
            )
        else:
            resolved_bottleneck_summary = get_bottleneck_summary(run_data, filter_params=filter_params)
        cache[cache_key] = query_dashboard_summary(
            run_data["prepared_parquet_path"],
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
            variant_items=get_variant_items(
                run_data,
                filter_params=filter_params,
                variant_pattern=variant_pattern,
            )[:10],
            bottleneck_summary=resolved_bottleneck_summary,
            coverage_limit=10,
        )

    return cache[cache_key]


def get_root_cause_summary(run_data, filter_params=None, prepared_df=None, variant_pattern=None):
    cache_key = (build_filter_cache_key(filter_params), str(variant_pattern or "").strip() or None)
    cache = run_data.setdefault("root_cause_cache", {})

    if cache_key not in cache:
        cache[cache_key] = query_root_cause_summary(
            run_data["prepared_parquet_path"],
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
            limit=10,
        )

    return cache[cache_key]


def get_impact_summary(run_data, filter_params=None, prepared_df=None, variant_pattern=None):
    cache_key = (build_filter_cache_key(filter_params), str(variant_pattern or "").strip() or None)
    cache = run_data.setdefault("impact_cache", {})

    if cache_key not in cache:
        cache[cache_key] = query_impact_summary(
            run_data["prepared_parquet_path"],
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=variant_pattern,
            limit=None,
        )

    return cache[cache_key]


def get_rule_based_insights_summary(
    run_data,
    analysis_key,
    analysis_rows=None,
    filter_params=None,
    prepared_df=None,
    variant_pattern=None,
    dashboard_summary=None,
    bottleneck_summary=None,
    impact_summary=None,
):
    cache_key = (
        str(analysis_key or "").strip().lower(),
        build_filter_cache_key(filter_params),
        str(variant_pattern or "").strip() or None,
    )
    cache = run_data.setdefault("insights_cache", {})

    if cache_key not in cache:
        normalized_analysis_key = str(analysis_key or "").strip().lower()
        filtered_df = prepared_df
        if filtered_df is None:
            filtered_df = pd.DataFrame(columns=["case_id", "activity", "duration_sec"])

        resolved_bottleneck_summary = bottleneck_summary
        resolved_impact_summary = impact_summary
        if normalized_analysis_key in {"frequency", "transition", "pattern"}:
            if resolved_bottleneck_summary is None:
                resolved_bottleneck_summary = {
                    "activity_bottlenecks": [],
                    "transition_bottlenecks": [],
                    "activity_heatmap": {},
                    "transition_heatmap": {},
                }
            if resolved_impact_summary is None:
                resolved_impact_summary = {
                    "has_data": False,
                    "rows": [],
                }
        else:
            if resolved_bottleneck_summary is None:
                resolved_bottleneck_summary = get_bottleneck_summary(
                    run_data,
                    filter_params=filter_params,
                )
            if resolved_impact_summary is None:
                resolved_impact_summary = get_impact_summary(
                    run_data,
                    filter_params=filter_params,
                    variant_pattern=variant_pattern,
                )

        cache[cache_key] = create_rule_based_insights(
            filtered_df,
            analysis_key=analysis_key,
            analysis_rows=analysis_rows,
            dashboard_summary=dashboard_summary or get_dashboard_summary(
                run_data,
                filter_params=filter_params,
                variant_pattern=variant_pattern,
            ),
            bottleneck_summary=resolved_bottleneck_summary,
            impact_summary=resolved_impact_summary,
            max_items=5,
        )

    return cache[cache_key]


def get_pattern_flow_snapshot(
    run_data,
    pattern_percent,
    pattern_count,
    activity_percent,
    connection_percent,
    variant_id=None,
    filter_params=None,
):
    filter_cache_key = build_filter_cache_key(filter_params)
    if variant_id is None:
        cache_key = (
            "pattern",
            int(pattern_percent),
            None if pattern_count is None else int(pattern_count),
            int(activity_percent),
            int(connection_percent),
            filter_cache_key,
        )
    else:
        cache_key = (
            "variant",
            int(variant_id),
            int(activity_percent),
            int(connection_percent),
            filter_cache_key,
        )
    cache = run_data.setdefault("pattern_flow_cache", OrderedDict())

    cached_snapshot = cache.get(cache_key)
    if cached_snapshot is not None:
        cache.move_to_end(cache_key)
        return cached_snapshot

    analyses = run_data["result"]["analyses"]
    variant_item = None
    variant_pattern = None
    if variant_id is not None:
        variant_item = get_variant_item(run_data, variant_id, filter_params=filter_params)
        variant_pattern = variant_item["pattern"]

    scoped_meta = get_filtered_meta_for_run(
        run_data,
        filter_params=filter_params,
        variant_pattern=variant_pattern,
    )
    scoped_event_count = int(scoped_meta.get("event_count") or 0)
    use_lightweight_flow = scoped_event_count >= LARGE_DATASET_FLOW_FAST_PATH_THRESHOLD

    if variant_id is None:
        pattern_analysis = get_analysis_data(run_data, "pattern", filter_params=filter_params)
        frequency_analysis = (
            analyses.get("frequency")
            if is_unfiltered_request(filter_cache_key)
            else get_analysis_data(run_data, "frequency", filter_params=filter_params)
        )
        selected_transition_rows = []
        pattern_selection = select_pattern_rows_for_flow(
            pattern_analysis["rows"],
            pattern_percent=pattern_percent,
            pattern_count=pattern_count,
            pattern_cap=PROCESS_FLOW_PATTERN_CAP,
        )
        selected_patterns = []
        for row in pattern_selection["selected_pattern_rows"]:
            selected_pattern = extract_pattern_text_from_row(row)
            if selected_pattern:
                selected_patterns.append(selected_pattern)
        if not use_lightweight_flow and selected_patterns:
            selected_transition_rows = query_transition_records_for_patterns(
                run_data["prepared_parquet_path"],
                selected_patterns,
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
            )

        snapshot = create_pattern_flow_snapshot(
            pattern_rows=pattern_analysis["rows"],
            prepared_df=None,
            transition_rows=selected_transition_rows,
            frequency_rows=(frequency_analysis or {}).get("rows", []),
            pattern_percent=pattern_percent,
            pattern_count=pattern_count,
            activity_percent=activity_percent,
            connection_percent=connection_percent,
            pattern_cap=PROCESS_FLOW_PATTERN_CAP,
        )
    else:
        if use_lightweight_flow:
            snapshot = create_pattern_flow_snapshot(
                pattern_rows=[
                    {
                        FLOW_PATTERN_CASE_COUNT_COLUMN: int(variant_item["count"]),
                        FLOW_PATTERN_COLUMN: variant_pattern,
                    }
                ],
                prepared_df=None,
                frequency_rows=[],
                pattern_percent=100,
                pattern_count=1,
                activity_percent=activity_percent,
                connection_percent=connection_percent,
                pattern_cap=1,
            )
        else:
            selected_transition_rows = query_transition_records_for_patterns(
                run_data["prepared_parquet_path"],
                [variant_pattern],
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
            )
            snapshot = create_pattern_flow_snapshot(
                pattern_rows=[
                    {
                        FLOW_PATTERN_CASE_COUNT_COLUMN: int(variant_item["count"]),
                        FLOW_PATTERN_COLUMN: variant_pattern,
                    }
                ],
                prepared_df=None,
                transition_rows=selected_transition_rows,
                frequency_rows=[],
                pattern_percent=100,
                pattern_count=1,
                activity_percent=activity_percent,
                connection_percent=connection_percent,
                pattern_cap=1,
            )
        snapshot["selected_variant"] = build_variant_response_item(variant_item, run_data=run_data)

    if use_lightweight_flow:
        snapshot["is_large_dataset_optimized"] = True

    cache[cache_key] = snapshot
    cache.move_to_end(cache_key)
    while len(cache) > MAX_PATTERN_FLOW_CACHE:
        cache.popitem(last=False)

    return snapshot


def build_preview_response(run_id, source_file_name, selected_analysis_keys, result, run_data):
    return {
        "run_id": run_id,
        "source_file_name": source_file_name,
        "selected_analysis_keys": selected_analysis_keys,
        "case_count": result["case_count"],
        "event_count": result["event_count"],
        "group_columns": result.get("group_columns", []),
        "group_mode": result.get("group_mode", False),
        "group_summary": result.get("group_summary", {}),
        "applied_filters": run_data.get("base_filter_params"),
        "column_settings": build_column_settings_payload(run_data.get("column_settings")),
        "filter_options": get_filter_options_payload(run_data),
        "analyses": {
            analysis_key: build_analysis_payload(analysis, PREVIEW_ROW_COUNT)
            for analysis_key, analysis in result["analyses"].items()
        },
    }


def build_log_profile_payload(
    raw_df,
    source_file_name,
    case_id_column="",
    activity_column="",
    timestamp_column="",
    filter_column_settings=None,
    include_diagnostics=False,
):
    headers = [str(column_name) for column_name in raw_df.columns.tolist()]
    selection_payload = build_column_selection_payload(headers)
    resolved_case_id_column = suggest_column_name(headers, "case_id_column", case_id_column)
    resolved_activity_column = suggest_column_name(headers, "activity_column", activity_column)
    resolved_timestamp_column = suggest_column_name(headers, "timestamp_column", timestamp_column)
    normalized_filter_column_settings = normalize_filter_column_settings(**(filter_column_settings or {}))

    return {
        "source_file_name": source_file_name,
        **selection_payload,
        "column_settings": build_column_settings_payload(
            {
                "case_id_column": resolved_case_id_column,
                "activity_column": resolved_activity_column,
                "timestamp_column": resolved_timestamp_column,
                **normalized_filter_column_settings,
            }
        ),
        "filter_options": get_filter_options(
            raw_df,
            filter_column_settings=normalized_filter_column_settings,
        ),
        "diagnostics": (
            create_log_diagnostics(
                raw_df,
                case_id_column=resolved_case_id_column,
                activity_column=resolved_activity_column,
                timestamp_column=resolved_timestamp_column,
                filter_column_settings=normalized_filter_column_settings,
            )
            if include_diagnostics
            else None
        ),
    }


def resolve_log_diagnostic_sample_row_limit(raw_value):
    try:
        sample_row_limit = int(str(raw_value or "").strip() or DEFAULT_LOG_DIAGNOSTIC_SAMPLE_ROW_LIMIT)
    except (TypeError, ValueError):
        sample_row_limit = DEFAULT_LOG_DIAGNOSTIC_SAMPLE_ROW_LIMIT

    if sample_row_limit <= 0:
        sample_row_limit = DEFAULT_LOG_DIAGNOSTIC_SAMPLE_ROW_LIMIT

    return min(MAX_LOG_DIAGNOSTIC_SAMPLE_ROW_LIMIT, sample_row_limit)


def build_log_diagnostic_period_text(diagnostics):
    time_range = (diagnostics or {}).get("time_range") or {}
    min_time = str(time_range.get("min") or "").strip()
    max_time = str(time_range.get("max") or "").strip()
    if not min_time or not max_time:
        return "ケースID / アクティビティ / タイムスタンプ列を選択すると表示します。"
    return f"{min_time} 〜 {max_time}"


def build_log_diagnostic_missing_count_text(diagnostics):
    missing_counts = (diagnostics or {}).get("missing_counts") or {}
    return (
        f"ケースID {missing_counts.get('case_id', '-') if missing_counts.get('case_id') is not None else '-'}"
        f" / アクティビティ {missing_counts.get('activity', '-') if missing_counts.get('activity') is not None else '-'}"
        f" / タイムスタンプ {missing_counts.get('timestamp', '-') if missing_counts.get('timestamp') is not None else '-'}"
    )


def build_log_diagnostic_duplicate_rate_text(diagnostics):
    duplicate_rate = float((diagnostics or {}).get("duplicate_rate") or 0.0)
    return f"{duplicate_rate * 100:.1f}%"


def build_log_diagnostic_filter_rows(profile_payload, preview_limit=30):
    diagnostics = (profile_payload or {}).get("diagnostics") or {}
    diagnostics_filter_rows = {
        str(row.get("slot") or ""): row
        for row in (diagnostics.get("filters") or [])
        if str(row.get("slot") or "").strip()
    }
    column_settings = (profile_payload or {}).get("column_settings") or {}
    filter_definitions = column_settings.get("filters") or []
    rows = []

    for filter_definition in filter_definitions:
        slot = str(filter_definition.get("slot") or "").strip()
        diagnostics_row = diagnostics_filter_rows.get(slot, {})
        options = diagnostics_row.get("options") or []
        option_preview_values = [str(option) for option in options[:preview_limit]]
        option_preview = ", ".join(option_preview_values) if option_preview_values else "-"
        if len(options) > preview_limit:
            option_preview = f"{option_preview} ... 他 {len(options) - preview_limit} 件"

        rows.append(
            {
                "スロット": slot or "-",
                "表示名": str(filter_definition.get("label") or "").strip() or "-",
                "対象列": str(filter_definition.get("column_name") or "").strip() or "未設定",
                "候補数": int(len(options)),
                "候補一覧": option_preview,
            }
        )

    return rows


def build_log_diagnostic_sample_rows(raw_df, sample_row_limit):
    headers = [str(column_name) for column_name in raw_df.columns.tolist()]
    sampled_df = raw_df.head(max(0, int(sample_row_limit))).copy()
    sample_rows = []

    for row_number, (original_index, row) in enumerate(sampled_df.iterrows(), start=1):
        row_payload = {
            "レコード順": int(original_index) + 1,
        }
        for header in headers:
            row_payload[header] = row.get(header, "")
        sample_rows.append(row_payload)

    return sample_rows, ["レコード順", *headers]


def build_log_diagnostic_workbook_bytes(profile_payload, raw_df, sample_row_limit):
    diagnostics = (profile_payload or {}).get("diagnostics") or {}
    column_settings = (profile_payload or {}).get("column_settings") or {}
    sample_rows, sample_headers = build_log_diagnostic_sample_rows(raw_df, sample_row_limit)
    sample_row_count = len(sample_rows)
    total_record_count = int((diagnostics or {}).get("record_count") or len(raw_df))
    omitted_row_count = max(0, total_record_count - sample_row_count)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = sanitize_workbook_sheet_name(LOG_DIAGNOSTIC_SHEET_NAMES["summary"])
    initialize_excel_worksheet(summary_sheet)

    summary_rows = [
        ("元ファイル名", profile_payload.get("source_file_name") or ""),
        ("ケースID列", column_settings.get("case_id_column") or "未設定"),
        ("アクティビティ列", column_settings.get("activity_column") or "未設定"),
        ("タイムスタンプ列", column_settings.get("timestamp_column") or "未設定"),
        ("ログレコード数", diagnostics.get("record_count", "-")),
        ("総ケース数", diagnostics.get("case_count", "-")),
        ("アクティビティ種類数", diagnostics.get("activity_type_count", "-")),
        ("ログ期間", build_log_diagnostic_period_text(diagnostics)),
        ("欠損件数", build_log_diagnostic_missing_count_text(diagnostics)),
        ("重複行数", diagnostics.get("duplicate_row_count", 0)),
        ("重複あり/なし", diagnostics.get("duplicate_status", "なし")),
        ("重複除外後レコード数", diagnostics.get("deduplicated_record_count", "-")),
        ("重複率", build_log_diagnostic_duplicate_rate_text(diagnostics)),
        ("ヘッダー一覧", ", ".join(diagnostics.get("headers") or [])),
        ("ログサンプル出力上限", int(sample_row_limit)),
        ("ログサンプル出力件数", sample_row_count),
    ]
    next_row = append_key_value_rows(
        summary_sheet,
        "ログ診断サマリー",
        summary_rows,
        description="トップ画面のログ診断に表示する件数・期間・欠損・重複をまとめています。",
    )

    column_summary_rows = [
        {
            "列名": str(column.get("name") or ""),
            "サンプル値": ", ".join(column.get("sample_values") or []) or "-",
            "ユニーク件数": int(column.get("unique_count") or 0),
            "欠損件数": int(column.get("missing_count") or 0),
        }
        for column in (diagnostics.get("columns") or [])
    ]
    next_row = append_table_to_worksheet(
        summary_sheet,
        "列サマリー",
        column_summary_rows,
        ["列名", "サンプル値", "ユニーク件数", "欠損件数"],
        start_row=next_row,
        description="トップ画面に表示する列ごとのサンプル値・ユニーク件数・欠損件数です。",
        min_column_widths={"列名": 24, "サンプル値": 48},
    )

    sample_sheet = workbook.create_sheet(sanitize_workbook_sheet_name(LOG_DIAGNOSTIC_SHEET_NAMES["sample"]))
    initialize_excel_worksheet(sample_sheet)
    sample_description = (
        f"ログサンプルとして先頭 {sample_row_count} 件を掲載しています。"
        if omitted_row_count <= 0
        else f"ログサンプルとして先頭 {sample_row_count} 件を掲載しています。残り {omitted_row_count} 件は省略しています。"
    )
    append_table_to_worksheet(
        sample_sheet,
        "ログサンプル",
        sample_rows,
        sample_headers,
        description=sample_description,
        min_column_widths={"レコード順": 14},
    )

    for worksheet in workbook.worksheets:
        autosize_worksheet_columns(worksheet)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()


def get_analysis_options():
    analysis_definitions = get_available_analysis_definitions()
    analysis_options = []

    for analysis_key in DEFAULT_ANALYSIS_KEYS:
        analysis_options.append(
            {
                "key": analysis_key,
                "label": analysis_definitions[analysis_key]["config"]["analysis_name"],
            }
        )

    return analysis_options


@app.get("/")
def index(request: Request):
    sample_profile_payload = build_log_profile_payload(
        raw_df=read_raw_log_dataframe(SAMPLE_FILE),
        source_file_name=SAMPLE_FILE.name,
        include_diagnostics=False,
    )

    return _template_response(
        request,
        "index.html",
        {
            "analysis_options": get_analysis_options(),
            "default_headers": DEFAULT_HEADERS,
            "sample_profile_payload": sample_profile_payload,
            "sample_file_name": SAMPLE_FILE.name,
            "static_version": get_static_version(),
        },
    )


@app.get("/analysis/patterns/{pattern_index}")
def pattern_detail_page(request: Request, pattern_index: int):
    return _template_response(
        request,
        "pattern_detail.html",
        {
            "pattern_index": pattern_index,
            "static_version": get_static_version(),
        },
    )


@app.get("/analysis/{analysis_key}")
def analysis_detail(request: Request, analysis_key):
    analysis_definitions = get_available_analysis_definitions()

    if analysis_key not in analysis_definitions:
        raise HTTPException(status_code=404, detail="分析種別が見つかりません。")

    return _template_response(
        request,
        "analysis_detail.html",
        {
            "analysis_key": analysis_key,
            "analysis_name": analysis_definitions[analysis_key]["config"]["analysis_name"],
            "static_version": get_static_version(),
        },
    )


@app.get("/api/runs/{run_id}/patterns/{pattern_index}")
def pattern_detail_api(run_id: str, pattern_index: int):
    run_data = get_run_data(run_id)
    pattern_analysis, summary_row, _, pattern = get_pattern_summary_row(run_data, pattern_index)

    detail = query_pattern_bottleneck_details(
        run_data["prepared_parquet_path"],
        pattern,
        filter_column_settings=run_data.get("column_settings"),
    )
    return JSONResponse(
        content={
            "run_id": run_id,
            "pattern_index": pattern_index,
            "source_file_name": run_data["source_file_name"],
            "analysis_name": pattern_analysis["analysis_name"],
            "summary_row": summary_row,
            "repeat_flag": summary_row.get("繰り返し", ""),
            "repeat_count": summary_row.get("繰り返し回数", 0),
            "repeat_rate_pct": summary_row.get("繰り返し率(%)", 0),
            "repeat_rate_band": summary_row.get("繰り返し率区分", ""),
            "review_flag": summary_row.get("確認区分", ""),
            "avg_case_duration_diff_min": summary_row.get("平均処理時間差分(分)", 0),
            "improvement_priority_score": summary_row.get("改善優先度スコア", 0),
            "overall_impact_pct": summary_row.get("全体影響度(%)", 0),
            "fastest_pattern_flag": summary_row.get("最短処理", ""),
            "simple_comment": summary_row.get("簡易コメント", ""),
            **detail,
        }
    )


@app.get("/api/runs/{run_id}/analyses/{analysis_key}")
def analysis_detail_api(
    request: Request,
    run_id: str,
    analysis_key: str,
    row_limit: int | None = None,
    row_offset: int = 0,
    include_dashboard: bool = True,
    include_impact: bool = True,
    include_root_cause: bool = True,
    include_insights: bool = True,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    analysis = get_analysis_data(run_data, analysis_key, filter_params=filter_params)
    filtered_df = None
    filtered_meta = get_filtered_meta_for_run(
        run_data,
        filter_params=filter_params,
    )
    deferred_sections = []

    response_analyses = {
        analysis_key: build_analysis_payload(
            analysis,
            row_limit=row_limit,
            row_offset=row_offset,
        )
    }
    dashboard_summary = None
    if include_dashboard:
        dashboard_summary = get_dashboard_summary(
            run_data,
            filter_params=filter_params,
            prepared_df=filtered_df,
        )
    else:
        deferred_sections.append("dashboard")

    impact_summary = None
    if include_impact:
        impact_summary = get_impact_summary(
            run_data,
            filter_params=filter_params,
            prepared_df=filtered_df,
        )
    else:
        deferred_sections.append("impact")

    root_cause_summary = None
    if include_root_cause:
        root_cause_summary = get_root_cause_summary(
            run_data,
            filter_params=filter_params,
            prepared_df=filtered_df,
        )
    else:
        deferred_sections.append("root_cause")

    insights_summary = None
    if include_insights:
        insights_summary = get_rule_based_insights_summary(
            run_data,
            analysis_key=analysis_key,
            analysis_rows=analysis.get("rows"),
            filter_params=filter_params,
            prepared_df=filtered_df,
            dashboard_summary=dashboard_summary,
            impact_summary=impact_summary,
        )
    else:
        deferred_sections.append("insights")

    return JSONResponse(
        content={
            "run_id": run_id,
            "source_file_name": run_data["source_file_name"],
            "selected_analysis_keys": run_data["selected_analysis_keys"],
            "case_count": filtered_meta["case_count"],
            "event_count": filtered_meta["event_count"],
            "dashboard": dashboard_summary,
            "impact": impact_summary,
            "insights": insights_summary,
            "root_cause": root_cause_summary,
            "deferred_sections": deferred_sections,
            "applied_filters": filter_params,
            "column_settings": build_column_settings_payload(run_data.get("column_settings")),
            "analyses": response_analyses,
        }
    )


@app.get("/api/runs/{run_id}/ai-insights/{analysis_key}")
def ai_insights_state_api(request: Request, run_id: str, analysis_key: str):
    run_data = get_run_data(run_id)
    analysis_definitions = get_available_analysis_definitions()
    normalized_analysis_key = str(analysis_key or "").strip().lower()

    if normalized_analysis_key not in analysis_definitions:
        raise HTTPException(status_code=404, detail="分析種別が見つかりません。")

    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    cached_summary = get_cached_ai_summary(
        run_data,
        normalized_analysis_key,
        filter_params=filter_params,
    )
    if cached_summary is not None:
        return JSONResponse(content=cached_summary)

    return JSONResponse(
        content=build_empty_ai_summary(
            normalized_analysis_key,
            analysis_definitions[normalized_analysis_key]["config"]["analysis_name"],
        )
    )


@app.post("/api/runs/{run_id}/ai-insights/{analysis_key}")
def ai_insights_generate_api(
    request: Request,
    run_id: str,
    analysis_key: str,
    force_refresh: bool = False,
):
    run_data = get_run_data(run_id)
    analysis_definitions = get_available_analysis_definitions()
    normalized_analysis_key = str(analysis_key or "").strip().lower()

    if normalized_analysis_key not in analysis_definitions:
        raise HTTPException(status_code=404, detail="分析種別が見つかりません。")

    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    payload = build_ai_insights_summary(
        run_data=run_data,
        analysis_key=normalized_analysis_key,
        filter_params=filter_params,
        force_refresh=force_refresh,
    )

    return JSONResponse(content=payload)


@app.get("/api/runs/{run_id}/excel-files/{analysis_key}")
def analysis_excel_file_api(run_id: str, analysis_key: str):
    run_data = get_run_data(run_id)
    analyses = run_data["result"]["analyses"]
    analysis = analyses.get(analysis_key)

    if not analysis:
        raise HTTPException(status_code=404, detail="分析データが見つかりません。")

    excel_df = pd.DataFrame(analysis["rows"])
    excel_bytes = build_excel_bytes(excel_df, analysis["sheet_name"])
    output_file_name = build_analysis_excel_file_name(
        run_data["source_file_name"],
        analysis_key,
        analysis.get("analysis_name", ""),
    )

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_file_name)}",
        },
    )


@app.get("/api/runs/{run_id}/excel-archive")
def analysis_excel_archive_api(run_id: str):
    run_data = get_run_data(run_id)
    analyses = run_data["result"]["analyses"]

    archive_buffer = BytesIO()
    with ZipFile(archive_buffer, mode="w", compression=ZIP_DEFLATED) as archive_file:
        for current_analysis_key, analysis in analyses.items():
            excel_df = pd.DataFrame(analysis["rows"])
            output_file_name = build_analysis_excel_file_name(
                run_data["source_file_name"],
                current_analysis_key,
                analysis.get("analysis_name", ""),
            )
            excel_bytes = build_excel_bytes(excel_df, analysis["sheet_name"])
            archive_file.writestr(output_file_name, excel_bytes)

    archive_file_name = f"{Path(run_data['source_file_name']).stem}_analysis_excels.zip"

    return Response(
        content=archive_buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(archive_file_name)}",
        },
    )


@app.get("/api/runs/{run_id}/detail-excel")
@app.get("/api/runs/{run_id}/report-excel")
def detail_excel_export_api(
    request: Request,
    run_id: str,
    analysis_key: str,
    pattern_display_limit: str = "10",
    variant_id: int | None = None,
    selected_activity: str = "",
    selected_transition_key: str = "",
    case_id: str = "",
    drilldown_limit: int = 20,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    excel_bytes = build_detail_export_workbook_bytes(
        run_id=run_id,
        run_data=run_data,
        analysis_key=analysis_key,
        filter_params=filter_params,
        pattern_display_limit=pattern_display_limit,
        variant_id=variant_id,
        selected_activity=selected_activity,
        selected_transition_key=selected_transition_key,
        case_id=case_id,
        drilldown_limit=drilldown_limit,
    )
    analysis_name = resolve_analysis_display_name(analysis_key)
    output_file_name = build_analysis_excel_file_name(
        run_data["source_file_name"],
        analysis_key,
        analysis_name,
        suffix="レポート",
    )

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_file_name)}",
        },
    )


@app.get("/api/runs/{run_id}/filter-options")
def filter_options_api(run_id: str):
    run_data = get_run_data(run_id)

    return JSONResponse(
        content={
            "run_id": run_id,
            "options": get_filter_options_payload(run_data),
            "applied_filters": run_data.get("base_filter_params"),
            "column_settings": build_column_settings_payload(run_data.get("column_settings")),
        }
    )


@app.get("/api/runs/{run_id}/pattern-flow")
def pattern_flow_api(
    request: Request,
    run_id: str,
    pattern_percent: int = 10,
    pattern_count: int | None = None,
    activity_percent: int = 40,
    connection_percent: int = 30,
    variant_id: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    pattern_analysis = get_analysis_data(run_data, "pattern", filter_params=filter_params)

    if not pattern_analysis and variant_id is None:
        raise HTTPException(status_code=400, detail="処理順パターン分析を利用できません。")

    snapshot = get_pattern_flow_snapshot(
        run_data=run_data,
        pattern_percent=pattern_percent,
        pattern_count=pattern_count,
        activity_percent=activity_percent,
        connection_percent=connection_percent,
        variant_id=variant_id,
        filter_params=filter_params,
    )
    filtered_meta = get_filtered_meta_for_run(
        run_data,
        filter_params=filter_params,
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "filtered_case_count": filtered_meta["case_count"],
            "filtered_event_count": filtered_meta["event_count"],
            "applied_filters": filter_params,
            **snapshot,
        }
    )


@app.get("/api/runs/{run_id}/variants")
def variant_list_api(request: Request, run_id: str, limit: int = 10):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    safe_limit = max(0, int(limit))
    filtered_meta = get_filtered_meta_for_run(
        run_data,
        filter_params=filter_params,
    )
    all_variant_items = get_variant_items(run_data, filter_params=filter_params)
    variant_items = all_variant_items if safe_limit == 0 else all_variant_items[:safe_limit]

    return JSONResponse(
        content={
            "run_id": run_id,
            "variants": [
                build_variant_response_item(variant_item, run_data=run_data)
                for variant_item in variant_items
            ],
            "coverage": build_variant_coverage_payload(
                total_case_count=filtered_meta["case_count"],
                variant_items=variant_items,
            ),
            "filtered_case_count": filtered_meta["case_count"],
            "filtered_event_count": filtered_meta["event_count"],
            "applied_filters": filter_params,
        }
    )


@app.get("/api/runs/{run_id}/bottlenecks")
def bottleneck_list_api(
    request: Request,
    run_id: str,
    limit: int = 5,
    variant_id: int | None = None,
    pattern_index: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    safe_limit = max(0, int(limit))
    variant_pattern = get_run_variant_pattern(
        run_data,
        variant_id=variant_id,
        pattern_index=pattern_index,
        filter_params=filter_params,
    )
    bottleneck_summary = get_bottleneck_summary(
        run_data,
        variant_id=variant_id,
        pattern_index=pattern_index,
        filter_params=filter_params,
    )
    filtered_meta = get_filtered_meta_for_run(
        run_data,
        filter_params=filter_params,
        variant_pattern=variant_pattern,
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "limit": safe_limit,
            "variant_id": variant_id,
            "pattern_index": pattern_index,
            "filtered_case_count": filtered_meta["case_count"],
            "filtered_event_count": filtered_meta["event_count"],
            "applied_filters": filter_params,
            "activity_bottlenecks": bottleneck_summary["activity_bottlenecks"][:safe_limit],
            "transition_bottlenecks": bottleneck_summary["transition_bottlenecks"][:safe_limit],
            "activity_heatmap": bottleneck_summary["activity_heatmap"],
            "transition_heatmap": bottleneck_summary["transition_heatmap"],
        }
    )


@app.get("/api/runs/{run_id}/transition-cases")
def transition_case_drilldown_api(
    request: Request,
    run_id: str,
    from_activity: str,
    to_activity: str,
    limit: int = 20,
    variant_id: int | None = None,
    pattern_index: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    safe_limit = max(0, int(limit))
    case_rows = query_transition_case_drilldown(
        run_data["prepared_parquet_path"],
        from_activity=from_activity,
        to_activity=to_activity,
        limit=safe_limit,
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=get_run_variant_pattern(
            run_data,
            variant_id=variant_id,
            pattern_index=pattern_index,
            filter_params=filter_params,
        ),
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "variant_id": variant_id,
            "pattern_index": pattern_index,
            "from_activity": from_activity,
            "to_activity": to_activity,
            "transition_key": f"{from_activity}__TO__{to_activity}",
            "transition_label": f"{from_activity} → {to_activity}",
            "limit": safe_limit,
            "returned_case_count": len(case_rows),
            "applied_filters": filter_params,
            "cases": case_rows,
        }
    )


@app.get("/api/runs/{run_id}/activity-cases")
def activity_case_drilldown_api(
    request: Request,
    run_id: str,
    activity: str,
    limit: int = 20,
    variant_id: int | None = None,
    pattern_index: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    safe_limit = max(0, int(limit))
    case_rows = query_activity_case_drilldown(
        run_data["prepared_parquet_path"],
        activity=activity,
        limit=safe_limit,
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=get_run_variant_pattern(
            run_data,
            variant_id=variant_id,
            pattern_index=pattern_index,
            filter_params=filter_params,
        ),
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "variant_id": variant_id,
            "pattern_index": pattern_index,
            "activity": activity,
            "limit": safe_limit,
            "returned_case_count": len(case_rows),
            "applied_filters": filter_params,
            "cases": case_rows,
        }
    )


@app.get("/api/runs/{run_id}/cases/{case_id:path}")
def case_trace_api(run_id: str, case_id: str):
    run_data = get_run_data(run_id)
    normalized_case_id = str(case_id or "").strip()

    if not normalized_case_id:
        return JSONResponse(
            status_code=400,
            content={
                "run_id": run_id,
                "case_id": "",
                "found": False,
                "summary": None,
                "events": [],
                "error": "ケースIDが必要です。",
            },
        )

    case_trace = query_case_trace_details(
        run_data["prepared_parquet_path"],
        normalized_case_id,
    )
    return JSONResponse(
        content={
            "run_id": run_id,
            **case_trace,
        }
    )


@app.post("/api/csv-headers")
async def csv_headers(request: Request):
    form = await request.form()
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    filter_column_settings = get_form_filter_column_settings(form)
    file_source, source_file_name = resolve_profile_file_source(form)

    try:
        raw_df = read_raw_log_dataframe(file_source)
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "error": "CSVヘッダーを読み取れませんでした。ファイルの文字コードとヘッダー行を確認してください。",
                "detail": str(exc),
            },
        )

    return JSONResponse(
        content=build_log_profile_payload(
            raw_df=raw_df,
            source_file_name=source_file_name,
            case_id_column=str(raw_case_id_column or "").strip(),
            activity_column=str(raw_activity_column or "").strip(),
            timestamp_column=str(raw_timestamp_column or "").strip(),
            filter_column_settings=filter_column_settings,
            include_diagnostics=False,
        )
    )


@app.post("/api/log-diagnostics")
async def log_diagnostics(request: Request):
    form = await request.form()
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    filter_column_settings = get_form_filter_column_settings(form)
    file_source, source_file_name = resolve_profile_file_source(form)

    try:
        raw_df = read_raw_log_dataframe(file_source)
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "error": "ログ診断を読み取れませんでした。ファイルの文字コードとヘッダー行を確認してください。",
                "detail": str(exc),
            },
        )

    return JSONResponse(
        content=build_log_profile_payload(
            raw_df=raw_df,
            source_file_name=source_file_name,
            case_id_column=str(raw_case_id_column or "").strip(),
            activity_column=str(raw_activity_column or "").strip(),
            timestamp_column=str(raw_timestamp_column or "").strip(),
            filter_column_settings=filter_column_settings,
            include_diagnostics=True,
        )
    )


@app.post("/api/log-diagnostics-excel")
async def log_diagnostics_excel(request: Request):
    form = await request.form()
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    filter_column_settings = get_form_filter_column_settings(form)
    sample_row_limit = resolve_log_diagnostic_sample_row_limit(form.get("sample_row_limit"))
    file_source, source_file_name = resolve_profile_file_source(form)

    try:
        raw_df = read_raw_log_dataframe(file_source)
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "error": "ログ診断を読み取れませんでした。ファイルの文字コードとヘッダー行を確認してください。",
                "detail": str(exc),
            },
        )

    profile_payload = build_log_profile_payload(
        raw_df=raw_df,
        source_file_name=source_file_name,
        case_id_column=str(raw_case_id_column or "").strip(),
        activity_column=str(raw_activity_column or "").strip(),
        timestamp_column=str(raw_timestamp_column or "").strip(),
        filter_column_settings=filter_column_settings,
        include_diagnostics=True,
    )
    excel_bytes = build_log_diagnostic_workbook_bytes(
        profile_payload=profile_payload,
        raw_df=raw_df,
        sample_row_limit=sample_row_limit,
    )
    output_file_name = build_analysis_excel_file_name(
        source_file_name,
        "log_diagnostics",
        "ログ診断",
    )

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_file_name)}",
        },
    )


@app.post("/api/analyze")
async def analyze(request: Request):
    form = await request.form()
    uploaded_file = form.get("csv_file")
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    case_id_column = (
        DEFAULT_HEADERS["case_id_column"]
        if raw_case_id_column is None
        else str(raw_case_id_column).strip()
    )
    activity_column = (
        DEFAULT_HEADERS["activity_column"]
        if raw_activity_column is None
        else str(raw_activity_column).strip()
    )
    timestamp_column = (
        DEFAULT_HEADERS["timestamp_column"]
        if raw_timestamp_column is None
        else str(raw_timestamp_column).strip()
    )
    selected_analysis_keys = form.getlist("analysis_keys")
    filter_column_settings = get_form_filter_column_settings(form)
    base_filter_params = get_form_filter_params(form)

    if uploaded_file and uploaded_file.filename:
        uploaded_file.file.seek(0)
        file_source = uploaded_file.file
        source_file_name = uploaded_file.filename
    else:
        file_source = SAMPLE_FILE
        source_file_name = SAMPLE_FILE.name

    try:
        raw_df = read_raw_log_dataframe(file_source)
        headers = [str(column_name) for column_name in raw_df.columns.tolist()]
        case_id_column = resolve_required_column_name(
            headers,
            "case_id_column",
            case_id_column,
        )
        activity_column = resolve_required_column_name(
            headers,
            "activity_column",
            activity_column,
        )
        timestamp_column = resolve_required_column_name(
            headers,
            "timestamp_column",
            timestamp_column,
        )
        validate_selected_columns(
            case_id_column=case_id_column,
            activity_column=activity_column,
            timestamp_column=timestamp_column,
        )
        validate_filter_column_settings(filter_column_settings)
        prepared_df = load_prepared_event_log(
            file_source=file_source,
            case_id_column=case_id_column,
            activity_column=activity_column,
            timestamp_column=timestamp_column,
        )
        filtered_prepared_df = filter_prepared_df(
            prepared_df,
            base_filter_params,
            filter_column_settings=filter_column_settings,
        )
        group_columns = detect_group_columns(base_filter_params, filter_column_settings)
        result = analyze_prepared_event_log(
            prepared_df=filtered_prepared_df,
            selected_analysis_keys=selected_analysis_keys,
            output_root_dir=None,
            export_excel=False,
            group_columns=group_columns if group_columns else None,
        )
        if group_columns:
            result["group_summary"] = build_group_summary(filtered_prepared_df, group_columns)
        else:
            result["group_summary"] = {}
        run_id = save_run_data(
            source_file_name=source_file_name,
            selected_analysis_keys=selected_analysis_keys,
            prepared_df=prepared_df,
            result=result,
            column_settings={
                "case_id_column": case_id_column,
                "activity_column": activity_column,
                "timestamp_column": timestamp_column,
                **filter_column_settings,
            },
            base_filter_params=base_filter_params,
        )
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={
                "error": "分析に失敗しました。",
                "detail": str(exc),
            },
        )

    return JSONResponse(
        content=build_preview_response(
            run_id=run_id,
            source_file_name=source_file_name,
            selected_analysis_keys=selected_analysis_keys,
            result=result,
            run_data=get_run_data(run_id),
        )
    )


def build_bottleneck_prompt(data: dict) -> str:
    freq = data.get("frequency_top10", [])
    slow = sorted(
        freq,
        key=lambda x: x.get("平均処理時間(分)", x.get("平均時間(分)", 0)),
        reverse=True,
    )[:3]
    busy = sorted(freq, key=lambda x: x.get("イベント件数", 0), reverse=True)[:3]
    patterns = data.get("pattern_top10", [])[:3]

    return f"""あなたはプロセス改善の専門家です。
以下のプロセスマイニング分析結果をもとに、現場担当者が「明日から行動できる」レベルの解説をしてください。
数値の羅列ではなく、なぜそうなっているか・何をすべきかに焦点を当ててください。

## 分析データ
- 総ケース数: {data.get("total_cases", "不明")}
- 分析期間: {data.get("period", "不明")}
- 処理時間が長いアクティビティ上位3件: {slow}
- 件数が集中しているアクティビティ上位3件: {busy}
- 頻出プロセスパターン上位3件: {patterns}

## 回答形式（この5セクションで回答してください）

【1. 全体サマリー】
このプロセス全体の状態を2〜3文で総括してください。

【2. ボトルネックの特定】
最も改善効果が高いと考えられる箇所を1〜2つ、具体的な数値を使って説明してください。

【3. 考えられる原因】
なぜそこがボトルネックになっているか、現場でよくある原因を2〜3つ挙げてください。

【4. 改善アクション】
明日から実行できる具体的なアクションを2つ提案してください。
1つ目は「すぐできること（工数小）」、2つ目は「中期的な改善（工数大・効果大）」としてください。

【5. 次のステップ】
改善を進める上で、次に確認・分析すべきことを1つ提案してください。

回答は現場担当者に話しかけるような自然な文体で、専門用語は使わずにお願いします。
"""


@app.post("/api/ai-insights")
async def ai_insights(request: Request):
    import json as _json
    from fastapi.responses import StreamingResponse

    data = await request.json()
    prompt = build_bottleneck_prompt(data)

    async def generate():
        try:
            timeout = httpx.Timeout(connect=10.0, read=60.0, write=10.0, pool=10.0)
            async with httpx.AsyncClient(timeout=timeout) as client:
                async with client.stream(
                    "POST",
                    "http://localhost:11434/api/generate",
                    json={"model": "qwen2.5:7b", "prompt": prompt, "stream": True},
                ) as response:
                    response.raise_for_status()
                    async for line in response.aiter_lines():
                        if not line:
                            continue
                        chunk = _json.loads(line)
                        token = chunk.get("response", "")
                        if token:
                            yield f"data: {_json.dumps({'token': token}, ensure_ascii=False)}\n\n"
                        if chunk.get("done"):
                            yield f"data: {_json.dumps({'done': True})}\n\n"
                            return
        except httpx.ConnectError:
            yield f"data: {_json.dumps({'error': 'Ollamaが起動していません'})}\n\n"
        except Exception as exc:
            msg = str(exc) or f"{type(exc).__name__}（詳細なし）"
            yield f"data: {_json.dumps({'error': msg})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


if __name__ == "__main__":
    uvicorn.run("web_app:app", host="127.0.0.1", port=5000, reload=True)
