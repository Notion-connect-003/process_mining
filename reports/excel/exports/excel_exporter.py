from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from reports.excel.common import (
    EXCEL_ALT_ROW_FILL,
    EXCEL_BOLD_FONT,
    EXCEL_GROUP_SECTION_FILL,
    EXCEL_GROUP_SECTION_FONT,
    EXCEL_HEADER_FILL,
    EXCEL_THIN_BORDER,
    EXCEL_DISPLAY_COLUMN_RENAMES,
    autosize_worksheet_columns,
    initialize_excel_worksheet,
    rename_excel_display_columns,
    sanitize_workbook_sheet_name,
    style_excel_cell,
)


def _style_export_worksheet(worksheet):
    initialize_excel_worksheet(worksheet)

    if worksheet.max_row >= 1:
        for column_index in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=column_index)
            style_excel_cell(
                header_cell,
                font=EXCEL_BOLD_FONT,
                fill=EXCEL_HEADER_FILL,
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[1].height = 22
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    for row_index in range(2, worksheet.max_row + 1):
        row_fill = EXCEL_ALT_ROW_FILL if row_index % 2 == 0 else None
        for column_index in range(1, worksheet.max_column + 1):
            body_cell = worksheet.cell(row=row_index, column=column_index)
            style_excel_cell(
                body_cell,
                fill=row_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )

    autosize_worksheet_columns(worksheet, min_width=10, max_width=40)


def format_analysis_result(df, display_columns=None, group_columns=None):
    if display_columns is None:
        return df.copy()

    valid_group_cols = [col for col in (group_columns or []) if col in df.columns]
    ordered_columns = valid_group_cols + [
        column_name
        for column_name in display_columns.keys()
        if column_name in df.columns and column_name not in valid_group_cols
    ]
    formatted_df = df.loc[:, ordered_columns].copy() if ordered_columns else df.copy()
    rename_map = {k: v for k, v in display_columns.items() if k not in valid_group_cols}
    return formatted_df.rename(columns=rename_map)


GROUP_SUMMARY_META_KEY = "__meta__"
SUMMARY_SHEET_COLUMNS = [
    "グルーピング軸",
    "値",
    "ケース数",
    "ケース比率(%)",
    "イベント数",
    "イベント比率(%)",
    "平均所要時間(分)",
    "中央値所要時間(分)",
    "最大所要時間(分)",
    "合計所要時間(分)",
]


def insert_group_section_rows(worksheet, df, group_columns):
    if not group_columns:
        return

    first_group_col = group_columns[0]
    if first_group_col not in df.columns:
        return

    values = df[first_group_col].astype(str).tolist()
    change_indices = []
    prev = None
    for index, value in enumerate(values):
        if prev is not None and value != prev:
            change_indices.append((index, value))
        prev = value

    for df_idx, group_value in reversed(change_indices):
        excel_row = df_idx + 2
        worksheet.insert_rows(excel_row)
        section_cell = worksheet.cell(row=excel_row, column=1)
        section_cell.value = f"■ {group_value}"
        style_excel_cell(section_cell, font=EXCEL_GROUP_SECTION_FONT, fill=EXCEL_GROUP_SECTION_FILL)
        if worksheet.max_column > 1:
            worksheet.merge_cells(
                start_row=excel_row,
                start_column=1,
                end_row=excel_row,
                end_column=worksheet.max_column,
            )


def _format_group_axis_label(column_name, level, group_count):
    if group_count <= 1:
        return column_name
    suffix = "①②③"[level - 1] if level <= 3 else str(level)
    return f"{column_name}（グループ{suffix}）"


def _duration_cell_value(value):
    return value if value is not None else "—"


def build_summary_sheet_df(group_summary, group_columns):
    meta = group_summary.get(GROUP_SUMMARY_META_KEY, {})
    rows = []
    for level, col in enumerate(group_columns, start=1):
        col_data = group_summary.get(col, {})
        for value, stats in col_data.items():
            rows.append(
                {
                    "グルーピング軸": _format_group_axis_label(col, level, len(group_columns)),
                    "値": value,
                    "ケース数": stats.get("case_count", ""),
                    "ケース比率(%)": stats.get("case_ratio_pct", ""),
                    "イベント数": stats.get("event_count", ""),
                    "イベント比率(%)": stats.get("event_ratio_pct", ""),
                    "平均所要時間(分)": _duration_cell_value(stats.get("avg_duration_min")),
                    "中央値所要時間(分)": _duration_cell_value(stats.get("median_duration_min")),
                    "最大所要時間(分)": _duration_cell_value(stats.get("max_duration_min")),
                    "合計所要時間(分)": _duration_cell_value(stats.get("total_duration_min")),
                }
            )

    group_rows_df = pd.DataFrame(rows, columns=SUMMARY_SHEET_COLUMNS)
    if not group_rows_df.empty:
        group_rows_df = group_rows_df.sort_values(
            ["ケース数", "グルーピング軸", "値"],
            ascending=[False, True, True],
        ).reset_index(drop=True)

    total_row_df = pd.DataFrame(
        [
            {
                "グルーピング軸": "",
                "値": "全体",
                "ケース数": meta.get("total_case_count", ""),
                "ケース比率(%)": 100.0 if meta else "",
                "イベント数": meta.get("total_event_count", ""),
                "イベント比率(%)": 100.0 if meta else "",
                "平均所要時間(分)": _duration_cell_value(meta.get("avg_duration_min")) if meta else "—",
                "中央値所要時間(分)": _duration_cell_value(meta.get("median_duration_min")) if meta else "—",
                "最大所要時間(分)": _duration_cell_value(meta.get("max_duration_min")) if meta else "—",
                "合計所要時間(分)": _duration_cell_value(meta.get("total_duration_min")) if meta else "—",
            }
        ],
        columns=SUMMARY_SHEET_COLUMNS,
    )
    if group_rows_df.empty:
        return total_row_df if meta else pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    return pd.concat([total_row_df, group_rows_df], ignore_index=True)


def _build_summary_sheet_df(group_summary, group_columns):
    return build_summary_sheet_df(group_summary, group_columns)


def build_excel_bytes(df, sheet_name, group_columns=None, group_summary=None):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if group_columns and group_summary:
            summary_df = build_summary_sheet_df(group_summary, group_columns)
            summary_sheet_name = sanitize_workbook_sheet_name("サマリー")
            summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
            _style_export_worksheet(writer.sheets[summary_sheet_name])

        safe_sheet_name = sanitize_workbook_sheet_name(sheet_name)
        excel_df = (
            pd.DataFrame([rename_excel_display_columns(row) for row in df.to_dict(orient="records")])
            if not df.empty
            else df.rename(columns=EXCEL_DISPLAY_COLUMN_RENAMES)
        )
        excel_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        worksheet = writer.sheets[safe_sheet_name]
        if group_columns:
            insert_group_section_rows(worksheet, excel_df, group_columns)
        _style_export_worksheet(worksheet)

    return buffer.getvalue()


def export_dataframe_to_excel(df, output_file, sheet_name, group_columns=None, group_summary=None):
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        output_path.write_bytes(
            build_excel_bytes(df, sheet_name, group_columns=group_columns, group_summary=group_summary)
        )
    except PermissionError as exc:
        raise PermissionError(
            f"{output_path} に書き込めません。Excel で開いている場合は閉じてから再実行してください。"
        ) from exc

    return output_path


def export_analysis_to_excel(
    df,
    output_root_dir,
    analysis_name,
    output_file_name,
    sheet_name,
    display_columns=None,
    group_columns=None,
    group_summary=None,
):
    output_file = Path(output_root_dir) / analysis_name / output_file_name
    excel_df = format_analysis_result(df, display_columns, group_columns=group_columns)
    return export_dataframe_to_excel(
        excel_df,
        output_file,
        sheet_name,
        group_columns=group_columns,
        group_summary=group_summary,
    )


def convert_analysis_result_to_records(df, display_columns=None, group_columns=None):
    api_df = format_analysis_result(df, display_columns, group_columns=group_columns)
    return api_df.to_dict(orient="records")
