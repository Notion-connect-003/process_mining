from openpyxl.styles import Border, Font, PatternFill, Side


EXCEL_TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
EXCEL_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
EXCEL_SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="EFF5FB")
EXCEL_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E7F6")
EXCEL_GROUP_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
EXCEL_ASSUMPTION_SECTION_FILL = PatternFill(fill_type="solid", fgColor="E8EDF2")
EXCEL_MUTED_SECTION_FILL = PatternFill(fill_type="solid", fgColor="F0F0F0")
EXCEL_HEADER_FILL = PatternFill(fill_type="solid", fgColor="EDF2F7")
EXCEL_LABEL_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
EXCEL_ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="FBFDFF")
EXCEL_TEXT_BLOCK_FILL = PatternFill(fill_type="solid", fgColor="F7FAFE")
EXCEL_TERMINOLOGY_FILL = PatternFill(fill_type="solid", fgColor="F7F7F7")

EXCEL_TITLE_BORDER = Border(
    left=Side(style="thin", color="1F4E78"),
    right=Side(style="thin", color="1F4E78"),
    top=Side(style="thin", color="1F4E78"),
    bottom=Side(style="thin", color="1F4E78"),
)
EXCEL_THIN_BORDER = Border(
    left=Side(style="thin", color="D6DEE8"),
    right=Side(style="thin", color="D6DEE8"),
    top=Side(style="thin", color="D6DEE8"),
    bottom=Side(style="thin", color="D6DEE8"),
)
EXCEL_TERMINOLOGY_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

EXCEL_MUTED_FONT = Font(size=10, color="5B6B82")
EXCEL_NOTE_FONT = Font(size=9, color="5B6B82")
EXCEL_BODY_FONT = Font(size=10, color="1F2937")
EXCEL_BOLD_FONT = Font(bold=True, size=10, color="1F2937")
EXCEL_GROUP_SECTION_FONT = Font(bold=True, size=12, color="1F2937")
EXCEL_SECTION_HEADER_FONT = Font(bold=True, size=11, color="1F2937")


def style_excel_cell(cell, *, font=None, fill=None, alignment=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
