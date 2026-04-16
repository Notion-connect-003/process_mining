import unicodedata

from openpyxl.utils import get_column_letter

from reports.excel.common.cell_values import normalize_excel_cell_value


def merge_excel_row(worksheet, row_index, column_count):
    safe_column_count = max(1, int(column_count or 1))
    if safe_column_count > 1:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=safe_column_count,
        )


def estimate_wrapped_row_height(text, column_count=1, min_height=22, max_height=240):
    safe_text = str(text or "")
    if not safe_text:
        return min_height

    approx_chars_per_line = max(24, int(max(1, column_count)) * 18)
    logical_lines = 0
    for raw_line in safe_text.splitlines() or [""]:
        line_length = max(1, len(raw_line))
        logical_lines += max(1, (line_length + approx_chars_per_line - 1) // approx_chars_per_line)

    return max(min_height, min(max_height, 16 + logical_lines * 15))


def initialize_excel_worksheet(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90


def estimate_excel_text_width(value):
    normalized_value = str(normalize_excel_cell_value(value))
    if not normalized_value:
        return 0

    max_line_width = 0
    for line in normalized_value.splitlines() or [""]:
        line_width = 0
        for character in line:
            line_width += 2 if unicodedata.east_asian_width(character) in {"F", "W", "A"} else 1
        max_line_width = max(max_line_width, line_width)

    return max_line_width


def get_autosize_ignored_cells(worksheet):
    ignored_cells = set()
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_col == merged_range.max_col and merged_range.min_row == merged_range.max_row:
            continue

        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                ignored_cells.add((row_index, column_index))

    return ignored_cells


def autosize_worksheet_columns(worksheet, min_width=12, max_width=120):
    ignored_cells = get_autosize_ignored_cells(worksheet)
    min_width_overrides = getattr(worksheet, "_codex_min_column_widths", {})

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        measured_width = min_width

        for row_index in range(1, worksheet.max_row + 1):
            if (row_index, column_index) in ignored_cells:
                continue

            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value in (None, ""):
                continue
            measured_width = max(measured_width, estimate_excel_text_width(cell_value))

        worksheet.column_dimensions[column_letter].width = max(
            min_width,
            min(
                max_width,
                max(measured_width + 3, float(min_width_overrides.get(column_letter, min_width))),
            ),
        )
