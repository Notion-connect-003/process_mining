import re

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from reports.excel.common.cell_styles import (
    EXCEL_ALT_ROW_FILL,
    EXCEL_BODY_FONT,
    EXCEL_BOLD_FONT,
    EXCEL_HEADER_FILL,
    EXCEL_LABEL_FILL,
    EXCEL_MUTED_FONT,
    EXCEL_MUTED_SECTION_FILL,
    EXCEL_NOTE_FONT,
    EXCEL_SECTION_FILL,
    EXCEL_SECTION_HEADER_FONT,
    EXCEL_SUBTITLE_FILL,
    EXCEL_TERMINOLOGY_BORDER,
    EXCEL_TERMINOLOGY_FILL,
    EXCEL_TEXT_BLOCK_FILL,
    EXCEL_THIN_BORDER,
    EXCEL_TITLE_BORDER,
    EXCEL_TITLE_FILL,
    EXCEL_TITLE_FONT,
    style_excel_cell,
)
from reports.excel.common.cell_values import (
    build_rich_text_with_bold_numbers,
    normalize_excel_cell_value,
)
from reports.excel.common.sheet_layout import estimate_wrapped_row_height, merge_excel_row


def append_table_to_worksheet(
    worksheet,
    title,
    rows,
    headers,
    start_row=1,
    description="",
    no_wrap_headers=None,
    min_column_widths=None,
):
    current_row = start_row
    column_count = max(1, len(headers))
    normalized_no_wrap_headers = {str(header) for header in (no_wrap_headers or [])}
    merge_excel_row(worksheet, current_row, column_count)
    title_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        title_cell,
        font=EXCEL_TITLE_FONT,
        fill=EXCEL_TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_TITLE_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    if description:
        merge_excel_row(worksheet, current_row, column_count)
        description_cell = worksheet.cell(row=current_row, column=1, value=description)
        style_excel_cell(
            description_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_SUBTITLE_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(description, column_count)
        current_row += 1

    header_row = current_row
    for column_index, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=header not in normalized_no_wrap_headers),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    if not rows:
        merge_excel_row(worksheet, current_row, column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value="表示できるデータがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    data_start_row = current_row
    for row_index, row in enumerate(rows, start=0):
        for column_index, header in enumerate(headers, start=1):
            body_cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
            fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(wrap_text=header not in normalized_no_wrap_headers, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

    if min_column_widths:
        width_overrides = dict(getattr(worksheet, "_codex_min_column_widths", {}))
        for column_index, header in enumerate(headers, start=1):
            if header in min_column_widths:
                width_overrides[get_column_letter(column_index)] = max(
                    float(width_overrides.get(get_column_letter(column_index), 0)),
                    float(min_column_widths[header]),
                )
        worksheet._codex_min_column_widths = width_overrides

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(column_count)}{current_row - 1}"

    return current_row + 1


def append_key_value_rows(worksheet, title, rows, start_row=1, description=""):
    current_row = start_row
    merge_excel_row(worksheet, current_row, 2)
    title_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        title_cell,
        font=EXCEL_TITLE_FONT,
        fill=EXCEL_TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_TITLE_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    if description:
        merge_excel_row(worksheet, current_row, 2)
        description_cell = worksheet.cell(row=current_row, column=1, value=description)
        style_excel_cell(
            description_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_SUBTITLE_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(description, 2)
        current_row += 1

    for column_index, header in enumerate(("項目", "値"), start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    data_start_row = current_row
    for row_index, row in enumerate(rows, start=0):
        if isinstance(row, dict):
            label = row.get("label", "")
            value = row.get("value", "")
            row_style = row.get("style", "default")
        else:
            label, value = row
            row_style = "default"
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        label_cell = worksheet.cell(row=current_row, column=1, value=label)
        value_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(value))
        if row_style == "note":
            note_fill = EXCEL_SUBTITLE_FILL if fill is None else fill
            style_excel_cell(
                label_cell,
                font=EXCEL_NOTE_FONT,
                fill=note_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
            style_excel_cell(
                value_cell,
                font=EXCEL_NOTE_FONT,
                fill=note_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        else:
            style_excel_cell(
                label_cell,
                font=EXCEL_BOLD_FONT,
                fill=EXCEL_LABEL_FILL if fill is None else fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
            style_excel_cell(
                value_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = max(
            20,
            estimate_wrapped_row_height(value, 1, min_height=20, max_height=120),
        )
        current_row += 1

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref and rows:
        worksheet.auto_filter.ref = f"A{data_start_row - 1}:B{current_row - 1}"

    return current_row + 1


def append_bullet_rows(worksheet, title, items, start_row=1, column_count=6, empty_text="表示できる要点がありません。", section_header_fill=None):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=section_header_fill or EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    if not items:
        merge_excel_row(worksheet, current_row, safe_column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value=empty_text)
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    for index, item in enumerate(items, start=1):
        fill = EXCEL_ALT_ROW_FILL if index % 2 == 0 else None
        bullet_cell = worksheet.cell(row=current_row, column=1, value=f"{index}.")
        style_excel_cell(
            bullet_cell,
            font=EXCEL_BOLD_FONT,
            fill=fill,
            alignment=Alignment(horizontal="center", vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        if safe_column_count > 1:
            worksheet.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=safe_column_count,
            )
        text_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(item))
        style_excel_cell(
            text_cell,
            font=EXCEL_BODY_FONT,
            fill=fill,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(item, safe_column_count - 1)
        current_row += 1

    return current_row + 1


def append_custom_text_section_to_worksheet(
    worksheet,
    title,
    text,
    start_row=1,
    column_count=6,
    header_fill=None,
    body_fill=None,
    empty_text="表示できる内容がありません。",
):
    current_row = start_row
    safe_column_count = max(1, int(column_count or 1))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=header_fill or EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    merge_excel_row(worksheet, current_row, safe_column_count)
    body_cell = worksheet.cell(row=current_row, column=1, value=normalize_excel_cell_value(text) or empty_text)
    style_excel_cell(
        body_cell,
        font=EXCEL_BODY_FONT,
        fill=body_fill or EXCEL_TEXT_BLOCK_FILL,
        alignment=Alignment(wrap_text=True, vertical="top"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(
        body_cell.value,
        safe_column_count,
        min_height=54,
        max_height=240,
    )

    return current_row + 2


def append_definition_table_to_worksheet(worksheet, title, rows, start_row=1, column_count=6, header_fill=None):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=header_fill or EXCEL_MUTED_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    for column_index, header in enumerate(("用語", "説明"), start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    for row in rows or []:
        term_cell = worksheet.cell(row=current_row, column=1, value=normalize_excel_cell_value(row.get("用語")))
        description_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(row.get("説明")))
        style_excel_cell(
            term_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_TERMINOLOGY_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_TERMINOLOGY_BORDER,
        )
        style_excel_cell(
            description_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_TERMINOLOGY_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_TERMINOLOGY_BORDER,
        )
        worksheet.row_dimensions[current_row].height = max(
            20,
            estimate_wrapped_row_height(row.get("説明"), 1, min_height=20, max_height=96),
        )
        current_row += 1

    return current_row + 1


def append_structured_text_block_to_worksheet(
    worksheet,
    section_title,
    full_text,
    start_row=1,
    column_count=6,
    section_header_fill=None,
    subsection_marker_start="【",
    subsection_marker_end="】",
):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    header_fill = section_header_fill or EXCEL_SECTION_FILL
    merge_excel_row(worksheet, current_row, safe_column_count)
    title_cell = worksheet.cell(row=current_row, column=1, value=section_title)
    style_excel_cell(
        title_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=header_fill,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    normalized_text = str(full_text or "").strip()
    if not normalized_text:
        merge_excel_row(worksheet, current_row, safe_column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value="解説テキストがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    marker_pattern = re.compile(
        rf"^{re.escape(subsection_marker_start)}(.+?){re.escape(subsection_marker_end)}\s*$",
        re.MULTILINE,
    )
    parts = marker_pattern.split(normalized_text)

    if len(parts) <= 1:
        merge_excel_row(worksheet, current_row, safe_column_count)
        text_cell = worksheet.cell(row=current_row, column=1, value=build_rich_text_with_bold_numbers(normalized_text))
        style_excel_cell(
            text_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_TEXT_BLOCK_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(normalized_text, safe_column_count)
        return current_row + 2

    preamble = parts[0].strip()
    if preamble:
        merge_excel_row(worksheet, current_row, safe_column_count)
        preamble_cell = worksheet.cell(row=current_row, column=1, value=build_rich_text_with_bold_numbers(preamble))
        style_excel_cell(
            preamble_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_TEXT_BLOCK_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(preamble, safe_column_count)
        current_row += 1

    for index in range(1, len(parts), 2):
        subsection_title = parts[index].strip()
        subsection_body = parts[index + 1].strip() if index + 1 < len(parts) else ""
        merge_excel_row(worksheet, current_row, safe_column_count)
        subtitle_cell = worksheet.cell(row=current_row, column=1, value=f"【{subsection_title}】")
        style_excel_cell(
            subtitle_cell,
            font=EXCEL_BOLD_FONT,
            alignment=Alignment(horizontal="left", vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

        if subsection_body:
            merge_excel_row(worksheet, current_row, safe_column_count)
            body_cell = worksheet.cell(row=current_row, column=1, value=build_rich_text_with_bold_numbers(subsection_body))
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=EXCEL_TEXT_BLOCK_FILL,
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
            worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(subsection_body, safe_column_count)
            current_row += 1

    return current_row + 1
