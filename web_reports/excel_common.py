from pathlib import Path
import unicodedata

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_SHEET_NAMES = {
    "summary": "サマリー",
    "ai_insights": "分析コメント",
    "pattern_conclusion": "結論サマリー",
    "pattern_dashboard": "サマリーダッシュボード",
    "frequency": "頻度分析",
    "transition": "前後処理分析",
    "pattern": "処理順パターン分析",
    "pattern_summary": "パターンサマリー",
    "variant": "バリアント分析",
    "bottleneck": "ボトルネック分析",
    "impact": "改善インパクト分析",
    "drilldown": "ドリルダウン",
    "case_trace": "ケース追跡",
}

LOG_DIAGNOSTIC_SHEET_NAMES = {
    "summary": "ログ診断",
    "sample": "ログサンプル",
}

REPORT_HEADER_LABELS = {
    "run_id": "実行ID",
    "analysis_key": "分析種別",
    "analysis_name": "分析名",
    "source_file_name": "元ファイル名",
    "analysis_executed_at": "分析実行日時",
    "exported_at": "出力日時",
    "case_count": "対象ケース数",
    "event_count": "対象イベント数",
    "applied_filters": "適用条件",
    "selected_variant": "選択中バリアント",
    "selected_activity": "選択中アクティビティ",
    "selected_transition": "選択中遷移",
    "selected_case_id": "選択中ケースID",
    "rank": "順位",
    "pattern_variant": "パターン / バリアント",
    "repeat_flag": "繰り返し",
    "repeat_count": "繰り返し回数",
    "repeat_rate_pct": "繰り返し率(%)",
    "repeat_rate_band": "繰り返し率区分",
    "review_flag": "確認区分",
    "avg_case_duration_diff_min": "平均処理時間差分(分)",
    "improvement_priority_score": "改善優先度スコア",
    "overall_impact_pct": "全体影響度(%)",
    "fastest_pattern_flag": "最短処理",
    "simple_comment": "簡易コメント",
    "variant_id": "バリアントID",
    "count": "件数",
    "case_count": "対象ケース数",
    "ratio": "比率",
    "cumulative_case_ratio_pct": "累積カバー率(%)",
    "pattern": "パターン",
    "activity_count": "アクティビティ数",
    "avg_case_duration": "平均ケース処理時間",
    "avg_case_duration_min": "平均ケース処理時間(分)",
    "std_case_duration_min": "標準偏差ケース処理時間(分)",
    "min_case_duration_min": "最短処理時間(分)",
    "max_case_duration_min": "最長処理時間(分)",
    "p75_case_duration_min": "75%点ケース処理時間(分)",
    "p90_case_duration_min": "90%点ケース処理時間(分)",
    "p95_case_duration_min": "95%点ケース処理時間(分)",
    "avg_duration": "平均所要時間",
    "avg_duration_text": "平均所要時間",
    "avg_duration_min": "平均所要時間(分)",
    "median_duration_min": "中央値所要時間(分)",
    "std_duration_min": "標準偏差(分)",
    "min_duration_min": "最小所要時間(分)",
    "median_duration_text": "中央値所要時間",
    "max_duration": "最大所要時間",
    "max_duration_text": "最大所要時間",
    "max_duration_min": "最大所要時間(分)",
    "total_duration_min": "合計所要時間(分)",
    "p75_duration_min": "75%点(分)",
    "p90_duration_min": "90%点(分)",
    "p95_duration_min": "95%点(分)",
    "impact_score": "改善インパクト",
    "impact_share_pct": "改善インパクト比率(%)",
    "wait_share_pct": "構成比(%)",
    "case_id": "ケースID",
    "from_time": "開始時刻",
    "to_time": "終了時刻",
    "activity": "アクティビティ",
    "next_activity": "次アクティビティ",
    "sequence_no": "ステップ順",
    "transition": "遷移",
    "transition_label": "遷移",
    "duration_text": "所要時間",
    "total_duration": "総処理時間",
    "start_time": "開始時刻",
    "end_time": "終了時刻",
}

APPLIED_FILTERS_NOTE_TEXT = "\n".join(
    [
        "※ 適用条件の種類:",
        "  • 期間フィルター: 開始日 / 終了日",
        "  • グループ/カテゴリーフィルター①②③: CSVの任意カラムで絞り込み（例: 部署=営業部）",
        "  • アクティビティフィルター: 特定アクティビティを含む/除外",
    ]
)

GROUPING_CONDITION_NOTE_TEXT = (
    "※ カラムを選択し値を未選択にすると、そのカラムがグルーピング軸（比較用）になります"
)

ANALYSIS_PRECONDITIONS_TEXT = "\n".join(
    [
        "• 処理時間は、同一ケース内で当該アクティビティの開始時刻から次のアクティビティの開始時刻までの差分として算出しています。",
        "• ケース内の最終アクティビティは、次のイベントが存在しないため処理時間が0分となります。",
        "• 処理時間が0分のイベントも集計対象に含まれています。統計値（平均・中央値等）に影響する点にご留意ください。",
        "• 本分析はフィルター適用後のデータに基づいています。適用条件の詳細はサマリーシートをご参照ください。",
    ]
)

TERMINOLOGY_ROWS = [
    {
        "用語": "ケース",
        "説明": "業務プロセスの1つの実行単位（例: 1件の注文、1件の申請）",
    },
    {
        "用語": "アクティビティ",
        "説明": "ケース内で実行される個々の作業ステップ（例: 申請、承認、支払）",
    },
    {
        "用語": "イベント",
        "説明": "特定のケースで特定のアクティビティが実行された1回の記録",
    },
    {
        "用語": "処理時間",
        "説明": "あるアクティビティの開始から次のアクティビティの開始までの所要時間",
    },
    {
        "用語": "イベント比率(%)",
        "説明": "全イベント数に対する当該アクティビティのイベント数の割合",
    },
]

EXCEL_TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
EXCEL_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
EXCEL_SUBTITLE_FILL = PatternFill(fill_type="solid", fgColor="EFF5FB")
EXCEL_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E7F6")
EXCEL_GROUP_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
EXCEL_ASSUMPTION_SECTION_FILL = PatternFill(fill_type="solid", fgColor="E8EDF2")
EXCEL_MUTED_SECTION_FILL = PatternFill(fill_type="solid", fgColor="F0F0F0")
EXCEL_HEADER_FILL = PatternFill(fill_type="solid", fgColor="EDF2F7")
EXCEL_LABEL_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
EXCEL_ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="FBFDFF")
EXCEL_TEXT_BLOCK_FILL = PatternFill(fill_type="solid", fgColor="F7FAFE")
EXCEL_TITLE_BORDER = Border(
    left=Side(style="thin", color="1F4E78"),
    right=Side(style="thin", color="1F4E78"),
    top=Side(style="thin", color="1F4E78"),
    bottom=Side(style="thin", color="1F4E78"),
)
EXCEL_THIN_BORDER = Border(
    left=Side(style="thin", color="D6DEE8"),
    right=Side(style="thin", color="D6DEE8"),
    top=Side(style="thin", color="D6DEE8"),
    bottom=Side(style="thin", color="D6DEE8"),
)
EXCEL_MUTED_FONT = Font(size=10, color="5B6B82")
EXCEL_NOTE_FONT = Font(size=9, color="5B6B82")
EXCEL_BODY_FONT = Font(size=10, color="1F2937")
EXCEL_BOLD_FONT = Font(bold=True, size=10, color="1F2937")
EXCEL_GROUP_SECTION_FONT = Font(bold=True, size=12, color="1F2937")
EXCEL_SECTION_HEADER_FONT = Font(bold=True, size=11, color="1F2937")
EXCEL_TERMINOLOGY_FILL = PatternFill(fill_type="solid", fgColor="F7F7F7")
EXCEL_TERMINOLOGY_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def sanitize_workbook_sheet_name(sheet_name):
    invalid_characters = set('[]:*?/\\')
    normalized_name = "".join("_" if character in invalid_characters else character for character in str(sheet_name or "").strip())
    normalized_name = normalized_name or "Sheet"
    return normalized_name[:31]


def sanitize_file_name_component(value):
    invalid_characters = set('<>:"/\\|?*')
    normalized_value = "".join("_" if character in invalid_characters else character for character in str(value or "").strip())
    return normalized_value.strip(" .") or "analysis"


def resolve_analysis_display_name(analysis_key, analysis_name=""):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    if normalized_analysis_key in REPORT_SHEET_NAMES:
        return REPORT_SHEET_NAMES[normalized_analysis_key]
    return str(analysis_name or normalized_analysis_key or "分析").strip() or "分析"


def build_analysis_excel_file_name(source_file_name, analysis_key, analysis_name="", suffix=""):
    source_stem = sanitize_file_name_component(Path(str(source_file_name or "analysis")).stem)
    display_name = sanitize_file_name_component(resolve_analysis_display_name(analysis_key, analysis_name))
    normalized_suffix = sanitize_file_name_component(suffix) if str(suffix or "").strip() else ""
    return f"{source_stem}_{display_name}{normalized_suffix}.xlsx"


def normalize_excel_cell_value(value):
    if value is None:
        return ""

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, (list, tuple, set)):
        return " / ".join(str(item) for item in value)

    if isinstance(value, dict):
        return ", ".join(f"{key}={normalize_excel_cell_value(item_value)}" for key, item_value in value.items())

    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass

    return value


def style_excel_cell(cell, *, font=None, fill=None, alignment=None, border=None):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def merge_excel_row(worksheet, row_index, column_count):
    safe_column_count = max(1, int(column_count or 1))
    if safe_column_count > 1:
        worksheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=safe_column_count,
        )


def estimate_wrapped_row_height(text, column_count=1, min_height=22, max_height=240):
    safe_text = str(text or "")
    if not safe_text:
        return min_height

    approx_chars_per_line = max(24, int(max(1, column_count)) * 18)
    logical_lines = 0
    for raw_line in safe_text.splitlines() or [""]:
        line_length = max(1, len(raw_line))
        logical_lines += max(1, (line_length + approx_chars_per_line - 1) // approx_chars_per_line)

    return max(min_height, min(max_height, 16 + logical_lines * 15))


def initialize_excel_worksheet(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90


def estimate_excel_text_width(value):
    normalized_value = str(normalize_excel_cell_value(value))
    if not normalized_value:
        return 0

    max_line_width = 0
    for line in normalized_value.splitlines() or [""]:
        line_width = 0
        for character in line:
            line_width += 2 if unicodedata.east_asian_width(character) in {"F", "W", "A"} else 1
        max_line_width = max(max_line_width, line_width)

    return max_line_width


def get_autosize_ignored_cells(worksheet):
    ignored_cells = set()
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_col == merged_range.max_col and merged_range.min_row == merged_range.max_row:
            continue

        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                ignored_cells.add((row_index, column_index))

    return ignored_cells


def autosize_worksheet_columns(worksheet, min_width=12, max_width=120):
    ignored_cells = get_autosize_ignored_cells(worksheet)
    min_width_overrides = getattr(worksheet, "_codex_min_column_widths", {})

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        measured_width = min_width

        for row_index in range(1, worksheet.max_row + 1):
            if (row_index, column_index) in ignored_cells:
                continue

            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value in (None, ""):
                continue
            measured_width = max(
                measured_width,
                estimate_excel_text_width(cell_value),
            )

        worksheet.column_dimensions[column_letter].width = max(
            min_width,
            min(max_width, max(measured_width + 3, float(min_width_overrides.get(column_letter, min_width)))),
        )


def append_table_to_worksheet(
    worksheet,
    title,
    rows,
    headers,
    start_row=1,
    description="",
    no_wrap_headers=None,
    min_column_widths=None,
):
    current_row = start_row
    column_count = max(1, len(headers))
    normalized_no_wrap_headers = {
        str(header)
        for header in (no_wrap_headers or [])
    }
    merge_excel_row(worksheet, current_row, column_count)
    title_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        title_cell,
        font=EXCEL_TITLE_FONT,
        fill=EXCEL_TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_TITLE_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    if description:
        merge_excel_row(worksheet, current_row, column_count)
        description_cell = worksheet.cell(row=current_row, column=1, value=description)
        style_excel_cell(
            description_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_SUBTITLE_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(description, column_count)
        current_row += 1

    header_row = current_row
    for column_index, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=header not in normalized_no_wrap_headers),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    if not rows:
        merge_excel_row(worksheet, current_row, column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value="表示できるデータがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    data_start_row = current_row
    for row_index, row in enumerate(rows, start=0):
        for column_index, header in enumerate(headers, start=1):
            body_cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
            fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(wrap_text=header not in normalized_no_wrap_headers, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

    if min_column_widths:
        width_overrides = dict(getattr(worksheet, "_codex_min_column_widths", {}))
        for column_index, header in enumerate(headers, start=1):
            if header in min_column_widths:
                width_overrides[get_column_letter(column_index)] = max(
                    float(width_overrides.get(get_column_letter(column_index), 0)),
                    float(min_column_widths[header]),
                )
        worksheet._codex_min_column_widths = width_overrides

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(column_count)}{current_row - 1}"

    return current_row + 1


def append_key_value_rows(worksheet, title, rows, start_row=1, description=""):
    current_row = start_row
    merge_excel_row(worksheet, current_row, 2)
    title_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        title_cell,
        font=EXCEL_TITLE_FONT,
        fill=EXCEL_TITLE_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_TITLE_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    if description:
        merge_excel_row(worksheet, current_row, 2)
        description_cell = worksheet.cell(row=current_row, column=1, value=description)
        style_excel_cell(
            description_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_SUBTITLE_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(description, 2)
        current_row += 1

    for column_index, header in enumerate(("項目", "値"), start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    data_start_row = current_row
    for row_index, row in enumerate(rows, start=0):
        if isinstance(row, dict):
            label = row.get("label", "")
            value = row.get("value", "")
            row_style = row.get("style", "default")
        else:
            label, value = row
            row_style = "default"
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        label_cell = worksheet.cell(row=current_row, column=1, value=label)
        value_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(value))
        if row_style == "note":
            note_fill = EXCEL_SUBTITLE_FILL if fill is None else fill
            style_excel_cell(
                label_cell,
                font=EXCEL_NOTE_FONT,
                fill=note_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
            style_excel_cell(
                value_cell,
                font=EXCEL_NOTE_FONT,
                fill=note_fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        else:
            style_excel_cell(
                label_cell,
                font=EXCEL_BOLD_FONT,
                fill=EXCEL_LABEL_FILL if fill is None else fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
            style_excel_cell(
                value_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(vertical="top", wrap_text=True),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = max(
            20,
            estimate_wrapped_row_height(value, 1, min_height=20, max_height=120),
        )
        current_row += 1

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref and rows:
        worksheet.auto_filter.ref = f"A{data_start_row - 1}:B{current_row - 1}"

    return current_row + 1


def append_bullet_rows(worksheet, title, items, start_row=1, column_count=6, empty_text="表示できる要点がありません。", section_header_fill=None):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=section_header_fill or EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    if not items:
        merge_excel_row(worksheet, current_row, safe_column_count)
        empty_cell = worksheet.cell(row=current_row, column=1, value=empty_text)
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    for index, item in enumerate(items, start=1):
        bullet_cell = worksheet.cell(row=current_row, column=1, value=f"{index}.")
        style_excel_cell(
            bullet_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_ALT_ROW_FILL if index % 2 == 0 else None,
            alignment=Alignment(horizontal="center", vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        if safe_column_count > 1:
            worksheet.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=current_row,
                end_column=safe_column_count,
            )
        text_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(item))
        style_excel_cell(
            text_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_ALT_ROW_FILL if index % 2 == 0 else None,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(item, safe_column_count - 1)
        current_row += 1

    return current_row + 1


def append_text_block_to_worksheet(worksheet, title, text, start_row=1, column_count=6):
    current_row = start_row
    safe_column_count = max(1, int(column_count or 1))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_BOLD_FONT,
        fill=EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    merge_excel_row(worksheet, current_row, safe_column_count)
    body_cell = worksheet.cell(row=current_row, column=1, value=normalize_excel_cell_value(text) or "解説を表示できませんでした。")
    style_excel_cell(
        body_cell,
        font=EXCEL_BODY_FONT,
        fill=EXCEL_TEXT_BLOCK_FILL,
        alignment=Alignment(wrap_text=True, vertical="top"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(body_cell.value, safe_column_count, min_height=88, max_height=360)

    return current_row + 2


def append_custom_text_section_to_worksheet(
    worksheet,
    title,
    text,
    start_row=1,
    column_count=6,
    header_fill=None,
    body_fill=None,
    empty_text="表示できる内容がありません。",
):
    current_row = start_row
    safe_column_count = max(1, int(column_count or 1))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=header_fill or EXCEL_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    merge_excel_row(worksheet, current_row, safe_column_count)
    body_cell = worksheet.cell(
        row=current_row,
        column=1,
        value=normalize_excel_cell_value(text) or empty_text,
    )
    style_excel_cell(
        body_cell,
        font=EXCEL_BODY_FONT,
        fill=body_fill or EXCEL_TEXT_BLOCK_FILL,
        alignment=Alignment(wrap_text=True, vertical="top"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(
        body_cell.value,
        safe_column_count,
        min_height=54,
        max_height=240,
    )

    return current_row + 2


def append_definition_table_to_worksheet(worksheet, title, rows, start_row=1, column_count=6, header_fill=None):
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    merge_excel_row(worksheet, current_row, safe_column_count)
    section_cell = worksheet.cell(row=current_row, column=1, value=title)
    style_excel_cell(
        section_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=header_fill or EXCEL_MUTED_SECTION_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    headers = ("用語", "説明")
    for column_index, header in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=current_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    for row_index, row in enumerate(rows or [], start=0):
        term_cell = worksheet.cell(row=current_row, column=1, value=normalize_excel_cell_value(row.get("用語")))
        description_cell = worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(row.get("説明")))
        style_excel_cell(
            term_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_TERMINOLOGY_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_TERMINOLOGY_BORDER,
        )
        style_excel_cell(
            description_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_TERMINOLOGY_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_TERMINOLOGY_BORDER,
        )
        worksheet.row_dimensions[current_row].height = max(
            20,
            estimate_wrapped_row_height(row.get("説明"), 1, min_height=20, max_height=96),
        )
        current_row += 1

    return current_row + 1


def build_rich_text_with_bold_numbers(text, base_color="1F2937", base_size=10):
    """テキスト内の数値・数値+単位を太字にした CellRichText を返す。

    数値が含まれない場合は元のテキスト文字列をそのまま返す。
    """
    import re
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    safe_text = str(text or "").strip()
    if not safe_text:
        return safe_text

    # 数値パターン: 数字で始まり、カンマ・ピリオド・数字が続く部分 + 直後の単位文字
    number_pattern = re.compile(
        r"(\d[\d,.]*\d?)"
        r"(%|件|分|秒|時間|日|月|年|回|個|ケース|イベント|種類|パターン|s|ms|min|h)?"
    )

    # 分割: マッチ部分とそれ以外に分ける
    parts = []
    last_end = 0
    for match in number_pattern.finditer(safe_text):
        if match.start() > last_end:
            parts.append(("text", safe_text[last_end : match.start()]))
        parts.append(("number", match.group(0)))
        last_end = match.end()
    if last_end < len(safe_text):
        parts.append(("text", safe_text[last_end:]))

    # 数値が見つからなかった場合はそのまま返す
    if not any(kind == "number" for kind, _ in parts):
        return safe_text

    normal_font = InlineFont(sz=base_size, color=base_color)
    bold_font = InlineFont(b=True, sz=base_size, color=base_color)

    rich_parts = []
    for kind, value in parts:
        if not value:
            continue
        if kind == "number":
            rich_parts.append(TextBlock(bold_font, value))
        else:
            rich_parts.append(TextBlock(normal_font, value))

    if not rich_parts:
        return safe_text

    return CellRichText(*rich_parts)


def append_structured_text_block_to_worksheet(
    worksheet,
    section_title,
    full_text,
    start_row=1,
    column_count=6,
    section_header_fill=None,
    subsection_marker_start="【",
    subsection_marker_end="】",
):
    """解説本文を【】マーカーでサブセクションに分割し、小見出し付きで出力する。

    【全体傾向】【注目ポイント】【ボトルネック示唆】のようなマーカーを検出し、
    各サブセクションを小見出し行（太字）+ テキストブロックで表示する。
    マーカーが見つからない場合は、従来通り1ブロックで出力する。
    """
    import re
    current_row = start_row
    safe_column_count = max(2, int(column_count or 2))
    header_fill = section_header_fill or EXCEL_SECTION_FILL

    # セクション見出し行
    merge_excel_row(worksheet, current_row, safe_column_count)
    title_cell = worksheet.cell(row=current_row, column=1, value=section_title)
    style_excel_cell(
        title_cell,
        font=EXCEL_SECTION_HEADER_FONT,
        fill=header_fill,
        alignment=Alignment(horizontal="left", vertical="center"),
        border=EXCEL_THIN_BORDER,
    )
    worksheet.row_dimensions[current_row].height = 22
    current_row += 1

    normalized_text = str(full_text or "").strip()
    if not normalized_text:
        merge_excel_row(worksheet, current_row, safe_column_count)
        empty_cell = worksheet.cell(
            row=current_row, column=1, value="解説テキストがありません。"
        )
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 22
        return current_row + 2

    # 【...】マーカーでテキストを分割
    marker_pattern = re.compile(
        rf"^{re.escape(subsection_marker_start)}(.+?){re.escape(subsection_marker_end)}\s*$",
        re.MULTILINE,
    )
    parts = marker_pattern.split(normalized_text)
    # parts: [前テキスト, マーカー1タイトル, マーカー1本文, マーカー2タイトル, マーカー2本文, ...]

    if len(parts) <= 1:
        # マーカーなし: 従来通り1ブロックで出力
        merge_excel_row(worksheet, current_row, safe_column_count)
        text_cell = worksheet.cell(
            row=current_row, column=1, value=build_rich_text_with_bold_numbers(normalized_text)
        )
        style_excel_cell(
            text_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_TEXT_BLOCK_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(
            normalized_text, safe_column_count
        )
        return current_row + 2

    # 最初の部分（マーカー前のテキスト）がある場合
    preamble = parts[0].strip()
    if preamble:
        merge_excel_row(worksheet, current_row, safe_column_count)
        preamble_cell = worksheet.cell(
            row=current_row, column=1, value=build_rich_text_with_bold_numbers(preamble)
        )
        style_excel_cell(
            preamble_cell,
            font=EXCEL_BODY_FONT,
            fill=EXCEL_TEXT_BLOCK_FILL,
            alignment=Alignment(wrap_text=True, vertical="top"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = estimate_wrapped_row_height(
            preamble, safe_column_count
        )
        current_row += 1

    # マーカー付きサブセクションを出力
    for i in range(1, len(parts), 2):
        subsection_title = parts[i].strip()
        subsection_body = parts[i + 1].strip() if i + 1 < len(parts) else ""

        # 小見出し行（太字、背景なし）
        merge_excel_row(worksheet, current_row, safe_column_count)
        subtitle_cell = worksheet.cell(
            row=current_row, column=1, value=f"▸ {subsection_title}"
        )
        style_excel_cell(
            subtitle_cell,
            font=EXCEL_BOLD_FONT,
            fill=None,
            alignment=Alignment(horizontal="left", vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

        # 本文テキスト
        if subsection_body:
            merge_excel_row(worksheet, current_row, safe_column_count)
            body_cell = worksheet.cell(
                row=current_row, column=1, value=build_rich_text_with_bold_numbers(subsection_body)
            )
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=EXCEL_TEXT_BLOCK_FILL,
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
            worksheet.row_dimensions[current_row].height = (
                estimate_wrapped_row_height(subsection_body, safe_column_count)
            )
            current_row += 1

    return current_row + 1

