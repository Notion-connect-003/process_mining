from io import BytesIO
from datetime import datetime, timezone
import duckdb

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from web_reports.excel_common import (
    ANALYSIS_PRECONDITIONS_TEXT,
    APPLIED_FILTERS_NOTE_TEXT,
    EXCEL_ALT_ROW_FILL,
    EXCEL_ASSUMPTION_SECTION_FILL,
    EXCEL_BODY_FONT,
    EXCEL_BOLD_FONT,
    EXCEL_GROUP_SECTION_FILL,
    EXCEL_GROUP_SECTION_FONT,
    EXCEL_HEADER_FILL,
    EXCEL_LABEL_FILL,
    EXCEL_MUTED_FONT,
    EXCEL_MUTED_SECTION_FILL,
    EXCEL_NOTE_FONT,
    EXCEL_SECTION_FILL,
    EXCEL_SUBTITLE_FILL,
    EXCEL_TEXT_BLOCK_FILL,
    EXCEL_THIN_BORDER,
    EXCEL_TITLE_BORDER,
    EXCEL_TITLE_FILL,
    EXCEL_TITLE_FONT,
    GROUPING_CONDITION_NOTE_TEXT,
    LOG_DIAGNOSTIC_SHEET_NAMES,
    REPORT_HEADER_LABELS,
    REPORT_SHEET_NAMES,
    TERMINOLOGY_ROWS,
    append_bullet_rows,
    append_custom_text_section_to_worksheet,
    append_definition_table_to_worksheet,
    append_key_value_rows,
    append_table_to_worksheet,
    append_text_block_to_worksheet,
    autosize_worksheet_columns,
    build_analysis_excel_file_name,
    estimate_wrapped_row_height,
    initialize_excel_worksheet,
    merge_excel_row,
    normalize_excel_cell_value,
    resolve_analysis_display_name,
    sanitize_workbook_sheet_name,
    style_excel_cell,
)

from 共通スクリプト.Excel出力.excel_exporter import (
    build_excel_bytes,
    build_summary_sheet_df,
    convert_analysis_result_to_records,
)

from 共通スクリプト.analysis_service import (
    DEFAULT_ANALYSIS_KEYS,
    FLOW_PATTERN_CASE_COUNT_COLUMN,
    FLOW_PATTERN_COLUMN,
    analyze_prepared_event_log,
    build_group_summary,
    detect_group_columns,
    create_analysis_records,
    create_bottleneck_summary,
    create_dashboard_summary,
    create_impact_summary,
    create_log_diagnostics,
    create_rule_based_insights,
    create_root_cause_summary,
    filter_prepared_df,
    filter_prepared_df_by_pattern,
    get_filter_options,
    create_variant_summary,
    create_pattern_flow_snapshot,
    create_pattern_bottleneck_details,
    get_available_analysis_definitions,
    load_prepared_event_log,
    merge_filter_params,
    normalize_filter_params,
    normalize_filter_column_settings,
    select_pattern_rows_for_flow,
)

from 共通スクリプト.duckdb_service import (
    _build_scoped_relation_cte,
    _format_stddev_column,
    _get_parquet_column_names,
    _quote_identifier,
    persist_prepared_parquet,
    query_activity_case_drilldown,
    query_analysis_records,
    query_bottleneck_summary,
    query_case_trace_details,
    query_dashboard_summary,
    query_filter_options,
    query_filtered_meta,
    query_group_summary,
    query_impact_summary,
    query_pattern_bottleneck_details,
    query_period_text,
    query_root_cause_summary,
    query_transition_records_for_patterns,
    query_transition_case_drilldown,
    query_variant_summary,
)

def build_ranked_rows(rows, rank_key="rank"):
    ranked_rows = []
    for index, row in enumerate(rows, start=1):
        ranked_rows.append({
            rank_key: index,
            **row,
        })
    return ranked_rows

def _write_section_header(worksheet, row_index, title, column_count=10):
    safe_column_count = max(10, int(column_count or 10))
    for column_index in range(1, safe_column_count + 1):
        header_cell = worksheet.cell(row=row_index, column=column_index)
        if column_index == 1:
            header_cell.value = f"═══ {title} ═══"
            font = EXCEL_GROUP_SECTION_FONT
            alignment = Alignment(horizontal="left", vertical="center")
        else:
            font = EXCEL_GROUP_SECTION_FONT
            alignment = Alignment(horizontal="left", vertical="center")
        style_excel_cell(
            header_cell,
            font=font,
            fill=EXCEL_GROUP_SECTION_FILL,
            alignment=alignment,
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[row_index].height = 22
    return row_index + 1

def _iter_groups_from_parquet(
    parquet_path,
    grouping_columns,
    filter_params=None,
    filter_column_settings=None,
    variant_pattern=None,
):
    valid_columns = [
        str(column_name)
        for column_name in (grouping_columns or [])
        if str(column_name) in set(_get_parquet_column_names(parquet_path))
    ]
    if not valid_columns:
        return

    connection = duckdb.connect()
    try:
        cte_sql, params, relation_name = _build_scoped_relation_cte(
            parquet_path,
            filter_params=filter_params,
            filter_column_settings=filter_column_settings,
            variant_pattern=variant_pattern,
        )
        case_group_columns_sql = ",\n                ".join(
            [
                f"COALESCE(MAX(NULLIF(TRIM(CAST({_quote_identifier(column_name)} AS VARCHAR)), '')), '(未分類)') AS {_quote_identifier(column_name)}"
                for column_name in valid_columns
            ]
        )
        group_select_sql = ", ".join(_quote_identifier(column_name) for column_name in valid_columns)
        group_order_sql = ", ".join(_quote_identifier(column_name) for column_name in valid_columns)
        group_list_rows = connection.execute(
            f"""
            {cte_sql},
            case_groups AS (
                SELECT
                    case_id,
                    {case_group_columns_sql}
                FROM {relation_name}
                GROUP BY case_id
            )
            SELECT
                {group_select_sql},
                COUNT(*) AS case_count
            FROM case_groups
            GROUP BY {group_select_sql}
            ORDER BY case_count DESC, {group_order_sql}
            """,
            params,
        ).fetchall()
        if not group_list_rows:
            return

        frequency_display_columns = (
            get_available_analysis_definitions()
            .get("frequency", {})
            .get("config", {})
            .get("display_columns", {})
        )

        for row in group_list_rows:
            group_values = [str(value or "").strip() or "(未分類)" for value in row[: len(valid_columns)]]
            if len(valid_columns) == 1:
                group_name = group_values[0]
            else:
                group_name = ", ".join(
                    f"{column_name}={group_value}"
                    for column_name, group_value in zip(valid_columns, group_values)
                )

            group_conditions_sql = " AND ".join(
                f"{_quote_identifier(column_name)} = ?"
                for column_name in valid_columns
            )
            group_df = connection.execute(
                f"""
                {cte_sql},
                case_groups AS (
                    SELECT
                        case_id,
                        {case_group_columns_sql}
                    FROM {relation_name}
                    GROUP BY case_id
                ),
                selected_cases AS (
                    SELECT case_id
                    FROM case_groups
                    WHERE {group_conditions_sql}
                ),
                group_scoped AS (
                    SELECT scoped.*
                    FROM {relation_name} AS scoped
                    INNER JOIN selected_cases USING (case_id)
                )
                SELECT
                    activity,
                    COUNT(*) AS event_count,
                    COUNT(DISTINCT case_id) AS case_count,
                    ROUND(SUM(duration_min), 2) AS total_duration_min,
                    ROUND(AVG(duration_min), 2) AS avg_duration_min,
                    ROUND(MEDIAN(duration_min), 2) AS median_duration_min,
                    CASE WHEN COUNT(*) > 1 THEN ROUND(STDDEV_SAMP(duration_min), 2) ELSE NULL END AS std_duration_min,
                    ROUND(MIN(duration_min), 2) AS min_duration_min,
                    ROUND(MAX(duration_min), 2) AS max_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.75), 2) AS p75_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.90), 2) AS p90_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.95), 2) AS p95_duration_min,
                    ROUND(
                        COUNT(*) * 100.0 / NULLIF((SELECT COUNT(*) FROM group_scoped), 0),
                        2
                    ) AS event_ratio_pct
                FROM group_scoped
                GROUP BY activity
                ORDER BY event_count DESC, activity ASC
                """,
                params + group_values,
            ).df()
            if "std_duration_min" in group_df.columns:
                group_df["std_duration_min"] = _format_stddev_column(group_df["std_duration_min"])
            yield group_name, convert_analysis_result_to_records(
                group_df,
                frequency_display_columns,
            )
    finally:
        connection.close()

def _write_frequency_data(worksheet, rows, start_row):
    frequency_rows = build_ranked_rows(rows or [], rank_key=REPORT_HEADER_LABELS["rank"])
    frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
    header_row = start_row
    for column_index, header in enumerate(frequency_headers, start=1):
        header_cell = worksheet.cell(row=header_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[header_row].height = 22

    if not frequency_rows:
        merge_excel_row(worksheet, header_row + 1, len(frequency_headers))
        empty_cell = worksheet.cell(row=header_row + 1, column=1, value="表示できるデータがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[header_row + 1].height = 22
        return header_row + 3

    data_start_row = header_row + 1
    current_row = data_start_row
    for row_index, row in enumerate(frequency_rows, start=0):
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        for column_index, header in enumerate(frequency_headers, start=1):
            body_cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(frequency_headers))}{current_row - 1}"

    return current_row + 1

def localize_report_headers(headers):
    return [REPORT_HEADER_LABELS.get(header, header) for header in headers]

def localize_report_rows(rows, headers):
    localized_headers = localize_report_headers(headers)
    localized_rows = []

    for row in rows:
        localized_rows.append(
            {
                localized_header: row.get(header)
                for header, localized_header in zip(headers, localized_headers)
            }
        )

    return localized_rows, localized_headers

def build_transition_display_label(row):
    if not row:
        return ""

    transition_label = str(row.get("transition_label") or row.get("transition") or "").strip()
    if transition_label:
        return transition_label

    from_activity = str(row.get("from_activity") or row.get("activity") or "").strip()
    to_activity = str(row.get("to_activity") or row.get("next_activity") or "").strip()
    if from_activity and to_activity:
        return f"{from_activity} → {to_activity}"
    return from_activity or to_activity

def format_duration_text_for_report(duration_sec):
    total_seconds = max(0, int(round(float(duration_sec or 0))))
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)
    parts = []

    if days:
        parts.append(f"{days}d")
    if hours or days:
        parts.append(f"{hours}h")
    if minutes or hours or days:
        parts.append(f"{minutes}m")
    parts.append(f"{seconds}s")
    return " ".join(parts)

def build_bottleneck_export_rows(rows, label_key):
    export_rows = []

    for index, row in enumerate(rows or [], start=1):
        export_rows.append(
            {
                "rank": index,
                label_key: (
                    build_transition_display_label(row)
                    if label_key == "transition_label"
                    else str(row.get(label_key) or "").strip()
                ),
                "count": int(row.get("count") or 0),
                "case_count": int(row.get("case_count") or 0),
                "avg_duration_text": format_duration_text_for_report(row.get("avg_duration_sec")),
                "median_duration_text": format_duration_text_for_report(row.get("median_duration_sec")),
                "max_duration_text": format_duration_text_for_report(row.get("max_duration_sec")),
            }
        )

    return export_rows

def resolve_pattern_detail_sheet_count(pattern_display_limit, available_count):
    normalized_limit = str(pattern_display_limit or "10").strip().lower()

    if normalized_limit == "all":
        requested_count = available_count
    else:
        try:
            requested_count = int(normalized_limit)
        except (TypeError, ValueError):
            requested_count = 10

    requested_count = max(0, requested_count)
    if requested_count > 20:
        requested_count = 20

    return min(available_count, requested_count)

def build_pattern_overview_rows(pattern_rows, variant_items, pattern_column_label, analysis_definitions):
    pattern_config = analysis_definitions.get("pattern", {}).get("config", {})
    pattern_display_columns = pattern_config.get("display_columns", {})
    repeat_flag_label = pattern_display_columns.get("repeat_flag", "繰り返し")
    repeat_count_label = pattern_display_columns.get("repeat_count", "繰り返し回数")
    repeat_rate_label = pattern_display_columns.get("repeat_rate_pct", "繰り返し率(%)")
    repeat_rate_band_label = pattern_display_columns.get("repeat_rate_band", "繰り返し率区分")
    review_flag_label = pattern_display_columns.get("review_flag", "確認区分")
    avg_case_duration_diff_label = pattern_display_columns.get("avg_case_duration_diff_min", "平均処理時間差分(分)")
    improvement_priority_score_label = pattern_display_columns.get("improvement_priority_score", "改善優先度スコア")
    overall_impact_pct_label = pattern_display_columns.get("overall_impact_pct", "全体影響度(%)")
    fastest_pattern_flag_label = pattern_display_columns.get("fastest_pattern_flag", "最短処理")
    simple_comment_label = pattern_display_columns.get("simple_comment", "簡易コメント")
    case_count_label = pattern_display_columns.get("case_count", "ケース数")
    case_ratio_label = pattern_display_columns.get("case_ratio_pct", "ケース比率(%)")
    cumulative_case_ratio_label = pattern_display_columns.get("cumulative_case_ratio_pct", "累積カバー率(%)")
    avg_case_duration_label = pattern_display_columns.get("avg_case_duration_min", "平均ケース処理時間(分)")
    std_case_duration_label = pattern_display_columns.get("std_case_duration_min", "標準偏差ケース処理時間(分)")
    min_case_duration_label = pattern_display_columns.get("min_case_duration_min", "最小ケース処理時間(分)")
    max_case_duration_label = pattern_display_columns.get("max_case_duration_min", "最大ケース処理時間(分)")
    p75_case_duration_label = pattern_display_columns.get("p75_case_duration_min", "75%点ケース処理時間(分)")
    p90_case_duration_label = pattern_display_columns.get("p90_case_duration_min", "90%点ケース処理時間(分)")
    p95_case_duration_label = pattern_display_columns.get("p95_case_duration_min", "95%点ケース処理時間(分)")
    variant_by_pattern = {
        str(variant_item.get("pattern") or "").strip(): variant_item
        for variant_item in (variant_items or [])
    }
    overview_rows = []

    for index, pattern_row in enumerate(pattern_rows or [], start=1):
        pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
        matched_variant = variant_by_pattern.get(pattern_text, {})
        variant_id = matched_variant.get("variant_id")
        overview_rows.append(
            {
                "rank": index,
                "pattern_variant": (
                    f"Pattern #{index} / Variant #{variant_id}"
                    if variant_id
                    else f"Pattern #{index}"
                ),
                "repeat_flag": pattern_row.get(repeat_flag_label, ""),
                "repeat_count": pattern_row.get(repeat_count_label, 0),
                "repeat_rate_pct": pattern_row.get(repeat_rate_label, 0),
                "repeat_rate_band": pattern_row.get(repeat_rate_band_label, ""),
                "review_flag": pattern_row.get(review_flag_label, ""),
                "avg_case_duration_diff_min": pattern_row.get(avg_case_duration_diff_label, 0),
                "improvement_priority_score": pattern_row.get(improvement_priority_score_label, 0),
                "overall_impact_pct": pattern_row.get(overall_impact_pct_label, 0),
                "fastest_pattern_flag": pattern_row.get(fastest_pattern_flag_label, ""),
                "simple_comment": pattern_row.get(simple_comment_label, ""),
                "count": pattern_row.get(case_count_label, 0),
                "ratio": pattern_row.get(case_ratio_label, 0),
                "cumulative_case_ratio_pct": pattern_row.get(cumulative_case_ratio_label, 0),
                "avg_case_duration_min": pattern_row.get(avg_case_duration_label, 0),
                "std_case_duration_min": pattern_row.get(std_case_duration_label, 0),
                "min_case_duration_min": pattern_row.get(min_case_duration_label, 0),
                "max_case_duration_min": pattern_row.get(max_case_duration_label, 0),
                "p75_case_duration_min": pattern_row.get(p75_case_duration_label, 0),
                "p90_case_duration_min": pattern_row.get(p90_case_duration_label, 0),
                "p95_case_duration_min": pattern_row.get(p95_case_duration_label, 0),
                "pattern": pattern_text,
            }
        )

    return overview_rows

def coerce_report_number(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)

def build_pattern_export_summary(pattern_rows, pattern_display_columns):
    repeat_flag_label = pattern_display_columns.get("repeat_flag", "繰り返し")
    repeat_count_label = pattern_display_columns.get("repeat_count", "繰り返し回数")
    repeat_rate_label = pattern_display_columns.get("repeat_rate_pct", "繰り返し率(%)")
    repeat_rate_band_label = pattern_display_columns.get("repeat_rate_band", "繰り返し率区分")
    review_flag_label = pattern_display_columns.get("review_flag", "確認区分")
    avg_case_duration_diff_label = pattern_display_columns.get("avg_case_duration_diff_min", "平均処理時間差分(分)")
    improvement_priority_score_label = pattern_display_columns.get("improvement_priority_score", "改善優先度スコア")
    overall_impact_pct_label = pattern_display_columns.get("overall_impact_pct", "全体影響度(%)")
    fastest_pattern_flag_label = pattern_display_columns.get("fastest_pattern_flag", "最短処理")
    simple_comment_label = pattern_display_columns.get("simple_comment", "簡易コメント")
    case_count_label = pattern_display_columns.get("case_count", "ケース数")
    case_ratio_label = pattern_display_columns.get("case_ratio_pct", "ケース比率(%)")
    cumulative_case_ratio_label = pattern_display_columns.get("cumulative_case_ratio_pct", "累積カバー率(%)")
    avg_case_duration_label = pattern_display_columns.get("avg_case_duration_min", "平均ケース処理時間(分)")
    min_case_duration_label = pattern_display_columns.get("min_case_duration_min", "最小ケース処理時間(分)")
    max_case_duration_label = pattern_display_columns.get("max_case_duration_min", "最大ケース処理時間(分)")
    pattern_label = pattern_display_columns.get("pattern", "処理順パターン")

    comparison_rows = []
    repeated_patterns = []
    improvement_targets = []
    for index, pattern_row in enumerate(pattern_rows or [], start=1):
        comparison_row = {
            "順位": index,
            "繰り返し": pattern_row.get(repeat_flag_label, ""),
            "繰り返し回数": pattern_row.get(repeat_count_label, 0),
            "繰り返し率(%)": pattern_row.get(repeat_rate_label, 0),
            "繰り返し率区分": pattern_row.get(repeat_rate_band_label, ""),
            "件数": pattern_row.get(case_count_label, 0),
            "全体比率(%)": pattern_row.get(case_ratio_label, 0),
            "平均処理時間(分)": pattern_row.get(avg_case_duration_label, 0),
            "平均処理時間差分(分)": pattern_row.get(avg_case_duration_diff_label, 0),
            "改善優先度スコア": pattern_row.get(improvement_priority_score_label, 0),
            "全体影響度(%)": pattern_row.get(overall_impact_pct_label, 0),
            "最短処理": pattern_row.get(fastest_pattern_flag_label, ""),
            "最短処理時間(分)": pattern_row.get(min_case_duration_label, 0),
            "最長処理時間(分)": pattern_row.get(max_case_duration_label, 0),
            "確認区分": pattern_row.get(review_flag_label, ""),
            "簡易コメント": pattern_row.get(simple_comment_label, ""),
            "パターン": pattern_row.get(pattern_label, ""),
        }
        comparison_rows.append(comparison_row)
        if str(pattern_row.get(review_flag_label, "")).strip() != "要確認":
            pass
        else:
            repeated_patterns.append(comparison_row)
        if (
            coerce_report_number(pattern_row.get(repeat_rate_label, 0)) >= 10.0
            and coerce_report_number(pattern_row.get(avg_case_duration_diff_label, 0)) > 0
        ):
            improvement_targets.append(comparison_row)

    top3_rows = comparison_rows[:3]
    top3_coverage_pct = (
        coerce_report_number((pattern_rows or [])[2].get(cumulative_case_ratio_label, 0))
        if len(pattern_rows or []) >= 3
        else (
            coerce_report_number((pattern_rows or [])[-1].get(cumulative_case_ratio_label, 0))
            if pattern_rows
            else 0.0
        )
    )
    top10_coverage_pct = (
        coerce_report_number((pattern_rows or [])[9].get(cumulative_case_ratio_label, 0))
        if len(pattern_rows or []) >= 10
        else (
            coerce_report_number((pattern_rows or [])[-1].get(cumulative_case_ratio_label, 0))
            if pattern_rows
            else 0.0
        )
    )
    repeated_case_ratio_pct = round(
        sum(coerce_report_number(problem_row["全体比率(%)"]) for problem_row in repeated_patterns),
        2,
    )
    improvement_targets = sorted(
        improvement_targets,
        key=lambda row: (
            -coerce_report_number(row.get("改善優先度スコア"), 0),
            -coerce_report_number(row.get("繰り返し率(%)"), 0),
            -coerce_report_number(row.get("平均処理時間差分(分)"), 0),
            -coerce_report_number(row.get("件数"), 0),
            row.get("パターン", ""),
        ),
    )[:3]
    fastest_pattern = min(
        comparison_rows,
        key=lambda row: (
            coerce_report_number(row.get("平均処理時間(分)"), float("inf")),
            row.get("順位", 0),
        ),
        default=None,
    )
    coverage_summary_text = (
        f"上位3パターンで {round(top3_coverage_pct, 2):.2f}%、"
        f"上位10パターンで {round(top10_coverage_pct, 2):.2f}% をカバーしています。"
    )

    return {
        "top_patterns": top3_rows,
        "comparison_rows": comparison_rows[:10],
        "repeated_patterns": repeated_patterns,
        "top3_coverage_pct": round(top3_coverage_pct, 2),
        "top10_coverage_pct": round(top10_coverage_pct, 2),
        "coverage_summary_text": coverage_summary_text,
        "repeated_case_ratio_pct": repeated_case_ratio_pct,
        "fastest_pattern": fastest_pattern or {},
        "improvement_targets": improvement_targets,
    }

def calculate_pattern_time_impact_minutes(pattern_row):
    return round(
        max(0.0, coerce_report_number(pattern_row.get("平均処理時間差分(分)"), 0.0))
        * max(0.0, coerce_report_number(pattern_row.get("件数"), 0.0)),
        2,
    )

def build_pattern_issue_row(pattern_row):
    repeat_rate_pct = coerce_report_number(pattern_row.get("繰り返し率(%)"), 0.0)
    duration_diff_min = coerce_report_number(pattern_row.get("平均処理時間差分(分)"), 0.0)
    pattern_text = str(pattern_row.get("パターン") or "").strip()

    if repeat_rate_pct >= 30:
        issue_text = f"繰り返し率が {repeat_rate_pct:.2f}% と高く、手戻りが多いパターンです。"
        cause_text = "差戻しや再確認が発生しやすく、同じ工程を複数回通過している可能性があります。"
        action_text = "差戻し発生条件を洗い出し、一次判定や入力チェックの前倒しを検討してください。"
    elif repeat_rate_pct >= 10:
        issue_text = f"繰り返し率が {repeat_rate_pct:.2f}% あり、再作業が混在しています。"
        cause_text = "一部ケースで確認や承認のやり直しが発生し、処理が伸びている可能性があります。"
        action_text = "繰り返しが起きる工程の条件分岐を整理し、再実行の発生源を減らしてください。"
    else:
        issue_text = f"繰り返しは少ないものの、平均処理時間が全体平均より {duration_diff_min:.2f} 分長いパターンです。"
        cause_text = "特定工程の待ちや滞留により、パターン全体の処理時間が長くなっている可能性があります。"
        action_text = "ボトルネック工程の担当・承認・待機条件を見直し、滞留時間の短縮を優先してください。"

    return {
        "問題点": issue_text,
        "原因": cause_text,
        "改善案": action_text,
        "期待効果（時間短縮・分）": calculate_pattern_time_impact_minutes(pattern_row),
        "対象パターン": pattern_text,
    }

def build_pattern_conclusion_summary(pattern_summary):
    comparison_rows = pattern_summary.get("comparison_rows", [])
    improvement_targets = pattern_summary.get("improvement_targets", [])
    repeated_patterns = pattern_summary.get("repeated_patterns", [])
    fastest_pattern = pattern_summary.get("fastest_pattern", {}) or {}
    top_issue_candidates = improvement_targets[:3] if improvement_targets else comparison_rows[:3]
    issue_rows = [build_pattern_issue_row(row) for row in top_issue_candidates]
    total_impact_minutes = round(
        sum(calculate_pattern_time_impact_minutes(row) for row in improvement_targets),
        2,
    )
    overall_summary = (
        f"上位10パターンで {coerce_report_number(pattern_summary.get('top10_coverage_pct', 0.0)):.2f}% をカバーし、"
        f"要確認パターンは {len(repeated_patterns)} 件、改善対象TOP3で約 "
        f"{round(sum(calculate_pattern_time_impact_minutes(row) for row in improvement_targets[:3]), 2):.2f} 分の短縮余地があります。"
    )

    return {
        "overall_summary": overall_summary,
        "issue_rows": issue_rows,
        "total_impact_minutes": total_impact_minutes,
        "total_impact_hours": round(total_impact_minutes / 60.0, 2),
        "fastest_pattern": fastest_pattern,
        "improvement_targets": improvement_targets[:3],
    }

def build_pattern_dashboard_summary(pattern_summary, pattern_conclusion):
    top3_rows = pattern_summary.get("improvement_targets") or pattern_summary.get("comparison_rows", [])[:3]
    problem_points = [
        issue_row.get("問題点")
        for issue_row in pattern_conclusion.get("issue_rows", [])
        if str(issue_row.get("問題点") or "").strip()
    ]
    if not problem_points and pattern_summary.get("comparison_rows"):
        problem_points = [
            str(row.get("簡易コメント") or "").strip()
            for row in pattern_summary.get("comparison_rows", [])[:3]
            if str(row.get("簡易コメント") or "").strip()
        ]

    return {
        "overall_summary": pattern_conclusion.get("overall_summary", ""),
        "top3_rows": top3_rows,
        "problem_points": problem_points[:3],
        "top10_coverage_pct": pattern_summary.get("top10_coverage_pct", 0),
        "total_impact_minutes": pattern_conclusion.get("total_impact_minutes", 0),
    }

def _set_chart_str_categories(chart, labels_ref):
    """Set chart categories as strRef so Excel treats text labels correctly."""
    label_formula = str(labels_ref)
    for series in chart.ser:
        series.cat = AxDataSource(strRef=StrRef(f=label_formula))

def _ensure_chart_data_sheet(workbook):
    sheet_name = "_chart_data"
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    data_sheet = workbook.create_sheet(title=sheet_name)
    data_sheet.sheet_state = "hidden"
    return data_sheet

def _write_chart_data_block(workbook, block_label, comparison_rows, columns):
    data_sheet = _ensure_chart_data_sheet(workbook)
    start_row = (data_sheet.max_row or 0) + 2

    data_sheet.cell(row=start_row, column=1, value=block_label)
    header_row = start_row + 1
    for col_offset, col_name in enumerate(columns):
        data_sheet.cell(row=header_row, column=1 + col_offset, value=col_name)

    for index, row in enumerate(comparison_rows, start=1):
        data_row = header_row + index
        data_sheet.cell(row=data_row, column=1, value=f"Pattern #{row.get('順位', index)}")
        for col_offset, col_name in enumerate(columns):
            if col_offset == 0:
                continue
            data_sheet.cell(
                row=data_row, column=1 + col_offset,
                value=coerce_report_number(row.get(col_name), 0.0),
            )

    return data_sheet, header_row, header_row + len(comparison_rows)

def build_excel_anchor(column_letter, row_number):
    return f"{column_letter}{max(1, int(row_number or 1))}"

def sort_pattern_rows_by_avg_duration_desc(rows):
    return sorted(
        rows,
        key=lambda row: (
            -coerce_report_number(row.get("平均処理時間(分)"), 0.0),
            row.get("順位", 0),
        ),
    )

def append_pattern_dashboard_pie_chart(workbook, worksheet, comparison_rows, anchor="A1"):
    if not comparison_rows:
        return

    data_sheet, header_row, max_row = _write_chart_data_block(
        workbook, "dashboard_pie", comparison_rows, ["パターン", "件数"],
    )

    pie_chart = PieChart()
    pie_chart.title = "上位10パターンの割合"
    pie_chart.style = 10
    pie_chart.height = 14.0
    pie_chart.width = 16.0
    pie_data = Reference(data_sheet, min_col=2, min_row=header_row, max_row=max_row)
    pie_labels = Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    _set_chart_str_categories(pie_chart, pie_labels)
    pie_chart.dLbls = DataLabelList(
        showCatName=False,
        showVal=False,
        showPercent=True,
        showSerName=False,
        showLegendKey=False,
        showLeaderLines=True,
        dLblPos="bestFit",
        separator="\n",
    )
    worksheet.add_chart(pie_chart, anchor)

def append_pattern_conclusion_charts(workbook, worksheet, comparison_rows, pie_anchor="A1", bar_anchor="A20"):
    if not comparison_rows:
        return

    data_sheet, header_row, max_row = _write_chart_data_block(
        workbook, "conclusion_charts", comparison_rows,
        ["パターン", "件数"],
    )
    bar_rows = sort_pattern_rows_by_avg_duration_desc(comparison_rows)
    bar_data_sheet, bar_header_row, bar_max_row = _write_chart_data_block(
        workbook,
        "conclusion_duration_desc",
        bar_rows,
        ["パターン", "平均処理時間(分)"],
    )

    pie_chart = PieChart()
    pie_chart.title = "上位10パターンの割合"
    pie_chart.style = 10
    pie_chart.height = 14.0
    pie_chart.width = 16.0
    pie_data = Reference(data_sheet, min_col=2, min_row=header_row, max_row=max_row)
    pie_labels = Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    _set_chart_str_categories(pie_chart, pie_labels)
    pie_chart.dLbls = DataLabelList(
        showCatName=False,
        showVal=False,
        showPercent=True,
        showSerName=False,
        showLegendKey=False,
        showLeaderLines=True,
        dLblPos="bestFit",
        separator="\n",
    )
    worksheet.add_chart(pie_chart, pie_anchor)

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 10
    bar_chart.title = "平均処理時間の比較（長い順）"
    bar_chart.height = 14.0
    bar_chart.width = 16.0
    bar_chart.legend = None
    bar_chart.varyColors = False
    bar_chart.gapWidth = 60
    bar_chart.x_axis.scaling.orientation = "maxMin"
    bar_chart.x_axis.tickLblPos = "low"
    bar_chart.y_axis.crosses = "autoZero"
    bar_chart.y_axis.delete = False
    bar_data = Reference(bar_data_sheet, min_col=2, min_row=bar_header_row, max_row=bar_max_row)
    bar_labels = Reference(bar_data_sheet, min_col=1, min_row=bar_header_row + 1, max_row=bar_max_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    _set_chart_str_categories(bar_chart, bar_labels)
    bar_chart.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        dLblPos="outEnd",
    )
    worksheet.add_chart(bar_chart, bar_anchor)

def append_pattern_detail_sheet(
    workbook,
    filtered_df,
    pattern_row,
    pattern_rank,
    pattern_column_label,
    analysis_definitions,
    variant_item=None,
    pattern_detail=None,
):
    pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
    if not pattern_text:
        return

    detail = pattern_detail or create_pattern_bottleneck_details(filtered_df, pattern_text)
    sheet_name = sanitize_workbook_sheet_name(f"パターン{pattern_rank:02d}詳細")
    detail_sheet = workbook.create_sheet(title=sheet_name)
    initialize_excel_worksheet(detail_sheet)

    variant_label = (
        f"Variant #{variant_item['variant_id']}"
        if variant_item and variant_item.get("variant_id")
        else "該当なし"
    )
    bottleneck_transition = detail.get("bottleneck_transition") or {}
    next_row = append_key_value_rows(
        detail_sheet,
        f"Pattern #{pattern_rank} 詳細",
        [
            ("パターン / バリアント", f"Pattern #{pattern_rank} / {variant_label}"),
            ("繰り返し", pattern_row.get("繰り返し", "")),
            ("繰り返し回数", pattern_row.get("繰り返し回数", 0)),
            ("繰り返し率(%)", pattern_row.get("繰り返し率(%)", 0)),
            ("繰り返し率区分", pattern_row.get("繰り返し率区分", "")),
            ("確認区分", pattern_row.get("確認区分", "")),
            ("平均処理時間差分(分)", pattern_row.get("平均処理時間差分(分)", 0)),
            ("改善優先度スコア", pattern_row.get("改善優先度スコア", 0)),
            ("全体影響度(%)", pattern_row.get("全体影響度(%)", 0)),
            ("最短処理", pattern_row.get("最短処理", "")),
            ("簡易コメント", pattern_row.get("簡易コメント", "")),
            ("ケース数", detail.get("case_count", 0)),
            ("ケース比率(%)", detail.get("case_ratio_pct", 0)),
            ("平均ケース処理時間(分)", detail.get("avg_case_duration_min", 0)),
            ("中央値ケース処理時間(分)", detail.get("median_case_duration_min", 0)),
            ("最小ケース処理時間(分)", detail.get("min_case_duration_min", 0)),
            ("最大ケース処理時間(分)", detail.get("max_case_duration_min", 0)),
            ("代表ルート", pattern_text),
            ("ボトルネック遷移", bottleneck_transition.get("transition_label", "該当なし")),
            ("ボトルネック平均所要時間(分)", bottleneck_transition.get("avg_duration_min", 0)),
        ],
        description="上位パターンのケース概要、代表ルート、ボトルネック遷移をまとめています。",
    )

    step_metric_rows, step_metric_headers = localize_report_rows(
        [
            {
                "rank": index,
                "sequence_no": step_metric["sequence_no"],
                "transition": step_metric["transition_label"],
                "case_count": step_metric["case_count"],
                "avg_duration_min": step_metric["avg_duration_min"],
                "median_duration_min": step_metric["median_duration_min"],
                "min_duration_min": step_metric["min_duration_min"],
                "max_duration_min": step_metric["max_duration_min"],
                "total_duration_min": step_metric["total_duration_min"],
                "wait_share_pct": step_metric["wait_share_pct"],
            }
            for index, step_metric in enumerate(detail.get("step_metrics", []), start=1)
        ],
        [
            "rank",
            "sequence_no",
            "transition",
            "case_count",
            "avg_duration_min",
            "median_duration_min",
            "min_duration_min",
            "max_duration_min",
            "total_duration_min",
            "wait_share_pct",
        ],
    )
    next_row = append_table_to_worksheet(
        detail_sheet,
        "ステップ別所要時間",
        step_metric_rows,
        step_metric_headers,
        start_row=next_row,
        description="各ステップの所要時間と全体に占める比率を比較できます。",
        no_wrap_headers=["遷移"],
        min_column_widths={"遷移": 32},
    )

    case_example_rows, case_example_headers = localize_report_rows(
        [
            {
                "rank": index,
                "case_id": case_example["case_id"],
                "start_time": case_example["start_time"],
                "end_time": case_example["end_time"],
                "total_duration_min": case_example["case_total_duration_min"],
            }
            for index, case_example in enumerate(detail.get("case_examples", []), start=1)
        ],
        ["rank", "case_id", "start_time", "end_time", "total_duration_min"],
    )
    append_table_to_worksheet(
        detail_sheet,
        "代表ケース",
        case_example_rows,
        case_example_headers,
        start_row=next_row,
        description="このパターンに属するケースのうち、総処理時間が長いケースを上位から掲載しています。",
    )

def build_detail_summary_kpi_rows(
    analysis_key,
    analysis_rows,
    dashboard_summary,
    impact_summary,
    bottleneck_summary,
    prepared_df=None,
    variant_items=None,
):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    top_row = analysis_rows[0] if analysis_rows else {}
    top_transition_bottleneck = (
        bottleneck_summary["transition_bottlenecks"][0]
        if bottleneck_summary.get("transition_bottlenecks")
        else None
    )
    top_transition_bottleneck_label = build_transition_display_label(top_transition_bottleneck) or "該当なし"
    top_impact_row = impact_summary["rows"][0] if impact_summary.get("rows") else None

    common_rows = [
        ("平均ケース処理時間", dashboard_summary.get("avg_case_duration_text", "0s")),
        ("中央値ケース処理時間", dashboard_summary.get("median_case_duration_text", "0s")),
    ]

    if normalized_analysis_key == "frequency":
        frequency_rows = common_rows + [
            ("最大ケース処理時間", dashboard_summary.get("max_case_duration_text", "0s")),
            ("最多アクティビティ", top_row.get("アクティビティ", "該当なし") if top_row else "該当なし"),
            ("最多アクティビティ件数", normalize_excel_cell_value(top_row.get("イベント件数", 0)) if top_row else 0),
            ("上位10バリアントカバー率", f"{float(dashboard_summary.get('top10_variant_coverage_pct', 0.0)):.2f}%"),
        ]
        if variant_items is not None:
            frequency_rows.append(("バリアント総数", len(variant_items)))

        unique_activity_count = dashboard_summary.get("unique_activity_count")
        if unique_activity_count in (None, ""):
            unique_activity_count = dashboard_summary.get("activity_type_count")
        if unique_activity_count in (None, "") and prepared_df is not None and "activity" in prepared_df.columns:
            unique_activity_count = int(prepared_df["activity"].nunique())
        if unique_activity_count not in (None, ""):
            frequency_rows.append(("ユニークアクティビティ数", int(unique_activity_count)))

        total_cases = int(dashboard_summary.get("total_cases", 0) or 0)
        total_records = int(dashboard_summary.get("total_records", 0) or 0)
        if total_cases > 0:
            frequency_rows.append(("平均ケースあたりイベント数", round(total_records / total_cases, 2)))

        return frequency_rows

    if normalized_analysis_key == "transition":
        top_transition_label = (
            f"{top_row.get('前処理アクティビティ名', '')} → {top_row.get('後処理アクティビティ名', '')}".strip(" →")
            if top_row
            else "該当なし"
        )
        return common_rows + [
            ("主要遷移", top_transition_label or "該当なし"),
            ("最大ボトルネック遷移", top_transition_bottleneck_label),
            ("最大改善インパクト遷移", top_impact_row.get("transition_label", "該当なし") if top_impact_row else "該当なし"),
        ]

    if normalized_analysis_key == "pattern":
        return common_rows + [
            ("最頻出パターン", top_row.get("処理順パターン", top_row.get("パターン", "該当なし")) if top_row else "該当なし"),
            ("最頻出パターン比率", f"{float(top_row.get('ケース比率(%)', 0.0)):.2f}%" if top_row else "0.00%"),
            ("上位10バリアントカバー率", f"{float(dashboard_summary.get('top10_variant_coverage_pct', 0.0)):.2f}%"),
        ]

    return common_rows + [
        ("最大ケース処理時間", dashboard_summary.get("max_case_duration_text", "0s")),
        ("最大ボトルネック遷移", top_transition_bottleneck_label),
    ]

def _create_report_sheet(workbook, title, *, use_active=False):
    worksheet = workbook.active if use_active else workbook.create_sheet()
    worksheet.title = sanitize_workbook_sheet_name(title)
    initialize_excel_worksheet(worksheet)
    return worksheet

def _serialize_workbook_bytes(workbook):
    for worksheet in workbook.worksheets:
        autosize_worksheet_columns(worksheet)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()

def _build_detail_export_summary_rows(
    context,
    run_data,
    analysis_key,
    variant_id=None,
    selected_activity="",
    case_id="",
):
    group_columns = context["group_columns"]
    grouping_text = "、".join(group_columns) if group_columns else "なし"
    summary_rows = [
        (REPORT_HEADER_LABELS["analysis_key"], analysis_key),
        (REPORT_HEADER_LABELS["analysis_name"], context["analysis_name"]),
        (REPORT_HEADER_LABELS["source_file_name"], run_data["source_file_name"]),
        (REPORT_HEADER_LABELS["analysis_executed_at"], run_data.get("created_at", "")),
        (REPORT_HEADER_LABELS["exported_at"], datetime.now(timezone.utc).isoformat()),
        (REPORT_HEADER_LABELS["case_count"], context["filtered_meta"]["case_count"]),
        (REPORT_HEADER_LABELS["event_count"], context["filtered_meta"]["event_count"]),
        (REPORT_HEADER_LABELS["applied_filters"], context.get("applied_filter_summary_text", "未適用")),
        {"label": "", "value": APPLIED_FILTERS_NOTE_TEXT, "style": "note"},
        ("分析期間", context["period_text"]),
        ("グルーピング条件", grouping_text),
        {"label": "", "value": GROUPING_CONDITION_NOTE_TEXT, "style": "note"},
    ]
    if variant_id is not None:
        summary_rows.append((REPORT_HEADER_LABELS["selected_variant"], f"#{variant_id}"))
    if analysis_key != "frequency":
        summary_rows.extend(
            [
                (REPORT_HEADER_LABELS["selected_activity"], selected_activity or "未選択"),
                (REPORT_HEADER_LABELS["selected_transition"], context["selected_transition_label"] or "未選択"),
                (REPORT_HEADER_LABELS["selected_case_id"], case_id or "未選択"),
            ]
        )
    return summary_rows

def _append_detail_export_summary_sheet(
    summary_sheet,
    context,
    run_data,
    analysis_key,
    filter_params,
    variant_id=None,
    selected_activity="",
    case_id="",
):
    summary_rows = _build_detail_export_summary_rows(
        context,
        run_data,
        analysis_key,
        variant_id=variant_id,
        selected_activity=selected_activity,
        case_id=case_id,
    )
    next_row = append_key_value_rows(
        summary_sheet,
        REPORT_SHEET_NAMES["summary"],
        summary_rows,
        description="対象範囲、選択条件、出力時点の情報をまとめています。",
    )
    kpi_rows = build_detail_summary_kpi_rows(
        analysis_key,
        context["selected_analysis"].get("rows", []),
        context["dashboard_summary"],
        context["impact_summary"],
        context["bottleneck_summary"],
        prepared_df=None,
        variant_items=context["export_variant_items"],
    )
    next_row = append_key_value_rows(
        summary_sheet,
        "主要KPI",
        kpi_rows,
        start_row=next_row,
        description=f"{context['analysis_name']} で優先して見たい代表値をまとめています。",
    )
    next_row = append_bullet_rows(
        summary_sheet,
        "分析ハイライト",
        context["ai_summary"].get("highlights", []),
        start_row=next_row,
        column_count=4,
    )

    if not context["group_columns"]:
        return

    group_summary = query_group_summary(
        run_data["prepared_parquet_path"],
        context["group_columns"],
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=context["variant_pattern"],
    )
    if not group_summary:
        return

    group_summary_df = build_summary_sheet_df(group_summary, context["group_columns"])
    append_table_to_worksheet(
        summary_sheet,
        "グループ別比較",
        group_summary_df.to_dict(orient="records"),
        list(group_summary_df.columns),
        start_row=next_row + 2,
        description="グルーピング条件ごとのケース数・処理時間の比較です。",
    )

def _append_detail_export_ai_sheet(workbook, context):
    ai_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["ai_insights"])
    ai_meta_rows = [
        ("対象分析", context["analysis_name"]),
        ("分析期間", context["ai_summary"].get("period", "不明")),
        ("出力時刻", context["ai_summary"].get("generated_at", "")),
    ]
    next_row = append_key_value_rows(
        ai_sheet,
        REPORT_SHEET_NAMES["ai_insights"],
        ai_meta_rows,
        description="現在の分析条件に対応する分析コメント、または既存集計からの要約を掲載します。",
    )
    next_row = append_custom_text_section_to_worksheet(
        ai_sheet,
        "分析前提",
        ANALYSIS_PRECONDITIONS_TEXT,
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_ASSUMPTION_SECTION_FILL,
    )
    next_row = append_text_block_to_worksheet(
        ai_sheet,
        "解説本文",
        context["ai_summary"].get("text", ""),
        start_row=next_row,
        column_count=6,
    )
    next_row = append_definition_table_to_worksheet(
        ai_sheet,
        "用語説明",
        TERMINOLOGY_ROWS,
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_MUTED_SECTION_FILL,
    )
    append_custom_text_section_to_worksheet(
        ai_sheet,
        "補足・免責事項",
        context["ai_summary"].get("note", ""),
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_MUTED_SECTION_FILL,
        body_fill=EXCEL_LABEL_FILL,
    )

def _append_frequency_export_sheet(workbook, context, run_data, filter_params):
    frequency_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["frequency"])
    group_columns = context["group_columns"]
    selected_analysis = context["selected_analysis"]
    variant_pattern = context["variant_pattern"]

    if not group_columns:
        frequency_rows = build_ranked_rows(selected_analysis["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
        frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
        append_table_to_worksheet(
            frequency_sheet,
            REPORT_SHEET_NAMES["frequency"],
            frequency_rows,
            frequency_headers,
            description="アクティビティごとの件数、ケース数、処理時間の代表値を確認できます。",
        )
        return

    overall_frequency_rows = query_analysis_records(
        run_data["prepared_parquet_path"],
        "frequency",
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    )["rows"]

    current_row = 1
    max_frequency_columns = max(10, len(overall_frequency_rows[0]) + 1 if overall_frequency_rows else 10)
    current_row = _write_section_header(
        frequency_sheet,
        current_row,
        "全体",
        column_count=max_frequency_columns,
    )
    current_row = _write_frequency_data(
        frequency_sheet,
        overall_frequency_rows,
        current_row,
    )

    for group_name, group_frequency_rows in _iter_groups_from_parquet(
        run_data["prepared_parquet_path"],
        group_columns,
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    ) or []:
        current_row += 3
        group_column_count = max(10, len(group_frequency_rows[0]) + 1 if group_frequency_rows else max_frequency_columns)
        current_row = _write_section_header(
            frequency_sheet,
            current_row,
            f"グループ: {group_name}",
            column_count=group_column_count,
        )
        current_row = _write_frequency_data(
            frequency_sheet,
            group_frequency_rows,
            current_row,
        )

def _append_transition_export_sheet(workbook, context):
    transition_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["transition"])
    transition_rows = build_ranked_rows(context["selected_analysis"]["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
    transition_headers = list(transition_rows[0].keys()) if transition_rows else [REPORT_HEADER_LABELS["rank"]]
    append_table_to_worksheet(
        transition_sheet,
        REPORT_SHEET_NAMES["transition"],
        transition_rows,
        transition_headers,
        description="前後遷移ごとの件数、ケース数、平均所要時間を確認できます。",
    )

def build_detail_export_workbook_bytes(
    run_data,
    analysis_key,
    context,
    filter_params,
    pattern_display_limit="10",
    variant_id=None,
    selected_activity="",
    case_id="",
    drilldown_limit=20,
):
    workbook = Workbook()
    summary_sheet = _create_report_sheet(
        workbook,
        REPORT_SHEET_NAMES["summary"],
        use_active=True,
    )
    _append_detail_export_summary_sheet(
        summary_sheet,
        context,
        run_data,
        analysis_key,
        filter_params=filter_params,
        variant_id=variant_id,
        selected_activity=selected_activity,
        case_id=case_id,
    )
    _append_detail_export_ai_sheet(workbook, context)
    _append_detail_export_analysis_sheets(
        workbook,
        context,
        run_data,
        filter_params,
        pattern_display_limit=pattern_display_limit,
    )
    _append_drilldown_export_sheet(
        workbook,
        context,
        run_data,
        filter_params,
        selected_activity=selected_activity,
        drilldown_limit=drilldown_limit,
    )
    _append_case_trace_export_sheet(workbook, run_data, case_id=case_id)
    return _serialize_workbook_bytes(workbook)

def _append_pattern_export_sheets(workbook, context, run_data, filter_params, pattern_display_limit="10"):
    analysis_definitions = context["analysis_definitions"]
    selected_analysis = context["selected_analysis"]
    export_variant_items = context["export_variant_items"]
    pattern_config = analysis_definitions.get("pattern", {}).get("config", {})
    pattern_display_columns = pattern_config.get("display_columns", {})
    pattern_column_label = pattern_display_columns.get("pattern", "処理順パターン")
    pattern_summary = build_pattern_export_summary(
        selected_analysis["rows"],
        pattern_display_columns,
    )
    pattern_conclusion = build_pattern_conclusion_summary(pattern_summary)
    pattern_dashboard = build_pattern_dashboard_summary(pattern_summary, pattern_conclusion)

    conclusion_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["pattern_conclusion"])
    conclusion_next_row = append_key_value_rows(
        conclusion_sheet,
        REPORT_SHEET_NAMES["pattern_conclusion"],
        [
            ("全体要約", pattern_conclusion["overall_summary"]),
            ("改善による時間インパクト(分)", pattern_conclusion["total_impact_minutes"]),
            ("改善による時間インパクト(時間)", pattern_conclusion["total_impact_hours"]),
            ("最短処理パターン", pattern_conclusion["fastest_pattern"].get("パターン", "該当なし")),
            ("最短処理パターン平均処理時間(分)", pattern_conclusion["fastest_pattern"].get("平均処理時間(分)", 0)),
        ],
        description="処理順パターン分析の結論、改善優先度、想定時間効果をまとめています。",
    )
    conclusion_next_row = append_table_to_worksheet(
        conclusion_sheet,
        "問題点3つ",
        pattern_conclusion["issue_rows"],
        ["問題点", "原因", "改善案", "期待効果（時間短縮・分）", "対象パターン"],
        start_row=conclusion_next_row,
        description="改善対象パターンTOP3を中心に、問題点・原因・改善案・期待効果を整理しています。",
        no_wrap_headers=["対象パターン"],
        min_column_widths={"問題点": 40, "原因": 38, "改善案": 42, "対象パターン": 72},
    )
    append_pattern_conclusion_charts(
        workbook,
        conclusion_sheet,
        pattern_summary["comparison_rows"],
        pie_anchor=build_excel_anchor("A", conclusion_next_row + 1),
        bar_anchor=build_excel_anchor("C", conclusion_next_row + 1),
    )

    dashboard_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["pattern_dashboard"])
    dashboard_next_row = append_key_value_rows(
        dashboard_sheet,
        REPORT_SHEET_NAMES["pattern_dashboard"],
        [
            ("全体要約", pattern_dashboard["overall_summary"]),
            ("上位10パターン累積カバー率(%)", pattern_dashboard["top10_coverage_pct"]),
            ("改善による時間インパクト(分)", pattern_dashboard["total_impact_minutes"]),
        ],
        description="処理順パターン分析の主要サマリーをダッシュボード形式でまとめています。",
    )
    dashboard_next_row = append_table_to_worksheet(
        dashboard_sheet,
        "改善優先TOP3",
        pattern_dashboard["top3_rows"],
        ["順位", "パターン", "改善優先度スコア", "全体影響度(%)", "繰り返し率(%)", "平均処理時間差分(分)", "簡易コメント"],
        start_row=dashboard_next_row,
        description="改善優先度スコアが高い上位3パターンです。",
        no_wrap_headers=["パターン", "簡易コメント"],
        min_column_widths={"パターン": 72, "簡易コメント": 36},
    )
    append_bullet_rows(
        dashboard_sheet,
        "問題点",
        pattern_dashboard["problem_points"],
        start_row=dashboard_next_row,
        column_count=6,
        empty_text="抽出できる問題点はありません。",
    )

    if "pattern_summary" in context["export_sheet_keys"]:
        pattern_summary_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["pattern_summary"])
        next_row = append_key_value_rows(
            pattern_summary_sheet,
            REPORT_SHEET_NAMES["pattern_summary"],
            [
                ("上位3パターン累積カバー率(%)", pattern_summary["top3_coverage_pct"]),
                ("上位10パターン累積カバー率(%)", pattern_summary["top10_coverage_pct"]),
                ("カバー率要約", pattern_summary["coverage_summary_text"]),
                ("最短処理パターン", pattern_summary["fastest_pattern"].get("パターン", "該当なし")),
                ("最短処理パターン平均処理時間(分)", pattern_summary["fastest_pattern"].get("平均処理時間(分)", 0)),
                ("要確認パターン数", len(pattern_summary["repeated_patterns"])),
                ("要確認パターン影響比率(%)", pattern_summary["repeated_case_ratio_pct"]),
                ("要確認判定基準", "繰り返し率 30%以上"),
                ("改善対象抽出基準", "繰り返し率 10%以上 かつ 平均処理時間差分がプラス"),
            ],
            description="処理順パターンの上位カバー率と、繰り返し率が高いパターンをまとめています。",
        )
        next_row = append_table_to_worksheet(
            pattern_summary_sheet,
            "上位10パターン",
            pattern_summary["comparison_rows"],
            ["順位", "繰り返し", "繰り返し回数", "繰り返し率(%)", "繰り返し率区分", "件数", "全体比率(%)", "平均処理時間(分)", "平均処理時間差分(分)", "改善優先度スコア", "全体影響度(%)", "最短処理", "最短処理時間(分)", "最長処理時間(分)", "確認区分", "簡易コメント", "パターン"],
            start_row=next_row,
            description="件数上位10パターンの比率・処理時間・繰り返し率を比較できます。",
            no_wrap_headers=["パターン", "簡易コメント"],
            min_column_widths={"パターン": 72, "簡易コメント": 48},
        )
        next_row = append_table_to_worksheet(
            pattern_summary_sheet,
            "要確認パターン一覧",
            pattern_summary["repeated_patterns"],
            ["順位", "繰り返し", "繰り返し回数", "繰り返し率(%)", "繰り返し率区分", "件数", "全体比率(%)", "平均処理時間(分)", "平均処理時間差分(分)", "改善優先度スコア", "全体影響度(%)", "最短処理", "最短処理時間(分)", "最長処理時間(分)", "確認区分", "簡易コメント", "パターン"],
            start_row=next_row,
            description="繰り返し率が高く、確認を優先したいパターンを一覧化しています。",
            no_wrap_headers=["パターン", "簡易コメント"],
            min_column_widths={"パターン": 72, "簡易コメント": 48},
        )
        append_table_to_worksheet(
            pattern_summary_sheet,
            "改善対象パターンTOP3",
            pattern_summary["improvement_targets"],
            ["順位", "繰り返し", "繰り返し回数", "繰り返し率(%)", "繰り返し率区分", "件数", "全体比率(%)", "平均処理時間(分)", "平均処理時間差分(分)", "改善優先度スコア", "全体影響度(%)", "最短処理", "確認区分", "簡易コメント", "パターン"],
            start_row=next_row,
            description="繰り返し率が一定以上で、平均処理時間も全体平均より長い改善候補パターンです。",
            no_wrap_headers=["パターン", "簡易コメント"],
            min_column_widths={"パターン": 72, "簡易コメント": 48},
        )

    pattern_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["pattern"])
    pattern_rows, pattern_headers = localize_report_rows(
        build_pattern_overview_rows(
            selected_analysis["rows"],
            export_variant_items,
            pattern_column_label,
            analysis_definitions,
        ),
        [
            "rank",
            "pattern_variant",
            "repeat_flag",
            "repeat_count",
            "repeat_rate_pct",
            "repeat_rate_band",
            "review_flag",
            "count",
            "ratio",
            "cumulative_case_ratio_pct",
            "avg_case_duration_min",
            "avg_case_duration_diff_min",
            "improvement_priority_score",
            "overall_impact_pct",
            "fastest_pattern_flag",
            "std_case_duration_min",
            "min_case_duration_min",
            "max_case_duration_min",
            "p75_case_duration_min",
            "p90_case_duration_min",
            "p95_case_duration_min",
            "simple_comment",
            "pattern",
        ],
    )
    append_table_to_worksheet(
        pattern_sheet,
        REPORT_SHEET_NAMES["pattern"],
        pattern_rows,
        pattern_headers,
        description="パターン / バリアントを 1 つの一覧にまとめ、繰り返し回数・繰り返し率・確認区分・件数・比率・累積カバー率・処理時間の代表値と代表ルートを比較できます。",
        no_wrap_headers=["パターン", "簡易コメント"],
        min_column_widths={"パターン": 80, "簡易コメント": 48},
    )

    variant_by_pattern = {
        str(variant_item.get("pattern") or "").strip(): variant_item
        for variant_item in export_variant_items
    }
    pattern_detail_count = resolve_pattern_detail_sheet_count(
        pattern_display_limit,
        len(selected_analysis.get("rows", [])),
    )
    for pattern_rank, pattern_row in enumerate(selected_analysis.get("rows", [])[:pattern_detail_count], start=1):
        pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
        pattern_detail = None
        if pattern_text:
            pattern_detail = query_pattern_bottleneck_details(
                run_data["prepared_parquet_path"],
                pattern_text,
                filter_params=filter_params,
                filter_column_settings=run_data.get("column_settings"),
                scope_variant_pattern=context["variant_pattern"],
            )
        append_pattern_detail_sheet(
            workbook,
            None,
            pattern_row,
            pattern_rank,
            pattern_column_label,
            analysis_definitions,
            variant_item=variant_by_pattern.get(str(pattern_row.get(pattern_column_label) or "").strip()),
            pattern_detail=pattern_detail,
        )

def _append_bottleneck_export_sheet(workbook, context):
    bottleneck_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["bottleneck"])
    activity_bottleneck_rows, activity_bottleneck_headers = localize_report_rows(
        build_bottleneck_export_rows(
            context["bottleneck_summary"]["activity_bottlenecks"],
            "activity",
        ),
        ["rank", "activity", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
    )
    next_row = append_table_to_worksheet(
        bottleneck_sheet,
        "アクティビティボトルネック",
        activity_bottleneck_rows,
        activity_bottleneck_headers,
        description="アクティビティ単位で所要時間が大きい箇所を並べています。",
    )
    transition_bottleneck_rows, transition_bottleneck_headers = localize_report_rows(
        build_bottleneck_export_rows(
            context["bottleneck_summary"]["transition_bottlenecks"],
            "transition_label",
        ),
        ["rank", "transition_label", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
    )
    append_table_to_worksheet(
        bottleneck_sheet,
        "遷移ボトルネック",
        transition_bottleneck_rows,
        transition_bottleneck_headers,
        start_row=next_row,
        description="前後遷移ごとの平均所要時間・中央値・最大値を比較できます。",
    )

def _append_impact_export_sheet(workbook, context):
    impact_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["impact"])
    impact_rows, impact_headers = localize_report_rows(
        [
            {
                "rank": impact_row["rank"],
                "transition": impact_row["transition_label"],
                "case_count": impact_row["case_count"],
                "avg_duration": impact_row["avg_duration_text"],
                "max_duration": impact_row["max_duration_text"],
                "impact_score": impact_row["impact_score"],
                "impact_share_pct": impact_row["impact_share_pct"],
            }
            for impact_row in context["impact_summary"]["rows"]
        ],
        ["rank", "transition", "case_count", "avg_duration", "max_duration", "impact_score", "impact_share_pct"],
    )
    append_table_to_worksheet(
        impact_sheet,
        REPORT_SHEET_NAMES["impact"],
        impact_rows,
        impact_headers,
        description="改善インパクトが高い遷移を優先順位付きで確認できます。",
    )

def _append_drilldown_export_sheet(
    workbook,
    context,
    run_data,
    filter_params,
    selected_activity="",
    drilldown_limit=20,
):
    selected_activity_name = str(selected_activity or "").strip()
    drilldown_rows = []
    drilldown_title = REPORT_SHEET_NAMES["drilldown"]

    if context["from_activity"] and context["to_activity"]:
        drilldown_title = f"遷移ドリルダウン: {context['from_activity']} → {context['to_activity']}"
        drilldown_rows = query_transition_case_drilldown(
            run_data["prepared_parquet_path"],
            from_activity=context["from_activity"],
            to_activity=context["to_activity"],
            limit=max(0, int(drilldown_limit)),
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=context["variant_pattern"],
        )
    elif selected_activity_name:
        drilldown_title = f"アクティビティドリルダウン: {selected_activity_name}"
        drilldown_rows = query_activity_case_drilldown(
            run_data["prepared_parquet_path"],
            activity=selected_activity_name,
            limit=max(0, int(drilldown_limit)),
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=context["variant_pattern"],
        )

    if not drilldown_rows:
        return

    drilldown_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["drilldown"])
    drilldown_rows, drilldown_headers = localize_report_rows(
        drilldown_rows,
        ["case_id", "activity", "next_activity", "duration_text", "from_time", "to_time"],
    )
    append_table_to_worksheet(
        drilldown_sheet,
        drilldown_title,
        drilldown_rows,
        drilldown_headers,
        description="選択中アクティビティ / 遷移に該当するケースの明細です。",
    )

def _append_case_trace_export_sheet(workbook, run_data, case_id=""):
    normalized_case_id = str(case_id or "").strip()
    if not normalized_case_id:
        return

    case_trace = query_case_trace_details(
        run_data["prepared_parquet_path"],
        normalized_case_id,
    )
    if not case_trace.get("found"):
        return

    case_trace_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["case_trace"])
    next_row = append_key_value_rows(
        case_trace_sheet,
        "ケース概要",
        [
            (REPORT_HEADER_LABELS["case_id"], case_trace["case_id"]),
            (REPORT_HEADER_LABELS["event_count"], case_trace["summary"]["event_count"]),
            (REPORT_HEADER_LABELS["total_duration"], case_trace["summary"]["total_duration_text"]),
            (REPORT_HEADER_LABELS["start_time"], case_trace["summary"]["start_time"]),
            (REPORT_HEADER_LABELS["end_time"], case_trace["summary"]["end_time"]),
        ],
        description="指定したケースの概要情報です。",
    )
    case_trace_event_rows, case_trace_event_headers = localize_report_rows(
        case_trace["events"],
        ["case_id", "activity", "next_activity", "start_time", "end_time", "duration_text"],
    )
    append_table_to_worksheet(
        case_trace_sheet,
        "通過イベント",
        case_trace_event_rows,
        case_trace_event_headers,
        start_row=next_row,
        description="ケース内で通過したイベントを時系列順に並べています。",
    )

def _append_detail_export_analysis_sheets(workbook, context, run_data, filter_params, pattern_display_limit="10"):
    export_sheet_keys = context["export_sheet_keys"]
    if "frequency" in export_sheet_keys:
        _append_frequency_export_sheet(workbook, context, run_data, filter_params)
    if "transition" in export_sheet_keys:
        _append_transition_export_sheet(workbook, context)
    if "pattern" in export_sheet_keys:
        _append_pattern_export_sheets(
            workbook,
            context,
            run_data,
            filter_params,
            pattern_display_limit=pattern_display_limit,
        )
    if "bottleneck" in export_sheet_keys:
        _append_bottleneck_export_sheet(workbook, context)
    if "impact" in export_sheet_keys:
        _append_impact_export_sheet(workbook, context)
