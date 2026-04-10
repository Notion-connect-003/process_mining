from io import BytesIO
from datetime import datetime, timezone
import duckdb

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from web_reports.excel_common import (
    ANALYSIS_PRECONDITIONS_TEXT,
    APPLIED_FILTERS_NOTE_TEXT,
    EXCEL_ALT_ROW_FILL,
    EXCEL_ASSUMPTION_SECTION_FILL,
    EXCEL_BODY_FONT,
    EXCEL_BOLD_FONT,
    EXCEL_GROUP_SECTION_FILL,
    EXCEL_GROUP_SECTION_FONT,
    EXCEL_HEADER_FILL,
    EXCEL_LABEL_FILL,
    EXCEL_MUTED_FONT,
    EXCEL_MUTED_SECTION_FILL,
    EXCEL_NOTE_FONT,
    EXCEL_SECTION_FILL,
    EXCEL_SUBTITLE_FILL,
    EXCEL_TEXT_BLOCK_FILL,
    EXCEL_THIN_BORDER,
    EXCEL_TITLE_BORDER,
    EXCEL_TITLE_FILL,
    EXCEL_TITLE_FONT,
    GROUPING_CONDITION_NOTE_TEXT,
    LOG_DIAGNOSTIC_SHEET_NAMES,
    REPORT_HEADER_LABELS,
    REPORT_SHEET_NAMES,
    TERMINOLOGY_ROWS,
    append_bullet_rows,
    append_custom_text_section_to_worksheet,
    append_definition_table_to_worksheet,
    append_key_value_rows,
    append_table_to_worksheet,
    append_text_block_to_worksheet,
    autosize_worksheet_columns,
    build_analysis_excel_file_name,
    estimate_wrapped_row_height,
    initialize_excel_worksheet,
    merge_excel_row,
    normalize_excel_cell_value,
    resolve_analysis_display_name,
    sanitize_workbook_sheet_name,
    style_excel_cell,
)

from web_reports.detail_report_helpers import *
from web_reports.detail_report_helpers import (
    _build_scoped_relation_cte,
    _format_stddev_column,
    _get_parquet_column_names,
    _quote_identifier,
)
from web_reports.detail_report_pattern_sheets import _append_pattern_export_sheets


def _write_section_header(worksheet, row_index, title, column_count=10):
    safe_column_count = max(10, int(column_count or 10))
    for column_index in range(1, safe_column_count + 1):
        header_cell = worksheet.cell(row=row_index, column=column_index)
        if column_index == 1:
            header_cell.value = f"═══ {title} ═══"
            font = EXCEL_GROUP_SECTION_FONT
            alignment = Alignment(horizontal="left", vertical="center")
        else:
            font = EXCEL_GROUP_SECTION_FONT
            alignment = Alignment(horizontal="left", vertical="center")
        style_excel_cell(
            header_cell,
            font=font,
            fill=EXCEL_GROUP_SECTION_FILL,
            alignment=alignment,
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[row_index].height = 22
    return row_index + 1

def _iter_groups_from_parquet(
    parquet_path,
    grouping_columns,
    filter_params=None,
    filter_column_settings=None,
    variant_pattern=None,
):
    valid_columns = [
        str(column_name)
        for column_name in (grouping_columns or [])
        if str(column_name) in set(_get_parquet_column_names(parquet_path))
    ]
    if not valid_columns:
        return

    connection = duckdb.connect()
    try:
        cte_sql, params, relation_name = _build_scoped_relation_cte(
            parquet_path,
            filter_params=filter_params,
            filter_column_settings=filter_column_settings,
            variant_pattern=variant_pattern,
        )
        case_group_columns_sql = ",\n                ".join(
            [
                f"COALESCE(MAX(NULLIF(TRIM(CAST({_quote_identifier(column_name)} AS VARCHAR)), '')), '(未分類)') AS {_quote_identifier(column_name)}"
                for column_name in valid_columns
            ]
        )
        group_select_sql = ", ".join(_quote_identifier(column_name) for column_name in valid_columns)
        group_order_sql = ", ".join(_quote_identifier(column_name) for column_name in valid_columns)
        group_list_rows = connection.execute(
            f"""
            {cte_sql},
            case_groups AS (
                SELECT
                    case_id,
                    {case_group_columns_sql}
                FROM {relation_name}
                GROUP BY case_id
            )
            SELECT
                {group_select_sql},
                COUNT(*) AS case_count
            FROM case_groups
            GROUP BY {group_select_sql}
            ORDER BY case_count DESC, {group_order_sql}
            """,
            params,
        ).fetchall()
        if not group_list_rows:
            return

        frequency_display_columns = (
            get_available_analysis_definitions()
            .get("frequency", {})
            .get("config", {})
            .get("display_columns", {})
        )

        for row in group_list_rows:
            group_values = [str(value or "").strip() or "(未分類)" for value in row[: len(valid_columns)]]
            if len(valid_columns) == 1:
                group_name = group_values[0]
            else:
                group_name = ", ".join(
                    f"{column_name}={group_value}"
                    for column_name, group_value in zip(valid_columns, group_values)
                )

            group_conditions_sql = " AND ".join(
                f"{_quote_identifier(column_name)} = ?"
                for column_name in valid_columns
            )
            group_df = connection.execute(
                f"""
                {cte_sql},
                case_groups AS (
                    SELECT
                        case_id,
                        {case_group_columns_sql}
                    FROM {relation_name}
                    GROUP BY case_id
                ),
                selected_cases AS (
                    SELECT case_id
                    FROM case_groups
                    WHERE {group_conditions_sql}
                ),
                group_scoped AS (
                    SELECT scoped.*
                    FROM {relation_name} AS scoped
                    INNER JOIN selected_cases USING (case_id)
                )
                SELECT
                    activity,
                    COUNT(*) AS event_count,
                    COUNT(DISTINCT case_id) AS case_count,
                    ROUND(SUM(duration_min), 2) AS total_duration_min,
                    ROUND(AVG(duration_min), 2) AS avg_duration_min,
                    ROUND(MEDIAN(duration_min), 2) AS median_duration_min,
                    CASE WHEN COUNT(*) > 1 THEN ROUND(STDDEV_SAMP(duration_min), 2) ELSE NULL END AS std_duration_min,
                    ROUND(MIN(duration_min), 2) AS min_duration_min,
                    ROUND(MAX(duration_min), 2) AS max_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.75), 2) AS p75_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.90), 2) AS p90_duration_min,
                    ROUND(QUANTILE_CONT(duration_min, 0.95), 2) AS p95_duration_min,
                    ROUND(
                        COUNT(*) * 100.0 / NULLIF((SELECT COUNT(*) FROM group_scoped), 0),
                        2
                    ) AS event_ratio_pct
                FROM group_scoped
                GROUP BY activity
                ORDER BY event_count DESC, activity ASC
                """,
                params + group_values,
            ).df()
            if "std_duration_min" in group_df.columns:
                group_df["std_duration_min"] = _format_stddev_column(group_df["std_duration_min"])
            yield group_name, convert_analysis_result_to_records(
                group_df,
                frequency_display_columns,
            )
    finally:
        connection.close()

def _write_frequency_data(worksheet, rows, start_row):
    frequency_rows = build_ranked_rows(rows or [], rank_key=REPORT_HEADER_LABELS["rank"])
    frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
    header_row = start_row
    for column_index, header in enumerate(frequency_headers, start=1):
        header_cell = worksheet.cell(row=header_row, column=column_index, value=header)
        style_excel_cell(
            header_cell,
            font=EXCEL_BOLD_FONT,
            fill=EXCEL_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=EXCEL_THIN_BORDER,
        )
    worksheet.row_dimensions[header_row].height = 22

    if not frequency_rows:
        merge_excel_row(worksheet, header_row + 1, len(frequency_headers))
        empty_cell = worksheet.cell(row=header_row + 1, column=1, value="表示できるデータがありません。")
        style_excel_cell(
            empty_cell,
            font=EXCEL_MUTED_FONT,
            fill=EXCEL_LABEL_FILL,
            alignment=Alignment(wrap_text=True, vertical="center"),
            border=EXCEL_THIN_BORDER,
        )
        worksheet.row_dimensions[header_row + 1].height = 22
        return header_row + 3

    data_start_row = header_row + 1
    current_row = data_start_row
    for row_index, row in enumerate(frequency_rows, start=0):
        fill = EXCEL_ALT_ROW_FILL if row_index % 2 else None
        for column_index, header in enumerate(frequency_headers, start=1):
            body_cell = worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
            style_excel_cell(
                body_cell,
                font=EXCEL_BODY_FONT,
                fill=fill,
                alignment=Alignment(wrap_text=True, vertical="top"),
                border=EXCEL_THIN_BORDER,
            )
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

    if not worksheet.freeze_panes:
        worksheet.freeze_panes = f"A{data_start_row}"
    if not worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(frequency_headers))}{current_row - 1}"

    return current_row + 1

def _set_chart_str_categories(chart, labels_ref):
    """Set chart categories as strRef so Excel treats text labels correctly."""
    label_formula = str(labels_ref)
    for series in chart.ser:
        series.cat = AxDataSource(strRef=StrRef(f=label_formula))

def _ensure_chart_data_sheet(workbook):
    sheet_name = "_chart_data"
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    data_sheet = workbook.create_sheet(title=sheet_name)
    data_sheet.sheet_state = "hidden"
    return data_sheet

def _write_chart_data_block(workbook, block_label, comparison_rows, columns):
    data_sheet = _ensure_chart_data_sheet(workbook)
    start_row = (data_sheet.max_row or 0) + 2

    data_sheet.cell(row=start_row, column=1, value=block_label)
    header_row = start_row + 1
    for col_offset, col_name in enumerate(columns):
        data_sheet.cell(row=header_row, column=1 + col_offset, value=col_name)

    for index, row in enumerate(comparison_rows, start=1):
        data_row = header_row + index
        data_sheet.cell(row=data_row, column=1, value=f"Pattern #{row.get('順位', index)}")
        for col_offset, col_name in enumerate(columns):
            if col_offset == 0:
                continue
            data_sheet.cell(
                row=data_row, column=1 + col_offset,
                value=coerce_report_number(row.get(col_name), 0.0),
            )

    return data_sheet, header_row, header_row + len(comparison_rows)

def build_excel_anchor(column_letter, row_number):
    return f"{column_letter}{max(1, int(row_number or 1))}"

def sort_pattern_rows_by_avg_duration_desc(rows):
    return sorted(
        rows,
        key=lambda row: (
            -coerce_report_number(row.get("平均処理時間(分)"), 0.0),
            row.get("順位", 0),
        ),
    )

def append_pattern_dashboard_pie_chart(workbook, worksheet, comparison_rows, anchor="A1"):
    if not comparison_rows:
        return

    data_sheet, header_row, max_row = _write_chart_data_block(
        workbook, "dashboard_pie", comparison_rows, ["パターン", "件数"],
    )

    pie_chart = PieChart()
    pie_chart.title = "上位10パターンの割合"
    pie_chart.style = 10
    pie_chart.height = 14.0
    pie_chart.width = 16.0
    pie_data = Reference(data_sheet, min_col=2, min_row=header_row, max_row=max_row)
    pie_labels = Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    _set_chart_str_categories(pie_chart, pie_labels)
    pie_chart.dLbls = DataLabelList(
        showCatName=False,
        showVal=False,
        showPercent=True,
        showSerName=False,
        showLegendKey=False,
        showLeaderLines=True,
        dLblPos="bestFit",
        separator="\n",
    )
    worksheet.add_chart(pie_chart, anchor)

def append_pattern_conclusion_charts(workbook, worksheet, comparison_rows, pie_anchor="A1", bar_anchor="A20"):
    if not comparison_rows:
        return

    data_sheet, header_row, max_row = _write_chart_data_block(
        workbook, "conclusion_charts", comparison_rows,
        ["パターン", "件数"],
    )
    bar_rows = sort_pattern_rows_by_avg_duration_desc(comparison_rows)
    bar_data_sheet, bar_header_row, bar_max_row = _write_chart_data_block(
        workbook,
        "conclusion_duration_desc",
        bar_rows,
        ["パターン", "平均処理時間(分)"],
    )

    pie_chart = PieChart()
    pie_chart.title = "上位10パターンの割合"
    pie_chart.style = 10
    pie_chart.height = 14.0
    pie_chart.width = 16.0
    pie_data = Reference(data_sheet, min_col=2, min_row=header_row, max_row=max_row)
    pie_labels = Reference(data_sheet, min_col=1, min_row=header_row + 1, max_row=max_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    _set_chart_str_categories(pie_chart, pie_labels)
    pie_chart.dLbls = DataLabelList(
        showCatName=False,
        showVal=False,
        showPercent=True,
        showSerName=False,
        showLegendKey=False,
        showLeaderLines=True,
        dLblPos="bestFit",
        separator="\n",
    )
    worksheet.add_chart(pie_chart, pie_anchor)

    bar_chart = BarChart()
    bar_chart.type = "bar"
    bar_chart.style = 10
    bar_chart.title = "平均処理時間の比較（長い順）"
    bar_chart.height = 14.0
    bar_chart.width = 16.0
    bar_chart.legend = None
    bar_chart.varyColors = False
    bar_chart.gapWidth = 60
    bar_chart.x_axis.scaling.orientation = "maxMin"
    bar_chart.x_axis.tickLblPos = "low"
    bar_chart.y_axis.crosses = "autoZero"
    bar_chart.y_axis.delete = False
    bar_data = Reference(bar_data_sheet, min_col=2, min_row=bar_header_row, max_row=bar_max_row)
    bar_labels = Reference(bar_data_sheet, min_col=1, min_row=bar_header_row + 1, max_row=bar_max_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    _set_chart_str_categories(bar_chart, bar_labels)
    bar_chart.dLbls = DataLabelList(
        showVal=True,
        showSerName=False,
        showCatName=False,
        dLblPos="outEnd",
    )
    worksheet.add_chart(bar_chart, bar_anchor)

def append_pattern_detail_sheet(
    workbook,
    filtered_df,
    pattern_row,
    pattern_rank,
    pattern_column_label,
    analysis_definitions,
    variant_item=None,
    pattern_detail=None,
):
    pattern_text = str(pattern_row.get(pattern_column_label) or "").strip()
    if not pattern_text:
        return

    detail = pattern_detail or create_pattern_bottleneck_details(filtered_df, pattern_text)
    sheet_name = sanitize_workbook_sheet_name(f"パターン{pattern_rank:02d}詳細")
    detail_sheet = workbook.create_sheet(title=sheet_name)
    initialize_excel_worksheet(detail_sheet)

    variant_label = (
        f"Variant #{variant_item['variant_id']}"
        if variant_item and variant_item.get("variant_id")
        else "該当なし"
    )
    bottleneck_transition = detail.get("bottleneck_transition") or {}
    next_row = append_key_value_rows(
        detail_sheet,
        f"Pattern #{pattern_rank} 詳細",
        [
            ("パターン / バリアント", f"Pattern #{pattern_rank} / {variant_label}"),
            ("繰り返し", pattern_row.get("繰り返し", "")),
            ("繰り返し回数", pattern_row.get("繰り返し回数", 0)),
            ("繰り返し率(%)", pattern_row.get("繰り返し率(%)", 0)),
            ("繰り返し率区分", pattern_row.get("繰り返し率区分", "")),
            ("確認区分", pattern_row.get("確認区分", "")),
            ("平均処理時間差分(分)", pattern_row.get("平均処理時間差分(分)", 0)),
            ("改善優先度スコア", pattern_row.get("改善優先度スコア", 0)),
            ("全体影響度(%)", pattern_row.get("全体影響度(%)", 0)),
            ("最短処理", pattern_row.get("最短処理", "")),
            ("簡易コメント", pattern_row.get("簡易コメント", "")),
            ("ケース数", detail.get("case_count", 0)),
            ("ケース比率(%)", detail.get("case_ratio_pct", 0)),
            ("平均ケース処理時間(分)", detail.get("avg_case_duration_min", 0)),
            ("中央値ケース処理時間(分)", detail.get("median_case_duration_min", 0)),
            ("最小ケース処理時間(分)", detail.get("min_case_duration_min", 0)),
            ("最大ケース処理時間(分)", detail.get("max_case_duration_min", 0)),
            ("代表ルート", pattern_text),
            ("ボトルネック遷移", bottleneck_transition.get("transition_label", "該当なし")),
            ("ボトルネック平均所要時間(分)", bottleneck_transition.get("avg_duration_min", 0)),
        ],
        description="上位パターンのケース概要、代表ルート、ボトルネック遷移をまとめています。",
    )

    step_metric_rows, step_metric_headers = localize_report_rows(
        [
            {
                "rank": index,
                "sequence_no": step_metric["sequence_no"],
                "transition": step_metric["transition_label"],
                "case_count": step_metric["case_count"],
                "avg_duration_min": step_metric["avg_duration_min"],
                "median_duration_min": step_metric["median_duration_min"],
                "min_duration_min": step_metric["min_duration_min"],
                "max_duration_min": step_metric["max_duration_min"],
                "total_duration_min": step_metric["total_duration_min"],
                "wait_share_pct": step_metric["wait_share_pct"],
            }
            for index, step_metric in enumerate(detail.get("step_metrics", []), start=1)
        ],
        [
            "rank",
            "sequence_no",
            "transition",
            "case_count",
            "avg_duration_min",
            "median_duration_min",
            "min_duration_min",
            "max_duration_min",
            "total_duration_min",
            "wait_share_pct",
        ],
    )
    next_row = append_table_to_worksheet(
        detail_sheet,
        "ステップ別所要時間",
        step_metric_rows,
        step_metric_headers,
        start_row=next_row,
        description="各ステップの所要時間と全体に占める比率を比較できます。",
        no_wrap_headers=["遷移"],
        min_column_widths={"遷移": 32},
    )

    case_example_rows, case_example_headers = localize_report_rows(
        [
            {
                "rank": index,
                "case_id": case_example["case_id"],
                "start_time": case_example["start_time"],
                "end_time": case_example["end_time"],
                "total_duration_min": case_example["case_total_duration_min"],
            }
            for index, case_example in enumerate(detail.get("case_examples", []), start=1)
        ],
        ["rank", "case_id", "start_time", "end_time", "total_duration_min"],
    )
    append_table_to_worksheet(
        detail_sheet,
        "代表ケース",
        case_example_rows,
        case_example_headers,
        start_row=next_row,
        description="このパターンに属するケースのうち、総処理時間が長いケースを上位から掲載しています。",
    )

def build_detail_summary_kpi_rows(
    analysis_key,
    analysis_rows,
    dashboard_summary,
    impact_summary,
    bottleneck_summary,
    prepared_df=None,
    variant_items=None,
):
    normalized_analysis_key = str(analysis_key or "").strip().lower()
    top_row = analysis_rows[0] if analysis_rows else {}
    top_transition_bottleneck = (
        bottleneck_summary["transition_bottlenecks"][0]
        if bottleneck_summary.get("transition_bottlenecks")
        else None
    )
    top_transition_bottleneck_label = build_transition_display_label(top_transition_bottleneck) or "該当なし"
    top_impact_row = impact_summary["rows"][0] if impact_summary.get("rows") else None

    common_rows = [
        ("平均ケース処理時間", dashboard_summary.get("avg_case_duration_text", "0s")),
        ("中央値ケース処理時間", dashboard_summary.get("median_case_duration_text", "0s")),
    ]

    if normalized_analysis_key == "frequency":
        frequency_rows = common_rows + [
            ("最大ケース処理時間", dashboard_summary.get("max_case_duration_text", "0s")),
            ("最多アクティビティ", top_row.get("アクティビティ", "該当なし") if top_row else "該当なし"),
            ("最多アクティビティ件数", normalize_excel_cell_value(top_row.get("イベント件数", 0)) if top_row else 0),
            ("上位10バリアントカバー率", f"{float(dashboard_summary.get('top10_variant_coverage_pct', 0.0)):.2f}%"),
        ]
        if variant_items is not None:
            frequency_rows.append(("バリアント総数", len(variant_items)))

        unique_activity_count = dashboard_summary.get("unique_activity_count")
        if unique_activity_count in (None, ""):
            unique_activity_count = dashboard_summary.get("activity_type_count")
        if unique_activity_count in (None, "") and prepared_df is not None and "activity" in prepared_df.columns:
            unique_activity_count = int(prepared_df["activity"].nunique())
        if unique_activity_count not in (None, ""):
            frequency_rows.append(("ユニークアクティビティ数", int(unique_activity_count)))

        total_cases = int(dashboard_summary.get("total_cases", 0) or 0)
        total_records = int(dashboard_summary.get("total_records", 0) or 0)
        if total_cases > 0:
            frequency_rows.append(("平均ケースあたりイベント数", round(total_records / total_cases, 2)))

        return frequency_rows

    if normalized_analysis_key == "transition":
        top_transition_label = (
            f"{top_row.get('前処理アクティビティ名', '')} → {top_row.get('後処理アクティビティ名', '')}".strip(" →")
            if top_row
            else "該当なし"
        )
        return common_rows + [
            ("主要遷移", top_transition_label or "該当なし"),
            ("最大ボトルネック遷移", top_transition_bottleneck_label),
            ("最大改善インパクト遷移", top_impact_row.get("transition_label", "該当なし") if top_impact_row else "該当なし"),
        ]

    if normalized_analysis_key == "pattern":
        return common_rows + [
            ("最頻出パターン", top_row.get("処理順パターン", top_row.get("パターン", "該当なし")) if top_row else "該当なし"),
            ("最頻出パターン比率", f"{float(top_row.get('ケース比率(%)', 0.0)):.2f}%" if top_row else "0.00%"),
            ("上位10バリアントカバー率", f"{float(dashboard_summary.get('top10_variant_coverage_pct', 0.0)):.2f}%"),
        ]

    return common_rows + [
        ("最大ケース処理時間", dashboard_summary.get("max_case_duration_text", "0s")),
        ("最大ボトルネック遷移", top_transition_bottleneck_label),
    ]


