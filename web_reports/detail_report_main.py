from io import BytesIO
from datetime import datetime, timezone
import duckdb

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from web_reports.excel_common import (
    ANALYSIS_PRECONDITIONS_TEXT,
    APPLIED_FILTERS_NOTE_TEXT,
    EXCEL_ASSUMPTION_SECTION_FILL,
    EXCEL_LABEL_FILL,
    EXCEL_MUTED_SECTION_FILL,
    GROUPING_CONDITION_NOTE_TEXT,
    REPORT_HEADER_LABELS,
    REPORT_SHEET_NAMES,
    append_bullet_rows,
    append_custom_text_section_to_worksheet,
    append_definition_table_to_worksheet,
    append_key_value_rows,
    append_structured_text_block_to_worksheet,
    append_table_to_worksheet,
    autosize_worksheet_columns,
    initialize_excel_worksheet,
    get_terminology_rows,
    merge_excel_row,
    sanitize_workbook_sheet_name,
)

from web_reports.detail_report_helpers import *
from web_reports.detail_report_sections import *
from web_reports.detail_report_sections import (
    _append_pattern_export_sheets,
    _iter_groups_from_parquet,
    _write_frequency_data,
    _write_section_header,
)

def _create_report_sheet(workbook, title, *, use_active=False):
    worksheet = workbook.active if use_active else workbook.create_sheet()
    worksheet.title = sanitize_workbook_sheet_name(title)
    initialize_excel_worksheet(worksheet)
    return worksheet

def _serialize_workbook_bytes(workbook):
    for worksheet in workbook.worksheets:
        autosize_worksheet_columns(worksheet)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()

def _build_detail_export_summary_rows(
    context,
    run_data,
    analysis_key,
    variant_id=None,
    selected_activity="",
    case_id="",
):
    group_columns = context["group_columns"]
    grouping_text = "、".join(group_columns) if group_columns else "なし"
    summary_rows = [
        (REPORT_HEADER_LABELS["analysis_key"], analysis_key),
        (REPORT_HEADER_LABELS["analysis_name"], context["analysis_name"]),
        (REPORT_HEADER_LABELS["source_file_name"], run_data["source_file_name"]),
        (REPORT_HEADER_LABELS["analysis_executed_at"], run_data.get("created_at", "")),
        (REPORT_HEADER_LABELS["exported_at"], datetime.now(timezone.utc).isoformat()),
        (REPORT_HEADER_LABELS["case_count"], context["filtered_meta"]["case_count"]),
        (REPORT_HEADER_LABELS["event_count"], context["filtered_meta"]["event_count"]),
        (REPORT_HEADER_LABELS["applied_filters"], context.get("applied_filter_summary_text", "未適用")),
        {"label": "", "value": APPLIED_FILTERS_NOTE_TEXT, "style": "note"},
        ("分析期間", context["period_text"]),
        ("グルーピング条件", grouping_text),
        {"label": "", "value": GROUPING_CONDITION_NOTE_TEXT, "style": "note"},
    ]
    if variant_id is not None:
        summary_rows.append((REPORT_HEADER_LABELS["selected_variant"], f"#{variant_id}"))
    if analysis_key != "frequency":
        summary_rows.extend(
            [
                (REPORT_HEADER_LABELS["selected_activity"], selected_activity or "未選択"),
                (REPORT_HEADER_LABELS["selected_transition"], context["selected_transition_label"] or "未選択"),
                (REPORT_HEADER_LABELS["selected_case_id"], case_id or "未選択"),
            ]
        )
    return summary_rows

def _append_detail_export_summary_sheet(
    summary_sheet,
    context,
    run_data,
    analysis_key,
    filter_params,
    variant_id=None,
    selected_activity="",
    case_id="",
):
    summary_rows = _build_detail_export_summary_rows(
        context,
        run_data,
        analysis_key,
        variant_id=variant_id,
        selected_activity=selected_activity,
        case_id=case_id,
    )
    next_row = append_key_value_rows(
        summary_sheet,
        REPORT_SHEET_NAMES["summary"],
        summary_rows,
        description="対象範囲、選択条件、出力時点の情報をまとめています。",
    )
    kpi_rows = build_detail_summary_kpi_rows(
        analysis_key,
        context["selected_analysis"].get("rows", []),
        context["dashboard_summary"],
        context["impact_summary"],
        context["bottleneck_summary"],
        prepared_df=None,
        variant_items=context["export_variant_items"],
    )
    next_row = append_key_value_rows(
        summary_sheet,
        "主要KPI",
        kpi_rows,
        start_row=next_row,
        description=f"{context['analysis_name']} で優先して見たい代表値をまとめています。",
    )
    next_row = append_bullet_rows(
        summary_sheet,
        "分析ハイライト",
        context["ai_summary"].get("highlights", []),
        start_row=next_row,
        column_count=4,
    )

    if not context["group_columns"]:
        return

    group_summary = query_group_summary(
        run_data["prepared_parquet_path"],
        context["group_columns"],
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=context["variant_pattern"],
    )
    if not group_summary:
        return

    group_summary_df = build_summary_sheet_df(group_summary, context["group_columns"])
    append_table_to_worksheet(
        summary_sheet,
        "グループ別比較",
        group_summary_df.to_dict(orient="records"),
        list(group_summary_df.columns),
        start_row=next_row + 2,
        description="グルーピング条件ごとのケース数・処理時間の比較です。",
    )

def _insert_spacer_row(worksheet, row_index, column_count=6, height=8):
    """セクション間に空のスペーサー行を挿入する。"""
    merge_excel_row(worksheet, row_index, column_count)
    worksheet.row_dimensions[row_index].height = height
    return row_index + 1


def _append_detail_export_ai_sheet(workbook, context, analysis_key=""):
    ai_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["ai_insights"])
    ai_meta_rows = [
        ("対象分析", context["analysis_name"]),
        ("分析期間", context["ai_summary"].get("period", "不明")),
        ("出力時刻", context["ai_summary"].get("generated_at", "")),
    ]
    next_row = append_key_value_rows(
        ai_sheet,
        REPORT_SHEET_NAMES["ai_insights"],
        ai_meta_rows,
        description="現在の分析条件に対応する分析コメント、または既存集計からの要約を掲載します。",
    )
    next_row = _insert_spacer_row(ai_sheet, next_row)
    next_row = append_custom_text_section_to_worksheet(
        ai_sheet,
        "分析前提",
        ANALYSIS_PRECONDITIONS_TEXT,
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_ASSUMPTION_SECTION_FILL,
    )
    next_row = _insert_spacer_row(ai_sheet, next_row)
    next_row = append_structured_text_block_to_worksheet(
        ai_sheet,
        "解説本文",
        context["ai_summary"].get("text", ""),
        start_row=next_row,
        column_count=6,
    )
    next_row = _insert_spacer_row(ai_sheet, next_row)
    next_row = append_bullet_rows(
        ai_sheet,
        "推奨アクション",
        context["ai_summary"].get("recommended_actions", []),
        start_row=next_row,
        column_count=6,
        empty_text="推奨アクションはありません。",
        section_header_fill=EXCEL_ASSUMPTION_SECTION_FILL,
    )
    next_row = _insert_spacer_row(ai_sheet, next_row)
    next_row = append_definition_table_to_worksheet(
        ai_sheet,
        "用語説明",
        get_terminology_rows(analysis_key),
        start_row=next_row,
        column_count=6,
        header_fill=EXCEL_MUTED_SECTION_FILL,
    )
    ai_sheet._codex_min_column_widths = {
        "A": 18,
        "B": 50,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
    }

def _append_frequency_export_sheet(workbook, context, run_data, filter_params):
    frequency_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["frequency"])
    group_columns = context["group_columns"]
    selected_analysis = context["selected_analysis"]
    variant_pattern = context["variant_pattern"]

    if not group_columns:
        frequency_rows = build_ranked_rows(selected_analysis["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
        frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
        append_table_to_worksheet(
            frequency_sheet,
            REPORT_SHEET_NAMES["frequency"],
            frequency_rows,
            frequency_headers,
            description="アクティビティごとの件数、ケース数、処理時間の代表値を確認できます。",
        )
        return

    overall_frequency_rows = query_analysis_records(
        run_data["prepared_parquet_path"],
        "frequency",
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    )["rows"]

    current_row = 1
    max_frequency_columns = max(10, len(overall_frequency_rows[0]) + 1 if overall_frequency_rows else 10)
    current_row = _write_section_header(
        frequency_sheet,
        current_row,
        "全体",
        column_count=max_frequency_columns,
    )
    current_row = _write_frequency_data(
        frequency_sheet,
        overall_frequency_rows,
        current_row,
    )

    for group_name, group_frequency_rows in _iter_groups_from_parquet(
        run_data["prepared_parquet_path"],
        group_columns,
        filter_params=filter_params,
        filter_column_settings=run_data.get("column_settings"),
        variant_pattern=variant_pattern,
    ) or []:
        current_row += 3
        group_column_count = max(10, len(group_frequency_rows[0]) + 1 if group_frequency_rows else max_frequency_columns)
        current_row = _write_section_header(
            frequency_sheet,
            current_row,
            f"グループ: {group_name}",
            column_count=group_column_count,
        )
        current_row = _write_frequency_data(
            frequency_sheet,
            group_frequency_rows,
            current_row,
        )

def _append_transition_export_sheet(workbook, context):
    transition_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["transition"])
    transition_rows = build_ranked_rows(context["selected_analysis"]["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
    transition_headers = list(transition_rows[0].keys()) if transition_rows else [REPORT_HEADER_LABELS["rank"]]
    append_table_to_worksheet(
        transition_sheet,
        REPORT_SHEET_NAMES["transition"],
        transition_rows,
        transition_headers,
        description="前後遷移ごとの件数、ケース数、平均所要時間を確認できます。",
    )

def build_detail_export_workbook_bytes(
    run_data,
    analysis_key,
    context,
    filter_params,
    pattern_display_limit="10",
    variant_id=None,
    selected_activity="",
    case_id="",
    drilldown_limit=20,
):
    workbook = Workbook()
    summary_sheet = _create_report_sheet(
        workbook,
        REPORT_SHEET_NAMES["summary"],
        use_active=True,
    )
    _append_detail_export_summary_sheet(
        summary_sheet,
        context,
        run_data,
        analysis_key,
        filter_params=filter_params,
        variant_id=variant_id,
        selected_activity=selected_activity,
        case_id=case_id,
    )
    _append_detail_export_ai_sheet(workbook, context, analysis_key=analysis_key)
    _append_detail_export_analysis_sheets(
        workbook,
        context,
        run_data,
        filter_params,
        pattern_display_limit=pattern_display_limit,
    )
    _append_drilldown_export_sheet(
        workbook,
        context,
        run_data,
        filter_params,
        selected_activity=selected_activity,
        drilldown_limit=drilldown_limit,
    )
    _append_case_trace_export_sheet(workbook, run_data, case_id=case_id)
    return _serialize_workbook_bytes(workbook)

def _append_bottleneck_export_sheet(workbook, context):
    bottleneck_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["bottleneck"])
    activity_bottleneck_rows, activity_bottleneck_headers = localize_report_rows(
        build_bottleneck_export_rows(
            context["bottleneck_summary"]["activity_bottlenecks"],
            "activity",
        ),
        ["rank", "activity", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
    )
    next_row = append_table_to_worksheet(
        bottleneck_sheet,
        "アクティビティボトルネック",
        activity_bottleneck_rows,
        activity_bottleneck_headers,
        description="アクティビティ単位で所要時間が大きい箇所を並べています。",
    )
    transition_bottleneck_rows, transition_bottleneck_headers = localize_report_rows(
        build_bottleneck_export_rows(
            context["bottleneck_summary"]["transition_bottlenecks"],
            "transition_label",
        ),
        ["rank", "transition_label", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
    )
    append_table_to_worksheet(
        bottleneck_sheet,
        "遷移ボトルネック",
        transition_bottleneck_rows,
        transition_bottleneck_headers,
        start_row=next_row,
        description="前後遷移ごとの平均所要時間・中央値・最大値を比較できます。",
    )

def _append_impact_export_sheet(workbook, context):
    impact_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["impact"])
    impact_rows, impact_headers = localize_report_rows(
        [
            {
                "rank": impact_row["rank"],
                "transition": impact_row["transition_label"],
                "case_count": impact_row["case_count"],
                "avg_duration": impact_row["avg_duration_text"],
                "max_duration": impact_row["max_duration_text"],
                "impact_score": impact_row["impact_score"],
                "impact_share_pct": impact_row["impact_share_pct"],
            }
            for impact_row in context["impact_summary"]["rows"]
        ],
        ["rank", "transition", "case_count", "avg_duration", "max_duration", "impact_score", "impact_share_pct"],
    )
    append_table_to_worksheet(
        impact_sheet,
        REPORT_SHEET_NAMES["impact"],
        impact_rows,
        impact_headers,
        description="改善インパクトが高い遷移を優先順位付きで確認できます。",
    )

def _append_drilldown_export_sheet(
    workbook,
    context,
    run_data,
    filter_params,
    selected_activity="",
    drilldown_limit=20,
):
    selected_activity_name = str(selected_activity or "").strip()
    drilldown_rows = []
    drilldown_title = REPORT_SHEET_NAMES["drilldown"]

    if context["from_activity"] and context["to_activity"]:
        drilldown_title = f"遷移ドリルダウン: {context['from_activity']} → {context['to_activity']}"
        drilldown_rows = query_transition_case_drilldown(
            run_data["prepared_parquet_path"],
            from_activity=context["from_activity"],
            to_activity=context["to_activity"],
            limit=max(0, int(drilldown_limit)),
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=context["variant_pattern"],
        )
    elif selected_activity_name:
        drilldown_title = f"アクティビティドリルダウン: {selected_activity_name}"
        drilldown_rows = query_activity_case_drilldown(
            run_data["prepared_parquet_path"],
            activity=selected_activity_name,
            limit=max(0, int(drilldown_limit)),
            filter_params=filter_params,
            filter_column_settings=run_data.get("column_settings"),
            variant_pattern=context["variant_pattern"],
        )

    if not drilldown_rows:
        return

    drilldown_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["drilldown"])
    drilldown_rows, drilldown_headers = localize_report_rows(
        drilldown_rows,
        ["case_id", "activity", "next_activity", "duration_text", "from_time", "to_time"],
    )
    append_table_to_worksheet(
        drilldown_sheet,
        drilldown_title,
        drilldown_rows,
        drilldown_headers,
        description="選択中アクティビティ / 遷移に該当するケースの明細です。",
    )

def _append_case_trace_export_sheet(workbook, run_data, case_id=""):
    normalized_case_id = str(case_id or "").strip()
    if not normalized_case_id:
        return

    case_trace = query_case_trace_details(
        run_data["prepared_parquet_path"],
        normalized_case_id,
    )
    if not case_trace.get("found"):
        return

    case_trace_sheet = _create_report_sheet(workbook, REPORT_SHEET_NAMES["case_trace"])
    next_row = append_key_value_rows(
        case_trace_sheet,
        "ケース概要",
        [
            (REPORT_HEADER_LABELS["case_id"], case_trace["case_id"]),
            (REPORT_HEADER_LABELS["event_count"], case_trace["summary"]["event_count"]),
            (REPORT_HEADER_LABELS["total_duration"], case_trace["summary"]["total_duration_text"]),
            (REPORT_HEADER_LABELS["start_time"], case_trace["summary"]["start_time"]),
            (REPORT_HEADER_LABELS["end_time"], case_trace["summary"]["end_time"]),
        ],
        description="指定したケースの概要情報です。",
    )
    case_trace_event_rows, case_trace_event_headers = localize_report_rows(
        case_trace["events"],
        ["case_id", "activity", "next_activity", "start_time", "end_time", "duration_text"],
    )
    append_table_to_worksheet(
        case_trace_sheet,
        "通過イベント",
        case_trace_event_rows,
        case_trace_event_headers,
        start_row=next_row,
        description="ケース内で通過したイベントを時系列順に並べています。",
    )

def _append_detail_export_analysis_sheets(workbook, context, run_data, filter_params, pattern_display_limit="10"):
    export_sheet_keys = context["export_sheet_keys"]
    if "frequency" in export_sheet_keys:
        _append_frequency_export_sheet(workbook, context, run_data, filter_params)
    if "transition" in export_sheet_keys:
        _append_transition_export_sheet(workbook, context)
    if "pattern" in export_sheet_keys:
        _append_pattern_export_sheets(
            workbook,
            context,
            run_data,
            filter_params,
            pattern_display_limit=pattern_display_limit,
        )
    if "bottleneck" in export_sheet_keys:
        _append_bottleneck_export_sheet(workbook, context)
    if "impact" in export_sheet_keys:
        _append_impact_export_sheet(workbook, context)

