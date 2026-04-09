from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="EDF2F7")
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="FBFDFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D6DEE8"),
    right=Side(style="thin", color="D6DEE8"),
    top=Side(style="thin", color="D6DEE8"),
    bottom=Side(style="thin", color="D6DEE8"),
)


def _autosize_worksheet_columns(worksheet, min_width=10, max_width=40):
    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        measured_width = min_width
        for row_index in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value in (None, ""):
                continue
            measured_width = max(measured_width, len(str(cell_value)))
        worksheet.column_dimensions[column_letter].width = max(
            min_width,
            min(max_width, measured_width + 2),
        )


def _style_export_worksheet(worksheet):
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90

    if worksheet.max_row >= 1:
        for column_index in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=column_index)
            header_cell.font = Font(bold=True, color="1F2937")
            header_cell.fill = HEADER_FILL
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cell.border = THIN_BORDER
        worksheet.row_dimensions[1].height = 22
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

    for row_index in range(2, worksheet.max_row + 1):
        row_fill = ALT_ROW_FILL if row_index % 2 == 0 else None
        for column_index in range(1, worksheet.max_column + 1):
            body_cell = worksheet.cell(row=row_index, column=column_index)
            body_cell.alignment = Alignment(vertical="top", wrap_text=True)
            body_cell.border = THIN_BORDER
            if row_fill is not None:
                body_cell.fill = row_fill

    _autosize_worksheet_columns(worksheet)


def format_analysis_result(df, display_columns=None, group_columns=None):
    if display_columns is None:
        return df.copy()

    valid_group_cols = [col for col in (group_columns or []) if col in df.columns]
    ordered_columns = valid_group_cols + [
        column_name
        for column_name in display_columns.keys()
        if column_name in df.columns and column_name not in valid_group_cols
    ]
    formatted_df = df.loc[:, ordered_columns].copy() if ordered_columns else df.copy()
    rename_map = {k: v for k, v in display_columns.items() if k not in valid_group_cols}
    return formatted_df.rename(columns=rename_map)


GROUP_SECTION_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")
GROUP_SECTION_FONT = Font(bold=True, color="1F2937", size=11)
GROUP_SUMMARY_META_KEY = "__meta__"
SUMMARY_SHEET_COLUMNS = [
    "グルーピング軸",
    "値",
    "ケース数",
    "ケース比率(%)",
    "イベント数",
    "イベント比率(%)",
    "平均処理時間(分)",
    "中央値処理時間(分)",
    "最大処理時間(分)",
    "合計処理時間(分)",
]


def insert_group_section_rows(worksheet, df, group_columns):
    """
    グループ値が変わる箇所にセクション区切り行（背景色付き）を挿入する。
    group_columns[0] の値が切り替わる行の上に区切り行を入れる。
    """
    if not group_columns:
        return

    first_group_col = group_columns[0]
    if first_group_col not in df.columns:
        return

    # グループ値が変わる行インデックスを収集（データフレームの行番号、0始まり）
    values = df[first_group_col].astype(str).tolist()
    # header行(row=1)の分を+2してExcel行番号に変換
    change_indices = []  # (df_index, group_value) — グループが変わる最初の行
    prev = None
    for i, val in enumerate(values):
        if prev is not None and val != prev:
            change_indices.append((i, val))  # i は 0始まり
        prev = val

    # 下から挿入（行番号がずれないように逆順処理）
    for df_idx, group_value in reversed(change_indices):
        # Excel行番号 = df_idx + 2 (header=1行目なので+2)
        excel_row = df_idx + 2
        worksheet.insert_rows(excel_row)
        section_cell = worksheet.cell(row=excel_row, column=1)
        section_cell.value = f"▸ {group_value}"
        section_cell.font = GROUP_SECTION_FONT
        section_cell.fill = GROUP_SECTION_FILL
        if worksheet.max_column > 1:
            worksheet.merge_cells(
                start_row=excel_row,
                start_column=1,
                end_row=excel_row,
                end_column=worksheet.max_column,
            )


def _format_group_axis_label(column_name, level, group_count):
    if group_count <= 1:
        return column_name
    suffix = "①②③"[level - 1] if level <= 3 else str(level)
    return f"{column_name}（グループ{suffix}）"


def _duration_cell_value(value):
    return value if value is not None else "—"


def build_summary_sheet_df(group_summary, group_columns):
    """サマリーシート用DataFrameを構築する。"""
    meta = group_summary.get(GROUP_SUMMARY_META_KEY, {})
    rows = []
    for level, col in enumerate(group_columns, start=1):
        col_data = group_summary.get(col, {})
        for value, stats in col_data.items():
            row = {
                "グルーピング軸": _format_group_axis_label(col, level, len(group_columns)),
                "値": value,
                "ケース数": stats.get("case_count", ""),
                "ケース比率(%)": stats.get("case_ratio_pct", ""),
                "イベント数": stats.get("event_count", ""),
                "イベント比率(%)": stats.get("event_ratio_pct", ""),
                "平均処理時間(分)": _duration_cell_value(stats.get("avg_duration_min")),
                "中央値処理時間(分)": _duration_cell_value(stats.get("median_duration_min")),
                "最大処理時間(分)": _duration_cell_value(stats.get("max_duration_min")),
                "合計処理時間(分)": _duration_cell_value(stats.get("total_duration_min")),
            }
            rows.append(row)

    group_rows_df = pd.DataFrame(rows, columns=SUMMARY_SHEET_COLUMNS)
    if not group_rows_df.empty:
        group_rows_df = group_rows_df.sort_values(
            ["ケース数", "グルーピング軸", "値"],
            ascending=[False, True, True],
        ).reset_index(drop=True)

    total_row_df = pd.DataFrame(
        [
            {
                "グルーピング軸": "",
                "値": "全体",
                "ケース数": meta.get("total_case_count", ""),
                "ケース比率(%)": 100.0 if meta else "",
                "イベント数": meta.get("total_event_count", ""),
                "イベント比率(%)": 100.0 if meta else "",
                "平均処理時間(分)": _duration_cell_value(meta.get("avg_duration_min")) if meta else "—",
                "中央値処理時間(分)": _duration_cell_value(meta.get("median_duration_min")) if meta else "—",
                "最大処理時間(分)": _duration_cell_value(meta.get("max_duration_min")) if meta else "—",
                "合計処理時間(分)": _duration_cell_value(meta.get("total_duration_min")) if meta else "—",
            }
        ],
        columns=SUMMARY_SHEET_COLUMNS,
    )
    if group_rows_df.empty:
        return total_row_df if meta else pd.DataFrame(columns=SUMMARY_SHEET_COLUMNS)

    return pd.concat([total_row_df, group_rows_df], ignore_index=True)


def _build_summary_sheet_df(group_summary, group_columns):
    return build_summary_sheet_df(group_summary, group_columns)


def build_excel_bytes(
    df,
    sheet_name,
    group_columns=None,
    group_summary=None,
):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # サマリーシート（グルーピングモード時のみ）
        if group_columns and group_summary:
            summary_df = build_summary_sheet_df(group_summary, group_columns)
            summary_df.to_excel(writer, sheet_name="サマリー", index=False)
            _style_export_worksheet(writer.sheets["サマリー"])

        # メインデータシート
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        if group_columns:
            insert_group_section_rows(ws, df, group_columns)
        _style_export_worksheet(ws)

    return buffer.getvalue()


def export_dataframe_to_excel(
    df,
    output_file,
    sheet_name,
    group_columns=None,
    group_summary=None,
):
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        output_path.write_bytes(
            build_excel_bytes(df, sheet_name, group_columns=group_columns, group_summary=group_summary)
        )
    except PermissionError as exc:
        raise PermissionError(
            f"{output_path} に書き込めません。Excel で開いている場合は閉じてから再実行してください。"
        ) from exc

    return output_path


def export_analysis_to_excel(
    df,
    output_root_dir,
    analysis_name,
    output_file_name,
    sheet_name,
    display_columns=None,
    group_columns=None,
    group_summary=None,
):
    output_file = Path(output_root_dir) / analysis_name / output_file_name
    excel_df = format_analysis_result(df, display_columns, group_columns=group_columns)
    return export_dataframe_to_excel(
        excel_df,
        output_file,
        sheet_name,
        group_columns=group_columns,
        group_summary=group_summary,
    )


def convert_analysis_result_to_records(
    df,
    display_columns=None,
    group_columns=None,
):
    api_df = format_analysis_result(df, display_columns, group_columns=group_columns)
    return api_df.to_dict(orient="records")
